from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User, Group
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.utils.timezone import localtime, now  # Add this import
from django.db import transaction
from django.db.models import Q, Count, Avg, Sum, F
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils.dateparse import parse_date
from django.template.loader import render_to_string
from django.contrib import messages
from datetime import datetime, timedelta, date
from decimal import Decimal
import json
import logging
import pytz
import sys
import traceback
import csv
import openpyxl

from .models import (UserSession, SystemError, Attendance, Support, FailedLoginAttempt,
                     PasswordChange, RoleAssignmentAudit, FeatureUsage, SystemUsage,
                     Timesheet, GlobalUpdate, UserDetails, ProjectUpdate, Presence,
                     PresenceStatus, Project, ClientProfile, ShiftMaster, ShiftAssignment,
                     Appraisal, AppraisalItem, AppraisalWorkflow)
from .helpers import is_user_in_group

# Set up logging
logger = logging.getLogger(__name__)

# Asia/Kolkata timezone
IST_TIMEZONE = pytz.timezone('Asia/Kolkata')

def get_current_time_ist():
    """Return current time in Asia/Kolkata timezone (aware)."""
    current_time = timezone.now().astimezone(IST_TIMEZONE)
    print(f"Current IST time: {current_time}")
    return current_time

def to_ist(dt):
    """Convert a datetime to Asia/Kolkata timezone (aware)."""
    if dt is None:
        print(f"Converting None to IST: None returned")
        return None
    if timezone.is_naive(dt):
        ist_time = IST_TIMEZONE.localize(dt)
    else:
        ist_time = dt.astimezone(IST_TIMEZONE)
    print(f"Converting {dt} to IST: {ist_time}")
    return ist_time

def to_utc(dt_ist):
    """Convert IST datetime to UTC for database storage."""
    if dt_ist is None:
        print(f"Converting None to UTC: None returned")
        return None
    if timezone.is_naive(dt_ist):
        dt_ist = IST_TIMEZONE.localize(dt_ist)
    utc_time = dt_ist.astimezone(pytz.UTC)
    print(f"Converting {dt_ist} to UTC: {utc_time}")
    return utc_time


# ==================== SESSION MANAGEMENT VIEWS ====================

@login_required
@require_POST
def session_heartbeat(request):
    """
    Handle session heartbeat with comprehensive tracking
    All times handled in Asia/Kolkata timezone
    """
    try:
        data = json.loads(request.body)

        # Extract session information
        tab_id = data.get('tabId') or request.headers.get('X-Tab-ID')
        parent_session_id = data.get('parentSessionId') or request.headers.get('X-Parent-Session-ID')
        device_fingerprint = data.get('deviceFingerprint') or request.headers.get('X-Device-Fingerprint')

        if not tab_id:
            return JsonResponse({'error': 'Tab ID required'}, status=400)

        # Get or create session
        user_session = _get_or_create_session(request.user, tab_id, parent_session_id, request, data)

        # Check for auto-logout
        if user_session.should_auto_logout():
            logger.info(f"Auto-logout triggered for user {request.user.username}")
            user_session.end_session()
            logout(request)
            return JsonResponse({
                'status': 'expired',
                'message': 'Your session has expired due to inactivity.',
                'redirect': '/login/?reason=timeout'
            })

        # Update session activity
        current_time_ist = get_current_time_ist()
        activity_data = {
            'is_idle': data.get('isIdle', False),
            'is_focused': data.get('isFocused', True),
            'is_online': data.get('isOnline', True),
            'url': data.get('url', request.get_full_path()),
            'performance_data': data.get('performanceMetrics', {}),
            'battery_level': data.get('batteryInfo', {}).get('level'),
            'connection_type': data.get('connectionInfo', {}).get('effectiveType'),
            'activity_count': data.get('activityCount', 0),
            'current_time_ist': current_time_ist
        }

        user_session.update_tab_activity(activity_data)

        # Check security anomalies
        warnings = []
        if device_fingerprint:
            anomalies = user_session.check_security_anomalies(device_fingerprint)
            if anomalies:
                warnings.extend([{
                    'type': 'security_anomaly',
                    'details': anomaly
                } for anomaly in anomalies])

        # Check for inactivity warning
        show_warning = user_session.should_show_inactivity_warning()
        if show_warning:
            user_session.record_inactivity_warning()
            warnings.append({
                'type': 'inactivity_warning',
                'message': 'Your session will expire soon due to inactivity'
            })

        # Get session summary
        summary = user_session.get_session_summary()

        return JsonResponse({
            'status': 'active',
            'session_id': user_session.parent_session_id,
            'tab_id': user_session.tab_id,
            'is_primary_tab': user_session.is_primary_tab,
            'idle': user_session.is_idle(),
            'productivity_score': user_session.productivity_score,
            'engagement_score': user_session.engagement_score,
            'warnings': warnings,
            'summary': summary,
            'server_time': current_time_ist.isoformat()
        })

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        logger.error(f"Error in session heartbeat: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)

@login_required
@require_POST
def log_activity(request):
    """
    Log user activities from client-side buffer
    All times handled in Asia/Kolkata timezone
    """
    try:
        data = json.loads(request.body)

        tab_id = data.get('tabId') or request.headers.get('X-Tab-ID')
        parent_session_id = data.get('parentSessionId') or request.headers.get('X-Parent-Session-ID')
        activities = data.get('activities', [])

        if not tab_id:
            return JsonResponse({'error': 'Tab ID required'}, status=400)

        # Get session
        user_session = UserSession.objects.filter(
            user=request.user,
            tab_id=tab_id,
            is_active=True
        ).first()

        if not user_session:
            # Try to find by parent session
            user_session = UserSession.objects.filter(
                user=request.user,
                parent_session_id=parent_session_id,
                is_active=True
            ).first()

        if not user_session:
            return JsonResponse({'error': 'Session not found'}, status=404)

        # Check if session expired
        if user_session.should_auto_logout():
            user_session.end_session()
            logout(request)
            return JsonResponse({
                'status': 'expired',
                'message': 'Your session has expired.',
                'redirect': '/login/'
            })

        # Process activities
        _process_activities(user_session, activities)

        # Update performance metrics if provided
        if data.get('performanceMetrics'):
            if not user_session.performance_metrics:
                user_session.performance_metrics = {}
            user_session.performance_metrics.update(data['performanceMetrics'])
            user_session.save(update_fields=['performance_metrics'])

        return JsonResponse({
            'status': 'success',
            'processed_activities': len(activities),
            'session_active': True
        })

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        logger.error(f"Error logging activity: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)

@login_required
@require_POST
def end_session(request):
    """
    End user session properly
    All times handled in Asia/Kolkata timezone
    """
    try:
        data = json.loads(request.body) if request.body else {}

        tab_id = data.get('tabId') or request.headers.get('X-Tab-ID')
        parent_session_id = data.get('parentSessionId') or request.headers.get('X-Parent-Session-ID')

        current_time_ist = get_current_time_ist()
        current_time_utc = to_utc(current_time_ist)

        # End specific tab session
        if tab_id:
            user_session = UserSession.objects.filter(
                user=request.user,
                tab_id=tab_id,
                is_active=True
            ).first()

            if user_session:
                user_session.end_session(logout_time=current_time_utc)
                logger.info(f"Ended tab session {tab_id} for user {request.user.username}")

        # End all sessions for this parent session
        if parent_session_id:
            sessions = UserSession.objects.filter(
                user=request.user,
                parent_session_id=parent_session_id,
                is_active=True
            )

            for session in sessions:
                session.end_session(logout_time=current_time_utc)

            logger.info(f"Ended parent session {parent_session_id} for user {request.user.username}")

        # If no specific session specified, end all active sessions
        if not tab_id and not parent_session_id:
            sessions = UserSession.objects.filter(
                user=request.user,
                is_active=True
            )

            for session in sessions:
                session.end_session(logout_time=current_time_utc)

            logger.info(f"Ended all sessions for user {request.user.username}")

        return JsonResponse({'status': 'success', 'message': 'Session ended successfully'})

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        logger.error(f"Error ending session: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)

@login_required
@require_GET
def session_status(request):
    """
    Get current session status and analytics
    All times returned in Asia/Kolkata timezone
    """
    try:
        tab_id = request.GET.get('tab_id') or request.headers.get('X-Tab-ID')

        # Get user's active sessions
        sessions = UserSession.objects.filter(
            user=request.user,
            is_active=True
        ).order_by('-login_time')

        if tab_id:
            # Get specific tab session
            current_session = sessions.filter(tab_id=tab_id).first()
        else:
            # Get primary session
            current_session = sessions.filter(is_primary_tab=True).first() or sessions.first()

        if not current_session:
            return JsonResponse({'status': 'no_active_session'})

        # Calculate session analytics
        total_tabs = sessions.count()
        current_time_ist = get_current_time_ist()
        login_time_ist = to_ist(current_session.login_time)
        session_duration = (current_time_ist - login_time_ist).total_seconds() / 60

        # Get multi-tab analytics
        analytics = UserSession.get_multi_tab_analytics(user=request.user, days=1)

        return JsonResponse({
            'status': 'active',
            'current_session': current_session.get_session_summary(),
            'total_active_tabs': total_tabs,
            'session_duration_minutes': session_duration,
            'analytics': analytics,
            'server_time': current_time_ist.isoformat()
        })

    except Exception as e:
        logger.error(f"Error getting session status: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)

@login_required
@require_GET
def session_analytics(request):
    """
    Get comprehensive session analytics
    All times handled in Asia/Kolkata timezone
    """
    try:
        days = int(request.GET.get('days', 7))
        days = min(days, 90)  # Limit to 90 days

        current_time_ist = get_current_time_ist()
        start_date_ist = current_time_ist - timedelta(days=days)
        start_date_utc = to_utc(start_date_ist)

        # Get user sessions
        sessions = UserSession.objects.filter(
            user=request.user,
            login_time__gte=start_date_utc
        )

        # Basic statistics
        stats = {
            'total_sessions': sessions.count(),
            'avg_session_duration': sessions.aggregate(
                avg_duration=Avg('session_duration')
            )['avg_duration'] or 0,
            'total_working_hours': sessions.aggregate(
                total_hours=Sum('working_hours')
            )['total_hours'] or timedelta(0),
            'total_idle_time': sessions.aggregate(
                total_idle=Sum('idle_time')
            )['total_idle'] or timedelta(0),
            'avg_productivity_score': sessions.filter(
                productivity_score__isnull=False
            ).aggregate(
                avg_score=Avg('productivity_score')
            )['avg_score'] or 0
        }

        # Multi-tab analytics
        multi_tab_stats = UserSession.get_multi_tab_analytics(user=request.user, days=days)

        # Device type breakdown
        device_stats = sessions.values('device_type').annotate(
            count=Count('id'),
            avg_duration=Avg('session_duration')
        ).order_by('-count')

        # Location breakdown
        location_stats = sessions.values('location').annotate(
            count=Count('id'),
            total_hours=Sum('working_hours')
        ).order_by('-count')

        # Daily activity pattern
        daily_activity = sessions.extra(
            select={'day': 'date(login_time)'}
        ).values('day').annotate(
            session_count=Count('id'),
            total_duration=Sum('session_duration'),
            avg_productivity=Avg('productivity_score')
        ).order_by('day')

        return JsonResponse({
            'period_days': days,
            'basic_stats': stats,
            'multi_tab_stats': multi_tab_stats,
            'device_breakdown': list(device_stats),
            'location_breakdown': list(location_stats),
            'daily_activity': list(daily_activity),
            'generated_at': current_time_ist.isoformat()
        })

    except Exception as e:
        logger.error(f"Error getting session analytics: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)

@login_required
@csrf_exempt
def update_last_activity(request):
    """
    View to handle activity updates from the client.
    Updates the user's last activity timestamp and tracks idle time.
    All times are handled and stored in Asia/Kolkata (IST) timezone.
    """
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Parse request data
                try:
                    data = json.loads(request.body)
                except ValueError:
                    data = {}

                # Get client IP
                x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
                ip_address = x_forwarded_for.split(',')[0] if x_forwarded_for else request.META.get('REMOTE_ADDR')

                # Get current time in IST
                current_time_ist = get_current_time_ist()
                current_time_utc = to_utc(current_time_ist)

                # Get or create user session using model logic
                user_session = UserSession.objects.filter(
                    user=request.user,
                    is_active=True
                ).select_for_update().first()

                if not user_session:
                    # Use model method to create session
                    user_session = UserSession.get_or_create_session(
                        user=request.user,
                        session_key=request.session.session_key,
                        ip_address=ip_address,
                        user_agent=request.META.get('HTTP_USER_AGENT', None)
                    )
                    return JsonResponse({
                        'status': 'success',
                        'message': 'New session created',
                        'session_id': user_session.id,
                        'server_time': current_time_ist.isoformat()
                    })

                # Ensure last_activity is in IST for comparison
                last_activity_ist = to_ist(user_session.last_activity)

                # Check for session timeout (5 minutes)
                if (current_time_ist - last_activity_ist) > timedelta(minutes=5):
                    user_session.end_session(current_time_utc, is_idle=True)
                    new_session = UserSession.get_or_create_session(
                        user=request.user,
                        session_key=request.session.session_key,
                        ip_address=ip_address,
                        user_agent=request.META.get('HTTP_USER_AGENT', None)
                    )
                    return JsonResponse({
                        'status': 'success',
                        'sessionExpired': True,
                        'message': 'Previous session timed out, new session created',
                        'session_id': new_session.id,
                        'server_time': current_time_ist.isoformat()
                    })

                # Update IP and location if changed
                if user_session.ip_address != ip_address:
                    user_session.ip_address = ip_address
                    user_session.location = user_session.determine_location()
                    user_session.save(update_fields=['ip_address', 'location'])

                # Handle client-reported idle state
                is_idle = data.get('isIdle', False)
                is_focused = data.get('isFocused', True)

                # Update activity with idle state
                user_session.update_activity(current_time_utc, is_idle=is_idle)

                user_session.refresh_from_db()

                # Always return times in IST
                last_activity_ist = to_ist(user_session.last_activity)
                working_hours = user_session.working_hours
                if working_hours is not None:
                    working_hours = str(working_hours)
                else:
                    working_hours = None

                return JsonResponse({
                    'status': 'success',
                    'last_activity': last_activity_ist.isoformat() if last_activity_ist else None,
                    'idle_time': str(user_session.idle_time),
                    'working_hours': working_hours,
                    'location': user_session.location,
                    'server_time': current_time_ist.isoformat()
                })

        except Exception as e:
            logger.error(f"Error updating last activity: {str(e)}")
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)

    return JsonResponse({
        'status': 'error',
        'message': 'Invalid request method'
    }, status=405)

@login_required
def get_session_status(request):
    """Get the current session status. All times are returned in Asia/Kolkata (IST) timezone."""
    try:
        user_session = UserSession.objects.filter(
            user=request.user,
            is_active=True
        ).first()

        if not user_session:
            return JsonResponse({
                'status': 'error',
                'message': 'No active session found'
            }, status=404)

        current_time_ist = get_current_time_ist()
        login_time_ist = to_ist(user_session.login_time)
        last_activity_ist = to_ist(user_session.last_activity)

        return JsonResponse({
            'status': 'success',
            'session_id': user_session.id,
            'login_time': login_time_ist.isoformat() if login_time_ist else None,
            'last_activity': last_activity_ist.isoformat() if last_activity_ist else None,
            'idle_time': str(user_session.idle_time),
            'location': user_session.location,
            'session_duration': user_session.get_session_duration_display(),
            'server_time': current_time_ist.isoformat()
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=400)

# ==================== HELPER FUNCTIONS ====================

def _get_or_create_session(user, tab_id, parent_session_id, request, data):
    """
    Get existing session or create new one with enhanced tracking
    All times handled in Asia/Kolkata timezone
    """
    with transaction.atomic():
        current_time_ist = get_current_time_ist()
        current_time_utc = to_utc(current_time_ist)

        # Look for existing session with this tab_id
        user_session = UserSession.objects.filter(
            user=user,
            tab_id=tab_id,
            is_active=True
        ).select_for_update().first()

        if user_session:
            # Update existing session
            user_session.last_activity = current_time_utc
            user_session.save(update_fields=['last_activity'])
            return user_session

        # Generate parent session ID if not provided
        if not parent_session_id:
            parent_session_id = f"{user.id}_{current_time_ist.strftime('%Y%m%d_%H%M%S')}"

        # Check if this is the first tab in the session
        existing_tabs_count = UserSession.objects.filter(
            user=user,
            parent_session_id=parent_session_id,
            is_active=True
        ).count()

        is_primary_tab = existing_tabs_count == 0

        # Extract client info
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        ip_address = x_forwarded_for.split(',')[0] if x_forwarded_for else request.META.get('REMOTE_ADDR')

        # Detect device type
        user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
        if any(mobile in user_agent for mobile in ['mobile', 'android', 'iphone']):
            device_type = 'mobile'
        elif 'ipad' in user_agent or 'tablet' in user_agent:
            device_type = 'tablet'
        else:
            device_type = 'desktop'

        # Create session
        user_session = UserSession.objects.create(
            user=user,
            session_key=request.session.session_key or UserSession.generate_session_key(),
            ip_address=ip_address,
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            login_time=current_time_utc,
            last_activity=current_time_utc,
            tab_id=tab_id,
            tab_opened_time=current_time_utc,
            tab_last_focus=current_time_utc,
            is_primary_tab=is_primary_tab,
            parent_session_id=parent_session_id,
            session_fingerprint=data.get('deviceFingerprint'),
            device_type=device_type,
            screen_resolution=data.get('sessionFingerprint', {}).get('screen'),
            timezone_offset=data.get('sessionFingerprint', {}).get('timezoneOffset'),
            language=data.get('sessionFingerprint', {}).get('language'),
            tab_url=data.get('url', request.get_full_path()),
            is_active=True
        )

        # Set location
        user_session.location = user_session.determine_location()
        user_session.save(update_fields=['location'])

        logger.info(f"Created new session for user {user.username}, tab_id: {tab_id}, primary: {is_primary_tab}")

        return user_session

def _process_activities(user_session, activities):
    """
    Process activity buffer from client
    All times handled in Asia/Kolkata timezone
    """
    if not activities:
        return

    current_time_ist = get_current_time_ist()
    current_time_utc = to_utc(current_time_ist)

    # Update last activity
    user_session.last_activity = current_time_utc

    # Process each activity
    for activity in activities:
        activity_type = activity.get('type')
        activity_timestamp = activity.get('timestamp')

        # Convert timestamp to IST if provided
        if activity_timestamp:
            try:
                activity_time_ist = datetime.fromtimestamp(activity_timestamp / 1000, tz=IST_TIMEZONE)
            except (ValueError, TypeError):
                activity_time_ist = current_time_ist
        else:
            activity_time_ist = current_time_ist

        if activity_type == 'click':
            if not user_session.click_events:
                user_session.click_events = []
            user_session.click_events.append({
                'timestamp': activity_time_ist.isoformat(),
                'target': activity.get('target'),
                'coordinates': activity.get('coordinates')
            })

        elif activity_type == 'scroll':
            if not user_session.scroll_events:
                user_session.scroll_events = []
            user_session.scroll_events.append({
                'timestamp': activity_time_ist.isoformat(),
                'scroll_position': activity.get('scrollPosition')
            })

        elif activity_type in ['keydown', 'keyup', 'keypress']:
            if not user_session.keyboard_events:
                user_session.keyboard_events = []
            user_session.keyboard_events.append({
                'timestamp': activity_time_ist.isoformat(),
                'key': activity.get('key')
            })

        elif activity_type == 'mousemove':
            user_session.mouse_movements += 1

        elif activity_type == 'page_change':
            user_session.tab_url = activity.get('url', user_session.tab_url)
            if not user_session.page_views:
                user_session.page_views = []
            user_session.page_views.append({
                'url': activity.get('url'),
                'timestamp': activity_time_ist.isoformat(),
                'referrer': activity.get('referrer', '')
            })

        elif activity_type in ['tab_focus', 'tab_blur']:
            if not user_session.tab_visibility_log:
                user_session.tab_visibility_log = []
            user_session.tab_visibility_log.append({
                'action': activity_type,
                'timestamp': activity_time_ist.isoformat(),
                'url': activity.get('url')
            })

            if activity_type == 'tab_focus':
                user_session.tab_switches += 1
                user_session.tab_last_focus = to_utc(activity_time_ist)

        elif activity_type == 'error':
            if not user_session.error_events:
                user_session.error_events = []
            user_session.error_events.append({
                'timestamp': activity_time_ist.isoformat(),
                'type': activity.get('error_type'),
                'message': activity.get('message'),
                'filename': activity.get('filename'),
                'lineno': activity.get('lineno')
            })

    # Trim arrays to prevent excessive growth
    for field in ['click_events', 'scroll_events', 'keyboard_events', 'page_views', 'tab_visibility_log', 'error_events']:
        events = getattr(user_session, field) or []
        if len(events) > 1000:  # Keep last 1000 events
            setattr(user_session, field, events[-1000:])

    # Update productivity and engagement scores
    user_session.calculate_productivity_score()
    user_session.calculate_engagement_score()

    user_session.save()

# ==================== BASIC VIEWS ====================

def home_view(request):
    return redirect('login')

def login_view(request):
    """Login view with session tracking"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)

            # Create initial session record
            current_time_ist = get_current_time_ist()
            current_time_utc = to_utc(current_time_ist)

            # Debug logging
            print(f"DEBUG: Current UTC time: {timezone.now()}")
            print(f"DEBUG: Current IST time: {current_time_ist}")

            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            ip_address = x_forwarded_for.split(',')[0] if x_forwarded_for else request.META.get('REMOTE_ADDR')

            try:
                session = UserSession.get_or_create_session(
                    user=user,
                    session_key=request.session.session_key,
                    ip_address=ip_address,
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )
                print(f"DEBUG: Session created: {session}")
                print(f"DEBUG: Session login time (UTC): {session.login_time}")
                print(f"DEBUG: Session login time (IST): {to_ist(session.login_time)}")

            except Exception as e:
                logger.error(f"Error creating session: {e}")
                print(f"ERROR: Session creation failed: {e}")

            # Process attendance if applicable
            try:
                attendance_status = process_login_attendance(user)
                if attendance_status:
                    print(f"DEBUG: Login processed - User: {user.username}, Status: {attendance_status['status']}, Clock in: {attendance_status.get('clock_in')}")
            except Exception as e:
                logger.error(f"Error processing attendance: {e}")

            messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password.')

    return render(request, 'login.html')


# # Login View
# def login_view(request):
#     if request.method == "POST":
#         username = request.POST.get('username')
#         password = request.POST.get('password')

#         try:
#             # Authenticate the user
#             user = authenticate(request, username=username, password=password)

#             if user is not None:
#                 login(request, user)

#                 # Check if the user is authenticated and create a session for them
#                 if request.user.is_authenticated:
#                     session = UserSession.get_or_create_session(
#                         user=request.user,
#                         session_key=request.session.session_key,
#                         ip_address=get_client_ip(request)  # Assuming you have a utility to get the user's IP
#                     )
#                 return redirect('dashboard')
#             else:
#                 # Show error if authentication fails
#                 error_message = 'Invalid username or password'
#                 return render(request, 'error.html', {'error': error_message})

#         except Exception as e:
#             # Handle any unexpected errors
#             error_message = f'An error occurred: {str(e)}'
#             return render(request, 'error.html', {'error': error_message})

#     return render(request, 'login.html')



# Logout View
from django.http import JsonResponse

def logout_view(request):
    if request.user.is_authenticated:
        try:
            # End any active sessions for this user
            active_session = UserSession.objects.filter(
                user=request.user,
                logout_time__isnull=True
            ).first()

            if active_session:
                active_session.end_session()  # Assume end_session() properly sets logout_time

            # Logout the user
            logout(request)

            return redirect('login')
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'An error occurred: {str(e)}'}, status=500)

    return redirect('login')

from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.models import User
from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages

def set_password_view(request, username):
    if request.method == "POST":
        try:
            # Get form data
            current_password = request.POST.get('current_pwd')
            new_password = request.POST.get('new_pwd')
            confirm_password = request.POST.get('confirm_pwd')

            # Validate required fields
            if not all([current_password, new_password, confirm_password]):
                messages.error(request, 'All password fields are required.')
                user = User.objects.get(username=username)
                return redirect('user_profile', user_id=user.id)

            # Check if passwords match
            if new_password != confirm_password:
                messages.error(request, 'New passwords do not match.')
                user = User.objects.get(username=username)
                return redirect('user_profile', user_id=user.id)

            # Get the user
            user = User.objects.get(username=username)

            # Verify current password
            if not user.check_password(current_password):
                messages.error(request, 'Current password is incorrect.')
                return redirect('user_profile', user_id=user.id)

            # Additional password validation
            if len(new_password) < 8:
                messages.error(request, 'Password must be at least 8 characters long.')
                return redirect('user_profile', user_id=user.id)

            # Set new password
            user.set_password(new_password)
            user.save()

            # IMPORTANT: Update session hash to keep user logged in
            update_session_auth_hash(request, user)

            # Add success message
            messages.success(request, 'Password updated successfully!')

            # Redirect back to profile
            return redirect('user_profile', user_id=user.id)

        except User.DoesNotExist:
            messages.error(request, 'User does not exist.')
            return redirect('dashboard')
        except Exception as e:
            messages.error(request, f'An error occurred: {str(e)}')
            return redirect('dashboard')

    # If GET request, redirect to profile
    try:
        user = User.objects.get(username=username)
        return redirect('user_profile', user_id=user.id)
    except User.DoesNotExist:
        return redirect('dashboard')


@login_required
def user_profile(request, user_id):
    """View to display user profile accessible to all logged-in users"""
    user_detail = get_object_or_404(UserDetails, user__id=user_id)
    return render(request, 'basic/user_profile.html', {
        'user_detail': user_detail,
        'role': user_detail.user.groups.first().name if user_detail.user.groups.exists() else 'User',
        'username': user_detail.user.username
    })

'''---------------------------------   DASHBOARD VIEW ----------------------------------'''
from django.shortcuts import render

from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Break
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required


def check_active_break(request):
    """
    Check if the authenticated user has an active break and remaining break count.
    """
    print(f"Checking active break for user: {request.user.username}")
    active_break = Break.objects.filter(user=request.user, end_time__isnull=True).first()

    # Get today's date
    today = timezone.now().date()

    # Get break counts for today
    break_counts = {}
    for break_type, _ in Break.BREAK_TYPES:
        total_allowed = Break.BREAK_LIMITS.get(break_type, 1)
        used_count = Break.objects.filter(
            user=request.user,
            break_type=break_type,
            start_time__date=today
        ).count()
        break_counts[break_type] = {
            'used': used_count,
            'remaining': total_allowed - used_count
        }

    if active_break and active_break.is_active:
        return JsonResponse({
            'status': 'success',
            'break_id': active_break.id,
            'break_type': active_break.break_type,
            'start_time': active_break.start_time,
            'is_active': active_break.is_active,
            'break_counts': break_counts
        })
    else:
        return JsonResponse({
            'status': 'error',
            'message': 'No active break found',
            'break_counts': break_counts
        })

def take_break(request):
    if request.method == 'POST':
        break_type = request.POST.get('break_type')

        # Ensure valid break type
        if not break_type or break_type not in dict(Break.BREAK_TYPES).keys():
            messages.error(request, "Invalid break type.")
            return redirect('dashboard')

        # Create new break
        new_break = Break(user=request.user, break_type=break_type)
        try:
            # Run validation to check for active breaks and limits
            new_break.clean()  # This will run all validations from the `clean` method
            new_break.save()
            messages.success(request, f"Started {break_type}")
        except ValidationError as e:
            messages.error(request, str(e))

        return redirect('dashboard')

from django.urls import reverse

@login_required
def end_break(request, break_id):
    if request.method == 'POST':
        try:
            print(f"Ending break with ID: {break_id} for user: {request.user.username}")
            active_break = get_object_or_404(Break, id=break_id, user=request.user, end_time__isnull=True)

            max_duration = Break.BREAK_DURATIONS.get(active_break.break_type, timedelta(minutes=15))
            # Use timezone.now() for consistent timezone-aware datetime
            elapsed_time = timezone.now() - active_break.start_time

            if elapsed_time > max_duration and not request.POST.get('reason'):
                return JsonResponse({
                    'status': 'error',
                    'message': 'Please provide a reason for the extended break.'
                })

            reason = request.POST.get('reason', '')
            # Use timezone.now() when setting end_time
            active_break.end_time = timezone.now()
            active_break.reason_for_extension = reason
            active_break.save()
            print(f"Break ended with reason: {reason}")

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success'})

            return redirect(f"{reverse('dashboard')}?success=Break ended successfully")

        except Exception as e:
            return redirect(f"{reverse('dashboard')}?error={str(e)}")

    return redirect(f"{reverse('dashboard')}?error=Invalid request method")


from django.utils.timezone import now  # Use timezone-aware datetime
from django.shortcuts import render
from django.db.models import Count

def get_attendance_stats(request):
    """Get attendance statistics for the current user."""
    user = request.user
    try:
        # Get current month and today's date
        current_time = now()  # Use timezone-aware datetime
        current_month = current_time.month
        current_year = current_time.year
        today_date = current_time.strftime('%B %d, %Y')  # Example: January 23, 2025
        print(f"Today's Date: {today_date}")

        # Get attendance records for the current month
        current_month = timezone.now().month
        current_year = timezone.now().year

        attendance_records = Attendance.objects.filter(
            user=request.user,
            date__month=current_month,
            date__year=current_year
        )
        print(f"Attendance Records for Current Month: {attendance_records.count()}")

        # Count present days (including WFH)
        present_days = attendance_records.filter(
            status__in=["Present", "Work From Home"]
        ).count()
        print(f"Total Present Days (including WFH): {present_days}")

        # Get leave requests for the current month
        leave_requests = Leave.objects.filter(
            user=user,
            start_date__month=current_month,
            start_date__year=current_year
        )
        print(f"Total Leave Requests (Current Month): {leave_requests.count()}")

        # Count pending and approved leaves
        leave_request_count = leave_requests.filter(status="Pending").count()
        approved_leave_count = leave_requests.filter(status="Approved").count()
        print(f"Pending Leave Requests: {leave_request_count}")
        print(f"Approved Leave Requests: {approved_leave_count}")

        # Format response data
        result = {
            'today_date': today_date,
            'current_month_name': current_time.strftime('%B'),  # Full month name, e.g., "January"
            'attendance': {
                'total_present': present_days,
                'leave_request_count': leave_request_count,
                'approved_leave_count': approved_leave_count,
            }
        }
        print(f"Formatted Attendance Data: {result}")
        return result

    except Exception as e:
        # Log the error
        # print(f"Error generating attendance stats: {str(e)}")

        # Return empty data on error
        return {
            'today_date': '',
            'current_month_name': '',
            'attendance': {
                'total_present': 0,
                'leave_request_count': 0,
                'approved_leave_count': 0,
            }
        }


from django.utils import timezone

@login_required
def dashboard_view(request):
    # Get today's date
    time = timezone.now()
    print(f"time: {time}")
    today = timezone.now().date()

    user = request.user

    # Get user's current session status
    user_session = UserSession.objects.filter(
        user=user,
        session_key=request.session.session_key,
        logout_time__isnull=True
    ).last()

    # Calculate user status
    def format_timedelta(td):
        """Helper function to format timedelta into a readable string"""
        total_seconds = int(td.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60

        if hours > 0:
            return f"{hours}h {minutes}m"
        return f"{minutes}m"

    # Calculate user status
    user_status = {
        'status': 'offline',
        'color': 'gray',
        'idle_time': None,
        'working_hours': None,
        'formatted_idle_time': None
    }

    if user_session:
        current_time = timezone.now()
        time_since_last_activity = current_time - user_session.last_activity

        # If last activity was less than 1 minute ago - user is active
        if time_since_last_activity < timedelta(minutes=1):
            user_status['status'] = 'active'
            user_status['color'] = 'green'
        # If last activity was between 1-5 minutes ago - user is idle
        elif time_since_last_activity < timedelta(minutes=5):
            user_status['status'] = 'idle'
            user_status['color'] = 'yellow'
        else:
            user_status['status'] = 'inactive'
            user_status['color'] = 'red'

        user_status['idle_time'] = user_session.idle_time
        user_status['working_hours'] = user_session.working_hours
        user_status['last_activity'] = user_session.last_activity if user_session.last_activity else None

        # Format idle time as a readable string
        if user_status['idle_time']:
            user_status['formatted_idle_time'] = format_timedelta(user_status['idle_time'])

    # Check if the user has the HR role
    is_hr = user.groups.filter(name='HR').exists()
    is_manager = user.groups.filter(name='Manager').exists()

    # Variables for attendance stats and active projects
    present_employees = absent_employees = active_projects = None

    # Get today's date using timezone-aware datetime
    today = timezone.now().date()

    # Get date range from request (default to today if not provided)
    start_date_str = request.GET.get('start_date', today)
    end_date_str = request.GET.get('end_date', today)

    # Convert string date inputs to date format
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if isinstance(start_date_str, str) else today
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if isinstance(end_date_str, str) else today
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'Invalid date format. Use YYYY-MM-DD.'})

    # Ensure the end date is inclusive
    end_date += timedelta(days=1)

    active_break = Break.objects.filter(user=user, end_time__isnull=True).first()
    break_data = None

    # Get today's date for break counts
    today = timezone.now().date()

    # Get break counts for today
    break_counts = {}
    for break_type, _ in Break.BREAK_TYPES:
        total_allowed = Break.DAILY_BREAK_LIMITS.get(break_type, 1)
        used_count = Break.objects.filter(
            user=user,
            break_type=break_type,
            start_time__date=today
        ).count()
        break_counts[break_type] = {
            'used': used_count,
            'remaining': total_allowed - used_count
        }

    # Always include break counts whether break is active or not
    break_data = {
        'break_counts': break_counts
    }

    if active_break and active_break.is_active:
        # Get break duration in minutes
        break_duration = Break.BREAK_DURATIONS.get(active_break.break_type, timedelta(minutes=15))

        # Ensure both times are timezone-aware
        now = timezone.now()
        elapsed_time = now - active_break.start_time
        remaining_time = max(timedelta(0), break_duration - elapsed_time)

        # Add active break info to break_data
        break_data.update({
            'break_id': active_break.id,
            'break_type': active_break.break_type,
            'start_time': active_break.start_time,
            'active_break': active_break.is_active,
            'remaining_minutes': int(remaining_time.total_seconds() / 60),
            'remaining_seconds': int(remaining_time.total_seconds() % 60)
        })


    if is_hr:
        # Get today's present employees
        present_employees = Attendance.objects.filter(
            status='Present',
            date=today
        ).count()

        # Get today's absent employees
        absent_employees = Attendance.objects.filter(
            status='Absent',
            date=today
        ).count()

        # Get active projects
        active_projects = Project.objects.filter(status='Active').count()

    # Retrieve assignments and projects for non-HR users
    assignments = ProjectAssignment.objects.filter(user=user)
    projects = [assignment.project for assignment in assignments]

    # Get project timelines for each project
    project_timelines = []
    for project in projects:
        timeline = project_timeline(request, project.id)  # Returns a dictionary with project info
        project_timelines.append(timeline['project'])

    # Retrieve global updates
    updates = GlobalUpdate.objects.all().order_by('-created_at')

    # Retrieve project team updates
    project_team_updates = ProjectUpdate.objects.all()

    # Check if we are editing an update
    update = None
    if 'update_id' in request.GET:
        update = GlobalUpdate.objects.filter(id=request.GET['update_id']).first()

    """Manager's Dashboard view."""

    # Get the manager's assigned users' breaks
    project_assignments = ProjectAssignment.objects.filter(
        project__projectassignment__user=request.user,
        project__projectassignment__role_in_project='Manager',
        is_active=True
    ).values_list('user', flat=True).distinct()

    if project_assignments.exists():
        active_breaks = Break.objects.filter(user__in=project_assignments, end_time__isnull=True).count()
    else:
        active_breaks = 0

    # Initialize present employees count
    present_employees_count = None

    if is_manager:
        # Get today's date
        today = timezone.now().date()

        # Get the manager's assigned projects
        project_assignments = ProjectAssignment.objects.filter(
            project__projectassignment__user=request.user,
            project__projectassignment__role_in_project='Manager',
            is_active=True
        )

        # Get the users assigned to the manager's projects
        users_in_projects = project_assignments.values_list('user', flat=True)

        # Count present employees
        present_employees_count = Attendance.objects.filter(
            user__in=users_in_projects,
            date=today,
            status='Present'
        ).count()
    # Context for the dashboard view

    """View to render the attendance dashboard card with current month's statistics"""

    # Get today's date and first day of current month
    today = timezone.now().date()
    first_day = today.replace(day=1)

    # Get today's attendance record
    today_attendance = Attendance.objects.filter(
        user=request.user,
        date=today
    ).first()

    # Get monthly statistics
    monthly_stats = {
        'present_count': Attendance.objects.filter(
            user=request.user,
            date__gte=first_day,
            date__lte=today,
            status='Present'
        ).count(),

        'late_count': Attendance.objects.filter(
            user=request.user,
            date__gte=first_day,
            date__lte=today,
            status__in=['Late', 'Present & Late']
        ).count(),

        'leave_count': Attendance.objects.filter(
            user=request.user,
            date__gte=first_day,
            date__lte=today,
            status__in=['On Leave', 'Half Day']
        ).count(),
    }
        # Attendance card data
    attendance_context = {
        'today_attendance': Attendance.objects.filter(
            user=request.user,
            date=today
        ).select_related('shift').first(),

        'monthly_stats': {
            'present_count': Attendance.objects.filter(
                user=request.user,
                date__gte=first_day,
                date__lte=today,
                status='Present'
            ).count(),

            'late_count': Attendance.objects.filter(
                user=request.user,
                date__gte=first_day,
                date__lte=today,
                status__in=['Late', 'Present & Late']
            ).count(),

            'leave_count': Attendance.objects.filter(
                user=request.user,
                date__gte=first_day,
                date__lte=today,
                status__in=['On Leave', 'Half Day']
            ).count(),
        }
    }# Attendance card data
    attendance_context = {
        'today_attendance': Attendance.objects.filter(
            user=request.user,
            date=today
        ).select_related('shift').first(),

        'monthly_stats': {
            'present_count': Attendance.objects.filter(
                user=request.user,
                date__gte=first_day,
                date__lte=today,
                status='Present'
            ).count(),

            'late_count': Attendance.objects.filter(
                user=request.user,
                date__gte=first_day,
                date__lte=today,
                status__in=['Late', 'Present & Late']
            ).count(),

            'leave_count': Attendance.objects.filter(
                user=request.user,
                date__gte=first_day,
                date__lte=today,
                status__in=['On Leave', 'Half Day']
            ).count(),
        }

    }

    context = {
        'active_breaks': active_breaks,
        'attendance': get_attendance_stats(request),  # Direct assignment
        'projects': projects,
        'project_timelines': project_timelines,
        'updates': updates,
        'is_hr': is_hr,
        'is_manager': is_manager,
        'projectTeamUpdates': project_team_updates,
        'update': update,
        'present_employees': present_employees,
        'absent_employees': absent_employees,
        'active_projects': active_projects,
        'start_date': start_date,
        'end_date': end_date - timedelta(days=1),
        'show_employee_directory': is_hr,
        'break_data': break_data,
        'break_types': dict(Break.BREAK_TYPES),
        'break_durations': {k: int(v.total_seconds() / 60) for k, v in Break.BREAK_DURATIONS.items()},
        'user_status': user_status,
        'present_employees_count': present_employees_count,
        'time': time,
        'today_attendance': today_attendance,
        'monthly_stats': monthly_stats,
        'current_month': today.strftime('%B %Y'),
        'attendance_data': attendance_context,

    }

    return render(request, 'dashboard.html', context)


def project_timeline(request, project_id):
    project = Project.objects.get(id=project_id)
    current_date = timezone.now().date()

    # Calculate total project duration and remaining time
    total_duration = project.deadline - project.start_date
    remaining_duration = project.deadline - current_date

    # Check if remaining time is within the last 25% of the total duration
    is_deadline_close = remaining_duration <= total_duration * 0.25


    return {
        'project': {
            'name': project.name,
            'start_date': project.start_date,
            'deadline': project.deadline,  # No need to include 'deadline' twice
            'is_deadline_close': is_deadline_close,
        }
    }



from django.utils.timezone import now, timezone

def is_hr(user):
    return user.is_authenticated and user.groups.filter(name='HR').exists()

@user_passes_test(is_hr)
@login_required
def employee_directory(request):
    # Check if the user has the HR role
    if not request.user.groups.filter(name='HR').exists():
        return JsonResponse({'error': 'Permission denied'}, status=403)

    # Fetch all employee details
    employees = UserDetails.objects.all().values(
        'id', 'user__username', 'user__first_name', 'user__last_name', 'contact_number_primary', 'personal_email'
    )

    # Convert queryset to list of dictionaries
    employee_data = list(employees)

    # Return the data as JSON
    return JsonResponse({'employees': employee_data})

# Create global update view
@login_required
@transaction.atomic
def hr_create_update(request):
    if not request.user.groups.filter(name='HR').exists():
        messages.error(request, "You do not have permission to manage global updates.")
        return redirect('dashboard')

    if request.method == 'POST':
        title = request.POST.get('title')
        description = request.POST.get('description')
        status = request.POST.get('status', 'upcoming')
        scheduled_date_str = request.POST.get('scheduled_date')

        if not title or not description:
            messages.error(request, "Title and description are required.")
            return redirect('dashboard')

        try:
            scheduled_date = None
            if scheduled_date_str:
                scheduled_date = datetime.strptime(scheduled_date_str, '%Y-%m-%dT%H:%M')
                scheduled_date = timezone.make_aware(scheduled_date)  # Make timezone-aware

            new_update = GlobalUpdate.objects.create(
                title=title,
                description=description,
                status=status,
                scheduled_date=scheduled_date,
                managed_by=request.user,
            )

            messages.success(request, "Global update created successfully.")
            return redirect('dashboard')
        except Exception as e:
            messages.error(request, "Error creating update. Please try again.")
            return redirect('dashboard')

    return redirect('dashboard')

# Get update data API for editing
@login_required
def get_update_data(request, update_id):
    """API endpoint to fetch update data for editing"""
    if not request.user.groups.filter(name='HR').exists():
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        update = get_object_or_404(GlobalUpdate, id=update_id)
        data = {
            'title': update.title,
            'description': update.description,
            'status': update.status,
            'scheduled_date': update.scheduled_date.isoformat() if update.scheduled_date else '',
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

# Edit global update view
@login_required
@transaction.atomic
def hr_edit_update(request, update_id):
    """View to handle update editing"""
    update = get_object_or_404(GlobalUpdate, id=update_id)

    # Check permissions
    if not request.user.groups.filter(name='HR').exists():
        messages.error(request, "You do not have permission to edit this update.")
        return JsonResponse({'error': 'Permission denied'}, status=403)

    if request.method == 'POST':
        try:
            # Get form data
            title = request.POST.get('title')
            description = request.POST.get('description')
            status = request.POST.get('status')
            scheduled_date_str = request.POST.get('scheduled_date')

            # Validate required fields
            if not title or not description:
                return JsonResponse({'error': 'Title and description are required'}, status=400)

            # Update fields
            update.title = title
            update.description = description
            update.status = status

            if scheduled_date_str:
                try:
                    scheduled_date = datetime.strptime(scheduled_date_str, '%Y-%m-%dT%H:%M')
                    update.scheduled_date = timezone.make_aware(scheduled_date)
                except ValueError:
                    return JsonResponse({'error': 'Invalid date format'}, status=400)
            else:
                update.scheduled_date = None

            update.save()
            messages.success(request, "Global update edited successfully.")
            return redirect('dashboard')  # Redirect to the dashboard after successful deletion

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)

# Delete global update view
@login_required
@transaction.atomic
def hr_delete_update(request, update_id):
    """View to handle update deletion"""
    if not request.user.groups.filter(name='HR').exists():
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        update = get_object_or_404(GlobalUpdate, id=update_id)
        update.delete()
        messages.success(request, "Global update deleted successfully.")
        return redirect('dashboard')  # Redirect to the dashboard after successful deletion

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

''' ------------------------------------------------------- BREAK AREA AREA --------------------------------------------------------- '''



''' ------------------------------------------------------- MANAGER TEAM PROJECT AREA --------------------------------------------------------- '''
def is_manager(user):
    return user.is_authenticated and user.groups.filter(name='Manager').exists()

@login_required
@user_passes_test(is_manager)
@transaction.atomic
def manager_create_project_update(request):
    """Manager creates an update for their project."""
    if request.method == 'POST':
        project_id = request.POST.get('project_id')
        title = request.POST.get('title')
        description = request.POST.get('description')
        status = request.POST.get('status', 'upcoming')
        scheduled_date_str = request.POST.get('scheduled_date')

        # Validate fields
        if not title or not description or not project_id:
            messages.error(request, "Title, description, and project are required.")
            return redirect('dashboard')

        try:
            # Fetch the project assigned to the manager
            project = get_object_or_404(Project, id=project_id)
            project_assignment = ProjectAssignment.objects.filter(project=project, user=request.user, is_active=True).first()
            if not project_assignment or project_assignment.role_in_project != 'Manager':
                messages.error(request, "You are not assigned as the manager for this project.")
                return redirect('dashboard')

            # Handle scheduled date
            scheduled_date = None
            if scheduled_date_str:
                scheduled_date = datetime.strptime(scheduled_date_str, '%Y-%m-%dT%H:%M')
                scheduled_date = timezone.make_aware(scheduled_date)

            # Create the project update
            new_update = ProjectUpdate.objects.create(
                project=project,
                created_by=request.user,
                title=title,
                description=description,
                status=status,
                scheduled_date=scheduled_date
            )

            messages.success(request, "Project update created successfully.")
            return redirect('dashboard')

        except Exception as e:
            messages.error(request, f"Error creating update: {str(e)}")
            return redirect('dashboard')

    return redirect('dashboard')

@login_required
@user_passes_test(is_manager)
@transaction.atomic
def manager_edit_project_update(request, update_id):
    """Manager edits an existing project update."""
    try:
        update = get_object_or_404(ProjectUpdate, id=update_id)
        if update.created_by != request.user:
            messages.error(request, "You do not have permission to edit this update.")
            return redirect('dashboard')

        if request.method == 'POST':
            title = request.POST.get('title')
            description = request.POST.get('description')
            status = request.POST.get('status', 'upcoming')
            scheduled_date_str = request.POST.get('scheduled_date')

            if not title or not description:
                return JsonResponse({'error': 'Title and description are required'}, status=400)

            scheduled_date = None
            if scheduled_date_str:
                try:
                    scheduled_date = datetime.strptime(scheduled_date_str, '%Y-%m-%dT%H:%M')
                    update.scheduled_date = timezone.make_aware(scheduled_date)
                except ValueError:
                    return JsonResponse({'error': 'Invalid date format'}, status=400)

            update.title = title
            update.description = description
            update.status = status
            update.save()

            messages.success(request, "Project update edited successfully.")
            return redirect('dashboard')  # Redirect to the dashboard after successful edit

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)

@login_required
@user_passes_test(is_manager)
@transaction.atomic
def manager_delete_project_update(request, update_id):
    """Manager deletes a project update."""
    try:
        update = get_object_or_404(ProjectUpdate, id=update_id)
        if update.created_by != request.user:
            messages.error(request, "You do not have permission to delete this update.")
            return redirect('dashboard')

        update.delete()
        messages.success(request, "Project update deleted successfully.")
        return redirect('dashboard')

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

''' --------------------------------------------------------- USER DETAILS AREA --------------------------------------------------------- '''

# aps/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group
from django.contrib import messages
from django.http import HttpResponseForbidden, HttpResponse, JsonResponse
from django.db import transaction
from django.db.models import Q, Count, Sum, Max, Avg
from django.core.paginator import Paginator
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.template.loader import render_to_string
from django.core.mail import EmailMessage, EmailMultiAlternatives
from django.utils import timezone
import logging
import re
import csv
import io
import random
import string
from datetime import date, datetime, timedelta
from .models import UserDetails, UserActionLog, UserSession, validate_pan, validate_aadhar

logger = logging.getLogger(__name__)

# Permission check functions
def is_hr(user):
    return user.groups.filter(name='HR').exists()

def is_manager(user):
    return user.groups.filter(name='Manager').exists()

def is_employee(user):
    return user.groups.filter(name='Employee').exists()


# Helper function to generate employee ID
def generate_employee_id(work_location=None, group_id=None):
    """
    Generate employee ID based on work location with role-based reserved ranges

    Reserved ranges for Management and Finance groups (IDs 7 and 8):
    - 1-15: First priority range
    - 301-400: Second priority range

    Regular employees use 101-300 and 401+ ranges
    """
    from django.db.models import Q
    from datetime import datetime
    import re

    # Check if user belongs to Finance or Management groups based on group_id
    is_reserved_role = False
    if group_id and group_id in ['7', '8']:  # Finance or Management
        is_reserved_role = True

    # Determine prefix based on location
    if work_location and work_location.lower() == 'betul':
        prefix = "ATS"
        separator = ""
        year_suffix = ""
    elif work_location and work_location.lower() == 'pune':
        prefix = "AT"
        separator = ""
        year_suffix = ""
    else:
        prefix = "EMP"
        current_year = str(datetime.now().year)[2:]
        year_suffix = current_year
        separator = "-"

    # Function to extract numeric ID from username
    def extract_id(username):
        # Extract numbers at the end of the string
        match = re.search(r'(\d+)$', username)
        if match:
            try:
                return int(match.group(1))
            except ValueError:
                return None
        return None

    # Set ID ranges based on role
    if is_reserved_role:
        # Check if there's a gap in priority range 1-15
        used_ids = []

        # Query all users with the prefix
        all_users = User.objects.filter(username__startswith=prefix)

        # Find all used IDs in the priority range
        for user in all_users:
            user_id = extract_id(user.username)
            if user_id and 1 <= user_id <= 15:
                used_ids.append(user_id)

        # Look for the first available ID in priority range
        for i in range(1, 16):
            if i not in used_ids:
                seq_num = i
                break
        else:
            # Priority range is full, check reserved range 301-400
            used_ids = []
            for user in all_users:
                user_id = extract_id(user.username)
                if user_id and 301 <= user_id <= 400:
                    used_ids.append(user_id)

            # Look for the first available ID in reserved range
            for i in range(301, 401):
                if i not in used_ids:
                    seq_num = i
                    break
            else:
                # Both ranges are full, generate fallback ID
                timestamp = int(datetime.now().timestamp())
                return f"{prefix}{separator}{timestamp}"
    else:
        # Regular employees use 101-300 and 401+
        all_users = User.objects.filter(username__startswith=prefix)
        highest_id = 100  # Start from 101

        # Find the highest used ID outside reserved ranges
        for user in all_users:
            user_id = extract_id(user.username)
            if user_id and user_id > highest_id and user_id not in range(1, 16) and user_id not in range(301, 401):
                highest_id = user_id

        # Start from highest + 1
        seq_num = highest_id + 1

        # Skip reserved ranges
        if 1 <= seq_num <= 15:
            seq_num = 101
        elif 301 <= seq_num <= 400:
            seq_num = 401

    # Format the sequence number and build the ID
    formatted_seq = f"{seq_num:04d}"
    if year_suffix:
        employee_id = f"{prefix}{year_suffix}{separator}{formatted_seq}"
    else:
        employee_id = f"{prefix}{formatted_seq}"

    # Final validation to ensure ID doesn't already exist
    if User.objects.filter(username=employee_id).exists():
        # If this ID is taken, recurse with a timestamp-based ID
        timestamp = int(datetime.now().timestamp())
        if year_suffix:
            return f"{prefix}{year_suffix}{separator}{timestamp}"
        else:
            return f"{prefix}{separator}{timestamp}"

    return employee_id

# Helper function to send welcome email
def send_welcome_email(user, password):
    """Send welcome email with login credentials"""
    from django.core.mail import EmailMessage, EmailMultiAlternatives
    from django.template.loader import render_to_string

    subject = "Welcome to Ardur Company Portal"

    # Plain text email body
    email_body = f"""
    Hello {user.first_name} {user.last_name},

    Welcome to Our Company! Your account has been created successfully.

    Here are your login details:
    Username: {user.username}
    Password: {password}

    Please log in at: https://home.ardurtechnology.com/login/

    For security reasons, we recommend changing your password after first login.

    Regards,
    HR Department
    """

    # Try to render HTML template, fall back to plain text if it fails
    try:
        html_message = render_to_string('components/hr/emails/welcome_email.html', {
            'user': user,
            'password': password,
            'login_url': 'https://home.ardurtechnology.com/login/'
        })

        # Send email with both HTML and plain text
        email = EmailMultiAlternatives(
            subject=subject,
            body=email_body,
            to=[user.email]
        )
        email.attach_alternative(html_message, "text/html")
    except Exception as e:
        logger.error(f"Error rendering HTML template: {str(e)}")
        # Fall back to plain text email
        email = EmailMessage(
            subject=subject,
            body=email_body,
            to=[user.email]
        )

    # Add logging before sending
    logger.info(f"Attempting to send welcome email to {user.email}")

    # Send the email
    email.send()
    logger.info(f"Welcome email sent successfully to {user.email}")

    return True

# HR Dashboard with enhanced features
@login_required
@user_passes_test(is_hr)
def hr_dashboard(request):
    """HR Dashboard with improved filtering, query optimization, and pagination"""
    # Get filter parameters with defaults
    filters = {
        'search': request.GET.get('search', ''),
        'department': request.GET.get('department', ''),
        'status': request.GET.get('status', ''),
        'work_location': request.GET.get('work_location', ''),
        'role': request.GET.get('role', ''),
        'employee_type': request.GET.get('employee_type', '')
    }

    # Start with optimized base queryset
    users = User.objects.select_related('profile').prefetch_related(
        'groups',
        Prefetch(
            'usersession_set',
            queryset=UserSession.objects.filter(
                login_time__date=datetime.now().date()
            ).order_by('-login_time'),
            to_attr='today_sessions'
        )
    )

    # Apply search filter
    if filters['search']:
        users = users.filter(
            Q(first_name__icontains=filters['search']) |
            Q(last_name__icontains=filters['search']) |
            Q(username__icontains=filters['search']) |
            Q(email__icontains=filters['search']) |
            Q(profile__job_description__icontains=filters['search']) |
            Q(profile__current_city__icontains=filters['search']) |
            Q(profile__current_state__icontains=filters['search'])
        ).distinct()

    # Apply other filters
    if filters['status']:
        users = users.filter(profile__employment_status=filters['status'])
    if filters['work_location']:
        users = users.filter(profile__work_location=filters['work_location'])
    if filters['role']:
        users = users.filter(groups__name=filters['role'])
    if filters['employee_type']:
        users = users.filter(profile__employee_type=filters['employee_type'])

    # Create missing profiles
    for user in users:
        if not hasattr(user, 'profile'):
            UserDetails.objects.create(user=user)

    # Get dashboard statistics using aggregation
    stats = {
        'total_users': users.count(),
        'active_users': users.filter(is_active=True).count(),
        'inactive_users': users.filter(is_active=False).count(),
        'status_counts': UserDetails.objects.values('employment_status').annotate(
            count=Count('id'),
            active_count=Count('user', filter=Q(user__is_active=True))
        ),
        'location_counts': UserDetails.objects.exclude(
            Q(work_location__isnull=True) | Q(work_location='')
        ).values('work_location').annotate(
            count=Count('id'),
            active_count=Count('user', filter=Q(user__is_active=True))
        ).order_by('-count'),
        'employee_type_counts': UserDetails.objects.values('employee_type').annotate(
            count=Count('id'),
            active_count=Count('user', filter=Q(user__is_active=True))
        ),
        'recent_users': User.objects.select_related('profile').filter(
            date_joined__gte=datetime.now().date() - timedelta(days=7)
        ).order_by('-date_joined')[:5]
    }

    # Handle pagination with remembered page size
    page_size = request.session.get('hr_dashboard_page_size', 20)
    if 'page_size' in request.GET:
        page_size = int(request.GET['page_size'])
        request.session['hr_dashboard_page_size'] = page_size

    paginator = Paginator(users, page_size)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Get filter options from model
    filter_options = {
        'employment_status_choices': UserDetails.EMPLOYMENT_STATUS_CHOICES,
        'employee_type_choices': UserDetails.EMPLOYEE_TYPE_CHOICES,
        'work_locations': UserDetails.objects.exclude(
            Q(work_location__isnull=True) | Q(work_location='')
        ).values_list('work_location', flat=True).distinct(),
        'roles': Group.objects.all()
    }

    context = {
        'page_obj': page_obj,
        'page_range': paginator.get_elided_page_range(
            page_obj.number,
            on_each_side=2,
            on_ends=1
        ),
        'page_size': page_size,
        'filters': filters,
        'filter_options': filter_options,
        'stats': stats,
        'role': 'HR'
    }

    return render(request, 'components/hr/hr_dashboard.html', context)

@login_required
@user_passes_test(is_hr)
def hr_user_detail(request, user_id):
    """Enhanced view for HR to view and edit comprehensive user details"""
    import re
    from datetime import date, datetime

    # Log the requested user_id to help diagnose issues
    logger.info(f"Accessing user detail for user_id: {user_id}")

    user = get_object_or_404(User, id=user_id)
    user_detail, created = UserDetails.objects.get_or_create(user=user)

    # Log the retrieved user information
    logger.info(f"Retrieved user: {user.username} ({user.first_name} {user.last_name}), user_id: {user.id}")

    # Get user action logs
    action_logs = UserActionLog.objects.filter(user=user).order_by('-timestamp')[:10]

    # Get recent sessions
    recent_sessions = UserSession.objects.filter(user=user).order_by('-login_time')[:5]

    if request.method == 'POST':
        try:
            logger.info(f"Processing POST request for user_id: {user_id}")

            # Extract and clean data
            data = request.POST.copy()

            # Check if we're only updating employment status
            is_status_only_update = False
            new_status = data.get('employment_status')

            # If we have a status change and it's not to 'active', we'll allow minimal validation
            if new_status and new_status != 'active' and user_detail.employment_status != new_status:
                is_status_only_update = True
                logger.info(f"Status-only update detected for user {user_id}: changing to {new_status}")

            # Regular validation logic for complete updates or when status is set to 'active'
            if not is_status_only_update:
                # Validate date of birth
                dob = data.get('dob')
                if dob:
                    dob_date = datetime.strptime(dob, '%Y-%m-%d').date()
                    today = date.today()
                    age = today.year - dob_date.year - ((today.month, today.day) < (dob_date.month, dob_date.day))
                    if age < 18:
                        raise ValueError('Employee must be at least 18 years old.')

                # Validate contact numbers
                def validate_contact(number, field_name):
                    if not number:
                        return None

                    # Remove any non-digit characters for validation
                    cleaned_number = ''.join(c for c in number if c.isdigit())

                    # Check if the number is valid (allowing for international format)
                    if len(cleaned_number) < 10 or len(cleaned_number) > 15:
                        raise ValueError(f'{field_name} must be between 10 and 15 digits.')

                    return number

                # Primary contact validation
                primary_number = data.get('contact_number_primary', '').strip()
                primary_contact = validate_contact(primary_number, 'Primary contact number')

                # Emergency contact validation
                emergency_number = data.get('emergency_contact_number', '').strip()
                emergency_contact = validate_contact(emergency_number, 'Emergency contact number')

                # Secondary emergency contact validation
                secondary_emergency_number = data.get('secondary_emergency_contact_number', '').strip()
                secondary_emergency_contact = validate_contact(secondary_emergency_number, 'Secondary emergency contact number')

                # Validate PAN using model validator
                pan = data.get('pan_number')
                if pan:
                    validate_pan(pan)

                # Validate Aadhar using model validator
                aadhar = data.get('aadhar_number', '').replace(' ', '')
                if aadhar:
                    validate_aadhar(aadhar)

                # Validate email
                email = data.get('personal_email')
                if email and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
                    raise ValueError('Invalid email format')

                company_email = data.get('company_email')
                if company_email and not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', company_email):
                    raise ValueError('Invalid company email format')

                # When setting status to 'active', require essential fields
                if new_status == 'active':
                    required_fields = {
                        'contact_number_primary': 'Primary contact number',
                        'emergency_contact_name': 'Emergency contact name',
                        'emergency_contact_number': 'Emergency contact number',
                        'employee_type': 'Employee type',
                        'hire_date': 'Hire date'
                    }

                    for field, label in required_fields.items():
                        if not data.get(field):
                            raise ValueError(f"{label} is required when employee status is Active")

            # Dictionary of fields to update based on UserDetails model fields
            fields_to_update = {
                # Only include these fields if we're not doing a status-only update
                # or if status is being set to 'active'
                'dob': None if is_status_only_update else (data.get('dob') or None),
                'blood_group': None if is_status_only_update else (data.get('blood_group') or None),
                'gender': None if is_status_only_update else (data.get('gender') or None),
                'marital_status': None if is_status_only_update else (data.get('marital_status') or None),

                'contact_number_primary': None if is_status_only_update else primary_contact,
                'personal_email': None if is_status_only_update else (email or None),
                'company_email': None if is_status_only_update else (company_email or None),

                'current_address_line1': None if is_status_only_update else (data.get('current_address_line1') or None),
                'current_address_line2': None if is_status_only_update else (data.get('current_address_line2') or None),
                'current_city': None if is_status_only_update else (data.get('current_city') or None),
                'current_state': None if is_status_only_update else (data.get('current_state') or None),
                'current_postal_code': None if is_status_only_update else (data.get('current_postal_code') or None),
                'current_country': None if is_status_only_update else (data.get('current_country') or None),

                'permanent_address_line1': None if is_status_only_update else (data.get('permanent_address_line1') or None),
                'permanent_address_line2': None if is_status_only_update else (data.get('permanent_address_line2') or None),
                'permanent_city': None if is_status_only_update else (data.get('permanent_city') or None),
                'permanent_state': None if is_status_only_update else (data.get('permanent_state') or None),
                'permanent_postal_code': None if is_status_only_update else (data.get('permanent_postal_code') or None),
                'permanent_country': None if is_status_only_update else (data.get('permanent_country') or None),
                'is_current_same_as_permanent': None if is_status_only_update else (data.get('is_current_same_as_permanent') == 'on'),

                'emergency_contact_name': None if is_status_only_update else (data.get('emergency_contact_name') or None),
                'emergency_contact_number': None if is_status_only_update else emergency_contact,
                'emergency_contact_relationship': None if is_status_only_update else (data.get('emergency_contact_relationship') or None),

                'secondary_emergency_contact_name': None if is_status_only_update else (data.get('secondary_emergency_contact_name') or None),
                'secondary_emergency_contact_number': None if is_status_only_update else secondary_emergency_contact,
                'secondary_emergency_contact_relationship': None if is_status_only_update else (data.get('secondary_emergency_contact_relationship') or None),

                # Employment Status is always included, as this is what we're changing
                'employment_status': new_status,

                # Only include these fields if we're not doing a status-only update
                'employee_type': None if is_status_only_update else (data.get('employee_type') or None),
                'reporting_manager_id': None if is_status_only_update else (data.get('reporting_manager') or None),
                'hire_date': None if is_status_only_update else (data.get('hire_date') or None),
                'start_date': None if is_status_only_update else (data.get('start_date') or None),
                'probation_end_date': None if is_status_only_update else (data.get('probation_end_date') or None),
                'notice_period_days': None if is_status_only_update else (data.get('notice_period_days') or 30),
                'job_description': None if is_status_only_update else (data.get('job_description') or None),
                'work_location': None if is_status_only_update else (data.get('work_location') or None),

                # For non-active statuses, we might want to capture exit date and reason
                'exit_date': None if new_status == 'active' else (data.get('exit_date') or (date.today() if new_status in ['terminated', 'resigned', 'absconding'] else None)),
                'exit_reason': None if new_status == 'active' else (data.get('exit_reason') or None),
                'rehire_eligibility': None if is_status_only_update else (data.get('rehire_eligibility') == 'on'),

                'salary_currency': None if is_status_only_update else (data.get('salary_currency') or 'INR'),
                'base_salary': None if is_status_only_update else (data.get('base_salary') or None),
                'salary_frequency': None if is_status_only_update else (data.get('salary_frequency') or 'monthly'),

                'pan_number': None if is_status_only_update else (pan or None),
                'aadhar_number': None if is_status_only_update else (aadhar or None),
                'passport_number': None if is_status_only_update else (data.get('passport_number') or None),
                'passport_expiry': None if is_status_only_update else (data.get('passport_expiry') or None),

                'bank_name': None if is_status_only_update else (data.get('bank_name') or None),
                'bank_account_number': None if is_status_only_update else (data.get('bank_account_number') or None),
                'bank_ifsc': None if is_status_only_update else (data.get('bank_ifsc') or None),

                'previous_company': None if is_status_only_update else (data.get('previous_company') or None),
                'previous_position': None if is_status_only_update else (data.get('previous_position') or None),
                'previous_experience_years': None if is_status_only_update else (data.get('previous_experience_years') or None),

                'skills': None if is_status_only_update else (data.get('skills') or None),

                # Only include confidential notes if user has permission and we're not doing a status-only update
                'confidential_notes': None if is_status_only_update else (
                    data.get('confidential_notes') if request.user.has_perm('view_confidential_notes') else user_detail.confidential_notes
                )
            }

            # If it's a status-only update, add a note about the status change
            if is_status_only_update and request.user.has_perm('view_confidential_notes'):
                status_note = f"Status changed to {dict(UserDetails.EMPLOYMENT_STATUS_CHOICES).get(new_status, new_status)} on {date.today()} by {request.user.get_full_name() or request.user.username}"

                existing_notes = user_detail.confidential_notes or ""
                if existing_notes:
                    fields_to_update['confidential_notes'] = f"{existing_notes}\n\n{status_note}"
                else:
                    fields_to_update['confidential_notes'] = status_note

            # Check if role/group is being updated (only for non-status-only updates)
            if not is_status_only_update:
                old_group = user.groups.first()
                new_group_id = data.get('group')
                if new_group_id:
                    new_group = Group.objects.get(id=new_group_id)

                    if not old_group or old_group.id != new_group.id:
                        user.groups.clear()
                        user.groups.add(new_group)

                        UserActionLog.objects.create(
                            user=user,
                            action_type='role_change',
                            action_by=request.user,
                            details=f"Role changed from {old_group.name if old_group else 'None'} to {new_group.name}"
                        )

            # Handle employment status changes based on model choices
            old_status = user_detail.employment_status

            if old_status != new_status and new_status:
                # Automatically deactivate user accounts for certain statuses
                if new_status in ['inactive', 'terminated', 'resigned', 'suspended', 'absconding']:
                    if user.is_active:
                        user.is_active = False
                        user.save()

                        UserActionLog.objects.create(
                            user=user,
                            action_type='deactivate',
                            action_by=request.user,
                            details=f"User account deactivated due to status change to {new_status}"
                        )
                # Reactivate accounts when changing to active
                elif new_status == 'active' and not user.is_active:
                    user.is_active = True
                    user.save()

                    UserActionLog.objects.create(
                        user=user,
                        action_type='activate',
                        action_by=request.user,
                        details="User account activated due to status change to active"
                    )

                # Log status change
                UserActionLog.objects.create(
                    user=user,
                    action_type='status_change',
                    action_by=request.user,
                    details=f"Employment status changed from {dict(UserDetails.EMPLOYMENT_STATUS_CHOICES).get(old_status, old_status) if old_status else 'None'} to {dict(UserDetails.EMPLOYMENT_STATUS_CHOICES).get(new_status, new_status)}"
                )

            # Remove None values (for fields that shouldn't be updated)
            fields_to_update = {k: v for k, v in fields_to_update.items() if v is not None}

            # Validate against model choices
            model_fields = UserDetails._meta.get_fields()
            for field_name, value in fields_to_update.items():
                field = next((f for f in model_fields if f.name == field_name), None)
                if hasattr(field, 'choices') and field.choices and value:
                    valid_choices = dict(field.choices)
                    if value not in valid_choices:
                        raise ValueError(f'Invalid value for {field_name}')

            # Update basic user details if provided and not doing a status-only update
            if not is_status_only_update:
                first_name = data.get('first_name')
                last_name = data.get('last_name')
                email = data.get('email')

                if first_name or last_name or email:
                    user.first_name = first_name or user.first_name
                    user.last_name = last_name or user.last_name
                    user.email = email or user.email
                    user.save()

                    UserActionLog.objects.create(
                        user=user,
                        action_type='update',
                        action_by=request.user,
                        details="Basic user information updated"
                    )

            # Perform atomic update
            with transaction.atomic():
                for field, value in fields_to_update.items():
                    setattr(user_detail, field, value)
                user_detail.save()

                if not is_status_only_update:
                    UserActionLog.objects.create(
                        user=user,
                        action_type='update',
                        action_by=request.user,
                        details="User details updated"
                    )

            messages.success(request, 'User details updated successfully.')
            return redirect('aps_hr:hr_user_detail', user_id=user.id)

        except ValueError as e:
            logger.warning(f"Validation Error for user {user_id}: {str(e)}")
            messages.error(request, str(e))
        except Exception as e:
            logger.error(f"Unexpected error for user {user_id}: {str(e)}", exc_info=True)
            messages.error(request, 'An unexpected error occurred while updating user details.')

    # Prepare context with user details and metadata
    context = {
        'user_obj': user,
        'user_detail': user_detail,
        'action_logs': action_logs,
        'recent_sessions': recent_sessions,
        'today': date.today(),
        'blood_group_choices': UserDetails.BLOOD_GROUP_CHOICES,
        'gender_choices': UserDetails.GENDER_CHOICES,
        'marital_status_choices': UserDetails.MARITAL_STATUS_CHOICES,
        'employment_status_choices': UserDetails.EMPLOYMENT_STATUS_CHOICES,
        'employee_type_choices': UserDetails.EMPLOYEE_TYPE_CHOICES,
        'groups': Group.objects.all(),
        'employment_duration': user_detail.employment_duration,
        'status_display': user_detail.status_display,
        'reporting_chain': user_detail.get_reporting_chain,
        'salary_frequencies': [('monthly', 'Monthly'), ('bi_weekly', 'Bi-Weekly'), ('weekly', 'Weekly')],
        'remaining_notice_period': user_detail.remaining_notice_period if user_detail.is_on_notice else None,
        'age': user_detail.age,
        'managers': User.objects.filter(groups__name='Manager').order_by('first_name', 'last_name')
    }

    return render(request, 'components/hr/hr_user_detail.html', context)

@login_required
@user_passes_test(is_hr)
def add_user(request):
    """Enhanced view to add a new user to the system with auto ID generation"""
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Extract basic user information
                email = request.POST.get('email', '').strip()
                first_name = request.POST.get('first_name')
                last_name = request.POST.get('last_name')
                group_id = request.POST.get('group')
                work_location = request.POST.get('work_location')

                # Validate required fields
                if not (first_name and last_name and group_id):
                    raise ValueError("First name, last name, and group are required fields")

                # Check if email already exists (only if provided)
                if email and User.objects.filter(email=email).exists():
                    raise ValueError(f"Email '{email}' already exists")

                # Generate employee ID for username
                employee_id = generate_employee_id(work_location, group_id)

                # Generate a random password
                password = ''.join(random.choices(string.ascii_letters + string.digits + string.punctuation, k=12))

                # Create new user
                user = User.objects.create_user(
                    username=employee_id,
                    email=email if email else None,
                    password=password,
                    first_name=first_name,
                    last_name=last_name
                )

                # Add user to group
                try:
                    group = Group.objects.get(id=int(group_id))
                    user.groups.add(group)
                except (ValueError, Group.DoesNotExist):
                    raise ValueError("Invalid group selected")

                # Create UserDetails with comprehensive information
                user_details = UserDetails.objects.create(
                    user=user,
                    hire_date=request.POST.get('hire_date') or datetime.now().date(),
                    start_date=request.POST.get('start_date') or datetime.now().date(),
                    employment_status='active',
                    work_location=work_location,
                    job_description=request.POST.get('job_description') or None,
                    dob=request.POST.get('dob') or None,
                    gender=request.POST.get('gender') or None,
                    blood_group=request.POST.get('blood_group') or None,
                    contact_number_primary=request.POST.get('contact_number_primary') or None,
                    personal_email=request.POST.get('personal_email') or None,
                    emergency_contact_name=request.POST.get('emergency_contact_name') or None,
                    emergency_contact_number=request.POST.get('emergency_contact_primary') or None,
                    emergency_contact_relationship=request.POST.get('emergency_contact_relationship') or None,
                    onboarded_by=request.user,
                    onboarding_date=timezone.now(),
                    employee_type=request.POST.get('employee_type') or None,
                    current_address_line1=request.POST.get('address_line1') or None,
                    current_address_line2=request.POST.get('address_line2') or None,
                    current_city=request.POST.get('city') or None,
                    current_state=request.POST.get('state') or None,
                    current_postal_code=request.POST.get('postal_code') or None,
                    current_country=request.POST.get('country') or None
                )

                # Log user creation
                UserActionLog.objects.create(
                    user=user,
                    action_type='create',
                    action_by=request.user,
                    details=f"User created with ID: {employee_id}, role: {group.name}"
                )

                # Send welcome email with better exception handling (only if email is provided)
                if email:
                    try:
                        send_welcome_email(user, password)
                        messages.success(request, f"Welcome email sent to {email} with login credentials.")
                    except ConnectionRefusedError:
                        logger.error("Email server connection refused. Check email server settings.", exc_info=True)
                        messages.warning(request, f"User created successfully, but welcome email could not be sent due to email server connection issues.")
                    except Exception as e:
                        logger.error(f"Error sending welcome email: {str(e)}", exc_info=True)
                        messages.warning(request, f"User created successfully, but welcome email could not be sent: {str(e)}")

                messages.success(request, f"User {employee_id} ({first_name} {last_name}) created successfully.")
                # Fix the URL name to include the namespace
                return redirect('aps_hr:hr_user_detail', user_id=user.id)

        except ValueError as e:
            messages.error(request, str(e))
        except Exception as e:
            logger.error(f"Error creating user: {str(e)}", exc_info=True)
            messages.error(request, f"Error creating user: {str(e)}")

    return render(request, 'components/hr/add_user.html', {
        'groups': Group.objects.all(),
        'today': date.today(),
        'blood_group_choices': UserDetails.BLOOD_GROUP_CHOICES,
        'gender_choices': UserDetails.GENDER_CHOICES,
        'employment_status_choices': UserDetails.EMPLOYMENT_STATUS_CHOICES,
        'employee_type_choices': UserDetails.EMPLOYEE_TYPE_CHOICES,
    })
@login_required
@user_passes_test(is_hr)
def bulk_add_users(request):
    """View to add multiple users to the system at once from a text file or pasted content"""

    if request.method == 'POST':
        user_data = request.POST.get('user_data', '').strip()
        selected_group_id = request.POST.get('group')
        work_location = request.POST.get('work_location')

        if not user_data:
            messages.error(request, "No user data provided. Please enter user data in the text area.")
            return redirect('aps_hr:bulk_add_users')

        if not selected_group_id:
            messages.error(request, "Please select a role/group for the users from the dropdown menu.")
            return redirect('aps_hr:bulk_add_users')

        try:
            group = Group.objects.get(id=int(selected_group_id))
        except (ValueError, Group.DoesNotExist):
            messages.error(request, "Invalid group selected. Please select a valid group from the dropdown.")
            return redirect('aps_hr:bulk_add_users')

        # Process the user data
        lines = user_data.strip().split('\n')
        success_count = 0
        skipped_count = 0
        error_messages = []

        with transaction.atomic():
            for line_number, line in enumerate(lines, 1):
                line = line.strip()
                if not line:
                    continue

                # Split by tab or space if tab not found
                parts = line.split('\t') if '\t' in line else line.split(maxsplit=1)

                if len(parts) < 2:
                    error_messages.append(f"Line {line_number}: Invalid format - '{line}'. Expected format: 'Username FirstName LastName'")
                    continue

                username = parts[0].strip()
                full_name = parts[1].strip()

                # Skip empty entries
                if not username or not full_name:
                    error_messages.append(f"Line {line_number}: Empty username or name - '{line}'")
                    skipped_count += 1
                    continue

                # Split full name into first and last name
                name_parts = full_name.split()
                if len(name_parts) > 1:
                    first_name = name_parts[0]
                    last_name = ' '.join(name_parts[1:])
                else:
                    first_name = full_name
                    last_name = ""

                # Check if user already exists
                if User.objects.filter(username=username).exists():
                    error_messages.append(f"Line {line_number}: Username '{username}' already exists in the system")
                    skipped_count += 1
                    continue

                try:
                    # Create user with standard password
                    user = User.objects.create_user(
                        username=username,
                        password="number@123",
                        first_name=first_name,
                        last_name=last_name
                    )

                    # Add user to the selected group
                    user.groups.add(group)

                    # Create UserDetails
                    UserDetails.objects.create(
                        user=user,
                        hire_date=datetime.now().date(),
                        start_date=datetime.now().date(),
                        employment_status='active',
                        work_location=work_location,
                        onboarded_by=request.user,
                        onboarding_date=timezone.now()
                    )

                    # Log user creation
                    UserActionLog.objects.create(
                        user=user,
                        action_type='create',
                        action_by=request.user,
                        details=f"User created in bulk import with ID: {username}, role: {group.name}"
                    )

                    success_count += 1

                except Exception as e:
                    error_messages.append(f"Line {line_number}: Error creating user '{username}' - {str(e)}")

        if success_count > 0:
            messages.success(request, f"Successfully added {success_count} users to the system.")

        if skipped_count > 0:
            messages.warning(request, f"Skipped {skipped_count} entries due to duplicates or invalid data.")

        if error_messages:
            messages.error(request, "The following errors occurred during import:")
            for msg in error_messages[:10]:  # Show first 10 errors
                messages.error(request, msg)

            if len(error_messages) > 10:
                messages.error(request, f"...and {len(error_messages) - 10} more errors. Please check your input data and try again.")

        return redirect('aps_hr:hr_dashboard')

    # For GET request, show the upload form
    return render(request, 'components/hr/bulk_add_users.html', {
        'groups': Group.objects.all(),
    })

@login_required
@user_passes_test(is_hr)
def import_errors(request):
    """View to display errors from bulk import"""
    errors = request.session.get('import_errors', [])
    if not errors:
        messages.info(request, "No import errors to display")
        return redirect('hr_dashboard')

    return render(request, 'components/hr/import_errors.html', {
        'errors': errors
    })


@login_required
@user_passes_test(is_hr)
def user_action_logs(request, user_id=None):
    """View all action logs or filtered by user"""
    if user_id:
        user = get_object_or_404(User, id=user_id)
        logs = UserActionLog.objects.filter(user=user).order_by('-timestamp')
        context = {
            'logs': logs,
            'user': user,
        }
        return render(request, 'components/hr/user_action_logs.html', context)
    else:
        # Get filter parameters
        action_type = request.GET.get('action_type', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        user_filter = request.GET.get('user', '')

        logs = UserActionLog.objects.all().order_by('-timestamp')

        # Apply filters
        if action_type:
            logs = logs.filter(action_type=action_type)

        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                logs = logs.filter(timestamp__date__gte=start_date)
            except ValueError:
                messages.error(request, 'Invalid start date format')

        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                logs = logs.filter(timestamp__date__lte=end_date)
            except ValueError:
                messages.error(request, 'Invalid end date format')

        if user_filter:
            logs = logs.filter(
                Q(user__username__icontains=user_filter) |
                Q(user__first_name__icontains=user_filter) |
                Q(user__last_name__icontains=user_filter)
            )

        # Pagination
        paginator = Paginator(logs, 20)  # Show 20 logs per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'page_obj': page_obj,
            'action_types': UserActionLog.ACTION_TYPES,
            'filter_action_type': action_type,
            'filter_start_date': start_date,
            'filter_end_date': end_date,
            'filter_user': user_filter,
        }
        return render(request, 'components/hr/all_action_logs.html', context)

@login_required
@user_passes_test(is_hr)
def session_logs(request, user_id=None):
    """View session logs for a specific user or all users"""
    if user_id:
        user = get_object_or_404(User, id=user_id)
        sessions = UserSession.objects.filter(user=user).order_by('-login_time')
        context = {
            'sessions': sessions,
            'user': user,
        }
        return render(request, 'components/hr/user_session_logs.html', context)
    else:
        # Get filter parameters
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        user_filter = request.GET.get('user', '')
        location_filter = request.GET.get('location', '')

        sessions = UserSession.objects.all().order_by('-login_time')

        # Apply filters
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                sessions = sessions.filter(login_time__date__gte=start_date)
            except ValueError:
                messages.error(request, 'Invalid start date format')

        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                sessions = sessions.filter(login_time__date__lte=end_date)
            except ValueError:
                messages.error(request, 'Invalid end date format')

        if user_filter:
            sessions = sessions.filter(
                Q(user__username__icontains=user_filter) |
                Q(user__first_name__icontains=user_filter) |
                Q(user__last_name__icontains=user_filter)
            )

        if location_filter:
            sessions = sessions.filter(location=location_filter)

        # Pagination
        paginator = Paginator(sessions, 20)  # Show 20 sessions per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Get unique locations for filter dropdown
        locations = UserSession.objects.values_list('location', flat=True).distinct()

        context = {
            'page_obj': page_obj,
            'locations': locations,
            'filter_start_date': start_date,
            'filter_end_date': end_date,
            'filter_user': user_filter,
            'filter_location': location_filter,
        }
        return render(request, 'components/hr/all_session_logs.html', context)

@login_required
@user_passes_test(is_hr)
def user_reports(request):
    """View to generate reports on user data"""
    report_type = request.GET.get('report_type', 'active_users')
    export_format = request.GET.get('export_format', '')

    if report_type == 'active_users':
        # Report on active vs inactive users
        active_count = User.objects.filter(is_active=True).count()
        inactive_count = User.objects.filter(is_active=False).count()

        # Users by employment status
        status_counts = UserDetails.objects.exclude(employment_status__isnull=True).values('employment_status').annotate(count=Count('employment_status'))

        # Convert to dict with display names
        status_choices = dict(UserDetails._meta.get_field('employment_status').choices)
        status_data = {status_choices.get(item['employment_status'], item['employment_status']): item['count'] for item in status_counts}

        context = {
            'report_type': 'active_users',
            'active_count': active_count,
            'inactive_count': inactive_count,
            'status_data': status_data,
        }

    elif report_type == 'location_distribution':
        # Report on work location distribution
        location_counts = UserDetails.objects.exclude(work_location__isnull=True).exclude(work_location='').values('work_location').annotate(count=Count('work_location')).order_by('-count')

        context = {
            'report_type': 'location_distribution',
            'location_counts': location_counts,
        }

    elif report_type == 'session_activity':
        # Report on session activity
        # Today's active users
        today = datetime.now().date()
        active_today = UserSession.objects.filter(login_time__date=today).values('user').distinct().count()

        # Last 7 days active users
        week_ago = today - timedelta(days=7)
        active_week = UserSession.objects.filter(login_time__date__gte=week_ago).values('user').distinct().count()

        # Average session duration
        avg_duration = UserSession.objects.filter(
            working_hours__isnull=False
        ).values('user').annotate(
            avg_duration=Avg('working_hours')
        )

        # Location breakdown
        location_sessions = UserSession.objects.filter(
            login_time__date__gte=week_ago
        ).values('location').annotate(
            count=Count('id')
        ).order_by('-count')

        context = {
            'report_type': 'session_activity',
            'active_today': active_today,
            'active_week': active_week,
            'avg_duration': avg_duration,
            'location_sessions': location_sessions,
        }

    elif report_type == 'role_distribution':
        # Report on role distribution
        role_counts = Group.objects.annotate(
            user_count=Count('user')
        ).values('name', 'user_count').order_by('-user_count')

        context = {
            'report_type': 'role_distribution',
            'role_counts': role_counts,
        }

    # Handle export if requested
    if export_format:
        if export_format == 'csv':
            return export_as_csv(request, report_type, context)
        # Comment out or remove undefined functions until they're implemented
        # elif export_format == 'excel':
        #     return export_as_excel(request, report_type, context)
        # elif export_format == 'pdf':
        #     return export_as_pdf(request, report_type, context)

    return render(request, 'components/hr/user_reports.html', context)

@login_required
@user_passes_test(is_hr)
def export_as_csv(request, report_type, context):
    """Export report data as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report.csv"'

    writer = csv.writer(response)

    if report_type == 'active_users':
        writer.writerow(['Status', 'Count'])
        writer.writerow(['Active', context['active_count']])
        writer.writerow(['Inactive', context['inactive_count']])

        writer.writerow([])  # Empty row as separator
        writer.writerow(['Employment Status', 'Count'])
        for status, count in context['status_data'].items():
            writer.writerow([status, count])

    elif report_type == 'location_distribution':
        writer.writerow(['Location', 'User Count'])
        for item in context['location_counts']:
            writer.writerow([item['work_location'], item['count']])

    elif report_type == 'session_activity':
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Active Users Today', context['active_today']])
        writer.writerow(['Active Users Last 7 Days', context['active_week']])

        writer.writerow([])  # Empty row as separator
        writer.writerow(['Location', 'Session Count'])
        for item in context['location_sessions']:
            writer.writerow([item['location'], item['count']])

    elif report_type == 'role_distribution':
        writer.writerow(['Role', 'User Count'])
        for item in context['role_counts']:
            writer.writerow([item['name'], item['user_count']])

    return response


@login_required
@user_passes_test(is_hr)
def reset_user_password(request, user_id):
    """Reset a user's password"""
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        reason = request.POST.get('reason')

        if not reason:
            messages.error(request, "Please provide a reason for password reset")
            return redirect('aps_hr:hr_user_detail', user_id=user_id)

        if not new_password or len(new_password) < 8:
            messages.error(request, "Password must be at least 8 characters")
            return redirect('aps_hr:hr_user_detail', user_id=user_id)

        if new_password != confirm_password:
            messages.error(request, "Passwords do not match")
            return redirect('aps_hr:hr_user_detail', user_id=user_id)

        try:
            # Set new password
            user.set_password(new_password)
            user.save()

            # Log password reset with reason
            UserActionLog.objects.create(
                user=user,
                action_type='password_reset',
                action_by=request.user,
                details=f"Password reset by HR. Reason: {reason}"
            )

            # Send email notification to user
            try:
                subject = "Your Password Has Been Reset"
                message = f"""
                Hello {user.first_name} {user.last_name},

                Your password has been reset by HR. Your new password is:

                {new_password}

                Please log in at http://yourcompanyportal.com/login/ and change your password immediately.

                Reason for reset: {reason}

                Regards,
                HR Department
                """

                user.email_user(subject, message, fail_silently=True)
                messages.success(request, f"Password for {user.username} has been reset and email sent to the user.")
            except Exception as e:
                logger.error(f"Error sending email: {str(e)}", exc_info=True)
                messages.warning(request, f"Password for {user.username} has been reset but email could not be sent.")

            return redirect('aps_hr:hr_user_detail', user_id=user.id)

        except Exception as e:
            logger.error(f"Error resetting password: {str(e)}", exc_info=True)
            messages.error(request, f"Error resetting password: {str(e)}")
            return redirect('aps_hr:hr_user_detail', user_id=user_id)

    # This fallback should also redirect to the user detail page
    messages.error(request, "Invalid request")
    return redirect('aps_hr:hr_user_detail', user_id=user_id)

@login_required
@user_passes_test(is_hr)
def change_user_status(request, user_id):
    """Change a user's status (activate/deactivate)"""
    # Debug information
    logger.info(f"change_user_status called with user_id={user_id}")
    logger.info(f"POST data: {request.POST}")

    # Get the user from the URL parameter
    user = get_object_or_404(User, id=user_id)
    user_detail, created = UserDetails.objects.get_or_create(user=user)

    if request.method == 'POST':
        # Extra safety check - verify the user_id in POST matches the URL
        form_user_id = request.POST.get('user_id')
        if form_user_id and int(form_user_id) != int(user_id):
            messages.error(request, "User ID mismatch detected. Operation aborted for security.")
            return redirect('aps_hr:hr_user_detail', user_id=user_id)

        new_status = request.POST.get('status')
        reason = request.POST.get('reason', '')

        if not new_status or new_status not in dict(UserDetails._meta.get_field('employment_status').choices):
            messages.error(request, "Invalid status provided")
            return redirect('aps_hr:hr_user_detail', user_id=user.id)

        try:
            with transaction.atomic():
                # Update UserDetails status
                old_status = user_detail.employment_status
                user_detail.employment_status = new_status
                user_detail.last_status_change = timezone.now()
                user_detail.save()

                # Update User.is_active based on status
                if new_status in ['inactive', 'terminated', 'resigned', 'suspended', 'absconding']:
                    user.is_active = False
                elif new_status == 'active':
                    user.is_active = True
                user.save()

                # Log status change
                UserActionLog.objects.create(
                    user=user,
                    action_type='status_change',
                    action_by=request.user,
                    details=f"Status changed from '{old_status or 'None'}' to '{new_status}'. Reason: {reason}"
                )

                messages.success(request, f"Status for {user.username} changed to {new_status}")
        except Exception as e:
            logger.error(f"Error changing user status: {str(e)}", exc_info=True)
            messages.error(request, f"Error changing user status: {str(e)}")

    return redirect('aps_hr:hr_user_detail', user_id=user.id)

@login_required
@user_passes_test(is_hr)
def change_user_role(request, user_id):
    """Change a user's role/group"""
    # Get the user from the URL parameter
    user = get_object_or_404(User, id=user_id)
    user_detail, created = UserDetails.objects.get_or_create(user=user)

    if request.method == 'POST':
        # Extra safety check - verify the user_id in POST matches the URL
        form_user_id = request.POST.get('user_id')
        if form_user_id and int(form_user_id) != int(user_id):
            messages.error(request, "User ID mismatch detected. Operation aborted for security.")
            return redirect('aps_hr:hr_user_detail', user_id=user_id)

        new_group_id = request.POST.get('group')
        reason = request.POST.get('reason', '')

        if not new_group_id:
            messages.error(request, "No role selected")
            return redirect('aps_hr:hr_user_detail', user_id=user.id)

        try:
            new_group = Group.objects.get(id=new_group_id)
            old_groups = list(user.groups.all())
            old_group_names = ', '.join([group.name for group in old_groups]) if old_groups else 'None'

            with transaction.atomic():
                # Remove from all current groups
                user.groups.clear()

                # Add to new group
                user.groups.add(new_group)

                # Update UserDetails
                user_detail.group = new_group
                user_detail.save()

                # Log role change
                UserActionLog.objects.create(
                    user=user,
                    action_type='role_change',
                    action_by=request.user,
                    details=f"Role changed from '{old_group_names}' to '{new_group.name}'. Reason: {reason}"
                )

                messages.success(request, f"Role for {user.username} changed to {new_group.name}")
        except Group.DoesNotExist:
            messages.error(request, "Invalid role selected")
        except Exception as e:
            logger.error(f"Error changing user role: {str(e)}", exc_info=True)
            messages.error(request, f"Error changing user role: {str(e)}")

    return redirect('aps_hr:hr_user_detail', user_id=user.id)

# Manager Views
@login_required
@user_passes_test(is_manager)
def manager_employee_profile(request):
    """Manager Dashboard to view team members"""
    manager_group = request.user.groups.first()
    team_members = UserDetails.objects.filter(group=manager_group).exclude(user=request.user)

    return render(request, 'components/manager/manager_dashboard.html', {
        'team_members': team_members,
        'role': 'Manager',
        'user_detail': request.user.userdetails,
    })

@login_required
@user_passes_test(is_manager)
def manager_user_detail(request, user_id):
    """Manager view to see (but not edit) user details"""
    user_detail = get_object_or_404(UserDetails, id=user_id)

    # Managers should only see users in their group
    if request.user.groups.first() != user_detail.group:
        return HttpResponseForbidden("You don't have permission to view this user's details")

    # Get recent sessions
    recent_sessions = UserSession.objects.filter(user=user_detail.user).order_by('-login_time')[:5]

    return render(request, 'components/manager/manager_user_detail.html', {
        'user_detail': user_detail,
        'recent_sessions': recent_sessions,
        'role': 'Manager'
    })

# Employee Views
@login_required
@user_passes_test(is_employee)
def employee_profile(request):
    """Employee Profile to view their own details"""
    try:
        user_detail = UserDetails.objects.get(user=request.user)
    except UserDetails.DoesNotExist:
        messages.error(request, 'Profile not found.')
        return redirect('home')

    # Get recent sessions
    recent_sessions = UserSession.objects.filter(
        user=request.user
    ).order_by('-login_time')[:5]

    return render(request, 'components/employee/employee_profile.html', {
        'user_detail': user_detail,
        'recent_sessions': recent_sessions,
        'role': 'Employee',
        'username': request.user.username
    })



# Alias for reset_user_password to maintain URL compatibility
# @login_required
# @user_passes_test(is_hr)
# def hr_reset_user_password(request, user_id):
#     """Alias for reset_user_password to maintain URL compatibility"""
#     return reset_user_password(request, user_id)
''' --------------------------------------------------------- ADMIN AREA --------------------------------------------------------- '''
# Helper function to check if the user belongs to the Admin group
def is_admin(user):
    """Check if the user belongs to the Admin group using Group model."""
    admin_group_id = 1  # Admin group ID from auth_group table
    return user.groups.filter(id=admin_group_id).exists()


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q
from django.utils import timezone
from django.contrib import messages
from datetime import datetime, timedelta

# Helper function to check if the user is an admin
def is_admin(user):
    return user.is_authenticated and user.is_staff

from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render

def is_admin(user):
    """Check if the user has admin privileges."""
    return user.groups.filter(name='Admin').exists()



@login_required
@user_passes_test(is_manager)
def manager_report_view(request):
    """Manager report dashboard view."""

    # Navigation items for the manager dashboard
    nav_items = [
        # {
        #     'id': 'team_breaks',
        #     'name': 'Team Breaks',
        #     'icon': 'fas fa-coffee',
        #     'description': 'Monitor team break patterns and durations.',
        #     'url': reverse('aps_manager:break_report_view_manager')
        # },
        {
            'id': 'team_attendance',
            'name': 'Team Attendance',
            'icon': 'fas fa-calendar-check',
            'description': 'Track team attendance and working hours.',
            'url': reverse('aps_manager:attendance_report_view_manager')
        },
        {
            'id': 'project_progress ',
            'name': 'Project Progress (Releasing Soon)',
            'icon': 'fas fa-tasks',
            'description': 'View progress of your team\'s projects.',
            # 'url': reverse('manager:project_progress_report')
        },
        {
            'id': 'leave_management',
            'name': 'Leave Management (Releasing Soon)',
            'icon': 'fas fa-user-clock',
            'description': 'Manage team leave requests and schedules.',
            # 'url': reverse('manager:leave_management')
        }
    ]

    # Detailed sections for the manager dashboard
    sections = [
        # {
        #     "title": "Team Break Analysis",
        #     "description": "Monitor and analyze your team's break patterns",
        #     "content": "View break statistics, patterns, and ensure policy compliance.",
        #     # "link": reverse('manager:team_breaks_report'),
        #     "metrics": [
        #         {"label": "Active Breaks", "value": "get_active_breaks_count()"},
        #         {"label": "Today's Total Breaks", "value": "get_today_breaks_count()"}
        #     ]
        # },
        {
            "title": "Team Attendance Overview",
            "description": "Track your team's attendance and working hours",
            "content": "Monitor check-ins, working hours, and attendance patterns.",
            # "link": reverse('manager:team_attendance_report'),
            "metrics": [
                {"label": "Team Present", "value": "get_present_count()"},
                {"label": "On Leave", "value": "get_on_leave_count()"}
            ]
        },
        {
            "title": "Project Status (Releasing Soon)",
            "description": "Current status of all projects under your management",
            "content": "Track project progress, deadlines, and resource allocation.",
            # "link": reverse('manager:project_progress_report'),
            "metrics": [
                {"label": "Active Projects", "value": "get_active_projects_count()"},
                {"label": "Upcoming Deadlines", "value": "get_upcoming_deadlines_count()"}
            ]
        },
        {
            "title": "Leave Management (Releasing Soon)",
            "description": "Manage team leave requests and schedules",
            "content": "Review and approve leave requests, plan team availability.",
            # "link": reverse('manager:leave_management'),
            "metrics": [
                {"label": "Pending Requests", "value": "get_pending_leaves_count()"},
                {"label": "Approved Leaves", "value": "get_approved_leaves_count()"}
            ]
        }
    ]

    context = {
        'nav_items': nav_items,
        'sections': sections,
        'page_title': 'Manager Dashboard',
        'manager_name': request.user.get_full_name() or request.user.username
    }

    return render(request, 'components/manager/report.html', context)

from django.db.models import F, ExpressionWrapper, DurationField
from django.db.models.functions import Coalesce
from django.utils import timezone

@login_required
@user_passes_test(is_manager)
def break_report_view_manager(request):
    """View for managers to see breaks taken by users assigned to their projects."""

    # Get filter parameters
    group_name = request.GET.get('group', '')
    break_type = request.GET.get('break_type', '')
    date_str = request.GET.get('date', '')

    # Start with base query
    breaks_query = Break.objects.select_related('user').all()

    # Filter based on the manager's team (users assigned to projects managed by the manager)
    project_assignments = ProjectAssignment.objects.filter(
        project__projectassignment__user=request.user,
        project__projectassignment__role_in_project='Manager',
        is_active=True
    ).values_list('user', flat=True).distinct()

    # Ensure only breaks from users assigned to the manager's projects are visible
    if project_assignments.exists():
        breaks_query = breaks_query.filter(user__in=project_assignments)
    else:
        # No assigned users for the manager
        breaks_query = Break.objects.none()

    # Apply additional filters
    if group_name:
        breaks_query = breaks_query.filter(user__groups__name=group_name)

    if break_type:
        breaks_query = breaks_query.filter(break_type=break_type)

    if date_str:
        try:
            filter_date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
            breaks_query = breaks_query.filter(start_time__date=filter_date)
        except ValueError:
            pass

    # Calculate duration for each break (in minutes), accounting for ongoing breaks
    breaks_query = breaks_query.annotate(
        duration=ExpressionWrapper(
            Coalesce(F('end_time'), timezone.now()) - F('start_time'),
            output_field=DurationField()
        )
    )

    # Convert the duration to minutes after the query
    for break_obj in breaks_query:
        if break_obj.duration:
            break_obj.duration_minutes = break_obj.duration.total_seconds() / 60
        else:
            break_obj.duration_minutes = None

    # Order breaks by start time (most recent first)
    breaks_query = breaks_query.order_by('-start_time')

    # Get available groups based on the manager's team
    groups = Group.objects.filter(user__in=project_assignments).distinct()

    # Pagination
    paginator = Paginator(breaks_query, 10)  # 10 breaks per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Add pagination info
    page_obj.start = (page_obj.number - 1) * paginator.per_page + 1
    page_obj.end = min(page_obj.start + paginator.per_page - 1, paginator.count)
    page_obj.total = paginator.count

    # Define break types
    break_types = [
        'Tea Break (10 mins)',
        'Lunch/Dinner Break (30 mins)',
        'Tea Break (15 mins)'
    ]

    # Add role-specific data for managers
    context = {
        'breaks': page_obj,
        'groups': groups,
        'break_types': break_types,
        'selected_group': group_name,
        'selected_break_type': break_type,
        'selected_date': date_str,
        'total_breaks': breaks_query.count(),
        'active_breaks': breaks_query.filter(end_time__isnull=True).count(),
    }

    return render(request, 'components/manager/break_report.html', context)


from django.db.models import Q, Count, Value, Sum, Avg
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator
from django.utils import timezone
from typing import Dict, Any
from django.db.models import Prefetch

def get_dynamic_report(
    request,
    model_class,
    template_name: str,
    filters: Dict[str, Any] = None,
    per_page: int = 10
) -> Dict[str, Any]:
    """
    Generic function to generate dynamic reports for any model with location awareness
    """
    # Initialize query with prefetch for employee instead of userdetails
    base_query = model_class.objects.select_related('user__employee').all()

    # Get filter parameters
    user_filter = request.GET.get('user', '')
    date_filter = request.GET.get('date', '')
    location_filter = request.GET.get('location', '')
    status_filter = request.GET.get('status', '')

    # Apply filters
    if filters:
        base_query = base_query.filter(**filters)

    if user_filter:
        base_query = base_query.filter(user__username=user_filter)

    if date_filter:
        try:
            filter_date = timezone.datetime.strptime(date_filter, '%Y-%m-%d').date()
            base_query = base_query.filter(date=filter_date)
        except ValueError:
            pass

    if location_filter:
        # Filter based on employee work_location instead of userdetails
        base_query = base_query.filter(user__employee__work_location=location_filter)

    if status_filter:
        base_query = base_query.filter(status=status_filter)

    # Add annotations for statistics by location
    location_stats = UserDetails.objects.values('work_location').annotate(
        present_count=Count(
            'user__attendance',
            filter=Q(user__attendance__status='Present')
        ),
        late_count=Count(
            'user__attendance',
            filter=Q(user__attendance__status='Late')
        ),
        absent_count=Count(
            'user__attendance',
            filter=Q(user__attendance__status='Absent')
        ),
        wfh_count=Count(
            'user__attendance',
            filter=Q(user__attendance__status='Work From Home')
        )
    )

    # Get filter options
    all_users = User.objects.select_related('employee').all()
    # Get unique work locations from Employee model
    all_locations = UserDetails.objects.values_list(
        'work_location', flat=True
    ).distinct().exclude(work_location__isnull=True)
    all_statuses = [status[0] for status in Attendance.STATUS_CHOICES]

    # Pagination
    paginator = Paginator(base_query, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Calculate pagination metadata
    page_obj.start = (page_obj.number - 1) * paginator.per_page + 1
    page_obj.end = min(page_obj.start + paginator.per_page - 1, paginator.count)
    page_obj.total = paginator.count

    context = {
        'records': page_obj,
        'all_users': all_users,
        'all_locations': all_locations,
        'all_statuses': all_statuses,
        'location_stats': location_stats,
        'selected_user': user_filter,
        'selected_date': date_filter,
        'selected_location': location_filter,
        'selected_status': status_filter,
        'model_name': model_class._meta.verbose_name.title(),
        'model_name_plural': model_class._meta.verbose_name_plural.title()
    }

    return context

@login_required
@user_passes_test(is_manager)
def attendance_report_view_manager(request):
    """View for managers to see attendance of all users with location filtering."""

    context = get_dynamic_report(
        request=request,
        model_class=Attendance,
        template_name='components/manager/attendance_report.html',
        per_page=15
    )

    # Add location-based attendance stats
    location_attendance_stats = UserDetails.objects.values('work_location').annotate(
        total_overtime=Coalesce(
            Sum('user__attendance__overtime_hours'),
            Value(0)
        ),
        avg_late_minutes=Coalesce(
            Avg('user__attendance__late_minutes'),
            Value(0)
        )
    )

    context.update({
        'location_attendance_stats': location_attendance_stats
    })

    return render(request, 'components/manager/attendance_report.html', context)




@login_required  # Ensure the user is logged in
@user_passes_test(is_admin)  # Ensure the user is an admin
def report_view(request):
    """Main report navigation view."""

    # Navigation items for the report dashboard
    nav_items = [
        {
            'id': 'breaks',
            'name': 'Break Report',
            'icon': 'fas fa-clock',
            'description': 'View breaks taken by all users.',
            'url': reverse('aps_admin:break_report_view')  # Add URL to the nav item
        },
        {
            'id': 'attendance',
            'name': 'Attendace Report',
            'icon': 'fas fa-clock',
            'description': 'View attendance taken by all users.',
            'url': reverse('aps_admin:attendance')  # Add URL to the nav item
        },
        {
            'id': 'projects',
            'name': 'Projects',
            'icon': 'fas fa-project-diagram',
            'description': 'Detailed overview of ongoing and completed projects.',
        },
        {
            'id': 'systemerrors',
            'name': 'System Errors',
            'icon': 'fas fa-exclamation-triangle',
            'description': 'Log and analyze system errors.',
        },
        {
            'id': 'systemusage',
            'name': 'System Usage',
            'icon': 'fas fa-desktop',
            'description': 'Track system performance metrics and user activity.',
        },
    ]

    # Additional sections for the report dashboard
    sections = [
        {
            "title": "Break Report",
            "description": "This section allows you to view all the breaks taken by users.",
            "content": "You can filter the breaks by group, break type, and date.",
            "link": "/aps/reports/breaks/",  # Link to the break report page
        },
        {
            "title": "Attendace Report",
            "description": "This section provides insights into how features are being used within the platform.",
            "content": "Coming soon...",
        },
        {
            "title": "Projects Report",
            "description": "Detailed overview of all ongoing and completed projects.",
            "content": "Coming soon...",
        },
        {
            "title": "System Errors",
            "description": "Log and analyze system errors to ensure smooth platform performance.",
            "content": "Coming soon...",
        },
        {
            "title": "System Usage",
            "description": "Track overall system usage, including performance metrics and user activity.",
            "content": "Coming soon...",
        },
    ]

    return render(request, 'components/admin/report.html', {'nav_items': nav_items, 'sections': sections})

# View for Feature Usage Information

from django.utils import timezone

@login_required
@user_passes_test(is_admin)
def admin_attendance_view(request):
    # Filter parameters
    username_filter = request.GET.get('username', '')
    status_filter = request.GET.get('status', '')
    date_filter = request.GET.get('date', '')
    date_range_start = request.GET.get('start_date', '')
    date_range_end = request.GET.get('end_date', '')

    # Query attendance records
    attendance_summary = Attendance.objects.all()

    if username_filter:
        attendance_summary = attendance_summary.filter(user__username__icontains=username_filter)
    if status_filter:
        attendance_summary = attendance_summary.filter(status=status_filter)
    if date_filter:
        try:
            date_obj = timezone.datetime.strptime(date_filter, '%Y-%m-%d').date()
            attendance_summary = attendance_summary.filter(date=date_obj)
        except ValueError:
            pass
    if date_range_start and date_range_end:
        try:
            start_date = timezone.datetime.strptime(date_range_start, '%Y-%m-%d').date()
            end_date = timezone.datetime.strptime(date_range_end, '%Y-%m-%d').date()
            attendance_summary = attendance_summary.filter(date__range=[start_date, end_date])
        except ValueError:
            pass

    # Fetch necessary fields and apply timezone conversion for clock-in/out times
    attendance_summary = attendance_summary.values(
        'user', 'user__first_name', 'user__last_name', 'user__username', 'status', 'date', 'total_hours',
        'clock_in_time', 'clock_out_time'
    ).order_by('-date')

    # Applying timezone conversion
    for record in attendance_summary:
        if record['clock_in_time']:
            # Convert clock_in_time to local time zone
            record['clock_in_time'] = timezone.localtime(record['clock_in_time'])
        if record['clock_out_time']:
            # Convert clock_out_time to local time zone
            record['clock_out_time'] = timezone.localtime(record['clock_out_time'])

    # Pagination
    paginator = Paginator(attendance_summary, 10)
    page = request.GET.get('page', 1)

    try:
        summary_records = paginator.get_page(page)
    except EmptyPage:
        summary_records = paginator.page(paginator.num_pages)
    except PageNotAnInteger:
        summary_records = paginator.page(1)

    return render(request, 'components/admin/attendance_report.html', {
        'summary': summary_records,
        'username_filter': username_filter,
        'status_filter': status_filter,
        'date_filter': date_filter,
        'date_range_start': date_range_start,
        'date_range_end': date_range_end,
    })


@login_required
@user_passes_test(is_admin)
def feature_usage_view(request):
    """View to display feature usage details."""
    try:
        feature_usages = FeatureUsage.objects.all().order_by('-usage_count')
        return render(request, 'components/admin/reports/feature_usage.html', {'feature_usages': feature_usages})

    except Exception as e:
        messages.error(request, f"An error occurred while fetching feature usage data: {str(e)}")
        return redirect('dashboard')


@login_required
@user_passes_test(is_admin)
def projects_report_view(request):
    """View to display projects report."""
    return render(request, 'components/admin/reports/projects_report.html')

@login_required
@user_passes_test(is_admin)
def break_report_view(request):
    """View for admin to see all breaks taken by all users."""

    # Get filter parameters
    group_name = request.GET.get('group', '')
    break_type = request.GET.get('break_type', '')
    date_str = request.GET.get('date', '')

    # Start with all breaks - no initial filter
    breaks_query = Break.objects.select_related('user').all()

    # Apply filters only if they're provided
    if group_name:
        breaks_query = breaks_query.filter(user__groups__name=group_name)

    if break_type:
        breaks_query = breaks_query.filter(break_type=break_type)

    if date_str:
        try:
            filter_date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
            breaks_query = breaks_query.filter(start_time__date=filter_date)
        except ValueError:
            pass

    # Calculate duration for each break
    for break_obj in breaks_query:
        if break_obj.end_time:
            duration = (break_obj.end_time - break_obj.start_time).total_seconds() / 60
            break_obj.duration = int(duration)  # Duration in minutes
        else:
            break_obj.duration = None  # If break is ongoing

    # Order breaks by start time (most recent first)
    breaks_query = breaks_query.order_by('-start_time')

    # Debug print
    print(f"Total breaks found: {breaks_query.count()}")
    for break_obj in breaks_query:
        print(f"Break: {break_obj.id}, User: {break_obj.user}, Type: {break_obj.break_type}")

    # Pagination
    paginator = Paginator(breaks_query, 10)  # 10 breaks per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Add pagination info
    page_obj.start = (page_obj.number - 1) * paginator.per_page + 1
    page_obj.end = min(page_obj.start + paginator.per_page - 1, paginator.count)
    page_obj.total = paginator.count

    # Get all groups for the filter dropdown
    groups = Group.objects.all()

    # Debug print groups
    print(f"Available groups: {[group.name for group in groups]}")

    # Define break types
    break_types = [
        'Tea Break (10 mins)',
        'Lunch/Dinner Break (30 mins)',
        'Tea Break (15 mins)'
    ]

    context = {
        'breaks': page_obj,
        'groups': groups,
        'break_types': break_types,
        'selected_group': group_name,
        'selected_break_type': break_type,
        'selected_date': date_str,
    }

    return render(request, 'components/admin/break_report.html', context)

@login_required
@user_passes_test(is_admin)
def system_error_view(request):
    """View to display system errors."""
    try:
        system_errors = SystemError.objects.all().order_by('-error_time')
        return render(request, 'components/admin/reports/system_error.html', {'system_errors': system_errors})

    except Exception as e:
        messages.error(request, f"An error occurred while fetching system errors: {str(e)}")
        return redirect('dashboard')


@login_required
@user_passes_test(is_admin)
def system_usage_view(request):
    """View to display system usage details."""
    try:
        system_usages = SystemUsage.objects.all().order_by('-peak_time_start')
        return render(request, 'components/admin/reports/system_usage.html', {'system_usages': system_usages})

    except Exception as e:
        messages.error(request, f"An error occurred while fetching system usage data: {str(e)}")
        return redirect('dashboard')


''' -------------- usersession ---------------'''
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db import transaction
from django.db.models import (Q, Count, Avg, Sum, F, Max, Min, Case, When,
                             IntegerField, FloatField, DurationField, Value)
from django.db.models.functions import (TruncDate, TruncHour, Extract,
                                       Coalesce, Cast)
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils.dateparse import parse_date, parse_datetime
from django.template.loader import render_to_string
from datetime import datetime, timedelta, date
from decimal import Decimal
import json
import logging
import pytz
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from collections import defaultdict, Counter

from .models import UserSession
from .helpers import is_user_in_group

# Set up logging
logger = logging.getLogger(__name__)

# Asia/Kolkata timezone
IST_TIMEZONE = pytz.timezone('Asia/Kolkata')

def is_admin(user):
    """Check if user is admin or has appropriate permissions"""
    return user.is_staff or user.is_superuser or is_user_in_group(user, 'Admin')

def get_current_time_ist():
    """Return current time in Asia/Kolkata timezone (aware)."""
    return timezone.now().astimezone(IST_TIMEZONE)

def to_ist(dt):
    """Convert a datetime to Asia/Kolkata timezone (aware)."""
    if dt is None:
        return None
    if timezone.is_naive(dt):
        return IST_TIMEZONE.localize(dt)
    return dt.astimezone(IST_TIMEZONE)

def to_utc(dt_ist):
    """Convert IST datetime to UTC for database storage."""
    if dt_ist is None:
        return None
    if timezone.is_naive(dt_ist):
        dt_ist = IST_TIMEZONE.localize(dt_ist)
    return dt_ist.astimezone(pytz.UTC)

@login_required
@user_passes_test(is_admin)
def user_sessions_view(request):
    # Handle export requests
    if request.method == 'POST' and request.GET.get('export'):
        export_format = request.GET.get('export')
        filters = _parse_session_filters(request)
        queryset = _get_filtered_sessions(filters)
        
        try:
            if export_format == 'csv':
                return _export_to_csv(queryset, filters)
            elif export_format == 'xlsx':
                return _export_to_excel(queryset, filters)
        except Exception as e:
            logger.error(f"Export error: {str(e)}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'error',
                    'message': 'Failed to generate export file'
                })
            messages.error(request, 'Failed to generate export file')
            return redirect(request.path)
    """
    Enhanced view to display user sessions with advanced filtering, analytics and insights
    for tracking employee productivity.
    """
    print(f"DEBUG: user_sessions_view called by user: {request.user.username}")

    try:
        # Get current time in IST
        print("DEBUG: Getting current time in IST")
        current_time_ist = get_current_time_ist()
        current_time_utc = timezone.now()
        print(f"DEBUG: Current time IST: {current_time_ist}, UTC: {current_time_utc}")

        # Parse filters from request
        print("DEBUG: Parsing session filters from request")
        filters = _parse_session_filters(request)
        date_from = filters['date_from']
        date_to = filters['date_to']
        user_filter = filters['user_filter']
        location_filter = filters['location_filter']
        device_filter = filters['device_filter']
        print(f"DEBUG: Filters parsed - date_from: {date_from}, date_to: {date_to}")
        print(f"DEBUG: Additional filters - user: {user_filter}, location: {location_filter}, device: {device_filter}")

        # Handle AJAX requests for live data
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            print("DEBUG: AJAX request detected")
            ajax_type = request.GET.get('type')
            print(f"DEBUG: AJAX type: {ajax_type}")

            if ajax_type == 'live_activity':
                print("DEBUG: Returning live activity data")
                return _get_live_activity_data(request)
            elif ajax_type == 'overview_metrics':
                print("DEBUG: Returning overview metrics")
                return _get_overview_metrics(request, date_from, date_to)
            elif ajax_type == 'session_analytics':
                print("DEBUG: Returning session analytics")
                return _get_session_analytics(request, date_from, date_to, filters)
            elif ajax_type == 'export':
                print("DEBUG: Exporting session data")
                return _export_session_data(request, filters)

        # Base queryset with filters
        print("DEBUG: Getting filtered sessions queryset")
        base_queryset = _get_filtered_sessions(filters)
        print(f"DEBUG: Base queryset count: {base_queryset.count()}")

        # ==================== HIGH-LEVEL OVERVIEW METRICS ====================
        print("DEBUG: Calculating overview metrics")
        overview_metrics = _calculate_overview_metrics(base_queryset, current_time_utc)
        print(f"DEBUG: Overview metrics calculated: {overview_metrics}")

        # ==================== LIVE ACTIVITY FEED ====================
        print("DEBUG: Getting live activity sessions")
        live_sessions = UserSession.objects.filter(
            is_active=True,
            last_activity__gte=current_time_utc - timedelta(minutes=5)
        ).select_related('user').order_by('-last_activity')[:50]
        print(f"DEBUG: Live sessions count: {live_sessions.count()}")

        live_activity_data = []
        for session in live_sessions:
            print(f"DEBUG: Processing live session {session.id} for user {session.user.username}")
            last_activity_ist = to_ist(session.last_activity)
            login_time_ist = to_ist(session.login_time)

            # Calculate session duration
            session_duration = (current_time_utc - session.login_time).total_seconds() / 60

            # Determine activity status
            time_since_activity = (current_time_utc - session.last_activity).total_seconds() / 60
            if time_since_activity <= 1:
                activity_status = 'active'
            elif time_since_activity <= 5:
                activity_status = 'idle'
            else:
                activity_status = 'inactive'

            live_activity_data.append({
                'user': session.user,
                'session_id': session.id,
                'tab_id': session.tab_id,
                'ip_address': session.ip_address,
                'location': session.location,
                'device_type': session.device_type,
                'is_primary_tab': session.is_primary_tab,
                'login_time_ist': login_time_ist,
                'last_activity_ist': last_activity_ist,
                'session_duration_minutes': round(session_duration, 1),
                'activity_status': activity_status,
                'productivity_score': session.productivity_score or 0,
                'engagement_score': session.engagement_score or 0,
                'tab_switches': session.tab_switches,
                'current_url': session.tab_url,
                'is_office_ip': session.ip_address in UserSession.OFFICE_IPS if session.ip_address else False
            })
        print(f"DEBUG: Live activity data prepared for {len(live_activity_data)} sessions")

        # ==================== USER-BASED INSIGHTS ====================
        print("DEBUG: Calculating user insights")
        user_insights = _calculate_user_insights(base_queryset, date_from, date_to)
        print(f"DEBUG: User insights calculated for {len(user_insights)} users")

        # ==================== TAB BEHAVIOR ANALYSIS ====================
        print("DEBUG: Calculating tab behavior analysis")
        tab_analysis = _calculate_tab_behavior_analysis(base_queryset)
        print(f"DEBUG: Tab analysis completed: {tab_analysis}")

        # ==================== SECURITY & DEVICE MONITORING ====================
        print("DEBUG: Calculating security insights")
        security_insights = _calculate_security_insights(base_queryset)
        print(f"DEBUG: Security insights calculated: {security_insights}")

        # ==================== LOCATION & ENVIRONMENT TRACKING ====================
        print("DEBUG: Calculating location insights")
        location_insights = _calculate_location_insights(base_queryset)
        print(f"DEBUG: Location insights calculated: {location_insights}")

        # ==================== PERFORMANCE MONITORING ====================
        print("DEBUG: Calculating performance insights")
        performance_insights = _calculate_performance_insights(base_queryset)
        print(f"DEBUG: Performance insights calculated: {performance_insights}")

        # ==================== SESSION QUALITY SCORING ====================
        print("DEBUG: Calculating quality insights")
        quality_insights = _calculate_quality_insights(base_queryset)
        print(f"DEBUG: Quality insights calculated: {quality_insights}")

        # ==================== PAGINATION FOR DETAILED SESSION LIST ====================
        print("DEBUG: Setting up pagination for sessions")
        sessions_list = base_queryset.select_related('user').order_by('-login_time')

        # Safely get and validate page number
        page_param = request.GET.get('page', '1')
        try:
            page_num = int(page_param)
            # Ensure page number is at least 1
            if page_num < 1:
                page_num = 1
        except (ValueError, TypeError):
            page_num = 1

        paginator = Paginator(sessions_list, 25)
        print(f"DEBUG: Paginator created with {paginator.count} total sessions")

        try:
            sessions_page = paginator.page(page_num)
            print(f"DEBUG: Page {page_num} loaded successfully")
        except PageNotAnInteger:
            print("DEBUG: PageNotAnInteger error, loading page 1")
            sessions_page = paginator.page(1)
        except EmptyPage:
            print("DEBUG: EmptyPage error, loading last page")
            # If the requested page is too high, go to the last page
            # num_pages is guaranteed to be at least 1
            sessions_page = paginator.page(paginator.num_pages)

        # Convert session times to IST for display
        print("DEBUG: Converting session times to IST for display")
        sessions_display = []
        for session in sessions_page:
            print(f"DEBUG: Processing session {session.id} for display")
            login_time_ist = to_ist(session.login_time)
            last_activity_ist = to_ist(session.last_activity)
            logout_time_ist = to_ist(session.logout_time) if session.logout_time else None

            # Calculate current session duration for active sessions
            if session.is_active:
                current_duration = (current_time_utc - session.login_time).total_seconds() / 60
            else:
                current_duration = session.session_duration or 0

            sessions_display.append({
                'session': session,
                'login_time_ist': login_time_ist,
                'last_activity_ist': last_activity_ist,
                'logout_time_ist': logout_time_ist,
                'current_duration_minutes': round(current_duration, 1),
                'working_hours_display': session.get_total_working_hours_display(),
                'idle_time_minutes': (session.idle_time.total_seconds() / 60) if session.idle_time else 0,
                'is_office_ip': session.ip_address in UserSession.OFFICE_IPS if session.ip_address else False,
                'page_views_count': len(session.page_views or []),
                'interactions_count': (
                    len(session.click_events or []) +
                    len(session.keyboard_events or []) +
                    len(session.scroll_events or [])
                ),
                'security_incidents_count': len(session.security_incidents or {})
            })
        print(f"DEBUG: Sessions display data prepared for {len(sessions_display)} sessions")

        # ==================== CONTEXT DATA ====================
        print("DEBUG: Preparing context data for template")
        context = {
            'current_time_ist': current_time_ist,
            'overview_metrics': overview_metrics,
            'live_activity_data': live_activity_data,
            'user_insights': user_insights,
            'tab_analysis': tab_analysis,
            'security_insights': security_insights,
            'location_insights': location_insights,
            'performance_insights': performance_insights,
            'quality_insights': quality_insights,
            'sessions_page': sessions_page,
            'sessions_display': sessions_display,
            'filters': filters,
            'users_list': User.objects.filter(is_active=True).order_by('username'),
            'date_range_options': [
                ('today', 'Today'),
                ('yesterday', 'Yesterday'),
                ('last_7_days', 'Last 7 Days'),
                ('last_30_days', 'Last 30 Days'),
                ('this_month', 'This Month'),
                ('last_month', 'Last Month'),
                ('custom', 'Custom Range')
            ],
            'device_types': ['desktop', 'mobile', 'tablet'],
            'locations': ['Office', 'Home', 'Unknown'],
            'session_qualities': ['high', 'medium', 'low']
        }
        print("DEBUG: Context data prepared successfully")

        print("DEBUG: Rendering user_sessions_dashboard.html template")
        return render(request, 'components/sessions/user_sessions_dashboard.html', context)
    except Exception as e:
            print(f"DEBUG ERROR: Exception in user_sessions_view: {str(e)}")
            print(f"DEBUG ERROR: Exception type: {type(e)}")
            import traceback
            print(f"DEBUG ERROR: Traceback: {traceback.format_exc()}")
            logger.error(f"Error in user_sessions_view: {str(e)}")
            raise

@login_required
@user_passes_test(is_admin)
def user_session_detail_view(request, user_id, date_str):
    """
    Detailed view of a user's sessions for a specific date.
    Shows session timeline and productivity metrics.
    """
    print(f"DEBUG: user_session_detail_view called for user_id: {user_id}, date: {date_str}")

    try:
        # Get user and parse date
        print(f"DEBUG: Getting user with ID: {user_id}")
        user = get_object_or_404(User, id=user_id)
        print(f"DEBUG: User found: {user.username}")

        try:
            print(f"DEBUG: Parsing date string: {date_str}")
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            print(f"DEBUG: Target date parsed: {target_date}")
        except ValueError as ve:
            print(f"DEBUG ERROR: Date parsing failed: {ve}")
            return render(request, 'components/sessions/error.html', {
                'error_message': 'Invalid date format. Please use YYYY-MM-DD.'
            })

        # Calculate date range in UTC for database queries
        print("DEBUG: Calculating date range in UTC for database queries")
        start_of_day_ist = IST_TIMEZONE.localize(datetime.combine(target_date, datetime.min.time()))
        end_of_day_ist = IST_TIMEZONE.localize(datetime.combine(target_date, datetime.max.time()))
        start_of_day_utc = to_utc(start_of_day_ist)
        end_of_day_utc = to_utc(end_of_day_ist)
        print(f"DEBUG: Date range - Start UTC: {start_of_day_utc}, End UTC: {end_of_day_utc}")

        # Get user's sessions for the specified date
        print("DEBUG: Querying user sessions for the specified date")
        sessions = UserSession.objects.filter(
            user=user,
            login_time__gte=start_of_day_utc,
            login_time__lte=end_of_day_utc
        ).order_by('login_time')
        print(f"DEBUG: Found {sessions.count()} sessions for user {user.username} on {target_date}")

        if not sessions.exists():
            print("DEBUG: No sessions found, returning no_sessions template")
            return render(request, 'components/sessions/user_session_detail.html', {
                'user': user,
                'target_date': target_date,
                'no_sessions': True
            })

        # ============== SESSION TIMELINE ANALYSIS ==============
        print("DEBUG: Starting session timeline analysis")
        timeline_data = []
        total_working_time = timedelta(0)
        total_idle_time = timedelta(0)
        total_session_time = timedelta(0)

        for session in sessions:
            print(f"DEBUG: Processing session {session.id} for timeline")
            login_time_ist = to_ist(session.login_time)
            logout_time_ist = to_ist(session.logout_time) if session.logout_time else None
            last_activity_ist = to_ist(session.last_activity)

            # Calculate session duration
            if session.is_active:
                session_end = timezone.now()
                current_duration = session_end - session.login_time
                print(f"DEBUG: Active session duration: {current_duration}")
            else:
                session_end = session.logout_time
                current_duration = timedelta(minutes=session.session_duration) if session.session_duration else timedelta(0)
                print(f"DEBUG: Inactive session duration: {current_duration}")

            total_session_time += current_duration

            if session.working_hours:
                total_working_time += session.working_hours
                print(f"DEBUG: Added working hours: {session.working_hours}")

            if session.idle_time:
                total_idle_time += session.idle_time
                print(f"DEBUG: Added idle time: {session.idle_time}")

            # Analyze tab behavior
            print("DEBUG: Analyzing tab focus patterns")
            tab_focus_analysis = _analyze_tab_focus_patterns(session)

            # Calculate hourly activity distribution
            print("DEBUG: Calculating hourly activity")
            hourly_activity = _calculate_hourly_activity(session)

            timeline_data.append({
                'session': session,
                'login_time_ist': login_time_ist,
                'logout_time_ist': logout_time_ist,
                'last_activity_ist': last_activity_ist,
                'duration_hours': current_duration.total_seconds() / 3600,
                'working_hours': session.working_hours.total_seconds() / 3600 if session.working_hours else 0,
                'idle_hours': session.idle_time.total_seconds() / 3600 if session.idle_time else 0,
                'productivity_score': session.productivity_score or 0,
                'engagement_score': session.engagement_score or 0,
                'tab_focus_analysis': tab_focus_analysis,
                'hourly_activity': hourly_activity,
                'page_views_count': len(session.page_views or []),
                'interactions_count': (
                    len(session.click_events or []) +
                    len(session.keyboard_events or []) +
                    len(session.scroll_events or [])
                ),
                'tab_switches': session.tab_switches,
                'mouse_movements': session.mouse_movements,
                'is_office_location': session.location == 'Office',
                'device_info': {
                    'type': session.device_type,
                    'screen_resolution': session.screen_resolution,
                    'browser_fingerprint': session.browser_fingerprint
                }
            })
        print(f"DEBUG: Timeline data prepared for {len(timeline_data)} sessions")

        # ============== DAILY SUMMARY METRICS ==============
        print("DEBUG: Calculating daily summary metrics")
        daily_summary = {
            'total_sessions': sessions.count(),
            'active_sessions': sessions.filter(is_active=True).count(),
            'multi_tab_sessions': sessions.filter(parent_session_id__isnull=False).values('parent_session_id').distinct().count(),
            'total_session_hours': total_session_time.total_seconds() / 3600,
            'total_working_hours': total_working_time.total_seconds() / 3600,
            'total_idle_hours': total_idle_time.total_seconds() / 3600,
            'avg_productivity_score': sessions.aggregate(Avg('productivity_score'))['productivity_score__avg'] or 0,
            'avg_engagement_score': sessions.aggregate(Avg('engagement_score'))['engagement_score__avg'] or 0,
            'office_sessions': sessions.filter(location='Office').count(),
            'home_sessions': sessions.filter(location='Home').count(),
            'desktop_sessions': sessions.filter(device_type='desktop').count(),
            'mobile_sessions': sessions.filter(device_type='mobile').count(),
            'tablet_sessions': sessions.filter(device_type='tablet').count()
        }
        print(f"DEBUG: Daily summary calculated: {daily_summary}")

        # Calculate productivity percentage
        if daily_summary['total_session_hours'] > 0:
            daily_summary['productivity_percentage'] = (
                daily_summary['total_working_hours'] / daily_summary['total_session_hours'] * 100
            )
        else:
            daily_summary['productivity_percentage'] = 0
        print(f"DEBUG: Productivity percentage: {daily_summary['productivity_percentage']}")

        # ============== PAGE VISIT ANALYSIS ==============
        print("DEBUG: Starting page visit analysis")
        page_visits = []
        url_counter = Counter()

        for session in sessions:
            print(f"DEBUG: Processing page views for session {session.id}")
            if session.page_views:
                for page_view in session.page_views:
                    url = page_view.get('url', 'Unknown')
                    url_counter[url] += 1

                    # Parse timestamp
                    try:
                        timestamp_str = page_view.get('timestamp', '')
                        if timestamp_str:
                            timestamp = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                            timestamp_ist = to_ist(timestamp)
                        else:
                            timestamp_ist = None
                    except Exception as ts_error:
                        print(f"DEBUG ERROR: Timestamp parsing failed: {ts_error}")
                        timestamp_ist = None

                    page_visits.append({
                        'url': url,
                        'title': page_view.get('title', 'Unknown'),
                        'timestamp_ist': timestamp_ist,
                        'session_id': session.id,
                        'tab_id': session.tab_id
                    })

        # Sort page visits by timestamp
        print("DEBUG: Sorting page visits by timestamp")
        page_visits.sort(key=lambda x: x['timestamp_ist'] or datetime.min.replace(tzinfo=IST_TIMEZONE))

        # Top visited pages
        top_pages = url_counter.most_common(10)
        print(f"DEBUG: Page visit analysis complete - {len(page_visits)} visits, {len(top_pages)} top pages")

        # ============== ACTIVITY HEATMAP DATA ==============
        print("DEBUG: Generating activity heatmap")
        activity_heatmap = _generate_activity_heatmap(sessions, target_date)
        print(f"DEBUG: Activity heatmap generated with {len(activity_heatmap)} hours")

        # ============== SECURITY ANALYSIS ==============
        print("DEBUG: Starting security analysis")
        security_analysis = {
            'unique_ips': list(set(s.ip_address for s in sessions if s.ip_address)),
            'unique_devices': list(set(s.device_type for s in sessions if s.device_type)),
            'security_incidents': [],
            'fingerprint_changes': 0
        }

        # Collect security incidents
        for session in sessions:
            if session.security_incidents:
                for incident_key, incidents in session.security_incidents.items():
                    security_analysis['security_incidents'].extend(incidents)

        # Check for fingerprint changes
        fingerprints = [s.session_fingerprint for s in sessions if s.session_fingerprint]
        security_analysis['fingerprint_changes'] = len(set(fingerprints)) - 1 if len(fingerprints) > 1 else 0
        print(f"DEBUG: Security analysis complete: {security_analysis}")

        # ============== CONTEXT DATA ==============
        print("DEBUG: Preparing context data for template")
        context = {
            'user': user,
            'target_date': target_date,
            'sessions': sessions,
            'timeline_data': timeline_data,
            'daily_summary': daily_summary,
            'page_visits': page_visits[:50],  # Limit to 50 most recent
            'top_pages': top_pages,
            'activity_heatmap': activity_heatmap,
            'security_analysis': security_analysis,
            'current_time_ist': get_current_time_ist()
        }
        print("DEBUG: Context data prepared successfully")

        print("DEBUG: Rendering user_session_detail.html template")
        return render(request, 'components/sessions/user_session_detail.html', context)

    except Exception as e:
        print(f"DEBUG ERROR: Exception in user_session_detail_view: {str(e)}")
        print(f"DEBUG ERROR: Exception type: {type(e)}")
        import traceback
        print(f"DEBUG ERROR: Traceback: {traceback.format_exc()}")
        logger.error(f"Error in user_session_detail_view: {str(e)}")
        raise

# ==================== HELPER FUNCTIONS ====================
#
#
import json
import math
from decimal import Decimal

def safe_json_parse(value, default=None):
    """
    Safely parse JSON from various input types
    """
    if value is None:
        return default

    # If it's already parsed (dict/list), return as-is
    if isinstance(value, (dict, list)):
        return value

    # If it's a string/bytes/bytearray, try to parse
    if isinstance(value, (str, bytes, bytearray)):
        try:
            return json.loads(value)
        except (json.JSONDecodeError, ValueError):
            return default

    # If it's a numeric type, don't try to parse as JSON
    if isinstance(value, (int, float, Decimal)):
        return default

    # For any other type, return default
    return default

def safe_json_serialize(data, context_name="Unknown"):
    """
    Helper function to safely serialize data and identify problematic values
    """
    def clean_value(obj, path="root", depth=0):
        # Prevent infinite recursion
        if depth > 10:
            return None

        if isinstance(obj, dict):
            cleaned = {}
            for key, value in obj.items():
                try:
                    cleaned[key] = clean_value(value, f"{path}.{key}", depth + 1)
                except Exception as e:
                    logger.error(f"Error at {path}.{key}: {e}")
                    cleaned[key] = None
            return cleaned
        elif isinstance(obj, list):
            cleaned = []
            for i, item in enumerate(obj):
                try:
                    cleaned.append(clean_value(item, f"{path}[{i}]", depth + 1))
                except Exception as e:
                    logger.error(f"Error at {path}[{i}]: {e}")
                    cleaned.append(None)
            return cleaned
        elif isinstance(obj, float):
            if math.isinf(obj) or math.isnan(obj):
                logger.warning(f"Found invalid float at {path}: {obj}")
                return 0.0
            return obj
        elif isinstance(obj, Decimal):
            try:
                float_val = float(obj)
                if math.isinf(float_val) or math.isnan(float_val):
                    logger.warning(f"Found invalid Decimal at {path}: {obj}")
                    return 0.0
                return float_val
            except (ValueError, OverflowError):
                return 0.0
        else:
            return obj

    try:
        cleaned_data = clean_value(data)
        json.dumps(cleaned_data)  # Test serialization
        return cleaned_data
    except Exception as e:
        logger.error(f"JSON serialization error in {context_name}: {e}")
        return {}

def safe_divide(numerator, denominator):
    """Safe division that returns 0 instead of infinity"""
    if denominator == 0 or denominator is None:
        return 0.0
    try:
        result = numerator / denominator
        if math.isinf(result) or math.isnan(result):
            return 0.0
        return result
    except (ZeroDivisionError, TypeError):
        return 0.0

def safe_percentage(part, total):
    """Safe percentage calculation"""
    if total == 0 or total is None or part is None:
        return 0.0
    try:
        result = (part / total) * 100
        if math.isinf(result) or math.isnan(result):
            return 0.0
        return round(result, 2)
    except (ZeroDivisionError, TypeError):
        return 0.0

def safe_average(values):
    """Safe average calculation"""
    if not values or len(values) == 0:
        return 0.0
    clean_values = []
    for v in values:
        if v is not None:
            try:
                if not (math.isinf(v) or math.isnan(v)):
                    clean_values.append(v)
            except TypeError:
                continue
    if not clean_values:
        return 0.0
    return sum(clean_values) / len(clean_values)

def _parse_session_filters(request):
    """Parse and validate session filters from request"""
    current_time_ist = get_current_time_ist()

    # Date range handling
    date_range = request.GET.get('date_range', 'last_7_days')
    custom_from = request.GET.get('date_from')
    custom_to = request.GET.get('date_to')

    if date_range == 'today':
        date_from = current_time_ist.date()
        date_to = current_time_ist.date()
    elif date_range == 'yesterday':
        yesterday = current_time_ist.date() - timedelta(days=1)
        date_from = yesterday
        date_to = yesterday
    elif date_range == 'last_7_days':
        date_from = current_time_ist.date() - timedelta(days=7)
        date_to = current_time_ist.date()
    elif date_range == 'last_30_days':
        date_from = current_time_ist.date() - timedelta(days=30)
        date_to = current_time_ist.date()
    elif date_range == 'this_month':
        date_from = current_time_ist.replace(day=1).date()
        date_to = current_time_ist.date()
    elif date_range == 'last_month':
        last_month = current_time_ist.replace(day=1) - timedelta(days=1)
        date_from = last_month.replace(day=1).date()
        date_to = last_month.date()
    elif date_range == 'custom' and custom_from and custom_to:
        try:
            date_from = parse_date(custom_from)
            date_to = parse_date(custom_to)
        except:
            date_from = current_time_ist.date() - timedelta(days=7)
            date_to = current_time_ist.date()
    else:
        date_from = current_time_ist.date() - timedelta(days=7)
        date_to = current_time_ist.date()

    return {
        'date_range': date_range,
        'date_from': date_from,
        'date_to': date_to,
        'user_filter': request.GET.get('user_filter'),
        'location_filter': request.GET.get('location_filter'),
        'device_filter': request.GET.get('device_filter'),
        'status_filter': request.GET.get('status_filter'),
        'ip_type_filter': request.GET.get('ip_type_filter'),
        'quality_filter': request.GET.get('quality_filter')
    }

def _get_filtered_sessions(filters):
    """Get filtered session queryset based on filters"""
    # Convert date range to UTC for database queries
    start_of_day_ist = IST_TIMEZONE.localize(datetime.combine(filters['date_from'], datetime.min.time()))
    end_of_day_ist = IST_TIMEZONE.localize(datetime.combine(filters['date_to'], datetime.max.time()))
    start_of_day_utc = to_utc(start_of_day_ist)
    end_of_day_utc = to_utc(end_of_day_ist)

    queryset = UserSession.objects.filter(
        login_time__gte=start_of_day_utc,
        login_time__lte=end_of_day_utc
    )

    # Apply filters
    if filters['user_filter']:
        try:
            user_id = int(filters['user_filter'])
            queryset = queryset.filter(user_id=user_id)
        except ValueError:
            pass

    if filters['location_filter']:
        queryset = queryset.filter(location=filters['location_filter'])

    if filters['device_filter']:
        queryset = queryset.filter(device_type=filters['device_filter'])

    if filters['status_filter']:
        if filters['status_filter'] == 'active':
            queryset = queryset.filter(is_active=True)
        elif filters['status_filter'] == 'inactive':
            queryset = queryset.filter(is_active=False)

    if filters['ip_type_filter']:
        if filters['ip_type_filter'] == 'office':
            queryset = queryset.filter(ip_address__in=UserSession.OFFICE_IPS)
        elif filters['ip_type_filter'] == 'remote':
            queryset = queryset.exclude(ip_address__in=UserSession.OFFICE_IPS)

    if filters['quality_filter']:
        queryset = queryset.filter(session_quality=filters['quality_filter'])

    return queryset

def _calculate_overview_metrics(queryset, current_time_utc):
    """Calculate high-level overview metrics"""
    # Basic counts
    total_sessions = queryset.count()
    active_sessions = queryset.filter(is_active=True).count()

    # Unique users
    unique_users = queryset.values('user').distinct().count()
    currently_online = queryset.filter(
        is_active=True,
        last_activity__gte=current_time_utc - timedelta(minutes=5)
    ).values('user').distinct().count()

    # Duration metrics with safe handling
    avg_session_duration = queryset.aggregate(
        avg_duration=Avg('session_duration')
    )['avg_duration'] or 0
    avg_session_duration = safe_divide(avg_session_duration, 1)

    avg_idle_time = queryset.aggregate(
        avg_idle=Avg('idle_time')
    )['avg_idle']
    avg_idle_minutes = avg_idle_time.total_seconds() / 60 if avg_idle_time else 0
    avg_idle_minutes = safe_divide(avg_idle_minutes, 1)

    # Productivity metrics with safe handling
    avg_productivity = queryset.filter(
        productivity_score__isnull=False
    ).aggregate(Avg('productivity_score'))['productivity_score__avg'] or 0
    avg_productivity = safe_divide(avg_productivity, 1)

    avg_engagement = queryset.filter(
        engagement_score__isnull=False
    ).aggregate(Avg('engagement_score'))['engagement_score__avg'] or 0
    avg_engagement = safe_divide(avg_engagement, 1)

    # Location breakdown
    location_stats = queryset.values('location').annotate(
        count=Count('id')
    ).order_by('-count')

    # Device breakdown
    device_stats = queryset.values('device_type').annotate(
        count=Count('id')
    ).order_by('-count')

    # Multi-tab sessions with safe percentage calculation
    multi_tab_sessions = queryset.filter(
        parent_session_id__isnull=False
    ).values('parent_session_id').distinct().count()

    multi_tab_percentage = safe_percentage(multi_tab_sessions, total_sessions)

    return {
        'total_sessions': total_sessions,
        'active_sessions': active_sessions,
        'unique_users': unique_users,
        'currently_online': currently_online,
        'avg_session_duration_minutes': round(avg_session_duration, 1),
        'avg_idle_time_minutes': round(avg_idle_minutes, 1),
        'avg_productivity_score': round(avg_productivity, 1),
        'avg_engagement_score': round(avg_engagement, 1),
        'location_breakdown': list(location_stats),
        'device_breakdown': list(device_stats),
        'multi_tab_sessions': multi_tab_sessions,
        'multi_tab_percentage': multi_tab_percentage
    }

def _calculate_user_insights(queryset, date_from, date_to):
    """Calculate detailed user-based insights"""
    user_stats = queryset.values('user__username', 'user__first_name', 'user__last_name', 'user_id').annotate(
        total_sessions=Count('id'),
        active_sessions=Count(Case(
            When(is_active=True, then=1),
            output_field=IntegerField()
        )),
        avg_session_duration=Avg('session_duration'),
        total_working_hours=Sum('working_hours'),
        total_idle_time=Sum('idle_time'),
        avg_productivity=Avg('productivity_score'),
        avg_engagement=Avg('engagement_score'),
        office_sessions=Count(Case(
            When(location='Office', then=1),
            output_field=IntegerField()
        )),
        home_sessions=Count(Case(
            When(location='Home', then=1),
            output_field=IntegerField()
        )),
        desktop_sessions=Count(Case(
            When(device_type='desktop', then=1),
            output_field=IntegerField()
        )),
        mobile_sessions=Count(Case(
            When(device_type='mobile', then=1),
            output_field=IntegerField()
        )),
        total_page_views=Sum(
            Case(
                When(page_views__isnull=False, then=F('page_views')),
                default=Value(0),
                output_field=IntegerField()
            )
        ),
        total_interactions=Sum(
            Coalesce(F('mouse_movements'), 0) +
            Case(
                When(click_events__isnull=False, then=F('click_events')),
                default=Value(0),
                output_field=IntegerField()
            )
        ),
        security_incidents=Count(Case(
            When(security_incidents__isnull=False, then=1),
            output_field=IntegerField()
        ))
    ).order_by('-total_sessions')

    # Calculate additional metrics for each user with safe calculations
    enhanced_user_stats = []
    for user_stat in user_stats:
        # Calculate productivity percentage safely
        avg_duration = user_stat['avg_session_duration'] or 0
        total_sessions = user_stat['total_sessions'] or 1

        if user_stat['total_working_hours']:
            working_hours_minutes = user_stat['total_working_hours'].total_seconds() / 60
        else:
            working_hours_minutes = 0

        session_hours_minutes = avg_duration * total_sessions
        productivity_percentage = safe_percentage(working_hours_minutes, session_hours_minutes)

        user_stat['productivity_percentage'] = productivity_percentage
        user_stat['avg_productivity'] = round(user_stat['avg_productivity'] or 0, 1)
        user_stat['avg_engagement'] = round(user_stat['avg_engagement'] or 0, 1)
        user_stat['avg_session_duration'] = round(avg_duration, 1)

        enhanced_user_stats.append(user_stat)

    return enhanced_user_stats[:20]  # Top 20 users

def _calculate_tab_behavior_analysis(queryset):
    """Analyze tab behavior patterns"""
    # Most visited URLs
    url_counter = Counter()
    tab_switch_stats = []

    for session in queryset:
        # Safely parse page_views
        page_views = safe_json_parse(session.page_views, [])
        if page_views:
            for page_view in page_views:
                if isinstance(page_view, dict):
                    url = page_view.get('url', 'Unknown')
                    url_counter[url] += 1

        tab_switches = session.tab_switches or 0
        if tab_switches > 0:
            tab_switch_stats.append({
                'session_id': session.id,
                'user': session.user.username,
                'tab_switches': tab_switches,
                'session_duration': session.session_duration or 0
            })

    # Calculate tab switch patterns safely
    avg_tab_switches = safe_average([stat['tab_switches'] for stat in tab_switch_stats])
    high_tab_switchers = [stat for stat in tab_switch_stats if stat['tab_switches'] > avg_tab_switches * 1.5]

    return {
        'most_visited_urls': url_counter.most_common(20),
        'avg_tab_switches_per_session': round(avg_tab_switches, 1),
        'high_tab_switchers_count': len(high_tab_switchers),
        'sessions_with_multiple_tabs': queryset.filter(parent_session_id__isnull=False).count(),
        'primary_tab_sessions': queryset.filter(is_primary_tab=True).count()
    }

def _calculate_security_insights(queryset):
    """Calculate security-related insights"""
    # IP address analysis
    ip_stats = queryset.values('ip_address').annotate(
        session_count=Count('id'),
        user_count=Count('user', distinct=True)
    ).order_by('-session_count')

    # Office vs Remote sessions
    office_sessions = queryset.filter(ip_address__in=UserSession.OFFICE_IPS).count()
    total_sessions = queryset.count()
    remote_sessions = total_sessions - office_sessions

    # Multiple IPs per user
    users_multiple_ips = queryset.values('user').annotate(
        ip_count=Count('ip_address', distinct=True)
    ).filter(ip_count__gt=2)

    # Sessions with security incidents
    sessions_with_incidents = queryset.filter(
        security_incidents__isnull=False
    ).count()

    # Device fingerprint changes
    fingerprint_changes = queryset.values('user').annotate(
        fingerprint_count=Count('session_fingerprint', distinct=True)
    ).filter(fingerprint_count__gt=1)

    office_percentage = safe_percentage(office_sessions, total_sessions)

    return {
        'unique_ips': len(ip_stats),
        'office_sessions': office_sessions,
        'remote_sessions': remote_sessions,
        'office_percentage': office_percentage,
        'users_with_multiple_ips': users_multiple_ips.count(),
        'sessions_with_security_incidents': sessions_with_incidents,
        'users_with_fingerprint_changes': fingerprint_changes.count(),
        'top_ips_by_usage': list(ip_stats[:10])
    }

def _calculate_location_insights(queryset):
    """Calculate location and environment insights"""
    location_stats = queryset.values('location').annotate(
        count=Count('id'),
        unique_users=Count('user', distinct=True),
        avg_productivity=Avg('productivity_score'),
        avg_session_duration=Avg('session_duration')
    ).order_by('-count')

    # Time zone analysis with safe datetime handling
    timezone_patterns = {}
    for session in queryset.select_related('user'):
        try:
            login_time_ist = to_ist(session.login_time)
            if login_time_ist:
                login_hour = login_time_ist.hour
                timezone_patterns[login_hour] = timezone_patterns.get(login_hour, 0) + 1
        except (AttributeError, TypeError):
            continue

    # Peak hours
    peak_hours = sorted(timezone_patterns.items(), key=lambda x: x[1], reverse=True)[:5]

    return {
        'location_breakdown': list(location_stats),
        'peak_login_hours': peak_hours,
        'timezone_distribution': timezone_patterns
    }

def _calculate_performance_insights(queryset):
    """Calculate performance-related insights"""
    # Session quality distribution
    quality_stats = queryset.values('session_quality').annotate(
        count=Count('id')
    ).order_by('-count')

    # Performance metrics - handle JSONField aggregation safely
    try:
        # For JSONField aggregation, we need to handle cases where the field might contain
        # non-JSON data (like raw floats) or be null
        performance_metrics = {
            'avg_page_load_time': 0,
            'avg_mouse_movements': 0,
            'avg_keyboard_events': 0,
            'avg_scroll_events': 0
        }

        # Calculate averages manually to avoid JSON aggregation issues
        valid_sessions = queryset.exclude(performance_metrics__isnull=True)

        if valid_sessions.exists():
            # Calculate mouse movements average safely
            mouse_movements_sum = 0
            mouse_movements_count = 0

            for session in queryset:
                if session.mouse_movements is not None:
                    mouse_movements_sum += session.mouse_movements
                    mouse_movements_count += 1

            if mouse_movements_count > 0:
                performance_metrics['avg_mouse_movements'] = mouse_movements_sum / mouse_movements_count

            # For performance_metrics JSONField, extract numeric values safely
            page_load_times = []
            keyboard_events = []
            scroll_events = []

            for session in valid_sessions:
                try:
                    # Safely parse performance_metrics JSON field
                    perf_data = safe_json_parse(session.performance_metrics, {})

                    # Extract page load time if it exists
                    if isinstance(perf_data, dict):
                        if 'page_load_time' in perf_data:
                            page_load_times.append(float(perf_data['page_load_time']))
                    elif isinstance(perf_data, (int, float)):
                        # If it's a raw number, treat it as page load time
                        page_load_times.append(float(perf_data))

                    # Extract keyboard events safely
                    keyboard_data = safe_json_parse(session.keyboard_events, [])
                    if isinstance(keyboard_data, list):
                        keyboard_events.append(len(keyboard_data))
                    elif isinstance(keyboard_data, (int, float)):
                        keyboard_events.append(int(keyboard_data))

                    # Extract scroll events safely
                    scroll_data = safe_json_parse(session.scroll_events, [])
                    if isinstance(scroll_data, list):
                        scroll_events.append(len(scroll_data))
                    elif isinstance(scroll_data, (int, float)):
                        scroll_events.append(int(scroll_data))

                except (ValueError, TypeError, AttributeError):
                    # Skip sessions with invalid data
                    continue

            # Calculate averages from collected data
            performance_metrics['avg_page_load_time'] = safe_average(page_load_times)
            performance_metrics['avg_keyboard_events'] = safe_average(keyboard_events)
            performance_metrics['avg_scroll_events'] = safe_average(scroll_events)

    except Exception as e:
        logger.error(f"Error calculating performance metrics: {str(e)}")
        performance_metrics = {
            'avg_page_load_time': 0,
            'avg_mouse_movements': 0,
            'avg_keyboard_events': 0,
            'avg_scroll_events': 0
        }

    # Slow sessions (longer than average + 1 std dev)
    avg_duration = queryset.aggregate(Avg('session_duration'))['session_duration__avg'] or 0
    slow_sessions = queryset.filter(session_duration__gt=avg_duration * 1.5).count()

    slow_sessions_percentage = safe_percentage(slow_sessions, queryset.count())

    return {
        'quality_distribution': list(quality_stats),
        'avg_page_load_time': round(performance_metrics['avg_page_load_time'] or 0, 2),
        'avg_mouse_movements': round(performance_metrics['avg_mouse_movements'] or 0, 1),
        'avg_interactions_per_session': round(
            (performance_metrics['avg_keyboard_events'] or 0) +
            (performance_metrics['avg_scroll_events'] or 0), 1
        ),
        'slow_sessions_count': slow_sessions,
        'slow_sessions_percentage': slow_sessions_percentage
    }

def _calculate_quality_insights(queryset):
    """Calculate session quality insights"""
    # Quality score distribution
    quality_ranges = {
        'excellent': queryset.filter(productivity_score__gte=80).count(),
        'good': queryset.filter(productivity_score__gte=60, productivity_score__lt=80).count(),
        'average': queryset.filter(productivity_score__gte=40, productivity_score__lt=60).count(),
        'poor': queryset.filter(productivity_score__lt=40).count()
    }

    # Engagement patterns
    high_engagement = queryset.filter(engagement_score__gte=70).count()
    low_engagement = queryset.filter(engagement_score__lt=30).count()

    # Idle time analysis
    high_idle_sessions = queryset.filter(
        idle_time__gt=timedelta(minutes=30)
    ).count()

    avg_quality = queryset.aggregate(Avg('productivity_score'))['productivity_score__avg'] or 0

    return {
        'quality_distribution': quality_ranges,
        'high_engagement_sessions': high_engagement,
        'low_engagement_sessions': low_engagement,
        'high_idle_sessions': high_idle_sessions,
        'avg_quality_score': round(avg_quality, 1)
    }

def _analyze_tab_focus_patterns(session):
    """Analyze tab focus patterns for a session"""
    page_views = safe_json_parse(session.page_views, [])

    if not page_views:
        return {
            'total_pages': 0,
            'unique_domains': 0,
            'focus_time_distribution': {},
            'most_visited_page': None
        }

    page_counter = Counter()
    domain_counter = Counter()

    for page_view in page_views:
        if isinstance(page_view, dict):
            url = page_view.get('url', '')
            page_counter[url] += 1

            # Extract domain
            try:
                from urllib.parse import urlparse
                domain = urlparse(url).netloc
                domain_counter[domain] += 1
            except:
                pass

    return {
        'total_pages': len(page_views),
        'unique_pages': len(page_counter),
        'unique_domains': len(domain_counter),
        'most_visited_page': page_counter.most_common(1)[0] if page_counter else None,
        'top_domains': domain_counter.most_common(5)
    }

def _calculate_hourly_activity(session):
    """Calculate hourly activity distribution for a session"""
    hourly_data = {i: 0 for i in range(24)}

    page_views = safe_json_parse(session.page_views, [])

    if page_views:
        for page_view in page_views:
            if isinstance(page_view, dict):
                try:
                    timestamp_str = page_view.get('timestamp', '')
                    if timestamp_str:
                        timestamp = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                        timestamp_ist = to_ist(timestamp)
                        if timestamp_ist:
                            hour = timestamp_ist.hour
                            hourly_data[hour] += 1
                except:
                    pass

    return hourly_data

def _generate_activity_heatmap(sessions, target_date):
    """Generate activity heatmap data for a specific date"""
    heatmap_data = []

    # Create 24-hour grid with 15-minute intervals
    for hour in range(24):
        hour_data = {'hour': hour, 'intervals': []}
        for quarter in range(4):  # 15-minute intervals
            interval_start = hour * 60 + quarter * 15  # minutes from start of day
            activity_count = 0

            # Count activities in this interval across all sessions
            for session in sessions:
                page_views = safe_json_parse(session.page_views, [])

                if page_views:
                    for page_view in page_views:
                        if isinstance(page_view, dict):
                            try:
                                timestamp_str = page_view.get('timestamp', '')
                                if timestamp_str:
                                    timestamp = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                                    timestamp_ist = to_ist(timestamp)

                                    if timestamp_ist and timestamp_ist.date() == target_date:
                                        minutes_from_start = timestamp_ist.hour * 60 + timestamp_ist.minute
                                        if interval_start <= minutes_from_start < interval_start + 15:
                                            activity_count += 1
                            except:
                                pass

            hour_data['intervals'].append({
                'quarter': quarter,
                'activity_count': activity_count,
                'time_label': f"{hour:02d}:{quarter*15:02d}"
            })

        heatmap_data.append(hour_data)

    return heatmap_data

def _get_live_activity_data(request):
    """Get live activity data for AJAX requests"""
    current_time_utc = timezone.now()

    # Get currently active sessions
    active_sessions = UserSession.objects.filter(
        is_active=True,
        last_activity__gte=current_time_utc - timedelta(minutes=10)
    ).select_related('user').order_by('-last_activity')[:20]

    activity_data = []
    for session in active_sessions:
        time_since_activity = (current_time_utc - session.last_activity).total_seconds() / 60

        last_activity_ist = to_ist(session.last_activity)
        last_activity_str = last_activity_ist.strftime('%H:%M:%S') if last_activity_ist else 'Unknown'

        activity_data.append({
            'user_id': session.user.id,
            'username': session.user.username,
            'last_activity': last_activity_str,
            'session_duration': round((current_time_utc - session.login_time).total_seconds() / 60, 1),
            'activity_status': 'active' if time_since_activity <= 2 else 'idle',
            'location': session.location,
            'device_type': session.device_type,
            'current_url': session.tab_url or 'Unknown'
        })

    current_time_ist = get_current_time_ist()
    timestamp_str = current_time_ist.strftime('%Y-%m-%d %H:%M:%S') if current_time_ist else 'Unknown'

    return {
        'status': 'success',
        'data': activity_data,
        'timestamp': timestamp_str
    }

def _get_overview_metrics(request, date_from, date_to):
    """Get overview metrics for AJAX requests"""
    filters = _parse_session_filters(request)
    queryset = _get_filtered_sessions(filters)
    metrics = _calculate_overview_metrics(queryset, timezone.now())

    return {
        'status': 'success',
        'data': metrics
    }

def _get_session_analytics(request, date_from, date_to, filters):
    """Get session analytics data for AJAX requests"""
    queryset = _get_filtered_sessions(filters)

    analytics_data = {
        'user_insights': _calculate_user_insights(queryset, date_from, date_to)[:10],
        'tab_analysis': _calculate_tab_behavior_analysis(queryset),
        'security_insights': _calculate_security_insights(queryset),
        'performance_insights': _calculate_performance_insights(queryset)
    }

    return {
        'status': 'success',
        'data': analytics_data
    }

def _export_session_data(request, filters):
    """Export session data to Excel"""
    try:
        queryset = _get_filtered_sessions(filters)
        export_format = request.GET.get('format', 'excel')

        if export_format == 'csv':
            return _export_to_csv(queryset, filters)
        else:
            return _export_to_excel(queryset, filters)

    except Exception as e:
        logger.error(f"Error exporting session data: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': 'Failed to export data'
        })

def _export_to_csv(queryset, filters):
    """Export session data to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="user_sessions_{filters["date_from"]}_to_{filters["date_to"]}.csv"'

    writer = csv.writer(response)

    # Write header
    writer.writerow([
        'User', 'Login Time (IST)', 'Logout Time (IST)', 'Duration (minutes)',
        'Location', 'Device Type', 'IP Address', 'Status', 'Productivity Score',
        'Engagement Score', 'Tab Switches', 'Page Views', 'Working Hours',
        'Idle Time (minutes)', 'Session Quality'
    ])

    # Write data
    for session in queryset.select_related('user'):
        login_time_ist = to_ist(session.login_time)
        logout_time_ist = to_ist(session.logout_time) if session.logout_time else None

        writer.writerow([
            session.user.username,
            login_time_ist.strftime('%Y-%m-%d %H:%M:%S'),
            logout_time_ist.strftime('%Y-%m-%d %H:%M:%S') if logout_time_ist else 'Active',
            session.session_duration or 0,
            session.location or 'Unknown',
            session.device_type or 'Unknown',
            session.ip_address or 'Unknown',
            'Active' if session.is_active else 'Inactive',
            session.productivity_score or 0,
            session.engagement_score or 0,
            session.tab_switches or 0,
            len(session.page_views) if session.page_views else 0,
            session.working_hours.total_seconds() / 3600 if session.working_hours else 0,
            session.idle_time.total_seconds() / 60 if session.idle_time else 0,
            session.session_quality or 'Unknown'
        ])

    return response

def _export_to_excel(queryset, filters):
    """Export session data to Excel with formatting"""
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Sessions"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = [
        'User', 'Login Time (IST)', 'Logout Time (IST)', 'Duration (minutes)',
        'Location', 'Device Type', 'IP Address', 'Status', 'Productivity Score',
        'Engagement Score', 'Tab Switches', 'Page Views', 'Working Hours',
        'Idle Time (minutes)', 'Session Quality'
    ]

    # Write and format headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data
    for row, session in enumerate(queryset.select_related('user'), 2):
        login_time_ist = to_ist(session.login_time)
        logout_time_ist = to_ist(session.logout_time) if session.logout_time else None

        data = [
            session.user.username,
            login_time_ist.strftime('%Y-%m-%d %H:%M:%S'),
            logout_time_ist.strftime('%Y-%m-%d %H:%M:%S') if logout_time_ist else 'Active',
            session.session_duration or 0,
            session.location or 'Unknown',
            session.device_type or 'Unknown',
            session.ip_address or 'Unknown',
            'Active' if session.is_active else 'Inactive',
            session.productivity_score or 0,
            session.engagement_score or 0,
            session.tab_switches or 0,
            len(session.page_views) if session.page_views else 0,
            round(session.working_hours.total_seconds() / 3600, 2) if session.working_hours else 0,
            round(session.idle_time.total_seconds() / 60, 2) if session.idle_time else 0,
            session.session_quality or 'Unknown'
        ]

        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="user_sessions_{filters["date_from"]}_to_{filters["date_to"]}.xlsx"'

    return response

@login_required
@user_passes_test(is_admin)
def session_analytics_api(request):
    """
    API endpoint for session analytics data
    """
    try:
        filters = _parse_session_filters(request)
        queryset = _get_filtered_sessions(filters)

        analytics_type = request.GET.get('type', 'overview')

        if analytics_type == 'overview':
            data = _calculate_overview_metrics(queryset, timezone.now())
        elif analytics_type == 'users':
            data = _calculate_user_insights(queryset, filters['date_from'], filters['date_to'])
        elif analytics_type == 'security':
            data = _calculate_security_insights(queryset)
        elif analytics_type == 'performance':
            data = _calculate_performance_insights(queryset)
        elif analytics_type == 'quality':
            data = _calculate_quality_insights(queryset)
        elif analytics_type == 'tabs':
            data = _calculate_tab_behavior_analysis(queryset)
        elif analytics_type == 'locations':
            data = _calculate_location_insights(queryset)
        else:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid analytics type'
            })

        return JsonResponse({
            'status': 'success',
            'data': data,
            'filters': filters,
            'timestamp': get_current_time_ist().isoformat()
        })

    except Exception as e:
        logger.error(f"Error in session_analytics_api: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': 'Failed to fetch analytics data'
        })

@login_required
@user_passes_test(is_admin)
def terminate_session(request, session_id):
    """
    Terminate a specific user session
    """
    if request.method != 'POST':
        return JsonResponse({
            'status': 'error',
            'message': 'POST method required'
        })

    try:
        session = get_object_or_404(UserSession, id=session_id)

        if not session.is_active:
            return JsonResponse({
                'status': 'error',
                'message': 'Session is already inactive'
            })

        # Mark session as inactive
        session.is_active = False
        session.logout_time = timezone.now()
        session.session_duration = (session.logout_time - session.login_time).total_seconds() / 60
        session.save()

        logger.info(f"Session {session_id} terminated by admin {request.user.username}")

        return JsonResponse({
            'status': 'success',
            'message': f'Session for {session.user.username} has been terminated'
        })

    except Exception as e:
        logger.error(f"Error terminating session {session_id}: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': 'Failed to terminate session'
        })

@login_required
@user_passes_test(is_admin)
def bulk_session_actions(request):
    """
    Perform bulk actions on multiple sessions
    """
    if request.method != 'POST':
        return JsonResponse({
            'status': 'error',
            'message': 'POST method required'
        })

    try:
        data = json.loads(request.body)
        action = data.get('action')
        session_ids = data.get('session_ids', [])

        if not session_ids:
            return JsonResponse({
                'status': 'error',
                'message': 'No sessions selected'
            })

        sessions = UserSession.objects.filter(id__in=session_ids)

        if action == 'terminate':
            active_sessions = sessions.filter(is_active=True)
            count = active_sessions.count()

            active_sessions.update(
                is_active=False,
                logout_time=timezone.now()
            )

            # Update session durations
            for session in active_sessions:
                session.session_duration = (timezone.now() - session.login_time).total_seconds() / 60
                session.save()

            logger.info(f"Bulk terminated {count} sessions by admin {request.user.username}")

            return JsonResponse({
                'status': 'success',
                'message': f'Successfully terminated {count} active sessions'
            })

        elif action == 'export':
            # This would trigger an export of selected sessions
            filters = {
                'date_from': datetime.now().date() - timedelta(days=30),
                'date_to': datetime.now().date(),
                'user_filter': None,
                'location_filter': None,
                'device_filter': None,
                'status_filter': None,
                'ip_type_filter': None,
                'quality_filter': None
            }

            return _export_to_excel(sessions, filters)

        else:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid action'
            })

    except json.JSONDecodeError:
        return JsonResponse({
            'status': 'error',
            'message': 'Invalid JSON data'
        })
    except Exception as e:
        logger.error(f"Error in bulk_session_actions: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': 'Failed to perform bulk action'
        })



''' ---------------------------------------- TIMESHEET AREA ---------------------------------------- '''
@login_required
@user_passes_test(is_employee)  # Only allow employees to access this view
def timesheet_view(request):
    if request.method == "POST":
        try:
            # Get the submitted data from the form
            week_start_date = request.POST.get('week_start_date')
            project_ids = request.POST.getlist('project_id[]')
            task_names = request.POST.getlist('task_name[]')
            task_descriptions = request.POST.getlist('task_description[]')
            hours = request.POST.getlist('hours[]')

            # Validate that all lists are the same length
            if len(project_ids) != len(task_names) or len(task_names) != len(hours) or len(task_descriptions) != len(hours):
                messages.error(request, "All form fields should have the same number of entries.")
                return redirect('aps_employee:timesheet')

            # Create the Timesheet objects and save them to the database
            for project_id, task_name, task_description, hour in zip(project_ids, task_names, task_descriptions, hours):
                try:
                    # Convert hours to float and validate
                    hours_float = float(hour)

                    # Get the project
                    project = Project.objects.get(id=project_id)

                    # Create new timesheet entry
                    timesheet = Timesheet(
                        user=request.user,
                        week_start_date=week_start_date,
                        project=project,
                        task_name=task_name,
                        task_description=task_description,
                        hours=hours_float
                    )

                    # This will run the clean method which validates hours limits
                    timesheet.save()

                except ValidationError as ve:
                    messages.error(request, f"Validation error: {str(ve)}")
                    return redirect('aps_employee:timesheet')
                except Project.DoesNotExist:
                    messages.error(request, f"Project with ID {project_id} does not exist.")
                    return redirect('aps_employee:timesheet')

            # Display success message
            messages.success(request, "Timesheet submitted successfully!")
            return redirect('aps_employee:timesheet')

        except Exception as e:
            # If an error occurs, show an error message
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect('aps_employee:timesheet')

    else:
        # If it's a GET request, show the current timesheet history
        today = timezone.now().date()

        # Calculate the start of the current week (Monday)
        current_week_start = today - timedelta(days=today.weekday())

        # Fetch the timesheet history for the logged-in employee, ordered by week start date and version
        timesheet_history = Timesheet.objects.filter(user=request.user).order_by('-week_start_date', '-version')

        # Fetch the list of projects the user is assigned to using the ProjectAssignment model
        assigned_projects = Project.objects.filter(projectassignment__user=request.user, projectassignment__is_active=True)

        # Get weekly hours summary
        week_end_date = current_week_start + timedelta(days=6)
        weekly_hours = Timesheet.objects.filter(
            user=request.user,
            week_start_date__gte=current_week_start,
            week_start_date__lte=week_end_date
        ).aggregate(total_hours=Sum('hours'))['total_hours'] or 0

        remaining_hours = 45 - weekly_hours  # Updated to 45 hours weekly limit

        # Render the timesheet page with the data
        return render(request, 'components/employee/timesheet.html', {
            'today': today,
            'current_week_start': current_week_start,
            'timesheet_history': timesheet_history,
            'assigned_projects': assigned_projects,
            'weekly_hours': weekly_hours,
            'remaining_hours': remaining_hours
        })

@login_required
@user_passes_test(is_employee)
def get_timesheet_details(request, week_start_date):
    """
    View to get detailed timesheet information for a specific week.
    Returns JSON data for AJAX requests to populate the timesheet modal.
    """
    try:
        # Try to parse the date with multiple possible formats
        start_date = None
        date_formats = ['%Y-%m-%d', '%B %d, %Y', '%m/%d/%Y', '%d-%m-%Y']

        for date_format in date_formats:
            try:
                start_date = timezone.datetime.strptime(week_start_date, date_format).date()
                break
            except ValueError:
                continue

        if start_date is None:
            raise ValueError(f"Could not parse date '{week_start_date}' with any supported format")

        # Get all timesheet entries for the specified week
        timesheet_entries = Timesheet.objects.filter(
            user=request.user,
            week_start_date=start_date
        ).select_related('project')

        if not timesheet_entries.exists():
            return JsonResponse({
                'success': False,
                'message': 'No timesheet entries found for this week.'
            })

        # Format the data for the response
        entries_data = []
        total_hours = 0

        for entry in timesheet_entries:
            entries_data.append({
                'id': entry.id,
                'project_name': entry.project.name,
                'project_id': entry.project.id,
                'task_name': entry.task_name,
                'task_description': entry.task_description,
                'hours': entry.hours,
                'approval_status': entry.approval_status,
                'comments': entry.manager_comments or '',
                'submitted_at': entry.submitted_at.isoformat() if hasattr(entry, 'submitted_at') else None,
                'last_modified': entry.reviewed_at.isoformat() if hasattr(entry, 'reviewed_at') else None
            })
            total_hours += entry.hours

        # Return the formatted data
        return JsonResponse({
            'success': True,
            'week_start': start_date.strftime('%Y-%m-%d'),
            'week_end': (start_date + timedelta(days=6)).strftime('%Y-%m-%d'),
            'total_hours': total_hours,
            'entries': entries_data
        })

    except Exception as e:
        logger.error(f"Error retrieving timesheet details: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f"An error occurred: {str(e)}"
        })

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Timesheet, Project, ProjectAssignment
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from datetime import timedelta
from django.http import JsonResponse
import logging
logger = logging.getLogger(__name__)

@login_required
@user_passes_test(is_manager)
def manager_view_timesheets(request):
    """
    View for managers to review and manage timesheets for their projects.
    Includes filtering, search, and pagination functionality.
    """
    # Get filter parameters from request
    time_filter = request.GET.get('time-filter', '7')
    search_query = request.GET.get('search', '')

    # Default to 7 days if custom filter not specified
    if time_filter != 'custom':
        filter_days = int(time_filter)
        start_date = timezone.now().date() - timedelta(days=filter_days)
    else:
        # Handle custom date range if implemented
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            start_date = timezone.now().date() - timedelta(days=30)  # Default fallback

    # Get projects managed by the current user
    managed_projects = Project.objects.filter(
        projectassignment__user=request.user,
        projectassignment__role_in_project='Manager',
        projectassignment__is_active=True
    ).distinct()

    managed_project_ids = managed_projects.values_list('id', flat=True)

    # Base queryset with prefetching for optimization
    timesheets = Timesheet.objects.select_related('project', 'user').filter(
        project_id__in=managed_project_ids
    )

    # Apply date filter
    timesheets = timesheets.filter(week_start_date__gte=start_date)

    # Apply search filter if provided
    if search_query:
        timesheets = timesheets.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(project__name__icontains=search_query) |
            Q(task_name__icontains=search_query) |
            Q(task_description__icontains=search_query)
        )

    # Calculate dashboard statistics
    total_hours = timesheets.aggregate(Sum('hours'))['hours__sum'] or 0
    active_projects = timesheets.values('project').distinct().count()
    pending_approvals = timesheets.filter(approval_status='Pending').count()
    completion_rate = calculate_completion_rate(timesheets)

    # Group timesheets by status for easier filtering in the template
    status_counts = {
        'pending': timesheets.filter(approval_status='Pending').count(),
        'approved': timesheets.filter(approval_status='Approved').count(),
        'rejected': timesheets.filter(approval_status='Rejected').count(),
        'clarification': timesheets.filter(approval_status='Clarification_Requested').count()
    }

    # Order by most recent first, then by employee name
    timesheets = timesheets.order_by('-week_start_date', 'user__first_name')

    # Paginate results
    paginator = Paginator(timesheets, 20)  # 20 items per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Prepare context for template
    context = {
        'page_obj': page_obj,
        'total_hours': total_hours,
        'active_projects': active_projects,
        'completion_rate': completion_rate,
        'pending_approvals': pending_approvals,
        'time_filter': time_filter,
        'search_query': search_query,
        'status_counts': status_counts,
        'managed_projects': managed_projects,
    }

    return render(request, 'components/manager/view_timesheets.html', context)


@login_required
@user_passes_test(is_manager)
def bulk_update_timesheet(request):
    """
    Handle bulk actions on timesheets (approve, reject, request clarification).
    Only allows managers to update timesheets for projects they manage.
    """
    if request.method != 'POST':
        messages.error(request, 'Invalid request method')
        return redirect('aps_manager:view_timesheets')

    # Get parameters from request
    timesheet_ids = request.POST.getlist('selected_timesheets[]')
    action = request.POST.get('action')
    rejection_reason = request.POST.get('rejection_reason')
    manager_comments = request.POST.get('manager_comments', '')

    # Validate input
    if not timesheet_ids:
        messages.error(request, 'No timesheets selected')
        return redirect('aps_manager:view_timesheets')

    if action not in ['approve', 'reject', 'request_clarification']:
        messages.error(request, 'Invalid action')
        return redirect('aps_manager:view_timesheets')

    # Map action to status
    status_map = {
        'approve': 'Approved',
        'reject': 'Rejected',
        'request_clarification': 'Clarification_Requested'
    }

    try:
        # Get projects managed by the current user
        managed_projects = Project.objects.filter(
            projectassignment__user=request.user,
            projectassignment__role_in_project='Manager',
            projectassignment__is_active=True
        ).values_list('id', flat=True)

        # Restrict timesheets to manager's projects
        timesheets = Timesheet.objects.filter(
            id__in=timesheet_ids,
            project_id__in=managed_projects
        )

        if not timesheets.exists():
            messages.error(request, 'You are not authorized to update the selected timesheets')
            return redirect('aps_manager:view_timesheets')

        # Validate rejection reason if rejecting
        if action == 'reject' and not rejection_reason:
            messages.error(request, 'Rejection reason is required')
            return redirect('aps_manager:view_timesheets')

        # Update timesheets
        update_count = 0
        for timesheet in timesheets:
            timesheet.approval_status = status_map[action]

            if action == 'reject':
                timesheet.rejection_reason = rejection_reason

            if manager_comments:
                timesheet.manager_comments = manager_comments

            # reviewed_at will be updated by the signal
            timesheet.save()
            update_count += 1

        # Create success message based on action
        action_display = 'approved' if action == 'approve' else 'rejected' if action == 'reject' else 'requested clarification for'
        messages.success(request, f'Successfully {action_display} {update_count} timesheet{"s" if update_count != 1 else ""}.')

        return redirect('aps_manager:view_timesheets')
    except Exception as e:
        logger.error(f"Error processing timesheets: {e}", exc_info=True)
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('aps_manager:view_timesheets')


@login_required
@user_passes_test(is_manager)
def timesheet_detail(request, timesheet_id):
    """
    View a single timesheet in detail, with history of previous versions.
    Only accessible to managers of the project.
    """
    try:
        # Get the timesheet with related data
        timesheet = Timesheet.objects.select_related('user', 'project').get(id=timesheet_id)

        # Check if user is manager of this project
        is_project_manager = ProjectAssignment.objects.filter(
            user=request.user,
            project=timesheet.project,
            role_in_project='Manager',
            is_active=True
        ).exists()

        if not is_project_manager:
            messages.error(request, "You don't have permission to view this timesheet.")
            return redirect('aps_manager:view_timesheets')

        # Get version history if this is a resubmission
        version_history = []
        if timesheet.original_submission_id:
            version_history = Timesheet.objects.filter(
                Q(id=timesheet.original_submission_id) |
                Q(original_submission_id=timesheet.original_submission_id)
            ).order_by('version')

        context = {
            'timesheet': timesheet,
            'version_history': version_history
        }

        return render(request, 'components/manager/timesheet_detail.html', context)

    except Timesheet.DoesNotExist:
        messages.error(request, "Timesheet not found.")
        return redirect('aps_manager:view_timesheets')


def calculate_completion_rate(timesheets):
    """Calculate the percentage of approved timesheets."""
    total_count = timesheets.count()
    if total_count == 0:
        return 0

    approved_count = timesheets.filter(approval_status='Approved').count()
    completion_rate = (approved_count / total_count) * 100
    return round(completion_rate, 1)


''' ---------------------------------------- LEAVE AREA ---------------------------------------- '''

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.db.models import Sum, Count, Q
from django.contrib import messages
from django.utils import timezone
from datetime import datetime, timedelta
import csv
import json

from .models import (
    LeavePolicy, LeaveType, LeaveAllocation, UserLeaveBalance,
    LeaveRequest, CompOffRequest,
)
from .forms import (
    LeavePolicyForm, LeaveTypeForm, LeaveAllocationForm,
    LeaveRequestForm, CompOffRequestForm, BulkLeaveAllocationForm,
    LeaveApprovalForm, LeaveRejectionForm
)

# Permission utilities
def is_hr(user):
    return user.groups.filter(name="HR").exists()

def is_manager(user):
    return user.groups.filter(name="Manager").exists()

def is_admin(user):
    return user.groups.filter(name="Admin").exists()

def is_finance(user):
    return user.groups.filter(name="Finance").exists()
def is_management(user):
    return {'is_management': request.user.groups.filter(name="Management").exists()} if request.user.is_authenticated else {'is_management': False}

def is_backoffice(user):
    return user.groups.filter(name="Backoffice").exists()

def is_client(user):
    return user.groups.filter(name="Client").exists()

def is_employee(user):
    return user.groups.filter(name="Employee").exists()

def can_approve_leave(user):
    """Check if user can approve leave requests"""
    return any([
        is_hr(user),
        is_manager(user),
        is_admin(user),
    ])

@login_required
@user_passes_test(is_hr)
def hr_leave_view(request):
    """View for HR to manage all leave-related functions"""

    # Get counts for quick stats
    total_leave_types = LeaveType.objects.count()
    total_leave_policies = LeavePolicy.objects.count()
    pending_leave_requests = LeaveRequest.objects.filter(status='Pending').count()

    # Get recent leave requests
    recent_leave_requests = LeaveRequest.objects.all().select_related(
        'user', 'leave_type', 'approver'
    ).order_by('-created_at')[:5]

    # Get HR's own leave balances
    user_balances = UserLeaveBalance.objects.filter(
        user=request.user,
        year=timezone.now().year
    ).select_related('leave_type')

    context = {
        'total_leave_types': total_leave_types,
        'total_leave_policies': total_leave_policies,
        'pending_leave_requests': pending_leave_requests,
        'recent_leave_requests': recent_leave_requests,
        'user_balances': user_balances,

        # URLs for leave type management
        'leave_type_urls': {
            'list': reverse('aps_leave:leave_type_list'),
            'create': reverse('aps_leave:leave_type_create'),
        },

        # URLs for leave policy management
        'leave_policy_urls': {
            'list': reverse('aps_leave:leave_policy_list'),
            'create': reverse('aps_leave:leave_policy_create'),
        },

        # URLs for HR's own leave management
        'leave_request_urls': {
            'create': reverse('aps_leave:leave_request_create'),
            'list': reverse('aps_leave:leave_request_list'),
        },
        'leave_balance_urls': {
            'bulk_create': reverse('aps_leave:bulk_leave_balance_create'),
            'list': reverse('aps_leave:leave_request_list'),
        },
    }

    return render(request, 'components/leave_management/hr_leave_dashboard.html', context)

@login_required
@user_passes_test(is_finance)
def finance_leave_view(request):
    """View for Finance to manage leave-related functions"""

    # Get approved leave requests for reporting
    approved_leaves = LeaveRequest.objects.filter(
        status='Approved'
    ).select_related(
        'user', 'leave_type', 'approver', 'user__profile'
    ).order_by('-start_date')[:10]

    # Get leave stats by department
    department_stats = LeaveRequest.objects.filter(
        status='Approved'
    ).values('user__profile__department').annotate(
        total_leaves=Count('id'),
        total_days=Sum('leave_days')
    )

    context = {
        'approved_leaves': approved_leaves,
        'department_stats': department_stats
    }

    return render(request, 'components/leave_management/finance_leave_dashboard.html', context)

@login_required
@user_passes_test(is_management)
def management_leave_view(request):
    """View for Management to oversee leave operations"""

    # Get overall leave stats
    total_employees = User.objects.filter(is_active=True).count()
    employees_on_leave = LeaveRequest.objects.filter(
        status='Approved',
        start_date__lte=timezone.now().date(),
        end_date__gte=timezone.now().date()
    ).count()

    # Get department-wise leave stats
    department_stats = LeaveRequest.objects.filter(
        status='Approved'
    ).values('user__profile__department').annotate(
        total_requests=Count('id'),
        total_days=Sum('leave_days')
    )

    # Get recent leave requests
    recent_requests = LeaveRequest.objects.all().select_related(
        'user', 'leave_type', 'approver', 'user__profile'
    ).order_by('-created_at')[:10]

    context = {
        'total_employees': total_employees,
        'employees_on_leave': employees_on_leave,
        'department_stats': department_stats,
        'recent_requests': recent_requests
    }

    return render(request, 'components/leave_management/management_leave_dashboard.html', context)



@login_required
@user_passes_test(is_manager)
def manager_leave_view(request):
    """View for Managers to handle all leave requests"""

    # Get all recent leave requests
    recent_requests = LeaveRequest.objects.all().select_related(
        'user', 'leave_type', 'approver'
    ).order_by('-created_at')[:10]

    # Get all pending approvals
    pending_approvals = LeaveRequest.objects.filter(
        status='Pending'
    ).select_related(
        'user', 'leave_type'
    ).order_by('-created_at')

    # Get all upcoming approved leaves
    upcoming_leaves = LeaveRequest.objects.filter(
        status='Approved',
        start_date__gte=timezone.now().date()
    ).select_related(
        'user', 'leave_type'
    ).order_by('start_date')[:10]

    context = {
        'recent_requests': recent_requests,
        'pending_approvals': pending_approvals,
        'upcoming_leaves': upcoming_leaves,
    }

    return render(request, 'components/leave_management/manager_leave_dashboard.html', context)

@login_required
def employee_leave_view(request):
    """View for employees to manage their leave requests"""

    # Get user's leave requests
    user_requests = LeaveRequest.objects.filter(
        user=request.user
    ).select_related(
        'leave_type', 'approver'
    ).order_by('-created_at')[:10]

    # Get pending requests
    pending_requests = LeaveRequest.objects.filter(
        user=request.user,
        status='Pending'
    ).select_related('leave_type')

    # Get upcoming approved leaves
    upcoming_leaves = LeaveRequest.objects.filter(
        user=request.user,
        status='Approved',
        start_date__gte=timezone.now().date()
    ).select_related('leave_type').order_by('start_date')[:10]

    # Get leave balances for current year
    leave_balances = UserLeaveBalance.objects.filter(
        user=request.user,
        year=timezone.now().year
    ).select_related('leave_type')

    context = {
        'user_requests': user_requests,
        'pending_requests': pending_requests,
        'upcoming_leaves': upcoming_leaves,
        'leave_balances': leave_balances
    }

    return render(request, 'components/leave_management/employee_leave_dashboard.html', context)


# ============= 1. LEAVE TYPE MANAGEMENT VIEWS (HR ONLY) =============

@login_required
@user_passes_test(is_hr)
def leave_type_list(request):
    """View to list all leave types"""
    leave_types = LeaveType.objects.all().order_by('name')
    return render(request, 'components/leave_management/leave_type_list.html', {
        'leave_types': leave_types
    })

@login_required
@user_passes_test(is_hr)
def leave_type_create(request):
    """View to create a new leave type"""
    if request.method == 'POST':
        form = LeaveTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Leave type created successfully!')
            return redirect('aps_leave:leave_type_list')
    else:
        form = LeaveTypeForm()

    return render(request, 'components/leave_management/leave_type_form.html', {
        'form': form,
        'action': 'Create'
    })

@login_required
@user_passes_test(is_hr)
def leave_type_update(request, pk):
    """View to update an existing leave type"""
    leave_type = get_object_or_404(LeaveType, pk=pk)

    if request.method == 'POST':
        form = LeaveTypeForm(request.POST, instance=leave_type)
        if form.is_valid():
            form.save()
            messages.success(request, 'Leave type updated successfully!')
            return redirect('aps_leave:leave_type_list')
    else:
        form = LeaveTypeForm(instance=leave_type)

    return render(request, 'components/leave_management/leave_type_form.html', {
        'form': form,
        'leave_type': leave_type,
        'action': 'Update'
    })

@login_required
@user_passes_test(is_hr)
def leave_type_delete(request, pk):
    """View to delete a leave type"""
    leave_type = get_object_or_404(LeaveType, pk=pk)

    if request.method == 'POST':
        # Check if there are any leave requests using this type
        if LeaveRequest.objects.filter(leave_type=leave_type).exists():
            messages.error(request, 'Cannot delete this leave type as it is being used in leave requests.')
            return redirect('aps_leave:leave_type_list')

        leave_type.delete()
        messages.success(request, 'Leave type deleted successfully!')
        return redirect('aps_leave:leave_type_list')

    return render(request, 'components/leave_management/leave_type_confirm_delete.html', {
        'leave_type': leave_type
    })

# ============= 2. LEAVE POLICY & ALLOCATION VIEWS (HR ONLY) =============

@login_required
@user_passes_test(is_hr)
def leave_policy_list(request):
    """View to list all leave policies"""
    policies = LeavePolicy.objects.all().order_by('group__name')
    return render(request, 'components/leave_management/leave_policy_list.html', {
        'policies': policies
    })

@login_required
@user_passes_test(is_hr)
def leave_policy_create(request):
    """View to create a new leave policy"""
    if request.method == 'POST':
        form = LeavePolicyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Leave policy created successfully!')
            return redirect('aps_leave:leave_policy_list')
    else:
        form = LeavePolicyForm()

    return render(request, 'components/leave_management/leave_policy_form.html', {
        'form': form,
        'action': 'Create'
    })

@login_required
@user_passes_test(is_hr)
def leave_policy_update(request, pk):
    """View to update an existing leave policy"""
    policy = get_object_or_404(LeavePolicy, pk=pk)

    if request.method == 'POST':
        form = LeavePolicyForm(request.POST, instance=policy)
        if form.is_valid():
            form.save()
            messages.success(request, 'Leave policy updated successfully!')
            return redirect('aps_leave:leave_policy_list')
    else:
        form = LeavePolicyForm(instance=policy)

    return render(request, 'components/leave_management/leave_policy_form.html', {
        'form': form,
        'policy': policy,
        'action': 'Update'
    })

@login_required
@user_passes_test(is_hr)
def leave_policy_delete(request, pk):
    """View to delete a leave policy"""
    policy = get_object_or_404(LeavePolicy, pk=pk)

    if request.method == 'POST':
        policy.delete()
        messages.success(request, 'Leave policy deleted successfully!')
        return redirect('aps_leave:leave_policy_list')

    return render(request, 'components/leave_management/leave_policy_confirm_delete.html', {
        'policy': policy
    })

@login_required
@user_passes_test(is_hr)
def leave_allocation_manage(request, policy_id):
    """View to manage leave allocations for a specific policy"""
    policy = get_object_or_404(LeavePolicy, pk=policy_id)
    leave_types = LeaveType.objects.filter(is_active=True)

    if request.method == 'POST':
        # Process the form data for multiple allocations
        for leave_type in leave_types:
            annual_days = request.POST.get(f'annual_days_{leave_type.id}', 0)
            carry_forward = request.POST.get(f'carry_forward_{leave_type.id}', 0)
            max_consecutive = request.POST.get(f'max_consecutive_{leave_type.id}', 0)
            advance_notice = request.POST.get(f'advance_notice_{leave_type.id}', 0)

            if annual_days:  # Only create/update if days are specified
                LeaveAllocation.objects.update_or_create(
                    policy=policy,
                    leave_type=leave_type,
                    defaults={
                        'annual_days': annual_days,
                        'carry_forward_limit': carry_forward,
                        'max_consecutive_days': max_consecutive,
                        'advance_notice_days': advance_notice
                    }
                )

        messages.success(request, 'Leave allocations updated successfully!')
        return redirect('aps_leave:leave_policy_list')

    # Get existing allocations as a queryset instead of dict
    allocations = LeaveAllocation.objects.filter(policy=policy)

    return render(request, 'components/leave_management/leave_allocation_form.html', {
        'policy': policy,
        'leave_types': leave_types,
        'allocations': allocations
    })

@login_required
@user_passes_test(is_hr)
def bulk_leave_balance_create(request):
    """View to create leave balances for all users based on their policy"""
    from django import forms

    class BulkLeaveAllocationForm(forms.Form):
        policy = forms.ModelChoiceField(
            queryset=LeavePolicy.objects.all(),
            label="Leave Policy"
        )
        year = forms.IntegerField(
            initial=timezone.now().year,
            min_value=2020,
            max_value=2050,
            label="Year"
        )
        include_existing = forms.BooleanField(
            required=False,
            initial=False,
            label="Update Existing Balances"
        )

    if request.method == 'POST':
        form = BulkLeaveAllocationForm(request.POST)
        if form.is_valid():
            policy = form.cleaned_data['policy']
            year = form.cleaned_data['year']
            include_existing = form.cleaned_data['include_existing']

            # Get all users in the policy's group
            users = User.objects.filter(groups=policy.group)

            # Get all allocations for this policy
            allocations = LeaveAllocation.objects.filter(policy=policy)

            created_count = 0
            updated_count = 0

            for user in users:
                for allocation in allocations:
                    leave_type = allocation.leave_type

                    # Check if balance already exists
                    balance = UserLeaveBalance.objects.filter(
                        user=user,
                        leave_type=leave_type,
                        year=year
                    ).first()

                    if balance:
                        if include_existing:
                            # Update existing balance
                            balance.allocated = allocation.annual_days
                            balance.save()
                            updated_count += 1
                    else:
                        # Create new balance
                        UserLeaveBalance.objects.create(
                            user=user,
                            leave_type=leave_type,
                            year=year,
                            allocated=allocation.annual_days,
                            used=0,
                            carried_forward=0,
                            additional=0
                        )
                        created_count += 1

            messages.success(
                request,
                f'Successfully created {created_count} new balances and updated {updated_count} existing balances.'
            )
            return redirect('aps_leave:hr_leave_view')
    else:
        form = BulkLeaveAllocationForm()

    return render(request, 'components/leave_management/bulk_leave_balance_form.html', {
        'form': form,
        'action': 'Create Bulk Balances'
    })

# ============= 3. LEAVE REQUEST VIEWS (ALL USERS) =============

@login_required
@user_passes_test(lambda u: True)  # All authenticated users can access
def leave_request_list(request):
    """View to list leave requests based on user role"""
    context = {}

    # Different views based on user role
    if can_approve_leave(request.user):
        # Show all leave requests if user can approve
        leave_requests = LeaveRequest.objects.all().order_by('-created_at')
        context['show_all'] = True
    else:
        # Show only user's own leave requests
        leave_requests = LeaveRequest.objects.filter(user=request.user).order_by('-created_at')

    # Allow filtering by status
    status_filter = request.GET.get('status', None)
    if status_filter:
        leave_requests = leave_requests.filter(status=status_filter)

    # Get leave types for filtering
    leave_types = LeaveType.objects.filter(is_active=True)

    # Get user's current leave balances
    current_year = timezone.now().year
    if not can_approve_leave(request.user):
        balances = UserLeaveBalance.objects.filter(
            user=request.user,
            year=current_year
        ).select_related('leave_type')
    else:
        balances = None

    context.update({
        'leave_requests': leave_requests,
        'leave_types': leave_types,
        'balances': balances,
        'status_choices': LeaveRequest.STATUS_CHOICES,
        'can_approve': can_approve_leave(request.user)
    })

    return render(request, 'components/leave_management/leave_request_list.html', context)


@login_required
@user_passes_test(lambda u: True)  # All authenticated users can access
def leave_request_create(request):
    """View to create a new leave request"""
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST, request.FILES)
        form.instance.user = request.user

        if form.is_valid():
            try:
                leave_request = form.save(commit=False)
                leave_request.clean()  # Run validation logic
                leave_request.save()

                messages.success(request, "Leave request created successfully.")
                return redirect('aps_leave:leave_request_detail', pk=leave_request.id)
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = LeaveRequestForm()

    # Get user's current leave balances for display
    current_year = timezone.now().year
    balances = UserLeaveBalance.objects.filter(
        user=request.user,
        year=current_year
    ).select_related('leave_type')

    context = {
        'form': form,
        'balances': balances,
    }

    return render(request, 'components/leave_management/leave_request_form.html', context)


@login_required
@user_passes_test(lambda u: True)  # All authenticated users can access
def leave_request_update(request, pk):
    """View to update a leave request (only if pending)"""
    leave_request = get_object_or_404(LeaveRequest, pk=pk)

    # Check if user owns this leave request
    if leave_request.user != request.user and not can_approve_leave(request.user):
        return HttpResponseForbidden("You don't have permission to update this leave request.")

    # Only pending leave requests can be updated
    if leave_request.status != 'Pending':
        messages.error(request, "Only pending leave requests can be updated.")
        return redirect('aps_leave:leave_request_detail', pk=leave_request.id)

    if request.method == 'POST':
        form = LeaveRequestForm(request.POST, request.FILES, instance=leave_request)

        if form.is_valid():
            try:
                leave_request = form.save(commit=False)
                leave_request.clean()  # Run validation logic
                leave_request.save()

                messages.success(request, "Leave request updated successfully.")
                return redirect('aps_leave:leave_request_detail', pk=leave_request.id)
            except Exception as e:
                messages.error(request, f"Error: {str(e)}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = LeaveRequestForm(instance=leave_request)

    # Get user's current leave balances for display
    current_year = timezone.now().year
    balances = UserLeaveBalance.objects.filter(
        user=leave_request.user,
        year=current_year
    ).select_related('leave_type')

    context = {
        'form': form,
        'leave_request': leave_request,
        'balances': balances,
    }

    return render(request, 'components/leave_management/leave_request_form.html', context)


@login_required
@user_passes_test(lambda u: True)  # All authenticated users can access
def leave_request_cancel(request, pk):
    """View to cancel a leave request"""
    leave_request = get_object_or_404(LeaveRequest, pk=pk)

    # Check if user owns this leave request
    if leave_request.user != request.user and not can_approve_leave(request.user):
        return HttpResponseForbidden("You don't have permission to cancel this leave request.")

    # Can't cancel already cancelled or rejected leave requests
    if leave_request.status in ['Cancelled', 'Rejected']:
        messages.error(request, f"Leave request is already {leave_request.status.lower()}.")
        return redirect('aps_leave:leave_request_detail', pk=leave_request.id)

    if request.method == 'POST':
        # If it was already approved, we need to revert the leave balance
        old_status = leave_request.status

        # Update status to cancelled
        leave_request.status = 'Cancelled'
        leave_request.save()

        messages.success(request, "Leave request cancelled successfully.")
        return redirect('aps_leave:leave_request_list')

    return render(request, 'components/leave_management/leave_request_cancel.html', {
        'leave_request': leave_request
    })


@login_required
@user_passes_test(lambda u: True)  # All authenticated users can access
def leave_request_detail(request, pk):
    """View to see details of a leave request"""
    leave_request = get_object_or_404(LeaveRequest, pk=pk)

    # Check if user owns this leave request or can approve leave
    if leave_request.user != request.user and not can_approve_leave(request.user):
        return HttpResponseForbidden("You don't have permission to view this leave request.")

    # Get leave balances for the user
    year = leave_request.start_date.year
    balance = UserLeaveBalance.objects.filter(
        user=leave_request.user,
        leave_type=leave_request.leave_type,
        year=year
    ).first()

    context = {
        'leave_request': leave_request,
        'balance': balance,
        'can_approve': can_approve_leave(request.user),
        'can_edit': leave_request.status == 'Pending' and leave_request.user == request.user,
        'can_cancel': leave_request.status in ['Pending', 'Approved'] and leave_request.user == request.user,
    }

    return render(request, 'components/leave_management/leave_request_detail.html', context)


# ============= 4. LEAVE APPROVAL VIEWS =============
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.utils import timezone
from django.db.models import Q
import json
from django.db import transaction
from decimal import Decimal
from .models import LeaveRequest, UserLeaveBalance


@login_required
@user_passes_test(can_approve_leave)
def leave_approval_list(request):
    """View to list leave requests pending approval"""
    # Get all pending leave requests
    pending_requests = LeaveRequest.objects.filter(status='Pending').order_by('start_date')

    # Filter by leave type
    leave_type_id = request.GET.get('leave_type', None)
    if leave_type_id:
        pending_requests = pending_requests.filter(leave_type_id=leave_type_id)

    # Filter by user or department if needed
    user_id = request.GET.get('user_id', None)
    if user_id:
        pending_requests = pending_requests.filter(user_id=user_id)

    leave_types = LeaveType.objects.filter(is_active=True)

    context = {
        'pending_requests': pending_requests,
        'leave_types': leave_types,
    }

    return render(request, 'components/leave_management/leave_approval_list.html', context)



@login_required
@user_passes_test(can_approve_leave)
def leave_request_approve(request, pk):
    """View to approve a leave request"""
    leave_request = get_object_or_404(LeaveRequest, pk=pk)

    # Can only approve pending requests
    if leave_request.status != 'Pending':
        messages.error(request, f"Leave request is already {leave_request.status.lower()}.")
        return redirect('aps_leave:leave_approval_list')

    if request.method == 'POST':
        form = LeaveApprovalForm(request.POST)

        if form.is_valid():
            with transaction.atomic():
                # Store previous status for reference
                previous_status = leave_request.status

                # Update leave request
                leave_request.status = 'Approved'
                leave_request.approver = request.user

                # Handle any notes from approver
                if form.cleaned_data.get('notes'):
                    leave_request.approver_notes = form.cleaned_data['notes']

                # Calculate leave days to ensure accuracy and convert to Decimal
                leave_days_calculated = leave_request.calculate_leave_days()
                leave_request.leave_days = Decimal(str(leave_days_calculated))

                # Save leave request first
                leave_request.save()

                # Explicitly update the leave balance to ensure it happens
                if leave_request.leave_type.is_paid and (previous_status != 'Approved'):
                    # Force balance update
                    year = leave_request.start_date.year
                    balance, created = UserLeaveBalance.objects.get_or_create(
                        user=leave_request.user,
                        leave_type=leave_request.leave_type,
                        year=year,
                        defaults={'allocated': 0}
                    )

                    # Use Decimal for precision
                    current_used = Decimal(str(balance.used))
                    leave_days = Decimal(str(leave_request.leave_days))

                    # Update balance
                    balance.used = current_used + leave_days
                    balance.save()

                    # Verify update
                    refreshed_balance = UserLeaveBalance.objects.get(
                        id=balance.id
                    )
                    print(f"[VERIFY] Balance after update - ID: {refreshed_balance.id}, Used: {refreshed_balance.used}")

                messages.success(request, f"Leave request for {leave_request.user.get_full_name()} has been approved.")
                return redirect('aps_leave:leave_approval_list')
    else:
        form = LeaveApprovalForm()

    # Check if user has sufficient balance
    has_balance = leave_request.has_sufficient_balance()

    context = {
        'leave_request': leave_request,
        'form': form,
        'has_balance': has_balance,
    }

    return render(request, 'components/leave_management/leave_request_approve.html', context)


@login_required
@user_passes_test(can_approve_leave)
def leave_request_reject(request, pk):
    """View to reject a leave request"""
    leave_request = get_object_or_404(LeaveRequest, pk=pk)

    # Can only reject pending requests
    if leave_request.status != 'Pending':
        messages.error(request, f"Leave request is already {leave_request.status.lower()}.")
        return redirect('aps_leave:leave_approval_list')

    if request.method == 'POST':
        form = LeaveRejectionForm(request.POST)

        if form.is_valid():
            leave_request.status = 'Rejected'
            leave_request.approver = request.user
            leave_request.rejection_reason = form.cleaned_data['rejection_reason']

            # Suggested alternative dates (optional)
            if form.cleaned_data.get('suggested_dates'):
                leave_request.suggested_dates = form.cleaned_data['suggested_dates']

            leave_request.save()

            # Notify the user of rejection (implementation depends on your notification system)
            # send_leave_notification(leave_request, 'rejected')

            messages.success(request, f"Leave request for {leave_request.user.get_full_name()} has been rejected.")
            return redirect('aps_leave:leave_approval_list')
    else:
        form = LeaveRejectionForm()

    context = {
        'leave_request': leave_request,
        'form': form,
    }

    return render(request, 'components/leave_management/leave_request_reject.html', context)

# @login_required
# @user_passes_test(can_approve_leave)
# def leave_approval_list(request):
#     """View to list leave requests pending approval"""
#     user = request.user
#     print(f"[DEBUG] User: {user.username}, Role Check - HR: {is_hr(user)}, Admin: {is_admin(user)}, Manager: {is_manager(user)}")

#     if is_hr(user) or is_admin(user):
#         # HR, Admin, and Management can see all pending leave requests
#         pending_requests = LeaveRequest.objects.filter(status='Pending').order_by('start_date')
#     elif is_manager(user):
#         # Managers only see other users' pending leave requests
#         # Since there's no department, fallback to only showing other users
#         pending_requests = LeaveRequest.objects.filter(
#             status='Pending'
#         ).exclude(user=user).order_by('start_date')
#     else:
#         # Shouldn't reach here due to @user_passes_test
#         print("[DEBUG] Invalid access: user does not have approval permissions")
#         pending_requests = LeaveRequest.objects.none()

#     print(f"[DEBUG] Total pending requests fetched: {pending_requests.count()}")

#     paginator = Paginator(pending_requests, 20)
#     page_number = request.GET.get('page')
#     page_obj = paginator.get_page(page_number)

#     return render(request, 'components/leave_management/leave_approval_list.html', {
#         'page_obj': page_obj
#     })

# from django.db import transaction

# @login_required
# @user_passes_test(can_approve_leave)
# def leave_request_approve(request, pk):
#     """View to approve a leave request and ensure balance updates"""
#     leave_request = get_object_or_404(LeaveRequest, pk=pk, status='Pending')
#     print(f"[DEBUG] Approving Leave ID: {pk} for User: {leave_request.user.username}")

#     if request.method == 'POST':
#         print("[DEBUG] POST request received for approval")

#         if not leave_request.has_sufficient_balance() and leave_request.leave_type.is_paid:
#             print("[DEBUG] Insufficient balance for paid leave")
#             if request.POST.get('auto_convert') == 'yes':
#                 print("[DEBUG] Attempting auto-convert to Loss of Pay")
#                 if leave_request.auto_convert_leave_type():
#                     messages.info(request, 'Leave request was automatically converted to Loss of Pay due to insufficient balance.')
#                 else:
#                     messages.error(request, 'Unable to approve leave request due to insufficient balance.')
#                     return redirect('aps_leave:leave_approval_list')
#             else:
#                 messages.error(request, 'Unable to approve leave request due to insufficient balance.')
#                 return redirect('aps_leave:leave_approval_list')

#         try:
#             with transaction.atomic():
#                 old_status = leave_request.status
#                 leave_request.status = 'Approved'
#                 leave_request.approver = request.user

#                 # Save first to ensure leave request is approved
#                 leave_request.save()

#                 # Explicitly update leave balance outside of the model's save method
#                 if leave_request.leave_type.is_paid:
#                     print(f"[DEBUG] Explicitly updating leave balance for {leave_request.user.username}, leave type: {leave_request.leave_type.name}, days: {leave_request.leave_days}")

#                     year = leave_request.start_date.year
#                     balance, created = UserLeaveBalance.objects.get_or_create(
#                         user=leave_request.user,
#                         leave_type=leave_request.leave_type,
#                         year=year,
#                         defaults={'allocated': 0}
#                     )

#                     print(f"[DEBUG] Before balance update - Used: {balance.used}, Leave days: {leave_request.leave_days}")
#                     balance.used = float(balance.used) + float(leave_request.leave_days)
#                     balance.save()
#                     print(f"[DEBUG] After balance update - Used: {balance.used}")

#                 # Verify balance was updated correctly
#                 try:
#                     current_balance = UserLeaveBalance.objects.get(
#                         user=leave_request.user,
#                         leave_type=leave_request.leave_type,
#                         year=leave_request.start_date.year
#                     )
#                     print(f"[DEBUG] Final balance verification - Used: {current_balance.used}, Available: {current_balance.available}")
#                 except UserLeaveBalance.DoesNotExist:
#                     print("[DEBUG] Warning: Could not find balance record to verify update")

#             print(f"[DEBUG] Leave request {pk} approved successfully")
#             messages.success(request, 'Leave request approved successfully!')
#         except Exception as e:
#             print(f"[ERROR] Failed to approve leave: {e}")
#             messages.error(request, 'Something went wrong while approving the leave request.')

#         return redirect('aps_leave:leave_approval_list')

#     has_balance = leave_request.has_sufficient_balance()
#     print(f"[DEBUG] Leave balance sufficient: {has_balance}")

#     return render(request, 'components/leave_management/leave_request_approve.html', {
#         'leave_request': leave_request,
#         'has_balance': has_balance
#     })

# @login_required
# @user_passes_test(can_approve_leave)
# def leave_request_reject(request, pk):
#     """View to reject a leave request"""
#     leave_request = get_object_or_404(LeaveRequest, pk=pk, status='Pending')
#     print(f"[DEBUG] Rejecting Leave ID: {pk} for User: {leave_request.user.username}")

#     if request.method == 'POST':
#         rejection_reason = request.POST.get('rejection_reason', '')
#         suggested_dates = request.POST.get('suggested_dates', '')
#         print(f"[DEBUG] Rejection reason: {rejection_reason}")
#         print(f"[DEBUG] Suggested dates (raw): {suggested_dates}")

#         leave_request.status = 'Rejected'
#         leave_request.approver = request.user
#         leave_request.rejection_reason = rejection_reason

#         if suggested_dates:
#             try:
#                 leave_request.suggested_dates = json.loads(suggested_dates)
#                 print("[DEBUG] Suggested dates parsed as JSON")
#             except json.JSONDecodeError:
#                 print("[DEBUG] Failed to parse suggested dates as JSON. Saving as plain text.")
#                 leave_request.suggested_dates = {'text': suggested_dates}

#         leave_request.save()
#         print(f"[DEBUG] Leave request {pk} rejected")

#         messages.success(request, 'Leave request rejected.')
#         return redirect('aps_leave:leave_approval_list')

#     return render(request, 'components/leave_management/leave_request_reject.html', {
#         'leave_request': leave_request
#     })


# ============= 5. COMP-OFF REQUEST VIEWS =============

@login_required
def comp_off_request_list(request):
    """View to list comp-off requests"""
    user = request.user

    if is_hr(user) or is_admin(user) or is_management(user):
        # HR, Admin, and Management can see all comp-off requests
        comp_off_requests = CompOffRequest.objects.all().order_by('-created_at')
    elif is_manager(user):
        # Managers can see comp-offs for their team members
        comp_off_requests = CompOffRequest.objects.filter(
            Q(user__profile__department=user.profile.department) |
            Q(user=user)
        ).order_by('-created_at')
    else:
        # Regular employees see only their own comp-offs
        comp_off_requests = CompOffRequest.objects.filter(user=user).order_by('-created_at')

    # Add pagination
    paginator = Paginator(comp_off_requests, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'components/leave_management/comp_off_request_list.html', {
        'page_obj': page_obj,
        'can_approve': can_approve_leave(user)
    })

@login_required
def comp_off_request_create(request):
    """View to create a new comp-off request"""
    if request.method == 'POST':
        form = CompOffRequestForm(request.POST)
        if form.is_valid():
            comp_off_request = form.save(commit=False)
            comp_off_request.user = request.user
            comp_off_request.save()

            messages.success(request, 'Comp-off request submitted successfully!')
            return redirect('aps_leave:comp_off_request_list')
    else:
        form = CompOffRequestForm()

    return render(request, 'components/leave_management/comp_off_request_form.html', {
        'form': form,
        'action': 'Create'
    })

@login_required
@user_passes_test(can_approve_leave)
def comp_off_request_approve(request, pk):
    """View to approve a comp-off request"""
    comp_off_request = get_object_or_404(CompOffRequest, pk=pk, status='Pending')

    if request.method == 'POST':
        comp_off_request.status = 'Approved'
        comp_off_request.approver = request.user
        comp_off_request.save()  # This will trigger the balance update in save method

        messages.success(request, 'Comp-off request approved!')
        return redirect('aps_leave:comp_off_request_list')

    return render(request, 'components/leave_management/comp_off_request_approve.html', {
        'comp_off_request': comp_off_request
    })

@login_required
@user_passes_test(can_approve_leave)
def comp_off_request_reject(request, pk):
    """View to reject a comp-off request"""
    comp_off_request = get_object_or_404(CompOffRequest, pk=pk, status='Pending')

    if request.method == 'POST':
        comp_off_request.status = 'Rejected'
        comp_off_request.approver = request.user
        comp_off_request.save()

        messages.success(request, 'Comp-off request rejected!')
        return redirect('aps_leave:comp_off_request_list')

    return render(request, 'components/leave_management/comp_off_request_reject.html', {
        'comp_off_request': comp_off_request
    })

# ============= 6. LEAVE DASHBOARD & REPORTS =============
# ============= 6. LEAVE DASHBOARD & REPORTS =============

@login_required
def leave_dashboard(request):
    """Main dashboard view for leave management"""
    user = request.user
    current_year = timezone.now().year

    # Get leave balances for current user
    user_balances = UserLeaveBalance.objects.filter(
        user=user,
        year=current_year
    )

    # Get pending leave requests for current user
    pending_leaves = LeaveRequest.objects.filter(
        user=user,
        status='Pending'
    ).order_by('start_date')[:5]

    # Get upcoming approved leaves
    upcoming_leaves = LeaveRequest.objects.filter(
        user=user,
        status='Approved',
        start_date__gte=timezone.now().date()
    ).order_by('start_date')[:5]

    # Additional data for managers, HR, and admins
    team_leaves = None
    pending_approvals = None
    leave_stats = None

    if can_approve_leave(user):
        # For managers - show team's upcoming leaves
        if is_manager(user):
            team_leaves = LeaveRequest.objects.filter(
                status='Approved',
                start_date__gte=timezone.now().date(),
                user__profile__department=user.profile.department
            ).exclude(user=user).order_by('start_date')[:10]

            pending_approvals = LeaveRequest.objects.filter(
                status='Pending',
                user__profile__department=user.profile.department
            ).exclude(user=user).order_by('start_date')[:10]

        # For HR, Admin, Management - show organizational stats
        if is_hr(user) or is_admin(user) or is_management(user):
            # Count leaves by type for the current month
            leave_stats = LeaveRequest.objects.filter(
                status='Approved',
                start_date__year=current_year,
                start_date__month=timezone.now().month
            ).values('leave_type__name').annotate(
                count=Count('id'),
                total_days=Sum('leave_days')
            )

            # Show all pending approvals
            pending_approvals = LeaveRequest.objects.filter(
                status='Pending'
            ).order_by('start_date')[:15]

    return render(request, 'components/leave_management/leave_dashboard.html', {
        'user_balances': user_balances,
        'pending_leaves': pending_leaves,
        'upcoming_leaves': upcoming_leaves,
        'team_leaves': team_leaves,
        'pending_approvals': pending_approvals,
        'leave_stats': leave_stats,
        'can_approve': can_approve_leave(user),
        'is_hr': is_hr(user),
        'is_admin': is_admin(user),
        'is_manager': is_manager(user)
    })

@login_required
@user_passes_test(lambda u: is_hr(u) or is_admin(u) or is_management(u) or is_finance(u))
def leave_balance_report(request):
    """Report showing leave balances for all employees"""
    year = int(request.GET.get('year', timezone.now().year))
    department = request.GET.get('department', '')
    export_format = request.GET.get('export', '')

    # Get all balances for the selected year
    balances = UserLeaveBalance.objects.filter(
        year=year
    ).select_related('user', 'leave_type')

    # Apply department filter if specified
    if department:
        balances = balances.filter(user__profile__department=department)

    # Organize data by user
    users_data = {}
    leave_types = set()

    for balance in balances:
        user_id = balance.user_id
        leave_type = balance.leave_type.name
        leave_types.add(leave_type)

        if user_id not in users_data:
            user = balance.user
            users_data[user_id] = {
                'user': user,
                'name': user.get_full_name() or user.username,
                'department': getattr(user.profile, 'department', '') if hasattr(user, 'profile') else '',
                'balances': {}
            }

        users_data[user_id]['balances'][leave_type] = {
            'allocated': balance.allocated,
            'used': balance.used,
            'carried_forward': balance.carried_forward,
            'additional': balance.additional,
            'available': balance.available
        }

    # Handle export
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="leave_balances_{year}.csv"'

        writer = csv.writer(response)

        # Create headers with leave types
        headers = ['Employee', 'Department']
        for lt in sorted(leave_types):
            headers.extend([f'{lt} - Allocated', f'{lt} - Used', f'{lt} - Available'])

        writer.writerow(headers)

        # Write data rows
        for user_data in users_data.values():
            row = [user_data['name'], user_data['department']]

            for lt in sorted(leave_types):
                if lt in user_data['balances']:
                    b = user_data['balances'][lt]
                    row.extend([b['allocated'], b['used'], b['available']])
                else:
                    row.extend(['0', '0', '0'])

            writer.writerow(row)

        return response

    # Get departments for filter
    departments = set([
        getattr(user.profile, 'department', '')
        for user in User.objects.filter(is_active=True)
        if hasattr(user, 'profile') and getattr(user.profile, 'department', '')
    ])

    return render(request, 'components/leave_management/leave_balance_report.html', {
        'users_data': users_data.values(),
        'leave_types': sorted(leave_types),
        'years': range(timezone.now().year - 2, timezone.now().year + 2),
        'departments': sorted(departments),
        'selected_year': year,
        'selected_department': department
    })

@login_required
@user_passes_test(lambda u: is_hr(u) or is_admin(u) or is_management(u) or is_finance(u))
def leave_report(request):
    """Advanced leave report with filters and export options"""
    year = int(request.GET.get('year', timezone.now().year))
    month = request.GET.get('month', '')
    leave_type = request.GET.get('leave_type', '')
    department = request.GET.get('department', '')
    status = request.GET.get('status', '')
    export_format = request.GET.get('export', '')

    # Start with all leave requests
    leave_requests = LeaveRequest.objects.filter(
        start_date__year=year
    ).select_related('user', 'leave_type')

    # Apply filters
    if month:
        leave_requests = leave_requests.filter(start_date__month=int(month))

    if leave_type:
        leave_requests = leave_requests.filter(leave_type_id=leave_type)

    if department:
        leave_requests = leave_requests.filter(user__profile__department=department)

    if status:
        leave_requests = leave_requests.filter(status=status)

    # Order by date
    leave_requests = leave_requests.order_by('start_date')

    # Handle export requests
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="leave_report.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Employee', 'Department', 'Leave Type', 'Start Date',
            'End Date', 'Days', 'Status', 'Approver'
        ])

        for leave in leave_requests:
            writer.writerow([
                leave.user.get_full_name() or leave.user.username,
                getattr(leave.user.profile, 'department', '') if hasattr(leave.user, 'profile') else '',
                leave.leave_type.name,
                leave.start_date.strftime('%Y-%m-%d'),
                leave.end_date.strftime('%Y-%m-%d'),
                leave.leave_days,
                leave.status,
                leave.approver.get_full_name() if leave.approver else ''
            ])

        return response

    # Get options for filters
    leave_types = LeaveType.objects.all()
    departments = set([
        getattr(user.profile, 'department', '')
        for user in User.objects.filter(is_active=True)
        if hasattr(user, 'profile') and getattr(user.profile, 'department', '')
    ])

    return render(request, 'components/leave_management/leave_report.html', {
        'leave_requests': leave_requests,
        'leave_types': leave_types,
        'departments': sorted(departments),
        'years': range(timezone.now().year - 2, timezone.now().year + 2),
        'months': range(1, 13),
        'selected_year': year,
        'selected_month': month,
        'selected_leave_type': leave_type,
        'selected_department': department,
        'selected_status': status
    })






from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import Http404
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from datetime import datetime





@login_required
@user_passes_test(lambda u: is_employee(u) or is_hr(u))  # Allow both employees and HR
def leave_view(request):
    """Handle multiple leave functionalities on one page."""
    # Get current year for filtering
    year = timezone.now().year

    try:
        # Get detailed leave balance info
        leave_balance = Leave.get_leave_balance(request.user)

        # Get all leave requests for the current user
        leave_requests = Leave.objects.filter(user=request.user).order_by('-created_at')

    except Exception as e:
        logger.error(f"Error calculating leave balance for user {request.user.username}: {str(e)}")
        logger.error(f"Full exception traceback:", exc_info=True)

        # Set default values if error occurs
        leave_balance = {
            'total_leaves': 0,
            'used_leaves': 0,
            'comp_off': 0,
            'loss_of_pay': 0
        }
        leave_requests = Leave.objects.none()
        messages.error(request, "Unable to calculate leave balance. Please contact HR.")

    # Handle leave request submission
    if request.method == 'POST' and 'request_leave' in request.POST:
        try:
            # Parse dates
            start_date = datetime.strptime(request.POST.get('start_date'), '%Y-%m-%d').date()
            end_date = datetime.strptime(request.POST.get('end_date'), '%Y-%m-%d').date()

            # Get half_day value properly from form
            half_day = request.POST.get('half_day') == 'true'

            # Get leave type with default
            leave_type = request.POST.get('leave_type')
            if not leave_type:
                messages.error(request, "Leave type is required")
                return redirect('aps_employee:leave_view')

            # Create new leave request
            leave = Leave(
                user=request.user,
                leave_type=leave_type,
                start_date=start_date,
                end_date=end_date,
                reason=request.POST.get('reason'),
                half_day=half_day,
                documentation=request.FILES.get('documentation')
            )

            # Auto convert leave type based on balance
            try:
                leave.auto_convert_leave_type()
            except Exception as e:
                messages.error(request, f"Error converting leave type: {str(e)}")
                return redirect('aps_employee:leave_view')

            # Run validations and save
            leave.full_clean()
            leave.save()

            messages.success(request, "Leave request submitted successfully.")
            return redirect('aps_employee:leave_view')

        except ValidationError as e:
            messages.error(request, str(e))
        except ValueError as e:
            messages.error(request, "Invalid date format")
        except Exception as e:
            messages.error(request, f"Error submitting leave request: {str(e)}")
        return redirect('aps_employee:leave_view')

    # Handle leave request updates
    if request.method == 'POST' and 'edit_leave' in request.POST:
        try:
            leave = Leave.objects.get(id=request.POST.get('leave_id'), user=request.user)

            if leave.status != 'Pending':
                messages.error(request, "Only pending leave requests can be edited.")
                return redirect('aps_employee:leave_view')

            # Update leave details
            leave.start_date = datetime.strptime(request.POST.get('start_date'), '%Y-%m-%d').date()
            leave.end_date = datetime.strptime(request.POST.get('end_date'), '%Y-%m-%d').date()
            leave.reason = request.POST.get('reason')
            leave.half_day = request.POST.get('half_day') == 'true'

            if 'documentation' in request.FILES:
                leave.documentation = request.FILES['documentation']

            # Auto convert leave type based on updated dates
            leave.auto_convert_leave_type()

            # Validate and save
            leave.full_clean()
            leave.save()

            messages.success(request, "Leave request updated successfully.")
            return redirect('aps_employee:leave_view')

        except (ValidationError, Leave.DoesNotExist) as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f"Error updating leave request: {str(e)}")
        return redirect('aps_employee:leave_view')

    # Handle leave cancellation
    if request.method == 'POST' and 'delete_leave' in request.POST:
        try:
            leave = Leave.objects.get(id=request.POST.get('leave_id'), user=request.user)

            if leave.status not in ['Pending', 'Approved']:
                messages.error(request, "This leave request cannot be cancelled.")
                return redirect('aps_employee:leave_view')

            leave.status = 'Cancelled'
            leave.save()

            messages.success(request, "Leave request cancelled successfully.")
            return redirect('aps_employee:leave_view')

        except Leave.DoesNotExist:
            messages.error(request, "Leave request not found.")
        except Exception as e:
            messages.error(request, f"Error cancelling leave request: {str(e)}")
        return redirect('aps_employee:leave_view')

    # Constants for display
    TOTAL_ANNUAL_LEAVES = 18.0

    return render(request, 'basic/leave.html', {
        'leave_balance': leave_balance,
        'leave_requests': leave_requests,
        'leave_types': Leave.LEAVE_TYPES,
        'total_annual_leaves': TOTAL_ANNUAL_LEAVES,
        'leaves_taken': leave_balance.get('used_leaves', 0.0),
        'remaining_leaves': leave_balance.get('total_leaves', TOTAL_ANNUAL_LEAVES),
        'loss_of_pay': leave_balance.get('loss_of_pay', 0.0)
    })


@login_required
@user_passes_test(is_hr)
def view_leave_requests_hr(request):
    """HR views all leave requests with comprehensive filtering and organizational hierarchy."""
    # Apply filters
    employee_filter = request.GET.get('employee', '')
    leave_type_filter = request.GET.get('leave_type', '')
    status_filter = request.GET.get('status', '')
    date_range = request.GET.get('date_range', '')

    # HR can see all leave requests except admin
    leave_requests = Leave.objects.all().select_related('user', 'approver').order_by('-created_at')

    # Apply filters
    if employee_filter:
        leave_requests = leave_requests.filter(
            Q(user__username__icontains=employee_filter) |
            Q(user__first_name__icontains=employee_filter) |
            Q(user__last_name__icontains=employee_filter)
        )

    if leave_type_filter:
        leave_requests = leave_requests.filter(leave_type=leave_type_filter)

    if status_filter:
        leave_requests = leave_requests.filter(status=status_filter)

    if date_range:
        try:
            date_parts = date_range.split(' - ')
            start_date = datetime.strptime(date_parts[0], '%Y-%m-%d').date()
            end_date = datetime.strptime(date_parts[1], '%Y-%m-%d').date()
            leave_requests = leave_requests.filter(
                Q(start_date__range=[start_date, end_date]) |
                Q(end_date__range=[start_date, end_date])
            )
        except (ValueError, IndexError):
            messages.error(request, "Invalid date range format")

    # Separate employee and manager requests that need HR approval
    employee_manager_requests = leave_requests.filter(
        Q(user__groups__name='Employee') |
        Q(user__groups__name='Manager'),
        status='Pending'
    )

    # Other requests for reference
    other_requests = leave_requests.exclude(
        Q(user__groups__name='Employee') |
        Q(user__groups__name='Manager'),
        status='Pending'
    )

    # Get organizational hierarchy data
    org_hierarchy = {
        'hr': User.objects.filter(groups__name='HR').count(),
        'managers': User.objects.filter(groups__name='Manager').count(),
        'employees': User.objects.filter(groups__name='Employee').count(),
    }

    # Get approval history for audit trail
    approval_history = Leave.objects.exclude(approver=None).values(
        'user__username', 'user__first_name', 'user__last_name',
        'approver__username', 'approver__first_name', 'approver__last_name',
        'status', 'created_at', 'updated_at', 'leave_type'
    ).order_by('-updated_at')[:100]  # Last 100 approvals

    return render(request, 'components/hr/view_leave_requests.html', {
        'employee_manager_requests': employee_manager_requests,
        'other_requests': other_requests,
        'leave_types': Leave.LEAVE_TYPES,
        'status_choices': Leave.STATUS_CHOICES,
        'org_hierarchy': org_hierarchy,
        'approval_history': approval_history,
        'filters': {
            'employee': employee_filter,
            'leave_type': leave_type_filter,
            'status': status_filter,
            'date_range': date_range
        }
    })

@login_required
@user_passes_test(is_hr)
def manage_leave_request_hr(request, leave_id, action):
    """HR approves or rejects employee and manager leave requests with comprehensive validation."""
    leave_request = get_object_or_404(Leave.objects.select_related('user', 'approver'), id=leave_id)

    # HR should only approve/reject employee and manager requests
    if not (is_employee(leave_request.user) or is_manager(leave_request.user)):
        messages.error(request, "HR should only approve employee and manager leave requests.")
        return redirect('aps_hr:view_leave_requests_hr')

    # Check if leave request is already processed
    if leave_request.status != 'Pending':
        messages.error(request, f"This leave request is already {leave_request.status.lower()}.")
        return redirect('aps_hr:view_leave_requests_hr')

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if action == 'approve':
                    # For managers, check team coverage
                    if is_manager(leave_request.user):
                        # Check if team has coverage during this period
                        team_size = User.objects.filter(employee__reporting_manager=leave_request.user).count()
                        if team_size > 0:
                            messages.info(request, f"Note: This manager has {team_size} direct reports.")

                    leave_request.status = 'Approved'
                    leave_request.approver = request.user
                    leave_request.approval_date = timezone.now()
                    leave_request.save()

                    # Log the approval
                    logger.info(f"Leave ID {leave_id} for {leave_request.user.username} approved by HR {request.user.username}")

                    messages.success(request, f"Leave for {leave_request.user.get_full_name() or leave_request.user.username} approved.")

                elif action == 'reject':
                    rejection_reason = request.POST.get('rejection_reason')
                    if not rejection_reason:
                        messages.error(request, "Rejection reason is required.")
                        return render(request, 'components/hr/manage_leave.html', {
                            'leave_request': leave_request,
                            'action': action.capitalize(),
                            'leave_balance': Leave.get_leave_balance(leave_request.user)
                        })

                    leave_request.status = 'Rejected'
                    leave_request.approver = request.user
                    leave_request.rejection_reason = rejection_reason
                    leave_request.approval_date = timezone.now()
                    leave_request.save()

                    # Log the rejection
                    logger.info(f"Leave ID {leave_id} for {leave_request.user.username} rejected by HR {request.user.username}")

                    messages.warning(request, f"Leave for {leave_request.user.get_full_name() or leave_request.user.username} rejected.")

                return redirect('aps_hr:view_leave_requests_hr')

        except Exception as e:
            logger.error(f"Error processing leave request: {str(e)}")
            messages.error(request, f"Error processing leave request: {str(e)}")
            return redirect('aps_hr:view_leave_requests_hr')

    # Get organizational context
    org_context = {}

    if is_manager(leave_request.user):
        org_context['role'] = 'Manager'
        org_context['team_size'] = User.objects.filter(employee__reporting_manager=leave_request.user).count()
        org_context['team_on_leave'] = Leave.objects.filter(
            user__employee__reporting_manager=leave_request.user,
            status='Approved',
            start_date__lte=leave_request.end_date,
            end_date__gte=leave_request.start_date
        ).count()

    # Get leave history for context
    leave_history = Leave.objects.filter(user=leave_request.user).exclude(id=leave_id).order_by('-created_at')[:10]

    return render(request, 'components/hr/manage_leave.html', {
        'leave_request': leave_request,
        'action': action.capitalize(),
        'leave_balance': Leave.get_leave_balance(leave_request.user),
        'leave_history': leave_history,
        'org_context': org_context
    })

@login_required
@user_passes_test(is_manager)
def view_leave_requests_manager(request):
    """Manager views team leave requests with comprehensive filtering."""
    # Apply filters
    employee_filter = request.GET.get('employee', '')
    leave_type_filter = request.GET.get('leave_type', '')
    status_filter = request.GET.get('status', '')
    date_range = request.GET.get('date_range', '')

    # Manager can only see their team's leave requests
    leave_requests = Leave.objects.filter(
        user__employee__reporting_manager=request.user
    ).select_related('user', 'approver').order_by('-created_at')

    # Apply filters
    if employee_filter:
        leave_requests = leave_requests.filter(
            Q(user__username__icontains=employee_filter) |
            Q(user__first_name__icontains=employee_filter) |
            Q(user__last_name__icontains=employee_filter)
        )

    if leave_type_filter:
        leave_requests = leave_requests.filter(leave_type=leave_type_filter)

    if status_filter:
        leave_requests = leave_requests.filter(status=status_filter)

    if date_range:
        try:
            date_parts = date_range.split(' - ')
            start_date = datetime.strptime(date_parts[0], '%Y-%m-%d').date()
            end_date = datetime.strptime(date_parts[1], '%Y-%m-%d').date()
            leave_requests = leave_requests.filter(
                Q(start_date__range=[start_date, end_date]) |
                Q(end_date__range=[start_date, end_date])
            )
        except (ValueError, IndexError):
            messages.error(request, "Invalid date range format")

    # Get team hierarchy data
    team_hierarchy = {
        'total_team': User.objects.filter(employee__reporting_manager=request.user).count(),
        'on_leave': Leave.objects.filter(
            user__employee__reporting_manager=request.user,
            status='Approved'
        ).count()
    }

    # Get approval history for audit trail
    approval_history = Leave.objects.filter(
        user__employee__reporting_manager=request.user
    ).exclude(approver=None).values(
        'user__username', 'user__first_name', 'user__last_name',
        'approver__username', 'approver__first_name', 'approver__last_name',
        'status', 'created_at', 'updated_at', 'leave_type'
    ).order_by('-updated_at')[:50]  # Last 50 approvals

    return render(request, 'components/manager/view_leave_requests.html', {
        'leave_requests': leave_requests,
        'leave_types': Leave.LEAVE_TYPES,
        'status_choices': Leave.STATUS_CHOICES,
        'team_hierarchy': team_hierarchy,
        'approval_history': approval_history,
        'filters': {
            'employee': employee_filter,
            'leave_type': leave_type_filter,
            'status': status_filter,
            'date_range': date_range
        }
    })

@login_required
@user_passes_test(is_manager)
def manage_leave_request_manager(request, leave_id, action):
    """Manager approves or rejects team leave requests with comprehensive validation."""
    leave_request = get_object_or_404(
        Leave.objects.select_related('user', 'approver'),
        id=leave_id,
        user__employee__reporting_manager=request.user
    )

    # Check if leave request is already processed
    if leave_request.status != 'Pending':
        messages.error(request, f"This leave request is already {leave_request.status.lower()}.")
        return redirect('aps_manager:view_leave_requests_manager')

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if action == 'approve':
                    # Check team coverage during leave period
                    team_on_leave = Leave.objects.filter(
                        user__employee__reporting_manager=request.user,
                        status='Approved',
                        start_date__lte=leave_request.end_date,
                        end_date__gte=leave_request.start_date
                    ).exclude(id=leave_id).count()

                    team_size = User.objects.filter(
                        employee__reporting_manager=request.user
                    ).count()

                    if team_on_leave >= (team_size / 2):
                        messages.warning(request, f"Warning: {team_on_leave} out of {team_size} team members will be on leave during this period.")

                    leave_request.status = 'Approved'
                    leave_request.approver = request.user
                    leave_request.approval_date = timezone.now()
                    leave_request.save()

                    # Log the approval
                    logger.info(f"Leave ID {leave_id} for {leave_request.user.username} approved by manager {request.user.username}")

                    messages.success(request, f"Leave for {leave_request.user.get_full_name() or leave_request.user.username} approved.")

                elif action == 'reject':
                    rejection_reason = request.POST.get('rejection_reason')
                    if not rejection_reason:
                        messages.error(request, "Rejection reason is required.")
                        return render(request, 'components/manager/manage_leave.html', {
                            'leave_request': leave_request,
                            'action': action.capitalize(),
                            'leave_balance': Leave.get_leave_balance(leave_request.user)
                        })

                    leave_request.status = 'Rejected'
                    leave_request.approver = request.user
                    leave_request.rejection_reason = rejection_reason
                    leave_request.approval_date = timezone.now()
                    leave_request.save()

                    # Log the rejection
                    logger.info(f"Leave ID {leave_id} for {leave_request.user.username} rejected by manager {request.user.username}")

                    messages.warning(request, f"Leave for {leave_request.user.get_full_name() or leave_request.user.username} rejected.")

                return redirect('aps_manager:view_leave_requests_manager')

        except Exception as e:
            logger.error(f"Error processing leave request: {str(e)}")
            messages.error(request, f"Error processing leave request: {str(e)}")
            return redirect('aps_manager:view_leave_requests_manager')

    # Get leave history for context
    leave_history = Leave.objects.filter(user=leave_request.user).exclude(id=leave_id).order_by('-created_at')[:10]

    return render(request, 'components/manager/manage_leave.html', {
        'leave_request': leave_request,
        'action': action.capitalize(),
        'leave_balance': Leave.get_leave_balance(leave_request.user),
        'leave_history': leave_history,
        'team_on_leave': Leave.objects.filter(
            user__employee__reporting_manager=request.user,
            status='Approved',
            start_date__lte=leave_request.end_date,
            end_date__gte=leave_request.start_date
        ).exclude(id=leave_id).count()
    })

@login_required
@user_passes_test(is_admin)
def view_leave_requests_admin(request):
    """Admin views all leave requests with comprehensive filtering and organizational hierarchy."""
    # Apply filters
    employee_filter = request.GET.get('employee', '')
    leave_type_filter = request.GET.get('leave_type', '')
    status_filter = request.GET.get('status', '')
    date_range = request.GET.get('date_range', '')

    # Admin can see all leave requests
    leave_requests = Leave.objects.all().select_related('user', 'approver').order_by('-created_at')

    # Apply filters
    if employee_filter:
        leave_requests = leave_requests.filter(
            Q(user__username__icontains=employee_filter) |
            Q(user__first_name__icontains=employee_filter) |
            Q(user__last_name__icontains=employee_filter)
        )

    if leave_type_filter:
        leave_requests = leave_requests.filter(leave_type=leave_type_filter)

    if status_filter:
        leave_requests = leave_requests.filter(status=status_filter)

    if date_range:
        try:
            date_parts = date_range.split(' - ')
            start_date = datetime.strptime(date_parts[0], '%Y-%m-%d').date()
            end_date = datetime.strptime(date_parts[1], '%Y-%m-%d').date()
            leave_requests = leave_requests.filter(
                Q(start_date__range=[start_date, end_date]) |
                Q(end_date__range=[start_date, end_date])
            )
        except (ValueError, IndexError):
            messages.error(request, "Invalid date range format")

    # Separate HR and manager requests that need admin approval
    hr_manager_requests = leave_requests.filter(
        Q(user__groups__name='HR') |
        Q(user__groups__name='Manager'),
        status='Pending'
    )

    # Other requests for reference
    other_requests = leave_requests.exclude(
        Q(user__groups__name='HR') |
        Q(user__groups__name='Manager'),
        status='Pending'
    )

    # Get organizational hierarchy data
    org_hierarchy = {
        'admin': User.objects.filter(groups__name='Admin').count(),
        'hr': User.objects.filter(groups__name='HR').count(),
        'managers': User.objects.filter(groups__name='Manager').count(),
        'employees': User.objects.filter(groups__name='Employee').count(),
    }

    # Get approval history for audit trail
    approval_history = Leave.objects.exclude(approver=None).values(
        'user__username', 'user__first_name', 'user__last_name',
        'approver__username', 'approver__first_name', 'approver__last_name',
        'status', 'created_at', 'updated_at', 'leave_type'
    ).order_by('-updated_at')[:100]  # Last 100 approvals

    return render(request, 'components/admin/view_leave_requests.html', {
        'hr_manager_requests': hr_manager_requests,
        'other_requests': other_requests,
        'leave_types': Leave.LEAVE_TYPES,
        'status_choices': Leave.STATUS_CHOICES,
        'org_hierarchy': org_hierarchy,
        'approval_history': approval_history,
        'filters': {
            'employee': employee_filter,
            'leave_type': leave_type_filter,
            'status': status_filter,
            'date_range': date_range
        }
    })

@login_required
@user_passes_test(is_admin)
def manage_leave_request_admin(request, leave_id, action):
    """Admin approves or rejects HR and manager leave requests with comprehensive validation."""
    leave_request = get_object_or_404(Leave.objects.select_related('user', 'approver'), id=leave_id)

    # Admin should only approve/reject HR and manager requests
    if not (is_hr(leave_request.user) or is_manager(leave_request.user)):
        messages.error(request, "Admin should only approve HR and Manager leave requests.")
        return redirect('aps_admin:view_leave_requests_admin')

    # Check if leave request is already processed
    if leave_request.status != 'Pending':
        messages.error(request, f"This leave request is already {leave_request.status.lower()}.")
        return redirect('aps_admin:view_leave_requests_admin')

    if request.method == 'POST':
        try:
            with transaction.atomic():
                if action == 'approve':
                    # For HR and managers, check organizational impact
                    if is_hr(leave_request.user):
                        # Check if other HR personnel are available during this period
                        other_hr_on_leave = Leave.objects.filter(
                            user__groups__name='HR',
                            status='Approved',
                            start_date__lte=leave_request.end_date,
                            end_date__gte=leave_request.start_date
                        ).exclude(id=leave_id).count()

                        total_hr = User.objects.filter(groups__name='HR').count()

                        if other_hr_on_leave >= (total_hr / 2):
                            messages.warning(request, f"Warning: {other_hr_on_leave} out of {total_hr} HR personnel will be on leave during this period.")

                    elif is_manager(leave_request.user):
                        # Check team coverage for manager's absence
                        team_size = User.objects.filter(employee__reporting_manager=leave_request.user).count()
                        if team_size > 0:
                            messages.info(request, f"Note: This manager has {team_size} direct reports.")

                    leave_request.status = 'Approved'
                    leave_request.approver = request.user
                    leave_request.approval_date = timezone.now()
                    leave_request.save()

                    # Log the approval
                    logger.info(f"Leave ID {leave_id} for {leave_request.user.username} approved by admin {request.user.username}")

                    messages.success(request, f"Leave for {leave_request.user.get_full_name() or leave_request.user.username} approved.")

                elif action == 'reject':
                    rejection_reason = request.POST.get('rejection_reason')
                    if not rejection_reason:
                        messages.error(request, "Rejection reason is required.")
                        return render(request, 'components/admin/manage_leave.html', {
                            'leave_request': leave_request,
                            'action': action.capitalize(),
                            'leave_balance': Leave.get_leave_balance(leave_request.user)
                        })

                    leave_request.status = 'Rejected'
                    leave_request.approver = request.user
                    leave_request.rejection_reason = rejection_reason
                    leave_request.approval_date = timezone.now()
                    leave_request.save()

                    # Log the rejection
                    logger.info(f"Leave ID {leave_id} for {leave_request.user.username} rejected by admin {request.user.username}")

                    messages.warning(request, f"Leave for {leave_request.user.get_full_name() or leave_request.user.username} rejected.")

                return redirect('aps_admin:view_leave_requests_admin')

        except Exception as e:
            logger.error(f"Error processing leave request: {str(e)}")
            messages.error(request, f"Error processing leave request: {str(e)}")
            return redirect('aps_admin:view_leave_requests_admin')

    # Get organizational context
    org_context = {}

    if is_hr(leave_request.user):
        org_context['role'] = 'HR'
        org_context['other_hr'] = User.objects.filter(groups__name='HR').exclude(id=leave_request.user.id).count()
        org_context['hr_on_leave'] = Leave.objects.filter(
            user__groups__name='HR',
            status='Approved',
            start_date__lte=leave_request.end_date,
            end_date__gte=leave_request.start_date
        ).exclude(user=leave_request.user).count()

    elif is_manager(leave_request.user):
        org_context['role'] = 'Manager'
        org_context['team_size'] = User.objects.filter(employee__reporting_manager=leave_request.user).count()
        org_context['team_on_leave'] = Leave.objects.filter(
            user__employee__reporting_manager=leave_request.user,
            status='Approved',
            start_date__lte=leave_request.end_date,
            end_date__gte=leave_request.start_date
        ).count()

    # Get leave history for context
    leave_history = Leave.objects.filter(user=leave_request.user).exclude(id=leave_id).order_by('-created_at')[:10]

    return render(request, 'components/admin/manage_leave.html', {
        'leave_request': leave_request,
        'action': action.capitalize(),
        'leave_balance': Leave.get_leave_balance(leave_request.user),
        'leave_history': leave_history,
        'org_context': org_context
    })



''' ------------------------------------------- PROJECT AREA ------------------------------------------- '''
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Sum
from django.core.exceptions import ValidationError
from .models import Project, ProjectAssignment, ClientParticipation, User
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.urls import reverse
import json
import logging
from django.contrib.auth.models import Group
from datetime import datetime

# Set up logging
logger = logging.getLogger(__name__)

class ProjectError(Exception):
    """Custom exception for project-related errors"""
    pass

def parse_request_data(request):
    """Helper method to parse request data consistently"""
    try:
        if request.content_type == 'application/json':
            return json.loads(request.body)
        return request.POST
    except json.JSONDecodeError as e:
        logger.error(f"JSON parsing error: {str(e)}")
        raise ProjectError("Invalid JSON data provided")

def validate_project_dates(start_date, deadline):
    """Validate project dates"""
    if start_date > deadline:
        raise ValidationError("Start date cannot be after deadline")
    if deadline < timezone.now().date():
        raise ValidationError("Deadline cannot be in the past")

def handle_assignment_changes(project, assignment, action='assign', role='Employee'):
    """Helper method to handle employee assignment changes"""
    try:
        if action == 'assign':
            if assignment:
                # Reactivate if previously deactivated
                assignment.is_active = True
                assignment.end_date = None
                assignment.role_in_project = role
                assignment.save()
                return False  # Not created, but updated
            return True  # New assignment created
        else:  # remove
            assignment.deactivate()
            return True
    except Exception as e:
        logger.error(f"Assignment change error: {str(e)}")
        raise ProjectError(f"Error {action}ing employee")

def get_users_from_group(group_name):
    """Fetch users dynamically from a given group."""
    try:
        group = Group.objects.get(name=group_name)
        return group.user_set.all()
    except Group.DoesNotExist:
        return User.objects.none()


# Assuming logger is set up
logger = logging.getLogger(__name__)

@login_required
def project_dashboard(request):
    try:
        today = date.today()

        # Get all active projects with related data
        projects = Project.objects.prefetch_related(
            'users',
            'clients',
            'projectassignment_set__user',
            'client_participations'
        ).all()

        # Print the project objects (you can adjust this as needed)
        print("Projects data:")
        for project in projects:
            print(f"Project ID: {project.id}, Project Name: {project.name}, Deadline: {project.deadline}")

            # Print the users related to the project (through projectassignment_set)
            print(f"Assigned Users for Project {project.name}:")
            for assignment in project.projectassignment_set.all():
                user = assignment.user
                print(f"  - User: {user.get_full_name()} (ID: {user.id}, Role: {assignment.get_role_in_project_display()})")

        for project in projects:
            # Calculate project duration and remaining days
            project_duration = (project.deadline - project.start_date).days
            remaining_days = (project.deadline - today).days

            remaining_percentage = max((remaining_days / project_duration) * 100, 0) if project_duration > 0 else 0

            # Set deadline status
            project.is_deadline_close = 0 <= remaining_percentage <= 10

            # Fetch active and removed assignments
            project.active_assignments = project.projectassignment_set.filter(is_active=True)
            project.removed_assignments = project.projectassignment_set.filter(is_active=False)

            # Print active assignments
            print(f"Project: {project.name} (Active Assignments)")
            for assignment in project.active_assignments:
                assignment.user.full_name = assignment.user.get_full_name()
                print(f"  - {assignment.user.full_name} (ID: {assignment.user.id}, Role: {assignment.get_role_in_project_display()})")

            # Print removed assignments
            print(f"Project: {project.name} (Removed Assignments)")
            for assignment in project.removed_assignments:
                assignment.user.full_name = assignment.user.get_full_name()
                print(f"  - {assignment.user.full_name} (ID: {assignment.user.id}, Ended: {assignment.end_date})")

        # Fetch users by group
        employees = get_users_from_group('Employee')
        managers = get_users_from_group('Manager')
        clients = get_users_from_group('Client')

        # Fetch role choices
        role_choices = dict(ProjectAssignment._meta.get_field('role_in_project').choices)

        # Context for rendering the template
        context = {
            'projects': projects,
            'employees': employees,
            'clients': clients,
            'managers': managers,
            'project_statuses': dict(Project._meta.get_field('status').choices),
            'role_choices': role_choices,

        }

        return render(request, 'components/admin/project_view.html', context)

    except Exception as e:
        # Capture exception details
        exc_type, exc_value, exc_tb = sys.exc_info()
        error_details = traceback.format_exception(exc_type, exc_value, exc_tb)

        # Log error and display error message
        logger.error(f"Dashboard error: {str(e)}")
        messages.error(request, "Error loading dashboard")

        # Provide detailed error information in the context for debugging
        context = {
            'error': str(e),
            'error_details': error_details,
        }
        return render(request, 'error.html', context)


@login_required
@require_http_methods(["POST"])
def project_create(request):
    """Handle project creation"""
    try:
        data = request.POST
        start_date_str = data.get('start_date')
        deadline_str = data.get('deadline')

        try:
            start_date = parse_date(start_date_str)
            deadline = parse_date(deadline_str)
            if not start_date or not deadline:
                raise ValidationError("Invalid date format. Expected 'YYYY-MM-DD'.")
        except ValueError:
            raise ValidationError("Invalid date format. Expected 'YYYY-MM-DD'.")

        validate_project_dates(start_date, deadline)

        with transaction.atomic():
            project = Project.objects.create(
                name=data.get('name'),
                description=data.get('description'),
                start_date=start_date,
                deadline=deadline,
                status='Pending'
            )

            client_ids = data.getlist('clients') if hasattr(data, 'getlist') else []
            if client_ids:
                project.clients.set(client_ids)
                for client_id in client_ids:
                    ClientParticipation.objects.create(project=project, client_id=client_id)

            manager_id = data.get('manager')
            if manager_id:
                ProjectAssignment.objects.create(project=project, user_id=manager_id, role_in_project='Manager')

            employee_ids = data.getlist('employees') if hasattr(data, 'getlist') else []
            for emp_id in employee_ids:
                ProjectAssignment.objects.create(project=project, user_id=emp_id, role_in_project='Employee')

        logger.info(f"Project created successfully: {project.name}")
        return redirect(reverse('aps_admin:project_dashboard'))

    except ValidationError as e:
        logger.warning(f"Validation error in project creation: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    except Exception as e:
        logger.error(f"Error creating project: {str(e)}")
        return JsonResponse({'status': 'error', 'message': 'Internal server error'}, status=500)

def update_project_status(project):
    """Automatically update project status based on dates"""
    today = datetime.now().date()

    # If the project is completed (Deadline passed)
    if project.deadline and today > project.deadline and project.status != 'Completed':
        project.status = 'Completed'

    # If the project is in progress (Start date has passed but deadline hasn't passed)
    elif project.start_date and today >= project.start_date and (not project.deadline or today <= project.deadline):
        project.status = 'In Progress'

    # If the project is on hold or any other condition you may define
    elif project.status != 'On Hold':  # Example condition
        project.status = 'On Hold'

    project.save()

@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def project_update(request, project_id):
    """Handle project updates"""
    print(f"Updating project with ID: {project_id}")  # Debug line

    try:
        project = get_object_or_404(Project, id=project_id)
        data = parse_request_data(request)

        # Validate status from the form (if provided)
        new_status = data.get('status')
        if new_status and new_status not in ['Completed', 'In Progress', 'Pending', 'On Hold']:
            raise ValidationError("Invalid project status")

        # Update project status explicitly
        if new_status:
            project.status = new_status

        # Convert deadline to a date object if provided and ensure it's a string
        # Handle deadline field
        if 'deadline' in data and data['deadline']:
            deadline = data.get('deadline')
            try:
                # Convert deadline to a date object only if it's not empty
                if deadline:
                    deadline = datetime.strptime(deadline, '%Y-%m-%d').date()

                    # Validate that deadline is not in the past and that it is later than start_date
                    if deadline < datetime.now().date():
                        raise ValidationError("Deadline cannot be in the past.")
                    if project.start_date and deadline < project.start_date:
                        raise ValidationError("Deadline cannot be earlier than the start date.")

                    project.deadline = deadline
            except ValueError:
                raise ValidationError("Invalid date format for deadline. Please use YYYY-MM-DD.")

            project.deadline = deadline

        with transaction.atomic():
            # Update project basic info
            project.name = data.get('name', project.name)
            project.description = data.get('description', project.description)

            project.save()

        return redirect(reverse('aps_admin:project_dashboard'))  # Adjust the name of the URL pattern if needed

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def parse_request_data(request):
    """Helper function to parse data from the request"""
    return {
        'name': request.POST.get('name'),
        'description': request.POST.get('description'),
        'start_date': request.POST.get('start_date'),
        'deadline': request.POST.get('deadline'),
        'status': request.POST.get('status'),
    }


@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def project_delete(request, project_id):
    """Handle project deletion"""
    try:
        project = get_object_or_404(Project, id=project_id)

        with transaction.atomic():
            # Soft delete all assignments
            ProjectAssignment.objects.filter(project=project).update(
                is_active=False,
                end_date=timezone.now().date()
            )

            # Soft delete all client participations
            ClientParticipation.objects.filter(project=project).update(is_active=False)

            # Delete the project
            project.delete()

        logger.info(f"Project deleted successfully: {project.name}")
        return redirect(reverse('aps_admin:project_dashboard'))  # This is fine after defining the URL

    except Exception as e:
        logger.error(f"Error deleting project: {str(e)}")
        return JsonResponse({'status': 'error', 'message': 'Internal server error'}, status=500)


@csrf_exempt
@login_required
@require_http_methods(["POST"])
def assign_employee(request, project_id):
    """Handle employee assignment to project dynamically, including deactivation"""
    try:
        with transaction.atomic():
            print(f"Transaction started for project_id: {project_id}")
            project = get_object_or_404(Project, id=project_id)
            print(f"Project found: {project.name} (ID: {project.id})")

            user_id = request.POST.get('user_id')
            role = request.POST.get('role', 'Employee')
            action = request.POST.get('action', 'assign')  # Action for remove or assign
            print(f"Received data - user_id: {user_id}, role: {role}, action: {action}")

            if not user_id:
                return JsonResponse({
                    'status': 'error',
                    'message': 'User ID is required'
                }, status=400)

            # Ensure role is valid
            role_choices = dict(ProjectAssignment._meta.get_field('role_in_project').choices)
            if role not in role_choices:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Invalid role. Available roles are {", ".join(role_choices.keys())}'
                }, status=400)

            user = get_object_or_404(User, id=user_id)
            print(f"User found: {user.username} (ID: {user.id})")

            if action == 'remove':
                assignment = project.projectassignment_set.filter(user=user, is_active=True).first()

                if not assignment:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'No active assignment found for employee {user.username} in this project'
                    }, status=404)

                # Soft delete by marking inactive
                assignment.is_active = False
                assignment.end_date = timezone.now().date()
                assignment.save()
                return JsonResponse({
                    'status': 'success',
                    'message': f'Employee {user.username} removed from the project'
                })

            # Check if the employee has been previously removed (soft deleted)
            assignment = project.projectassignment_set.filter(user=user, is_active=False).first()

            if assignment:
                # Reactivate the assignment if it was previously removed
                assignment.is_active = True
                assignment.role_in_project = role
                assignment.end_date = None  # Clear end_date if reactivating
                assignment.save()
                return JsonResponse({
                    'status': 'success',
                    'message': f'Employee {user.username} reactivated in the project'
                })
            else:
                # Handle assigning or updating an employee's role if not previously removed
                assignment, created = ProjectAssignment.objects.get_or_create(
                    project=project,
                    user=user
                )

                assignment.role_in_project = role
                assignment.is_active = True
                assignment.save()

                return JsonResponse({
                    'status': 'success',
                    'message': f'Employee {user.username} assigned to the project with role {role}'
                })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@csrf_exempt
@login_required
@require_http_methods(["POST"])
def change_role(request, project_id):
    """Handle changing the role of an assigned employee"""
    try:
        with transaction.atomic():
            project = get_object_or_404(Project, id=project_id)
            user_id = request.POST.get('user_id')
            new_role = request.POST.get('role', 'Employee')

            if not user_id:
                return JsonResponse({'status': 'error', 'message': 'User ID is required'}, status=400)

            # Ensure role is valid
            role_choices = dict(ProjectAssignment._meta.get_field('role_in_project').choices)
            if new_role not in role_choices:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Invalid role. Available roles are {", ".join(role_choices.keys())}'
                }, status=400)

            user = get_object_or_404(User, id=user_id)
            assignment = project.projectassignment_set.filter(user=user, is_active=True).first()

            if not assignment:
                return JsonResponse({'status': 'error', 'message': 'No active assignment found for this user'}, status=404)

            assignment.role_in_project = new_role
            assignment.save()

            return JsonResponse({'status': 'success', 'message': 'Employee role updated successfully'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@csrf_exempt
@login_required
@require_http_methods(["POST"])
def reactivate_employee(request, project_id):
    """Handle reactivating a previously removed employee"""
    try:
        with transaction.atomic():
            project = get_object_or_404(Project, id=project_id)
            user_id = request.POST.get('user_id')

            if not user_id:
                return JsonResponse({'status': 'error', 'message': 'User ID is required'}, status=400)

            user = get_object_or_404(User, id=user_id)

            # Find the most recent inactive assignment for this user in this project
            assignment = ProjectAssignment.objects.filter(
                project=project,
                user=user,
                is_active=False
            ).order_by('-end_date').first()

            if not assignment:
                return JsonResponse({'status': 'error', 'message': 'No removed assignment found for this user'}, status=404)

            # Reactivate the assignment
            assignment.is_active = True
            assignment.end_date = None  # Clear end date
            assignment.save()

            # Log the reactivation
            UserActionLog.objects.create(
                user=user,
                action_type='update',
                action_by=request.user,
                details=f"Reactivated in project: {project.name}"
            )

            return JsonResponse({
                'status': 'success',
                'message': 'Employee reactivated successfully',
                'user_name': user.get_full_name() or user.username,
                'role': assignment.role_in_project
            })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def update_hours(request, project_id):
    """Handle updating worked hours for an assignment"""
    try:
        project = get_object_or_404(Project, id=project_id)
        data = parse_request_data(request)

        user_id = data.get('user_id')
        hours = data.get('hours')

        if not user_id or hours is None:
            raise ValidationError("User ID and hours are required")

        try:
            hours = float(hours)
            if hours < 0:
                raise ValidationError("Hours cannot be negative")
        except ValueError:
            raise ValidationError("Invalid hours value")

        assignment = get_object_or_404(
            ProjectAssignment,
            project=project,
            user_id=user_id,
            is_active=True
        )

        assignment.update_hours(hours)

        logger.info(f"Hours updated successfully: {hours} hours for {user_id} in {project.name}")
        return JsonResponse({
            'status': 'success',
            'message': 'Hours updated successfully',
            'total_hours': assignment.get_total_hours()
        })

    except ValidationError as e:
        logger.warning(f"Validation error in hours update: {str(e)}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    except Exception as e:
        logger.error(f"Error updating hours: {str(e)}")
        return JsonResponse({'status': 'error', 'message': 'Internal server error'}, status=500)



def create_project(request):
    """Helper function to create a project."""
    name = request.POST.get('name')
    description = request.POST.get('description')
    start_date = request.POST.get('start_date')
    due_date = request.POST.get('due_date')
    client_ids = request.POST.getlist('clients')  # Ensure client_ids are being captured here
    print("Client IDs:", client_ids)  # Add this line for debugging

    project = Project.objects.create(
        name=name,
        description=description,
        start_date=start_date,
        deadline=due_date,
        status='Not Started'
    )

    assign_users_to_project(request, project, client_ids)  # Pass client_ids to the assign function
    return project


def update_project(request, project):
    """Helper function to update a project."""
    project.name = request.POST.get('name', project.name)
    project.description = request.POST.get('description', project.description)
    project.status = request.POST.get('status', project.status)
    project.start_date = request.POST.get('start_date', project.start_date)
    project.deadline = request.POST.get('deadline', project.deadline)
    project.save()

    # Reassign clients based on selected client_ids (Handle soft deletes if necessary)
    client_ids = request.POST.getlist('clients')
    # Clear current clients and reassign based on new client_ids
    project.clients.clear()
    for client_id in client_ids:
        client = User.objects.get(id=client_id)
        project.clients.add(client)

    return project


def assign_users_to_project(request, project, client_ids=None):
    """Helper function to assign users to a project."""
    # Assign manager
    manager_id = request.POST.get('manager')
    if manager_id:
        manager = User.objects.get(id=manager_id)
        ProjectAssignment.objects.update_or_create(
            project=project,
            user=manager,
            defaults={'role_in_project': 'Manager', 'hours_worked': 0.0}
        )

    # Assign employees
    employee_ids = request.POST.getlist('employees')
    for employee_id in employee_ids:
        employee = User.objects.get(id=employee_id)
        ProjectAssignment.objects.get_or_create(
            project=project,
            user=employee,
            defaults={'role_in_project': 'Employee', 'hours_worked': 0.0}
        )

    # Assign clients if client_ids are passed
    if client_ids:
        for client_id in client_ids:  # Iterate through the passed client_ids
            client = User.objects.get(id=client_id)
            ClientParticipation.objects.get_or_create(
                project=project,
                client=client,
                defaults={'feedback': '', 'approved': False}
            )


@login_required
@user_passes_test(is_manager)
def manager_project_view(request, action=None, project_id=None):
    """Manager view for managing projects."""

    # Get all managers and employees
    managers = User.objects.filter(groups__name='Manager')
    employees = User.objects.filter(groups__name='Employee')
    clients = User.objects.filter(groups__name='Client')

    # Action to list all projects
    if action == "list":
        # Get the current manager's projects
        assignments = ProjectAssignment.objects.filter(user=request.user, role_in_project='Manager', is_active=True)
        projects = [assignment.project for assignment in assignments]

        return render(request, 'components/manager/project_view.html', {
            'projects': projects,
            'managers': managers,
            'employees': employees,
            'clients': clients,
            'action': 'list'
        })

    # Action to view project details
    elif action == "detail" and project_id:
        project = get_object_or_404(Project, id=project_id)
        active_assignments = project.projectassignment_set.filter(is_active=True)
        # Get all inactive assignments, not just the most recent ones
        removed_assignments = project.projectassignment_set.filter(is_active=False).order_by('-end_date')
        client_participations = project.client_participations.all()

        context = {
            'project': project,
            'active_assignments': active_assignments,
            'removed_assignments': removed_assignments,
            'employees': employees,
            'clients': clients,
            'client_participations': client_participations,
            'role_choices': dict(ProjectAssignment._meta.get_field('role_in_project').choices),
            'action': 'detail'
        }
        return render(request, 'components/manager/project_view.html', context)

    # Action to create a new project
    elif action == "create":
        if request.method == 'POST':
            try:
                with transaction.atomic():
                    # Create the project
                    project = create_project(request)

                    # Log the action - removed log_user_action call

                    messages.success(request, "Project created successfully!")
                    return redirect('aps_manager:project_detail', project_id=project.id)

            except ValidationError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f"Error creating project: {str(e)}")
                logger.error(f"Project creation error: {str(e)}")

        return render(request, 'components/manager/project_view.html', {
            'employees': employees,
            'managers': managers,
            'clients': clients,
            'action': 'create'
        })

    # Action to update an existing project
    elif action == "update" and project_id:
        project = get_object_or_404(Project, id=project_id)

        # Verify manager has permission
        if not project.projectassignment_set.filter(user=request.user, role_in_project='Manager', is_active=True).exists():
            messages.error(request, "You don't have permission to update this project")
            return redirect('aps_manager:project_list')

        if request.method == 'POST':
            try:
                with transaction.atomic():
                    # Update the project
                    updated_project = update_project(request, project)

                    # Handle employee assignments
                    assign_users_to_project(request, updated_project)

                    # Log the action - removed log_user_action call

                    messages.success(request, "Project updated successfully!")
                    return redirect('aps_manager:project_detail', project_id=project.id)

            except ValidationError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f"Error updating project: {str(e)}")
                logger.error(f"Project update error: {str(e)}")

        # Get current project data for the form
        active_assignments = project.projectassignment_set.filter(is_active=True)
        # Include removed assignments for the update view as well
        removed_assignments = project.projectassignment_set.filter(is_active=False).order_by('-end_date')
        client_participations = project.client_participations.all()

        return render(request, 'components/manager/project_view.html', {
            'project': project,
            'employees': employees,
            'managers': managers,
            'clients': clients,
            'active_assignments': active_assignments,
            'removed_assignments': removed_assignments,
            'client_participations': client_participations,
            'action': 'update'
        })

    # Action to manage employees (assign/remove/change role)
    elif action == "manage_employees" and project_id:
        project = get_object_or_404(Project, id=project_id)

        # Verify manager has permission
        if not project.projectassignment_set.filter(user=request.user, role_in_project='Manager', is_active=True).exists():
            messages.error(request, "You don't have permission to manage employees for this project")
            return redirect('aps_manager:project_list')

        if request.method == 'POST':
            try:
                user_id = request.POST.get('user_id')
                role = request.POST.get('role', 'Employee')
                action_type = request.POST.get('action')
                hours = request.POST.get('hours')

                user = get_object_or_404(User, id=user_id)

                with transaction.atomic():
                    if action_type == 'assign':
                        # Check if there's an inactive assignment first
                        inactive_assignment = project.projectassignment_set.filter(
                            user=user,
                            is_active=False
                        ).order_by('-end_date').first()

                        if inactive_assignment:
                            # Reactivate the existing assignment
                            inactive_assignment.is_active = True
                            inactive_assignment.end_date = None
                            inactive_assignment.role_in_project = role
                            inactive_assignment.save()

                            # Log the reactivation
                            UserActionLog.objects.create(
                                user=user,
                                action_type='update',
                                action_by=request.user,
                                details=f"Reactivated in project: {project.name}"
                            )
                        else:
                            # Create new assignment
                            ProjectAssignment.objects.create(
                                project=project,
                                user=user,
                                role_in_project=role,
                                is_active=True
                            )

                    elif action_type == 'remove':
                        # Soft delete by marking inactive
                        assignment = project.projectassignment_set.filter(user=user, is_active=True).first()
                        if assignment:
                            assignment.is_active = False
                            assignment.end_date = timezone.now().date()
                            assignment.save()
                            # Removed log_user_action call

                    elif action_type == 'change_role':
                        # Update role
                        assignment = project.projectassignment_set.filter(user=user, is_active=True).first()
                        if assignment:
                            assignment.role_in_project = role
                            assignment.save()
                            # Removed log_user_action call

                    elif action_type == 'update_hours' and hours:
                        # Update worked hours
                        assignment = project.projectassignment_set.filter(user=user, is_active=True).first()
                        if assignment:
                            try:
                                hours_float = float(hours)
                                assignment.hours_worked = hours_float
                                assignment.save()
                                # Removed log_user_action call
                            except ValueError:
                                raise ValidationError("Invalid hours value")

                    elif action_type == 'reactivate':
                        # Find the most recent inactive assignment for this user
                        inactive_assignment = project.projectassignment_set.filter(
                            user=user,
                            is_active=False
                        ).order_by('-end_date').first()

                        if inactive_assignment:
                            # Reactivate the assignment
                            inactive_assignment.is_active = True
                            inactive_assignment.end_date = None
                            inactive_assignment.save()

                            # Log the reactivation
                            UserActionLog.objects.create(
                                user=user,
                                action_type='update',
                                action_by=request.user,
                                details=f"Reactivated in project: {project.name}"
                            )
                        else:
                            raise ValidationError("No inactive assignment found for this user")

                messages.success(request, "Employee assignment updated successfully!")

            except ValidationError as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, f"Error managing employees: {str(e)}")
                logger.error(f"Employee management error: {str(e)}")

        return redirect('aps_manager:project_detail', project_id=project.id)

    return redirect('aps_manager:project_list')

''' ------------------------------------------- ATTENDACE AREA ------------------------------------------- '''
# views.py
import pytz
from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse
from django.db.models import Q, Count, Avg
from django.utils import timezone
from datetime import datetime, timedelta
import json
from .services.attendance_service import AttendanceService
from .services.user_service import UserService
from .services.date_service import DateService

# Asia/Kolkata timezone
IST_TIMEZONE = pytz.timezone('Asia/Kolkata')

def get_current_time_ist():
    """Return current time in Asia/Kolkata timezone (aware)."""
    current_time = timezone.now().astimezone(IST_TIMEZONE)
    print(f"Current IST time: {current_time}")
    return current_time

def to_ist(dt):
    """Convert a datetime to Asia/Kolkata timezone (aware)."""
    if dt is None:
        print("Converting None to IST: None returned")
        return None
    if timezone.is_naive(dt):
        ist_time = IST_TIMEZONE.localize(dt)
    else:
        ist_time = dt.astimezone(IST_TIMEZONE)
    print(f"Converting {dt} to IST: {ist_time}")
    return ist_time

# Helper function to check if user is HR
def is_hr_check(user):
    return user.groups.filter(name__in=["HR", "Manager"]).exists()

# Helper function to check if user is HR or Admin
def is_hr_or_admin_check(user):
    return user.groups.filter(name__in=["HR", "Admin"]).exists()

# Helper function to check if user is Manager
def is_manager_check(user):
    return user.groups.filter(name="Manager").exists()


@login_required
@user_passes_test(is_hr_check)
def attendance_analytics(request):
    """
    Comprehensive attendance analytics dashboard with modal support
    """
    print("[attendance_analytics] Called attendance_analytics view.")

    # Initialize services
    print("[attendance_analytics] Initializing AttendanceService, UserService, DateService.")
    attendance_service = AttendanceService()
    user_service = UserService()
    date_service = DateService()

    # Get date parameters
    selected_date = request.GET.get('date')
    time_period = request.GET.get('time_period', 'today')
    print(f"[attendance_analytics] Received selected_date: {selected_date}, time_period: {time_period}")

    try:
        if selected_date:
            print(f"[attendance_analytics] Parsing selected_date: {selected_date}")
            start_date = end_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        else:
            print(f"[attendance_analytics] Getting date range for time_period: {time_period}")
            date_range = date_service.get_date_range(time_period)
            start_date = date_range['start_date']
            end_date = date_range['end_date']
    except ValueError:
        print("[attendance_analytics] ValueError in date parsing, falling back to current date.")
        current_date = date_service.get_current_date()
        start_date = end_date = current_date

    # Get search term for employee search
    search_term = request.GET.get('search', '').strip()
    print(f"[attendance_analytics] Search term: '{search_term}'")

    # Get all required data
    print(f"[attendance_analytics] Fetching overall attendance stats for {start_date} to {end_date}")
    overall_stats = attendance_service.get_attendance_overview(start_date, end_date)
    print(f"[attendance_analytics] Fetching location-wise attendance stats for {start_date} to {end_date}")
    location_stats = attendance_service.get_location_wise_attendance(start_date, end_date)
    print("[attendance_analytics] Fetching top absent users (30 days, limit 10)")
    top_absent_users = attendance_service.get_top_absent_users(days=30, limit=10)
    print("[attendance_analytics] Fetching top late users (30 days, limit 10)")
    top_late_users = attendance_service.get_top_late_users(days=30, limit=10)
    print(f"[attendance_analytics] Fetching yet to clock in users for {start_date}")
    yet_to_clock_in_users = attendance_service.get_yet_to_clock_in_users(start_date)

    # Get total employee statistics
    print("[attendance_analytics] Fetching total employee count")
    total_employees = user_service.get_total_employee_count()
    print("[attendance_analytics] Fetching all locations")
    all_locations = user_service.get_all_locations()

    # Calculate attendance percentages for total employees
    print("[attendance_analytics] Calculating attendance percentages")
    present_percentage = round((overall_stats['present_count'] / total_employees * 100), 2) if total_employees > 0 else 0
    absent_percentage = round((overall_stats['absent_count'] / total_employees * 100), 2) if total_employees > 0 else 0
    leave_percentage = round((overall_stats['leave_count'] / total_employees * 100), 2) if total_employees > 0 else 0

    # Search functionality
    search_results = []
    if search_term:
        print(f"[attendance_analytics] Performing user search for: {search_term}")
        search_results = user_service.search_users(search_term)

    # Get current IST time for display
    current_ist_time = get_current_time_ist()

    print("[attendance_analytics] Preparing context for template rendering")
    context = {
        # Date and time period
        'start_date': start_date,
        'end_date': end_date,
        'selected_date': selected_date or start_date.strftime('%Y-%m-%d'),
        'time_period': time_period,
        'current_ist_time': current_ist_time,

        # Overall statistics
        'overall_stats': overall_stats,
        'total_employees': total_employees,
        'present_percentage': present_percentage,
        'absent_percentage': absent_percentage,
        'leave_percentage': leave_percentage,

        # Location-wise data
        'location_stats': location_stats,
        'all_locations': all_locations,

        # Top users data
        'top_absent_users': top_absent_users,
        'top_late_users': top_late_users,
        'yet_to_clock_in_users': yet_to_clock_in_users,

        # Search
        'search_term': search_term,
        'search_results': search_results,

        # Flags for conditional rendering
        'has_location_data': len(location_stats) > 0,
        'has_status_data': len(overall_stats.get('status_counts', [])) > 0,
        'has_absent_data': len(top_absent_users) > 0,
        'has_late_data': len(top_late_users) > 0,
        'has_yet_to_clock_in_data': len(yet_to_clock_in_users) > 0,
    }
    print("[attendance_analytics] Rendering attendance_analytics.html with context.")
    return render(request, 'components/hr/attendance/attendance_analytics.html', context)


@login_required
@user_passes_test(is_hr_check)
def get_status_users_modal(request):
    """
    AJAX endpoint to get users by status for modal display
    """
    print("[get_status_users_modal] Called get_status_users_modal view.")
    status = request.GET.get('status', '')
    location = request.GET.get('location', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    print(f"[get_status_users_modal] Params - status: {status}, location: {location}, start_date: {start_date_str}, end_date: {end_date_str}")

    try:
        # Parse dates
        print("[get_status_users_modal] Parsing start and end dates.")
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            # Get current date in IST
            start_date = get_current_time_ist().date()

        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = start_date

        # Get attendance service
        print("[get_status_users_modal] Initializing AttendanceService.")
        attendance_service = AttendanceService()

        # Get users by status
        print(f"[get_status_users_modal] Fetching users by status: {status}, location: {location}, start_date: {start_date}, end_date: {end_date}")
        users_data = attendance_service.get_users_by_status(
            status=status,
            location=location if location else None,
            start_date=start_date,
            end_date=end_date
        )

        # Convert time fields to IST
        for user in users_data:
            if 'clock_in' in user and user['clock_in']:
                try:
                    clock_in_time = datetime.strptime(user['clock_in'], '%H:%M:%S')
                    clock_in_time = timezone.make_aware(clock_in_time, timezone.utc)
                    user['clock_in'] = to_ist(clock_in_time).strftime('%H:%M:%S')
                except (ValueError, TypeError):
                    pass

            if 'clock_out' in user and user['clock_out']:
                try:
                    clock_out_time = datetime.strptime(user['clock_out'], '%H:%M:%S')
                    clock_out_time = timezone.make_aware(clock_out_time, timezone.utc)
                    user['clock_out'] = to_ist(clock_out_time).strftime('%H:%M:%S')
                except (ValueError, TypeError):
                    pass

        # Determine table headers based on status
        print(f"[get_status_users_modal] Getting table headers for status: {status}")
        headers = get_table_headers(status)

        print("[get_status_users_modal] Returning users data as JsonResponse.")
        return JsonResponse({
            'success': True,
            'status': status,
            'location': location,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'users': users_data,
            'headers': headers,
            'total_count': len(users_data),
            'current_ist_time': get_current_time_ist().strftime('%Y-%m-%d %H:%M:%S')
        })

    except Exception as e:
        print(f"[get_status_users_modal] Error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e),
            'users': [],
            'headers': [],
            'total_count': 0
        })


def get_table_headers(status):
    """
    Get appropriate table headers based on attendance status
    """
    print(f"[get_table_headers] Called with status: {status}")
    base_headers = ['Name', 'Location']

    if status in ['Present', 'Present & Late']:
        print("[get_table_headers] Returning headers for Present/Present & Late")
        return base_headers + ['Clock In (IST)', 'Clock Out (IST)', 'Late By (mins)', 'Hours', 'Shift']
    elif status == 'On Leave':
        print("[get_table_headers] Returning headers for On Leave")
        return base_headers + ['Leave Type', 'Start Date', 'End Date', 'Days']
    elif status == 'Absent':
        print("[get_table_headers] Returning headers for Absent")
        return base_headers + ['Shift', 'Date']
    elif status == 'Yet to Clock In':
        print("[get_table_headers] Returning headers for Yet to Clock In")
        return base_headers + ['Shift', 'Shift Timing', 'Late By (mins)']
    else:
        print("[get_table_headers] Returning default headers")
        return base_headers + ['Status', 'Date']


@login_required
@user_passes_test(is_hr_check)
def get_attendance_by_date(request):
    """
    AJAX endpoint to get attendance data for a specific date
    """
    print("[get_attendance_by_date] Called get_attendance_by_date view.")
    date_str = request.GET.get('date', '')
    print(f"[get_attendance_by_date] Received date: {date_str}")

    try:
        # Parse date
        if date_str:
            print(f"[get_attendance_by_date] Parsing date: {date_str}")
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            print("[get_attendance_by_date] No date provided, using current date in IST.")
            selected_date = get_current_time_ist().date()

        # Get attendance service
        print("[get_attendance_by_date] Initializing AttendanceService.")
        attendance_service = AttendanceService()

        # Get attendance data for the date
        print(f"[get_attendance_by_date] Fetching attendance data for {selected_date}")
        attendance_data = attendance_service.get_attendance_by_date(selected_date)

        # Convert time fields to IST
        for record in attendance_data:
            if 'clock_in_time' in record and record['clock_in_time']:
                try:
                    clock_in_time = datetime.strptime(record['clock_in_time'], '%H:%M:%S')
                    clock_in_time = timezone.make_aware(clock_in_time, timezone.utc)
                    record['clock_in_time'] = to_ist(clock_in_time).strftime('%H:%M:%S')
                except (ValueError, TypeError):
                    pass

            if 'clock_out_time' in record and record['clock_out_time']:
                try:
                    clock_out_time = datetime.strptime(record['clock_out_time'], '%H:%M:%S')
                    clock_out_time = timezone.make_aware(clock_out_time, timezone.utc)
                    record['clock_out_time'] = to_ist(clock_out_time).strftime('%H:%M:%S')
                except (ValueError, TypeError):
                    pass

        print(f"[get_attendance_by_date] Fetching overall stats for {selected_date}")
        overall_stats = attendance_service.get_attendance_overview(selected_date, selected_date)
        print(f"[get_attendance_by_date] Fetching location stats for {selected_date}")
        location_stats = attendance_service.get_location_wise_attendance(selected_date, selected_date)

        print("[get_attendance_by_date] Returning attendance data as JsonResponse.")
        return JsonResponse({
            'success': True,
            'date': selected_date.strftime('%Y-%m-%d'),
            'current_ist_time': get_current_time_ist().strftime('%Y-%m-%d %H:%M:%S'),
            'attendance_data': attendance_data,
            'overall_stats': overall_stats,
            'location_stats': location_stats
        })

    except Exception as e:
        print(f"[get_attendance_by_date] Error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
@user_passes_test(is_hr_check)
def export_attendance_data(request):
    """
    Export attendance data to CSV/Excel
    """
    print("[export_attendance_data] Called export_attendance_data view.")
    from django.http import HttpResponse
    import csv

    status = request.GET.get('status', '')
    location = request.GET.get('location', '')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    export_format = request.GET.get('format', 'csv')
    print(f"[export_attendance_data] Params - status: {status}, location: {location}, start_date: {start_date_str}, end_date: {end_date_str}, format: {export_format}")

    try:
        # Parse dates
        print("[export_attendance_data] Parsing start and end dates.")
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else get_current_time_ist().date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else start_date

        # Get attendance service
        print("[export_attendance_data] Initializing AttendanceService.")
        attendance_service = AttendanceService()

        # Get users by status
        print(f"[export_attendance_data] Fetching users by status: {status}, location: {location}, start_date: {start_date}, end_date: {end_date}")
        users_data = attendance_service.get_users_by_status(
            status=status,
            location=location if location else None,
            start_date=start_date,
            end_date=end_date
        )

        # Create response
        print("[export_attendance_data] Creating HttpResponse for CSV export.")
        response = HttpResponse(content_type='text/csv')
        filename = f"attendance_{status.lower().replace(' ', '_')}_{start_date}_{end_date}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Write CSV data
        writer = csv.writer(response)

        # Write headers
        print(f"[export_attendance_data] Writing headers for status: {status}")
        headers = get_table_headers(status)
        writer.writerow(headers)

        # Write data rows
        print("[export_attendance_data] Writing user data rows.")
        for user in users_data:
            row = []
            for header in headers:
                key = header.lower().replace(' ', '_').replace('(', '').replace(')', '').replace('-', '_')
                if key == 'name':
                    row.append(user.get('name', ''))
                elif key == 'location':
                    row.append(user.get('location', ''))
                elif key == 'clock_in_ist':
                    row.append(user.get('clock_in', ''))
                elif key == 'clock_out_ist':
                    row.append(user.get('clock_out', ''))
                elif key == 'late_by_mins':
                    row.append(user.get('late_by_mins', 0))
                elif key == 'hours':
                    row.append(user.get('hours', ''))
                elif key == 'shift':
                    row.append(user.get('shift', ''))
                elif key == 'leave_type':
                    row.append(user.get('leave_type', ''))
                elif key == 'start_date':
                    row.append(user.get('start_date', ''))
                elif key == 'end_date':
                    row.append(user.get('end_date', ''))
                elif key == 'days':
                    row.append(user.get('days', ''))
                elif key == 'date':
                    row.append(user.get('date', ''))
                elif key == 'shift_timing':
                    row.append(user.get('shift_timing', ''))
                else:
                    row.append('')
            writer.writerow(row)

        print("[export_attendance_data] Returning CSV response.")
        return response

    except Exception as e:
        print(f"[export_attendance_data] Error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
@user_passes_test(is_hr_check)
def get_user_attendance_details(request):
    """
    AJAX endpoint to get detailed attendance information for a specific user
    """
    print("[get_user_attendance_details] Called get_user_attendance_details view.")
    user_id = request.GET.get('user_id')
    start_date_str = request.GET.get('start_date', '')
    end_date_str = request.GET.get('end_date', '')
    print(f"[get_user_attendance_details] Params - user_id: {user_id}, start_date: {start_date_str}, end_date: {end_date_str}")

    try:
        # Parse dates
        if start_date_str and end_date_str:
            print("[get_user_attendance_details] Parsing start and end dates.")
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            print("[get_user_attendance_details] No dates provided, defaulting to last 30 days.")
            end_date = get_current_time_ist().date()
            start_date = end_date - timedelta(days=30)

        # Get services
        print("[get_user_attendance_details] Initializing AttendanceService and UserService.")
        attendance_service = AttendanceService()
        user_service = UserService()

        # Get user basic info
        print(f"[get_user_attendance_details] Fetching basic info for user_id: {user_id}")
        user_info = user_service.get_user_basic_info(int(user_id))

        # Get user attendance history
        print(f"[get_user_attendance_details] Fetching attendance history for user_id: {user_id}, {start_date} to {end_date}")
        attendance_history = attendance_service.get_user_attendance_history(
            int(user_id), start_date, end_date
        )

        # Calculate attendance percentage
        print(f"[get_user_attendance_details] Calculating attendance percentage for user_id: {user_id}, {start_date} to {end_date}")
        attendance_percentage = attendance_service.calculate_attendance_percentage(
            int(user_id), start_date, end_date
        )

        print("[get_user_attendance_details] Returning user attendance details as JsonResponse.")
        return JsonResponse({
            'success': True,
            'user_info': user_info,
            'attendance_history': attendance_history,
            'attendance_percentage': attendance_percentage,
            'date_range': {
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d')
            }
        })

    except Exception as e:
        print(f"[get_user_attendance_details] Error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
@user_passes_test(is_hr_check)
def get_dashboard_stats(request):
    """
    AJAX endpoint to refresh dashboard statistics
    """
    print("[get_dashboard_stats] Called get_dashboard_stats view.")
    try:
        # Get date parameters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        print(f"[get_dashboard_stats] Params - start_date: {start_date_str}, end_date: {end_date_str}")

        # Parse dates
        if start_date_str and end_date_str:
            print("[get_dashboard_stats] Parsing start and end dates.")
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            print("[get_dashboard_stats] No dates provided, using current date.")
            start_date = end_date = get_current_time_ist().date()

        # Get services
        print("[get_dashboard_stats] Initializing AttendanceService and UserService.")
        attendance_service = AttendanceService()
        user_service = UserService()

        # Get fresh statistics
        print(f"[get_dashboard_stats] Fetching overall stats for {start_date} to {end_date}")
        overall_stats = attendance_service.get_attendance_overview(start_date, end_date)
        print(f"[get_dashboard_stats] Fetching location stats for {start_date} to {end_date}")
        location_stats = attendance_service.get_location_wise_attendance(start_date, end_date)
        print("[get_dashboard_stats] Fetching top absent users (30 days, limit 10)")
        top_absent_users = attendance_service.get_top_absent_users(days=30, limit=10)
        print("[get_dashboard_stats] Fetching top late users (30 days, limit 10)")
        top_late_users = attendance_service.get_top_late_users(days=30, limit=10)

        # Get total employees for percentage calculation
        print("[get_dashboard_stats] Fetching total employee count")
        total_employees = user_service.get_total_employee_count()

        print("[get_dashboard_stats] Returning dashboard stats as JsonResponse.")
        return JsonResponse({
            'success': True,
            'overall_stats': overall_stats,
            'location_stats': location_stats,
            'top_absent_users': top_absent_users,
            'top_late_users': top_late_users,
            'total_employees': total_employees,
            'date_range': {
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d')
            }
        })

    except Exception as e:
        print(f"[get_dashboard_stats] Error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })



'''---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------'''


@user_passes_test(is_hr_check)
def attendance_export(request):
    """
    Export attendance data based on filters in various formats (CSV, Excel)
    Supports global search filter.
    """
    from .models import Attendance, User, UserDetails
    from django.http import HttpResponse
    import csv
    import io
    import xlsxwriter

    # Get query parameters with defaults
    time_period = request.GET.get('time_period', 'today')
    custom_start = request.GET.get('start_date')
    custom_end = request.GET.get('end_date')
    location = request.GET.get('location')
    user_id = request.GET.get('user_id')
    status = request.GET.get('status')
    export_format = request.GET.get('format', 'csv')
    global_search = request.GET.get('search', '').strip()

    # Get appropriate date filters
    date_filters = get_time_filters(time_period)
    start_date = date_filters['start_date']
    end_date = date_filters['end_date']

    # Override with custom dates if provided
    if custom_start:
        try:
            start_date = parse_date(custom_start)
        except:
            pass

    if custom_end:
        try:
            end_date = parse_date(custom_end)
        except:
            pass

    # Base query for attendance records
    base_query = Attendance.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).select_related('user')

    # Apply filters if provided
    if location:
        user_ids = UserDetails.objects.filter(work_location=location).values_list('user_id', flat=True)
        base_query = base_query.filter(user_id__in=user_ids)

    if user_id:
        base_query = base_query.filter(user_id=user_id)

    if status:
        base_query = base_query.filter(status=status)

    # Global search filter
    if global_search:
        user_ids_from_details = UserDetails.objects.filter(
            Q(work_location__icontains=global_search)
        ).values_list('user_id', flat=True)
        base_query = base_query.filter(
            Q(user__username__icontains=global_search) |
            Q(user__first_name__icontains=global_search) |
            Q(user__last_name__icontains=global_search) |
            Q(user__id__icontains=global_search) |
            Q(user_id__in=user_ids_from_details)
        )

    # Prepare header and data for export
    headers = [
        'Employee ID', 'Username', 'Employee Name', 'Date', 'Status',
        'Clock In', 'Clock Out', 'Total Hours', 'Late Minutes',
        'Location', 'Leave Type', 'Regularization Status'
    ]

    # Generate the export file
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="attendance_export_{start_date}_to_{end_date}.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)

        for record in base_query:
            writer.writerow([
                record.user.id,
                record.user.username,
                f"{record.user.first_name} {record.user.last_name}",
                record.date,
                record.status,
                record.clock_in_time.strftime('%H:%M:%S') if record.clock_in_time else 'N/A',
                record.clock_out_time.strftime('%H:%M:%S') if record.clock_out_time else 'N/A',
                float(record.total_hours) if record.total_hours else 0,
                record.late_minutes,
                record.location,
                record.leave_type or 'N/A',
                record.regularization_status or 'N/A'
            ])

        return response

    elif export_format == 'excel':
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet()

        # Define styles
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'color': 'white',
            'border': 1
        })

        date_format = workbook.add_format({'num_format': 'yyyy-mm-dd'})
        time_format = workbook.add_format({'num_format': 'hh:mm:ss'})

        # Write headers
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)

        # Write data
        for row_num, record in enumerate(base_query, 1):
            worksheet.write(row_num, 0, record.user.id)
            worksheet.write(row_num, 1, record.user.username)
            worksheet.write(row_num, 2, f"{record.user.first_name} {record.user.last_name}")
            worksheet.write(row_num, 3, record.date, date_format)
            worksheet.write(row_num, 4, record.status)

            if record.clock_in_time:
                worksheet.write(row_num, 5, record.clock_in_time.strftime('%H:%M:%S'), time_format)
            else:
                worksheet.write(row_num, 5, 'N/A')

            if record.clock_out_time:
                worksheet.write(row_num, 6, record.clock_out_time.strftime('%H:%M:%S'), time_format)
            else:
                worksheet.write(row_num, 6, 'N/A')

            worksheet.write(row_num, 7, float(record.total_hours) if record.total_hours else 0)
            worksheet.write(row_num, 8, record.late_minutes)
            worksheet.write(row_num, 9, record.location)
            worksheet.write(row_num, 10, record.leave_type or 'N/A')
            worksheet.write(row_num, 11, record.regularization_status or 'N/A')

        # Auto-adjust column width
        for col_num in range(len(headers)):
            worksheet.set_column(col_num, col_num, 15)

        workbook.close()
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="attendance_export_{start_date}_to_{end_date}.xlsx"'
        return response

    return HttpResponse("Invalid export format specified", status=400)


@login_required
@user_passes_test(is_hr_check)
def attendance_detail_analysis(request, user_id=None):
    """
    Detailed attendance analysis for a specific user or all users
    Supports global search filter for all users view.
    """
    from .models import Attendance, User, UserDetails, LeaveRequest
    from django.db.models import Avg, Sum
    from django.http import JsonResponse
    from django.shortcuts import render

    # Get query parameters with defaults
    time_period = request.GET.get('time_period', 'this_month')
    custom_start = request.GET.get('start_date')
    custom_end = request.GET.get('end_date')
    global_search = request.GET.get('search', '').strip()

    # Get appropriate date filters
    date_filters = get_time_filters(time_period)
    start_date = date_filters['start_date']
    end_date = date_filters['end_date']

    # Override with custom dates if provided
    if custom_start:
        try:
            start_date = parse_date(custom_start)
        except:
            pass

    if custom_end:
        try:
            end_date = parse_date(custom_end)
        except:
            pass

    # Base query
    if user_id:
        user = User.objects.get(id=user_id)
        base_query = Attendance.objects.filter(
            user=user,
            date__gte=start_date,
            date__lte=end_date
        ).order_by('date')

        # Get user details
        try:
            user_details = UserDetails.objects.get(user=user)
            work_location = user_details.work_location
        except:
            work_location = "Not specified"

        # Get leave history
        leave_history = LeaveRequest.objects.filter(
            user=user,
            start_date__gte=start_date,
            end_date__lte=end_date
        ).order_by('-start_date')

        # Calculate attendance metrics
        attendance_metrics = {
            'total_days': (end_date - start_date).days + 1,
            'present_days': base_query.filter(status='Present').count(),
            'present_late_days': base_query.filter(status='Present & Late').count(),
            'absent_days': base_query.filter(status='Absent').count(),
            'leave_days': base_query.filter(status='On Leave').count(),
            'wfh_days': base_query.filter(status='Work From Home').count(),
            'weekend_days': base_query.filter(status='Weekend').count(),
            'holiday_days': base_query.filter(status='Holiday').count(),
            'avg_working_hours': base_query.filter(total_hours__isnull=False).aggregate(avg=Avg('total_hours'))['avg'] or 0,
            'total_late_minutes': base_query.aggregate(total=Sum('late_minutes'))['total'] or 0,
            'avg_late_minutes': base_query.filter(late_minutes__gt=0).aggregate(avg=Avg('late_minutes'))['avg'] or 0,
            'work_location': work_location
        }

        # Calculate attendance percentage
        working_days = attendance_metrics['total_days'] - attendance_metrics['weekend_days'] - attendance_metrics['holiday_days']
        present_days = attendance_metrics['present_days'] + attendance_metrics['present_late_days'] + attendance_metrics['wfh_days']

        if working_days > 0:
            attendance_metrics['attendance_percentage'] = round((present_days / working_days) * 100, 2)
        else:
            attendance_metrics['attendance_percentage'] = 0

        # Get daily trends for chart
        daily_stats = base_query.values('date', 'status', 'total_hours', 'late_minutes')

        context = {
            'user': user,
            'start_date': start_date,
            'end_date': end_date,
            'attendance_records': base_query,
            'attendance_metrics': attendance_metrics,
            'leave_history': leave_history,
            'time_period': time_period,
            'daily_stats': list(daily_stats)
        }

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'attendance_metrics': attendance_metrics,
                'attendance_records': list(base_query.values()),
                'leave_history': list(leave_history.values()),
                'daily_stats': list(daily_stats)
            })

        return render(request, 'components/hr/attendance/attendance_analytics.html', context)

    else:
        # Analysis for all users
        user_analysis = []
        users = User.objects.filter(is_active=True)
        if global_search:
            user_ids_from_details = UserDetails.objects.filter(
                Q(work_location__icontains=global_search)
            ).values_list('user_id', flat=True)
            users = users.filter(
                Q(username__icontains=global_search) |
                Q(first_name__icontains=global_search) |
                Q(last_name__icontains=global_search) |
                Q(id__icontains=global_search) |
                Q(id__in=user_ids_from_details)
            )

        for user in users:
            user_query = Attendance.objects.filter(
                user=user,
                date__gte=start_date,
                date__lte=end_date
            )

            # Calculate metrics for each user
            present_days = user_query.filter(status='Present').count()
            present_late_days = user_query.filter(status='Present & Late').count()
            absent_days = user_query.filter(status='Absent').count()
            leave_days = user_query.filter(status='On Leave').count()
            wfh_days = user_query.filter(status='Work From Home').count()
            weekend_days = user_query.filter(status='Weekend').count()
            holiday_days = user_query.filter(status='Holiday').count()

            # Get user's location
            try:
                user_details = UserDetails.objects.get(user=user)
                work_location = user_details.work_location
            except:
                work_location = "Not specified"

            # Calculate attendance percentage
            total_days = (end_date - start_date).days + 1
            working_days = total_days - weekend_days - holiday_days
            present_total = present_days + present_late_days + wfh_days

            attendance_percentage = round((present_total / max(working_days, 1)) * 100, 2)

            user_analysis.append({
                'user_id': user.id,
                'username': user.username,
                'name': f"{user.first_name} {user.last_name}",
                'present_days': present_days,
                'present_late_days': present_late_days,
                'absent_days': absent_days,
                'leave_days': leave_days,
                'wfh_days': wfh_days,
                'attendance_percentage': attendance_percentage,
                'work_location': work_location,
                'avg_working_hours': user_query.filter(total_hours__isnull=False).aggregate(avg=Avg('total_hours'))['avg'] or 0,
                'total_late_minutes': user_query.aggregate(total=Sum('late_minutes'))['total'] or 0
            })

        # Sort users by attendance percentage (descending)
        user_analysis.sort(key=lambda x: x['attendance_percentage'], reverse=True)

        context = {
            'start_date': start_date,
            'end_date': end_date,
            'user_analysis': user_analysis,
            'time_period': time_period,
            'global_search': global_search,
        }

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'user_analysis': user_analysis
            })

        return render(request, 'components/hr/attendance/attendance_analytics.html', context)


from django.shortcuts import render
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q, Count, Avg, Sum, Case, When, Value, IntegerField, F
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
from django.utils import timezone
from datetime import timedelta, date
from calendar import monthrange
from functools import lru_cache
import json

from .models import Attendance, UserDetails, LeaveRequest, User





@login_required
@user_passes_test(is_hr_check)
def hr_attendance_dashboard(request):
    """
    Enhanced Dashboard view for HR showing attendance statistics with dynamic filtering
    """
    today = timezone.localtime(timezone.now()).date()

    # Get filter parameters
    date_from = request.GET.get('date_from', str(today))
    date_to = request.GET.get('date_to', str(today))
    office_location = request.GET.get('location', None)
    user_query = request.GET.get('user_query', None)
    time_period = request.GET.get('time_period', 'daily')  # daily, weekly, monthly, quarterly, yearly

    try:
        date_from = date.fromisoformat(date_from)
        date_to = date.fromisoformat(date_to)
    except ValueError:
        date_from = today
        date_to = today

    # Get available locations for filter dropdown
    locations = UserDetails.objects.exclude(
        work_location__isnull=True
    ).exclude(
        work_location__exact=''
    ).values_list('work_location', flat=True).distinct().order_by('work_location')

    # Get counts for today (quick summary always showing today regardless of filters)
    present_count = Attendance.objects.filter(
        date=today,
        status__in=['Present', 'Present & Late', 'Work From Home']
    ).count()

    absent_count = Attendance.objects.filter(
        date=today,
        status='Absent'
    ).count()

    leave_count = Attendance.objects.filter(
        date=today,
        status__in=['On Leave', 'Half Day']
    ).count()

    late_count = Attendance.objects.filter(
        date=today,
        status='Present & Late'
    ).count()

    wfh_count = Attendance.objects.filter(
        date=today,
        status='Work From Home'
    ).count()

    comp_off_count = Attendance.objects.filter(
        date=today,
        status='Comp Off'
    ).count()

    not_marked_count = Attendance.objects.filter(
        date=today,
        status__in=['Not Marked', 'Yet to Clock In']
    ).count()

    # Get regularization requests pending approval
    pending_requests = Attendance.objects.filter(
        regularization_status='Pending'
    ).select_related('user').order_by('-date')[:10]

    # Get recent attendance records with optimized query
    recent_attendance = Attendance.objects.filter(
        date__lte=today
    ).select_related('user').order_by('-date')[:15]

    # Get upcoming leave requests
    upcoming_leaves = LeaveRequest.objects.filter(
        status='Approved',
        end_date__gte=today
    ).select_related('user', 'leave_type').order_by('start_date')[:10]

    location_counts = {}
    for loc in locations:
        # Get all users for this location
        users_in_location = UserDetails.objects.filter(work_location=loc).values_list('user_id', flat=True)

        # Get attendance counts for these users
        location_counts[loc] = {
            'Present': Attendance.objects.filter(
                date__range=[date_from, date_to],
                user_id__in=users_in_location,
                status__in=['Present', 'Present & Late', 'Work From Home']
            ).count(),
            'Absent': Attendance.objects.filter(
                date__range=[date_from, date_to],
                user_id__in=users_in_location,
                status='Absent'
            ).count(),
            'Leave': Attendance.objects.filter(
                date__range=[date_from, date_to],
                user_id__in=users_in_location,
                status__in=['On Leave', 'Half Day']
            ).count(),
            'Yet to Mark': Attendance.objects.filter(
                date__range=[date_from, date_to],
                user_id__in=users_in_location,
                status__in=['Not Marked', 'Yet to Clock In']
            ).count()
        }

    # Get comprehensive summary data based on filters
    summary_data = get_attendance_summary_data(
        date_from=date_from,
        date_to=date_to,
        office_location=office_location,
        user_query=user_query,
        time_period=time_period
    )

    if summary_data is None:
        summary_data = {}

    summary_data['location_counts'] = location_counts

    context = {
        'today': today,
        'present_count': present_count,
        'absent_count': absent_count,
        'leave_count': leave_count,
        'late_count': late_count,
        'wfh_count': wfh_count,
        'comp_off_count': comp_off_count,
        'not_marked_count': not_marked_count,
        'pending_requests': pending_requests,
        'recent_attendance': recent_attendance,
        'upcoming_leaves': upcoming_leaves,
        'summary_data': summary_data,
        'locations': locations,
        'selected_location': office_location,
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'user_query': user_query,
        'time_period': time_period,
    }

    return render(request, 'components/hr/attendance/hr_dashboard.html', context)
@login_required
@user_passes_test(is_hr_check)
def get_attendance_details(request):
    """
    Enhanced API endpoint to get user attendance details with advanced filtering
    """
    # Get parameters from request
    status = request.GET.get('status')
    location = request.GET.get('location')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    user_query = request.GET.get('user_query')

    today = timezone.localtime(timezone.now()).date()

    # If no location is specified, use the request parameter
    if not location:
        # Get user's location from their system profile as fallback
        user_details = UserDetails.objects.filter(user=request.user).first()
        location = user_details.work_location if user_details else None

    # Convert date strings to date objects
    try:
        if date_from:
            date_from = date.fromisoformat(date_from)
        else:
            date_from = today

        if date_to:
            date_to = date.fromisoformat(date_to)
        else:
            date_to = today
    except ValueError:
        date_from = today
        date_to = today

    if not status:
        return JsonResponse({'error': 'Status parameter is missing'}, status=400)

    # Map frontend status names to database status values
    status_mapping = {
        'Present': ['Present', 'Present & Late', 'Work From Home', 'Half Day'],
        'Absent': ['Absent'],
        'Leave': ['On Leave', 'Comp Off'],
        'Late': ['Present & Late'],
        'WFH': ['Work From Home'],
        'NotMarked': ['Not Marked', 'Yet to Clock In']
    }

    # Build the user filter based on location and other parameters
    user_filter = Q()

    if location:
        user_filter &= Q(profile__work_location=location)

    if user_query:
        user_filter &= (
            Q(username__icontains=user_query) |
            Q(first_name__icontains=user_query) |
            Q(last_name__icontains=user_query)
        )

    # Get users matching the filter
    user_ids = User.objects.filter(user_filter).values_list('id', flat=True)

    result = []

    # For Leave status, get data from LeaveRequest table
    if status == 'Leave':
        leave_requests = LeaveRequest.objects.filter(
            user_id__in=user_ids,
            status='Approved',
            start_date__lte=date_to,
            end_date__gte=date_from
        ).select_related('user', 'leave_type')

        for leave in leave_requests:
            user_details = UserDetails.objects.filter(user=leave.user).first()
            work_location = user_details.work_location if user_details else "N/A"

            leave_type_name = leave.leave_type.name if leave.leave_type else "N/A"
            half_day_status = "Half Day" if leave.half_day else "Full Day"

            result.append({
                'username': leave.user.username,
                'first_name': leave.user.first_name,
                'last_name': leave.user.last_name,
                'leave_type': leave_type_name,
                'start_date': leave.start_date.strftime('%Y-%m-%d'),
                'end_date': leave.end_date.strftime('%Y-%m-%d'),
                'leave_days': str(leave.leave_days),
                'half_day': half_day_status,
                'work_location': work_location,
                'status': 'On Leave',
                'active_status': 'On Leave'
            })
    else:
        # Get attendance records for these users with the specified status
        attendance_filter = Q(
            user_id__in=user_ids,
            date__gte=date_from,
            date__lte=date_to
        )

        if status in status_mapping:
            attendance_filter &= Q(status__in=status_mapping[status])
        else:
            attendance_filter &= Q(status=status)

        attendance_records = Attendance.objects.filter(attendance_filter).select_related('user')

        for record in attendance_records:
            user_details = UserDetails.objects.filter(user=record.user).first()
            work_location = user_details.work_location if user_details else "N/A"

            # Determine active status
            active_status = "Inactive"
            if record.clock_in_time and not record.clock_out_time:
                active_status = "Active"
            elif record.clock_in_time and record.clock_out_time:
                active_status = "Clocked Out"

            result.append({
                'username': record.user.username,
                'first_name': record.user.first_name,
                'last_name': record.user.last_name,
                'clock_in': record.clock_in_time.strftime('%H:%M:%S') if record.clock_in_time else None,
                'clock_out': record.clock_out_time.strftime('%H:%M:%S') if record.clock_out_time else None,
                'date': record.date.strftime('%Y-%m-%d'),
                'work_location': work_location,
                'late_minutes': record.late_minutes,
                'early_departure_minutes': record.early_departure_minutes,
                'total_hours': float(record.total_hours) if record.total_hours else 0,
                'active_status': active_status,
                'status': record.status
            })

    # Sort results by name for consistent display
    result = sorted(result, key=lambda x: (x['first_name'], x['last_name']))

    # Include the filter parameters in the response for reference
    response_data = {
        'location': location,
        'date_from': date_from.isoformat() if isinstance(date_from, date) else date_from,
        'date_to': date_to.isoformat() if isinstance(date_to, date) else date_to,
        'status': status,
        'count': len(result),
        'attendanceData': result
    }

    return JsonResponse(response_data, safe=False)


def get_attendance_summary_data(
    date_from=None,
    date_to=None,
    office_location=None,
    user_query=None,
    time_period='daily'
):
    """
    Enhanced and optimized function to generate summary data for attendance dashboard
    with dynamic filtering and time period selection.

    Args:
        date_from (date): Start date for filtering
        date_to (date): End date for filtering
        office_location (str): Office location filter
        user_query (str): Search query for user
        time_period (str): Period for aggregation (daily, weekly, monthly, quarterly, yearly)

    Returns:
        dict: Dictionary containing various attendance metrics and trends
    """
    from django.db.models import Prefetch, Q, Count, Avg, Sum, Case, When, Value, IntegerField, F
    from django.db.models.functions import TruncDate, TruncWeek, TruncMonth, TruncQuarter, TruncYear
    from functools import lru_cache
    from calendar import monthrange
    from django.utils import timezone
    from datetime import timedelta, date

    today = timezone.localtime(timezone.now()).date()

    # Set default date range if not provided
    if date_from is None:
        date_from = today
    if date_to is None:
        date_to = today

    # Calculate time period reference dates
    start_of_week = today - timedelta(days=today.weekday())
    start_of_month = today.replace(day=1)
    start_of_quarter = today.replace(month=((today.month-1)//3)*3+1, day=1)
    start_of_year = today.replace(month=1, day=1)

    # Calculate end of month
    _, last_day = monthrange(today.year, today.month)
    end_of_month = today.replace(day=last_day)

    # Determine date range for 3-month period
    three_months_ago = today - timedelta(days=90)

    # Create optimized base queryset
    attendance_qs = Attendance.objects.select_related(
        'user'
    ).only(
        'date', 'status', 'total_hours', 'late_minutes', 'overtime_hours',
        'user__username', 'user__first_name', 'user__last_name',
        'clock_in_time', 'clock_out_time', 'early_departure_minutes'
    )

    # Cache key for query optimization
    cache_key = f"attendance_summary_{date_from}_{date_to}_{office_location}_{user_query}_{time_period}"

    @lru_cache(maxsize=128)
    def get_filtered_queryset(key):
        """Get cached filtered queryset based on parameters"""
        # Apply date range filter
        qs = attendance_qs.filter(date__gte=date_from, date__lte=date_to)

        # Build user filter
        user_filter = Q()

        # Filter by office location with index
        if office_location:
            user_ids = UserDetails.objects.filter(
                work_location=office_location
            ).values_list('user_id', flat=True)
            user_filter &= Q(user_id__in=user_ids)

        # Apply user search filter
        if user_query:
            search_filter = (
                Q(user__username__icontains=user_query) |
                Q(user__first_name__icontains=user_query) |
                Q(user__last_name__icontains=user_query)
            )
            user_filter &= search_filter

        # Apply user filter if any
        if user_filter:
            qs = qs.filter(user_filter)

        return qs

    # Get filtered queryset
    filtered_qs = get_filtered_queryset(cache_key)

    # --- Helper functions for leave requests ---
    def get_leave_stats(start_date, end_date, location=None, user_query=None):
        """Get leave statistics for a given date range with filters"""
        leave_qs = LeaveRequest.objects.filter(
            start_date__lte=end_date,
            end_date__gte=start_date,
        )

        # Apply location filter
        if location:
            user_ids = UserDetails.objects.filter(
                work_location=location
            ).values_list('user_id', flat=True)
            leave_qs = leave_qs.filter(user_id__in=user_ids)

        # Apply user search filter
        if user_query:
            leave_qs = leave_qs.filter(
                Q(user__username__icontains=user_query) |
                Q(user__first_name__icontains=user_query) |
                Q(user__last_name__icontains=user_query)
            )

        # Get stats
        leave_stats = leave_qs.aggregate(
            pending_count=Count('id', filter=Q(status='Pending')),
            approved_count=Count('id', filter=Q(status='Approved')),
            rejected_count=Count('id', filter=Q(status='Rejected')),
            cancelled_count=Count('id', filter=Q(status='Cancelled')),
            total_leave_days=Sum('leave_days', filter=Q(status='Approved'))
        )

        return leave_stats

    def get_leave_type_distribution(start_date, end_date, location=None, user_query=None):
        """Get leave type distribution for a given date range with filters"""
        leave_qs = LeaveRequest.objects.filter(
            start_date__lte=end_date,
            end_date__gte=start_date,
            status='Approved'
        )

        # Apply location filter
        if location:
            user_ids = UserDetails.objects.filter(
                work_location=location
            ).values_list('user_id', flat=True)
            leave_qs = leave_qs.filter(user_id__in=user_ids)

        # Apply user search filter
        if user_query:
            leave_qs = leave_qs.filter(
                Q(user__username__icontains=user_query) |
                Q(user__first_name__icontains=user_query) |
                Q(user__last_name__icontains=user_query)
            )

        leave_types = leave_qs.values(
            'leave_type__name'
        ).annotate(
            count=Count('id'),
            days=Sum('leave_days')
        ).order_by('-count')

        return list(leave_types)

    # --- Get leave statistics for different time periods ---
    daily_leave_stats = get_leave_stats(
        today, today,
        location=office_location,
        user_query=user_query
    )

    weekly_leave_stats = get_leave_stats(
        start_of_week, today,
        location=office_location,
        user_query=user_query
    )

    monthly_leave_stats = get_leave_stats(
        start_of_month, today,
        location=office_location,
        user_query=user_query
    )

    quarterly_leave_stats = get_leave_stats(
        start_of_quarter, today,
        location=office_location,
        user_query=user_query
    )

    yearly_leave_stats = get_leave_stats(
        start_of_year, today,
        location=office_location,
        user_query=user_query
    )

    # Get leave type distribution for the selected time period
    if time_period == 'daily':
        period_leave_types = get_leave_type_distribution(
            date_from, date_to,
            location=office_location,
            user_query=user_query
        )
    elif time_period == 'weekly':
        period_leave_types = get_leave_type_distribution(
            start_of_week, start_of_week + timedelta(days=6),
            location=office_location,
            user_query=user_query
        )
    elif time_period == 'monthly':
        period_leave_types = get_leave_type_distribution(
            start_of_month, end_of_month,
            location=office_location,
            user_query=user_query
        )
    elif time_period == 'quarterly':
        period_leave_types = get_leave_type_distribution(
            start_of_quarter, today,
            location=office_location,
            user_query=user_query
        )
    else:  # yearly
        period_leave_types = get_leave_type_distribution(
            start_of_year, today,
            location=office_location,
            user_query=user_query
        )

    # Upcoming leaves in the next 14 days
    upcoming_leave_filter = Q(
        start_date__gte=today,
        start_date__lte=today + timedelta(days=14),
        status='Approved'
    )

    # Apply location filter for upcoming leaves
    if office_location:
        user_ids = UserDetails.objects.filter(
            work_location=office_location
        ).values_list('user_id', flat=True)
        upcoming_leave_filter &= Q(user_id__in=user_ids)

    # Apply user search filter for upcoming leaves
    if user_query:
        upcoming_leave_filter &= (
            Q(user__username__icontains=user_query) |
            Q(user__first_name__icontains=user_query) |
            Q(user__last_name__icontains=user_query)
        )

    upcoming_leaves = LeaveRequest.objects.filter(
        upcoming_leave_filter
    ).select_related(
        'user', 'leave_type'
    ).values(
        'user__username', 'user__first_name', 'user__last_name',
        'leave_type__name', 'start_date', 'end_date', 'leave_days'
    )

    # Top users with most leave days in the current period
    top_leave_filter = Q(
        status='Approved'
    )

    # Apply date filter based on time period
    if time_period == 'daily':
        top_leave_filter &= Q(start_date__lte=date_to, end_date__gte=date_from)
    elif time_period == 'weekly':
        top_leave_filter &= Q(start_date__lte=start_of_week + timedelta(days=6), end_date__gte=start_of_week)
    elif time_period == 'monthly':
        top_leave_filter &= Q(start_date__lte=end_of_month, end_date__gte=start_of_month)
    elif time_period == 'quarterly':
        top_leave_filter &= Q(start_date__lte=today, end_date__gte=start_of_quarter)
    else:  # yearly
        top_leave_filter &= Q(start_date__lte=today, end_date__gte=start_of_year)

    # Apply location filter for top leave users
    if office_location:
        user_ids = UserDetails.objects.filter(
            work_location=office_location
        ).values_list('user_id', flat=True)
        top_leave_filter &= Q(user_id__in=user_ids)

    # Apply user search filter for top leave users
    if user_query:
        top_leave_filter &= (
            Q(user__username__icontains=user_query) |
            Q(user__first_name__icontains=user_query) |
            Q(user__last_name__icontains=user_query)
        )

    top_leave_users = LeaveRequest.objects.filter(
        top_leave_filter
    ).values(
        'user_id'
    ).annotate(
        total_leave_days=Sum('leave_days')
    ).order_by('-total_leave_days')[:10]

    top_leaves = []
    for leave in top_leave_users:
        user_details = UserDetails.objects.filter(user_id=leave['user_id']).select_related('user').first()
        if user_details:
            top_leaves.append({
                'user__username': user_details.user.username,
                'user__first_name': user_details.user.first_name,
                'user__last_name': user_details.user.last_name,
                'work_location': user_details.work_location,
                'total_leave_days': leave['total_leave_days']
            })

    # --- Generate attendance summaries based on time period ---
    # Helper function for getting period-based statistics
    def get_period_stats(start_date, end_date, period_field=None):
        """Get attendance statistics for a specific period"""
        period_filter = Q(date__gte=start_date, date__lte=end_date)

        # Apply user filter if needed
        if office_location or user_query:
            user_filter = Q()

            if office_location:
                user_ids = UserDetails.objects.filter(
                    work_location=office_location
                ).values_list('user_id', flat=True)
                user_filter &= Q(user_id__in=user_ids)

            if user_query:
                user_filter &= (
                    Q(user__username__icontains=user_query) |
                    Q(user__first_name__icontains=user_query) |
                    Q(user__last_name__icontains=user_query)
                )

            period_filter &= user_filter

        # Get statistics
        stats = attendance_qs.filter(period_filter).aggregate(
            present=Count('id', filter=Q(status__in=['Present', 'Present & Late', 'Work From Home'])),
            absent=Count('id', filter=Q(status='Absent')),
            leave=Count('id', filter=Q(status__in=['On Leave', 'Half Day'])),
            late=Count('id', filter=Q(status='Present & Late')),
            wfh=Count('id', filter=Q(status='Work From Home')),
            comp_off=Count('id', filter=Q(status='Comp Off')),
            not_marked=Count('id', filter=Q(status__in=['Not Marked', 'Yet to Clock In'])),
            avg_hours=Avg('total_hours'),
            avg_late_minutes=Avg(Case(
                When(late_minutes__gt=0, then=F('late_minutes')),
                default=Value(0),
                output_field=IntegerField()
            )),
            early_departure=Count('id', filter=Q(left_early=True)),
            avg_early_minutes=Avg(Case(
                When(early_departure_minutes__gt=0, then=F('early_departure_minutes')),
                default=Value(0),
                output_field=IntegerField()
            )),
            total_overtime=Sum('overtime_hours')
        )

        # Add attendance percentage
        total_records = stats['present'] + stats['absent'] + stats['leave'] + stats['not_marked']
        stats['attendance_percentage'] = round((stats['present'] / total_records * 100), 2) if total_records > 0 else 0

        return stats

    # Get attendance statistics for different time periods
    daily_stats = get_period_stats(date_from, date_to)
    weekly_stats = get_period_stats(start_of_week, start_of_week + timedelta(days=6))
    monthly_stats = get_period_stats(start_of_month, end_of_month)
    quarterly_stats = get_period_stats(start_of_quarter, today)
    yearly_stats = get_period_stats(start_of_year, today)
    three_month_stats = get_period_stats(three_months_ago, today)

    # Get period-based statistics based on selected time period
    if time_period == 'daily':
        period_stats = daily_stats
        trend_start_date = date_from - timedelta(days=30)
        trend_end_date = date_to
        truncate_fn = TruncDate
        date_format = '%d %b'
    elif time_period == 'weekly':
        period_stats = weekly_stats
        trend_start_date = start_of_week - timedelta(days=90)
        trend_end_date = start_of_week + timedelta(days=6)
        truncate_fn = TruncWeek
        date_format = 'Week %W'
    elif time_period == 'monthly':
        period_stats = monthly_stats
        trend_start_date = start_of_month.replace(month=start_of_month.month - 12 if start_of_month.month > 12 else 1, year=start_of_month.year - 1 if start_of_month.month == 1 else start_of_month.year)
        trend_end_date = end_of_month
        truncate_fn = TruncMonth
        date_format = '%b %Y'
    elif time_period == 'quarterly':
        period_stats = quarterly_stats
        trend_start_date = start_of_quarter.replace(year=start_of_quarter.year - 2)
        trend_end_date = today
        truncate_fn = TruncQuarter
        date_format = 'Q%q %Y'
    else:  # yearly
        period_stats = yearly_stats
        trend_start_date = start_of_year.replace(year=start_of_year.year - 5)
        trend_end_date = today
        truncate_fn = TruncYear
        date_format = '%Y'

    # Generate attendance trends
    trend_filter = Q(date__gte=trend_start_date, date__lte=trend_end_date)

    # Apply user filters for trends
    if office_location or user_query:
        user_filter = Q()

        if office_location:
            user_ids = UserDetails.objects.filter(
                work_location=office_location
            ).values_list('user_id', flat=True)
            user_filter &= Q(user_id__in=user_ids)

        if user_query:
            user_filter &= (
                Q(user__username__icontains=user_query) |
                Q(user__first_name__icontains=user_query) |
                Q(user__last_name__icontains=user_query)
            )

        trend_filter &= user_filter

    # Get attendance trend data
    attendance_trend = attendance_qs.filter(trend_filter).annotate(
        period=truncate_fn('date')
    ).values(
        'period'
    ).annotate(
        present=Count('id', filter=Q(status__in=['Present', 'Present & Late', 'Work From Home'])),
        absent=Count('id', filter=Q(status='Absent')),
        leave=Count('id', filter=Q(status__in=['On Leave', 'Half Day'])),
        late=Count('id', filter=Q(status='Present & Late')),
        wfh=Count('id', filter=Q(status='Work From Home')),
        comp_off=Count('id', filter=Q(status='Comp Off')),
        not_marked=Count('id', filter=Q(status__in=['Not Marked', 'Yet to Clock In'])),
        avg_hours=Avg('total_hours'),
        avg_late_minutes=Avg(Case(
            When(late_minutes__gt=0, then=F('late_minutes')),
            default=Value(0),
            output_field=IntegerField()
        )),
        early_departure=Count('id', filter=Q(left_early=True)),
        avg_early_minutes=Avg(Case(
            When(early_departure_minutes__gt=0, then=F('early_departure_minutes')),
            default=Value(0),
            output_field=IntegerField()
        )),
        total_overtime=Sum('overtime_hours')
    ).order_by('period')

    # Format the trend data
    formatted_trend = []
    for item in attendance_trend:
        total_records = item['present'] + item['absent'] + item['leave'] + item['not_marked']
        attendance_percentage = round((item['present'] / total_records * 100), 2) if total_records > 0 else 0

        formatted_trend.append({
            'period': item['period'].strftime(date_format) if hasattr(item['period'], 'strftime') else str(item['period']),
            'present': item['present'],
            'absent': item['absent'],
            'leave': item['leave'],
            'late': item['late'],
            'wfh': item['wfh'],
            'comp_off': item['comp_off'],
            'not_marked': item['not_marked'],
            'avg_hours': round(item['avg_hours'], 2) if item['avg_hours'] is not None else 0,
            'avg_late_minutes': round(item['avg_late_minutes'], 2) if item['avg_late_minutes'] is not None else 0,
            'early_departure': item['early_departure'],
            'avg_early_minutes': round(item['avg_early_minutes'], 2) if item['avg_early_minutes'] is not None else 0,
            'total_overtime': round(item['total_overtime'], 2) if item['total_overtime'] is not None else 0,
            'attendance_percentage': attendance_percentage
        })

    # --- Get top attendance performers ---
    # Find users with highest attendance percentage
    top_attendance_filter = Q(date__gte=date_from, date__lte=date_to)

    # Apply location filter for top attendance
    if office_location:
        user_ids = UserDetails.objects.filter(
            work_location=office_location
        ).values_list('user_id', flat=True)
        top_attendance_filter &= Q(user_id__in=user_ids)

    # Apply user search filter for top attendance
    if user_query:
        top_attendance_filter &= (
            Q(user__username__icontains=user_query) |
            Q(user__first_name__icontains=user_query) |
            Q(user__last_name__icontains=user_query)
        )

    # Get top 10 users with highest attendance percentage
    top_attendance = attendance_qs.filter(top_attendance_filter).values(
        'user_id'
    ).annotate(
        total_days=Count('id'),
        present_days=Count('id', filter=Q(status__in=['Present', 'Present & Late', 'Work From Home']))
    ).annotate(
        attendance_percentage=F('present_days') * 100.0 / F('total_days')
    ).order_by('-attendance_percentage')[:10]

    # Format top attendance data
    top_attendance_data = []
    for item in top_attendance:
        user_details = UserDetails.objects.filter(user_id=item['user_id']).select_related('user').first()
        if user_details:
            top_attendance_data.append({
                'user__username': user_details.user.username,
                'user__first_name': user_details.user.first_name,
                'user__last_name': user_details.user.last_name,
                'work_location': user_details.work_location,
                'attendance_percentage': round(item['attendance_percentage'], 2),
                'present_days': item['present_days'],
                'total_days': item['total_days']
            })

    # --- Get top lateness defaulters ---
    # Find users with highest late instances
    top_late_filter = Q(date__gte=date_from, date__lte=date_to, status='Present & Late')

    # Apply location filter for top late users
    if office_location:
        user_ids = UserDetails.objects.filter(
            work_location=office_location
        ).values_list('user_id', flat=True)
        top_late_filter &= Q(user_id__in=user_ids)

    # Apply user search filter for top late users
    if user_query:
        top_late_filter &= (
            Q(user__username__icontains=user_query) |
            Q(user__first_name__icontains=user_query) |
            Q(user__last_name__icontains=user_query)
        )

    # Get top 10 users with highest lateness count
    top_late = attendance_qs.filter(top_late_filter).values(
        'user_id'
    ).annotate(
        late_count=Count('id'),
        avg_late_minutes=Avg('late_minutes')
    ).order_by('-late_count')[:10]

    # Format top late data
    top_late_data = []
    for item in top_late:
        user_details = UserDetails.objects.filter(user_id=item['user_id']).select_related('user').first()
        if user_details:
            top_late_data.append({
                'user__username': user_details.user.username,
                'user__first_name': user_details.user.first_name,
                'user__last_name': user_details.user.last_name,
                'work_location': user_details.work_location,
                'late_count': item['late_count'],
                'avg_late_minutes': round(item['avg_late_minutes'], 2)
            })

    # --- Get user counts by status for today ---
    today_filter = Q(date=today)

    # Apply location filter for today's counts
    if office_location:
        user_ids = UserDetails.objects.filter(
            work_location=office_location
        ).values_list('user_id', flat=True)
        today_filter &= Q(user_id__in=user_ids)

    # Apply user search filter for today's counts
    if user_query:
        today_filter &= (
            Q(user__username__icontains=user_query) |
            Q(user__first_name__icontains=user_query) |
            Q(user__last_name__icontains=user_query)
        )

    today_counts = attendance_qs.filter(today_filter).values(
        'status'
    ).annotate(
        count=Count('id')
    )

    # Format today's counts
    status_counts = {
        'Present': 0,
        'Present & Late': 0,
        'Absent': 0,
        'On Leave': 0,
        'Half Day': 0,
        'Work From Home': 0,
        'Comp Off': 0,
        'Not Marked': 0,
        'Yet to Clock In': 0
    }

    for item in today_counts:
        status_counts[item['status']] = item['count']

    # --- Prepare the final response ---
    response = {
        'daily_stats': daily_stats,
        'weekly_stats': weekly_stats,
        'monthly_stats': monthly_stats,
        'quarterly_stats': quarterly_stats,
        'yearly_stats': yearly_stats,
        'three_month_stats': three_month_stats,
        'period_stats': period_stats,
        'attendance_trend': formatted_trend,
        'top_attendance': top_attendance_data,
        'top_late': top_late_data,
        'status_counts': status_counts,
        'daily_leave_stats': daily_leave_stats,
        'weekly_leave_stats': weekly_leave_stats,
        'monthly_leave_stats': monthly_leave_stats,
        'quarterly_leave_stats': quarterly_leave_stats,
        'yearly_leave_stats': yearly_leave_stats,
        'period_leave_types': period_leave_types,
        'upcoming_leaves': list(upcoming_leaves),
        'top_leaves': top_leaves,
        'date_from': date_from.strftime('%Y-%m-%d'),
        'date_to': date_to.strftime('%Y-%m-%d'),
        'office_location': office_location,
        'user_query': user_query,
        'time_period': time_period
    }

    return response



from django.core.paginator import Paginator, EmptyPage
from django.db.models import Count, Q
from django.utils import timezone
from urllib.parse import urlencode

@login_required
@user_passes_test(is_hr_check)
def hr_attendance_list(request):
    """
    View for HR to see attendance records of all users with improved pagination and filtering
    """
    # Get filter parameters with defaults
    date_filter = request.GET.get('date', timezone.localtime(timezone.now()).date().strftime('%Y-%m-%d'))
    status_filter = request.GET.get('status', '')
    user_filter = request.GET.get('user', '')
    page_size = int(request.GET.get('page_size', 25))

    try:
        filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
    except ValueError:
        filter_date = timezone.localtime(timezone.now()).date()

    # Build base query with select_related to optimize performance
    attendance_query = Attendance.objects.select_related('user').filter(date=filter_date)

    # Apply filters
    if status_filter:
        attendance_query = attendance_query.filter(status=status_filter)

    if user_filter:
        attendance_query = attendance_query.filter(
            Q(user__username__icontains=user_filter) |
            Q(user__first_name__icontains=user_filter) |
            Q(user__last_name__icontains=user_filter)
        )

    # Order results
    attendance_list = attendance_query.order_by('user__username')

    # Pagination with proper error handling
    try:
        page_number = int(request.GET.get('page', 1))
    except ValueError:
        page_number = 1

    paginator = Paginator(attendance_list, page_size)

    try:
        page_obj = paginator.page(page_number)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # Calculate pagination context
    page_range = paginator.get_elided_page_range(
        page_obj.number,
        on_each_side=2,
        on_ends=1
    )

    # Prepare filter choices
    status_choices = [choice[0] for choice in Attendance.STATUS_CHOICES]

    # Build query parameters for pagination links
    query_params = {
        'date': date_filter,
        'status': status_filter,
        'user': user_filter,
        'page_size': page_size
    }
    query_string = urlencode({k: v for k, v in query_params.items() if v})

    context = {
        'page_obj': page_obj,
        'page_range': page_range,
        'date_filter': filter_date,
        'status_filter': status_filter,
        'user_filter': user_filter,
        'status_choices': status_choices,
        'query_string': query_string,
        'page_sizes': [10, 25, 50, 100],
        'current_page_size': page_size,
    }

    return render(request, 'components/hr/attendance/hr_attendance_list.html', context)

@login_required
@user_passes_test(is_hr_check)
def hr_edit_attendance(request, attendance_id):
    """
    View for HR to edit an attendance record
    """
    attendance = get_object_or_404(Attendance, id=attendance_id)

    if request.method == 'POST':
        # Get form data
        status = request.POST.get('status')
        clock_in_time_str = request.POST.get('clock_in_time')
        clock_out_time_str = request.POST.get('clock_out_time')
        leave_type = request.POST.get('leave_type', '')
        remarks = request.POST.get('remarks', '')

        # Process form data
        attendance.status = status
        attendance.regularization_reason = remarks
        attendance.modified_by = request.user

        # Process leave type
        if status in ['On Leave', 'Half Day']:
            attendance.leave_type = leave_type
            attendance.is_half_day = status == 'Half Day'
        else:
            attendance.leave_type = None
            attendance.is_half_day = False

        # Process clock in/out times
        date_only = attendance.date.strftime('%Y-%m-%d')
        if clock_in_time_str:
            try:
                clock_in_datetime = datetime.strptime(f"{date_only} {clock_in_time_str}", '%Y-%m-%d %H:%M')
                attendance.clock_in_time = timezone.make_aware(clock_in_datetime)
            except ValueError:
                messages.error(request, "Invalid clock-in time format.")

        if clock_out_time_str:
            try:
                clock_out_datetime = datetime.strptime(f"{date_only} {clock_out_time_str}", '%Y-%m-%d %H:%M')
                attendance.clock_out_time = timezone.make_aware(clock_out_datetime)
            except ValueError:
                messages.error(request, "Invalid clock-out time format.")

        # Save changes
        attendance.regularization_status = 'Approved'  # Auto-approve HR edits
        attendance.save()
        messages.success(request, f"Attendance record for {attendance.user.username} updated successfully.")

        # Redirect back to the attendance list
        return redirect('aps_attendance:hr_attendance_list')

    # Format times for form display
    clock_in_time = format_time(attendance.clock_in_time) if attendance.clock_in_time else ''
    clock_out_time = format_time(attendance.clock_out_time) if attendance.clock_out_time else ''

    # Get leave types
    leave_types = LeaveRequest.objects.values_list('leave_type__name', flat=True).distinct()

    context = {
        'attendance': attendance,
        'clock_in_time': clock_in_time,
        'clock_out_time': clock_out_time,
        'status_choices': Attendance.STATUS_CHOICES,
        'leave_types': leave_types
    }

    return render(request, 'components/hr/attendance/hr_edit_attendance.html', context)


def notify_hr_about_regularization(attendance):
    """
    Send notification to HR about a new regularization request

    Features:
    - Creates in-app notification
    - Sends email notification if enabled
    - Attaches relevant context data
    """
    try:
        # Get HR users
        hr_users = User.objects.filter(groups__name='HR')

        if not hr_users:
            logger.warning("No HR users found for regularization notification")
            return

        for hr_user in hr_users:
            # Create notification record
            Notification.objects.create(
                user=hr_user,
                title="New Regularization Request",
                message=f"{attendance.user.get_full_name() or attendance.user.username} has submitted an attendance regularization request for {attendance.date.strftime('%d-%b-%Y')}",
                url=reverse('aps_attendance:hr_attendance_regularization_requests'),
                category="regularization",
                related_object_id=attendance.id,
                related_object_type="attendance"
            )

            # Optional: Send email to HR if enabled
            if hasattr(hr_user, 'profile') and getattr(hr_user.profile, 'email_notifications_enabled', False):
                subject = f"New Regularization Request - {attendance.user.get_full_name() or attendance.user.username}"
                message = f"""
                Dear {hr_user.get_full_name() or hr_user.username},

                A new attendance regularization request requires your review:

                Employee: {attendance.user.get_full_name() or attendance.user.username}
                Date: {attendance.date.strftime('%d-%b-%Y')}
                Reason: {attendance.regularization_reason[:100]}{"..." if len(attendance.regularization_reason) > 100 else ""}

                Please review this request at your earliest convenience.

                Regards,
                HR Notification System
                """

                # Send email asynchronously
                send_email_async(hr_user.email, subject, message)

    except Exception as e:
        logger.error(f"Error notifying HR about regularization: {str(e)}")

def notify_employee_about_regularization_status(attendance):
    """
    Notify employee about regularization request status update

    Features:
    - Creates in-app notification
    - Sends email notification if enabled
    - Includes decision details and any remarks
    """
    try:
        status = attendance.regularization_status
        user = attendance.user

        # Define status-specific messages
        status_messages = {
            'Approved': 'has been approved',
            'Rejected': 'has been rejected',
            'Pending': 'is pending review',
            'Cancelled': 'has been cancelled'
        }

        message = f"Your attendance regularization request for {attendance.date.strftime('%d-%b-%Y')} {status_messages.get(status, 'has been updated')}"

        # Add remarks if present
        if attendance.regularization_remarks:
            message += f"\nRemarks: {attendance.regularization_remarks}"

        # Create in-app notification
        Notification.objects.create(
            user=user,
            title=f"Regularization Request {status}",
            message=message,
            url=reverse('aps_attendance:my_regularization_requests'),
            category="regularization",
            related_object_id=attendance.id,
            related_object_type="attendance"
        )

        # Send email if enabled
        if hasattr(user, 'profile') and getattr(user.profile, 'email_notifications_enabled', False):
            subject = f"Attendance Regularization {status} - {attendance.date.strftime('%d-%b-%Y')}"
            email_message = f"""
            Dear {user.get_full_name() or user.username},

            Your attendance regularization request has been updated:

            Date: {attendance.date.strftime('%d-%b-%Y')}
            Status: {status}
            {"Remarks: " + attendance.regularization_remarks if attendance.regularization_remarks else ""}

            {"Please check your attendance record for the updated status." if status == 'Approved' else ""}

            Regards,
            HR Team
            """

            # Send email asynchronously
            send_email_async(user.email, subject, email_message)

        logger.info(f"Regularization status notification sent to {user.username} for {attendance.date}")

    except Exception as e:
        logger.error(f"Error sending regularization status notification: {str(e)}")


def get_potential_regularization_dates(user, start_date=None, end_date=None):
    """
    Identify dates that may need regularization based on attendance patterns

    Features:
    - Checks for missing attendance
    - Identifies incomplete hours
    - Flags late arrivals and early departures
    - Considers holidays and leaves
    """
    try:
        # Default to last 30 days if no date range provided
        if not start_date:
            start_date = timezone.now().date() - timedelta(days=30)
        if not end_date:
            end_date = timezone.now().date()

        potential_dates = []
        current_date = start_date

        while current_date <= end_date:
            attendance = Attendance.objects.filter(
                user=user,
                date=current_date
            ).first()

            if attendance:
                # Check for conditions that might need regularization
                needs_regularization = False
                reason = []

                # Incomplete hours
                if attendance.total_hours and attendance.expected_hours:
                    if attendance.total_hours < attendance.expected_hours:
                        needs_regularization = True
                        reason.append("Incomplete hours")

                # Late arrival
                if attendance.late_minutes and attendance.late_minutes > 0:
                    needs_regularization = True
                    reason.append("Late arrival")

                # Early departure
                if attendance.left_early:
                    needs_regularization = True
                    reason.append("Early departure")

                # Missing clock in/out
                if not attendance.clock_in_time or not attendance.clock_out_time:
                    needs_regularization = True
                    reason.append("Missing clock in/out")

                if needs_regularization:
                    potential_dates.append({
                        'date': current_date,
                        'reasons': reason,
                        'attendance': attendance
                    })

            current_date += timedelta(days=1)

        return potential_dates

    except Exception as e:
        logger.error(f"Error identifying regularization dates: {str(e)}")
        return []


@login_required
@user_passes_test(is_hr_check)
def hr_attendance_regularization_requests(request):
    """
    Enhanced view for HR to see and process attendance regularization requests

    Features:
    - Filtering by status, date range, and employee
    - Sorting with pending requests first
    - Pagination for better UX
    """
    # Get filter parameters
    status_filter = request.GET.get('status', 'Pending')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    user_filter = request.GET.get('user', '')
    # department_filter = request.GET.get('department', '')

    # Build query for regularization requests
    requests_query = Attendance.objects.filter(regularization_status__isnull=False)

    # Apply status filter
    if status_filter and status_filter != 'All':
        requests_query = requests_query.filter(regularization_status=status_filter)

    # Apply date filters
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            requests_query = requests_query.filter(date__gte=date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            requests_query = requests_query.filter(date__lte=date_to_obj)
        except ValueError:
            pass

    # Apply user filter (search by username, first name, or last name)
    if user_filter:
        requests_query = requests_query.filter(
            Q(user__username__icontains=user_filter) |
            Q(user__first_name__icontains=user_filter) |
            Q(user__last_name__icontains=user_filter)
        )

    # Apply department filter if specified
    # if department_filter:
    #     requests_query = requests_query.filter(
    #         user__employee_profile__department__name=department_filter
    #     )

    # Order requests by priority (pending first) and date
    requests_list = requests_query.select_related('user').order_by(
        Case(
            When(regularization_status='Pending', then=Value(0)),
            When(regularization_status='Approved', then=Value(1)),
            When(regularization_status='Rejected', then=Value(2)),
            default=Value(3),
            output_field=IntegerField()
        ),
        '-date'
    )

    # Pagination
    paginator = Paginator(requests_list, 15)  # 15 items per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # For department filter dropdown
    # departments = Department.objects.all().order_by('name')

    # Get status choices for the form
    status_choices = [('All', 'All')] + list(Attendance.objects.values_list(
        'regularization_status', 'regularization_status'
    ).distinct().order_by('regularization_status'))

    context = {
        'page_obj': page_obj,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'user_filter': user_filter,
        'status_choices': status_choices,
        'attendance_status_choices': Attendance.STATUS_CHOICES,
    }

    return render(request, 'components/hr/attendance/hr_regularization_requests.html', context)


@require_POST
@login_required
@user_passes_test(is_hr_check)
def hr_process_regularization(request, attendance_id):
    """
    Enhanced HR process to approve/reject regularization and update status if needed

    Features:
    - Update attendance status if requested
    - Recalculate attendance metrics on approval
    - Add audit trail through remarks
    - Notify employee about decision
    """
    attendance = get_object_or_404(Attendance, id=attendance_id)
    action = request.POST.get('action')
    remarks = request.POST.get('remarks', '')
    new_status = request.POST.get('status')  # New attendance status

    # Check if there's a valid status to change to
    if new_status and new_status in dict(Attendance.STATUS_CHOICES):
        # Record the old status before changing
        old_status = attendance.status
        attendance.status = new_status
        status_changed = True
    else:
        # Check if there's a requested status from employee to apply
        if attendance.requested_status:
            if action == 'approve':
                old_status = attendance.status
                attendance.status = attendance.requested_status
                attendance.requested_status = None  # Clear the request
                status_changed = True
            else:
                status_changed = False
        else:
            status_changed = False

    if action == 'approve':
        attendance.regularization_status = 'Approved'
        attendance.modified_by = request.user
        attendance.is_hr_notified = True

        # Calculate updated attendance metrics if attendance is modified
        if attendance.clock_in_time and attendance.clock_out_time:
            # Recalculate total hours
            total_seconds = (attendance.clock_out_time - attendance.clock_in_time).total_seconds()
            attendance.total_hours = Decimal(total_seconds / 3600).quantize(Decimal('0.01'))

            # Update other metrics based on shift if available
            if attendance.shift:
                shift_start = combine_date_with_time(attendance.date, attendance.shift.start_time)
                shift_end = combine_date_with_time(attendance.date, attendance.shift.end_time)

                # Calculate late minutes if clocked in after shift start
                if attendance.clock_in_time > shift_start:
                    attendance.late_minutes = int((attendance.clock_in_time - shift_start).total_seconds() / 60)
                else:
                    attendance.late_minutes = 0

                # Calculate early departure minutes if clocked out before shift end
                if attendance.clock_out_time < shift_end:
                    attendance.early_departure_minutes = int((shift_end - attendance.clock_out_time).total_seconds() / 60)
                    attendance.left_early = attendance.early_departure_minutes > 0
                else:
                    attendance.early_departure_minutes = 0
                    attendance.left_early = False

                # Set expected hours based on shift
                start_time = attendance.shift.start_time
                end_time = attendance.shift.end_time

                # Handle shifts that cross midnight
                if end_time < start_time:
                    # Add a day to end_time
                    shift_hours = (24 - start_time.hour - start_time.minute / 60) + (end_time.hour + end_time.minute / 60)
                else:
                    shift_hours = (end_time.hour - start_time.hour) + (end_time.minute - start_time.minute) / 60

                attendance.expected_hours = Decimal(shift_hours).quantize(Decimal('0.01'))

        # Add remarks and status change info
        hr_name = request.user.get_full_name() or request.user.username
        remarks_text = f"\n\nApproved by HR ({hr_name}): {remarks}" if remarks else f"\n\nApproved by HR ({hr_name})"

        if status_changed:
            remarks_text += f"\nStatus changed from '{old_status}' to '{attendance.status}'"

        attendance.regularization_reason += remarks_text
        messages.success(request, f"Regularization approved for {attendance.user.get_full_name() or attendance.user.username}.")

        # Notify employee about approval
        notify_employee_about_regularization_status(attendance)

    elif action == 'reject':
        attendance.regularization_status = 'Rejected'
        attendance.modified_by = request.user
        attendance.is_hr_notified = True

        # Add remarks
        hr_name = request.user.get_full_name() or request.user.username
        remarks_text = f"\n\nRejected by HR ({hr_name}): {remarks}" if remarks else f"\n\nRejected by HR ({hr_name})"
        attendance.regularization_reason += remarks_text

        messages.warning(request, f"Regularization rejected for {attendance.user.get_full_name() or attendance.user.username}.")

        # Notify employee about rejection
        notify_employee_about_regularization_status(attendance)

    attendance.save()

    redirect_url = request.POST.get('next', reverse('aps_attendance:hr_attendance_regularization_requests'))
    return redirect(redirect_url)


@login_required
@user_passes_test(is_hr_check)
def regularization_analytics_dashboard(request):
    """
    Dashboard for HR to see analytics about regularization patterns

    Features:
    - Filter by date range and department
    - Show regularization trends
    - Display common reasons for regularization
    - Identify patterns of regularization by employee
    """
    # Date range for analytics
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    department_filter = request.GET.get('department', '')

    today = localtime(now()).date()

    # Default to current month if not specified
    if not date_from_str:
        date_from = today.replace(day=1)  # First day of current month
        date_from_str = date_from.strftime('%Y-%m-%d')
    else:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            date_from = today.replace(day=1)
            date_from_str = date_from.strftime('%Y-%m-%d')

    if not date_to_str:
        date_to = today
        date_to_str = date_to.strftime('%Y-%m-%d')
    else:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            date_to = today
            date_to_str = date_to.strftime('%Y-%m-%d')

    # Base query for regularization requests
    base_query = Attendance.objects.filter(
        regularization_status__isnull=False,
        date__range=[date_from, date_to]
    )

    # Filter by department if specified
    if department_filter:
        base_query = base_query.filter(
            user__employee_profile__department__name=department_filter
        )

    # Generate analytics data
    total_requests = base_query.count()
    approved_requests = base_query.filter(regularization_status='Approved').count()
    rejected_requests = base_query.filter(regularization_status='Rejected').count()
    pending_requests = base_query.filter(regularization_status='Pending').count()

    # Get reasons breakdown (top categories)
    reason_categories = [
        'Late arrival', 'Early departure', 'Forgot to clock in/out',
        'System issue', 'Work from home', 'Client meeting', 'Other'
    ]

    reasons_breakdown = []
    for category in reason_categories:
        count = base_query.filter(regularization_reason__icontains=category).count()
        if count > 0:
            reasons_breakdown.append({
                'category': category,
                'count': count,
                'percentage': round((count / total_requests * 100) if total_requests > 0 else 0, 1)
            })

    # Sort by count descending
    reasons_breakdown = sorted(reasons_breakdown, key=lambda x: x['count'], reverse=True)

    # Get top employees with regularization requests
    top_employees = base_query.values('user__username', 'user__first_name', 'user__last_name') \
                      .annotate(count=Count('id')) \
                      .order_by('-count')[:10]

    # Get departments list for filtering
    departments = Department.objects.all().order_by('name')

    # Data for regularization trend chart (by day)
    date_range = [(date_from + timedelta(days=x)) for x in range((date_to - date_from).days + 1)]

    trend_data = []
    for day in date_range:
        day_requests = base_query.filter(date=day).count()
        if day_requests > 0:
            trend_data.append({
                'date': day.strftime('%Y-%m-%d'),
                'count': day_requests
            })

    context = {
        'date_from': date_from_str,
        'date_to': date_to_str,
        'department_filter': department_filter,
        'departments': departments,
        'total_requests': total_requests,
        'approved_requests': approved_requests,
        'rejected_requests': rejected_requests,
        'pending_requests': pending_requests,
        'approval_rate': round((approved_requests / total_requests * 100) if total_requests > 0 else 0, 1),
        'rejection_rate': round((rejected_requests / total_requests * 100) if total_requests > 0 else 0, 1),
        'pending_rate': round((pending_requests / total_requests * 100) if total_requests > 0 else 0, 1),
        'reasons_breakdown': reasons_breakdown,
        'top_employees': top_employees,
        'trend_data': json.dumps(trend_data),
    }

    return render(request, 'components/hr/attendance/regularization_analytics.html', context)

@login_required
@user_passes_test(is_hr_check)
def hr_generate_report(request):
    """
    Enhanced view for HR to generate various types of attendance reports:
    - Daily reports (single date)
    - Weekly reports
    - Monthly reports
    - Custom date range reports
    - Individual user reports
    - Role-based reports

    With filtering capabilities and export options.
    """
    # Get filter parameters
    report_type = request.GET.get('report_type', 'all')  # all, user, role
    date_filter_type = request.GET.get('date_filter_type', 'custom')  # daily, weekly, monthly, custom
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    single_date = request.GET.get('single_date', '')
    month = request.GET.get('month', '')
    year = request.GET.get('year', '')
    week = request.GET.get('week', '')
    user_id = request.GET.get('user_id', '')
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')
    sort_by = request.GET.get('sort_by', 'username')
    export_format = request.GET.get('format', '')

    # Initialize with empty results
    report_data = []
    summary_data = {}
    from_date = None
    to_date = None

    # Process date filters based on selected type
    try:
        today = datetime.now().date()

        if date_filter_type == 'daily' and single_date:
            # Single day report
            from_date = datetime.strptime(single_date, '%Y-%m-%d').date()
            to_date = from_date

        elif date_filter_type == 'weekly' and week and year:
            # Weekly report
            year_num = int(year)
            week_num = int(week)
            from_date = datetime.strptime(f'{year_num}-W{week_num}-1', '%Y-W%W-%w').date()
            to_date = from_date + timedelta(days=6)

        elif date_filter_type == 'monthly' and month and year:
            # Monthly report
            year_num = int(year)
            month_num = int(month)
            from_date = date(year_num, month_num, 1)
            # Get last day of month
            if month_num == 12:
                to_date = date(year_num + 1, 1, 1) - timedelta(days=1)
            else:
                to_date = date(year_num, month_num + 1, 1) - timedelta(days=1)

        elif date_filter_type == 'custom' and date_from and date_to:
            # Custom date range
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()

        else:
            # Default to current month if no valid filter is provided
            from_date = date(today.year, today.month, 1)
            if today.month == 12:
                to_date = date(today.year + 1, 1, 1) - timedelta(days=1)
            else:
                to_date = date(today.year, today.month + 1, 1) - timedelta(days=1)

        # Build user query
        user_query = User.objects.filter(is_active=True)

        if report_type == 'user' and user_id:
            user_query = user_query.filter(id=user_id)
        elif report_type == 'role' and role_filter:
            user_query = user_query.filter(groups__name=role_filter)

        # Apply sorting
        if sort_by == 'name':
            user_query = user_query.order_by('first_name', 'last_name')
        elif sort_by == 'attendance':
            # We'll sort by attendance later after calculations
            user_query = user_query.order_by('username')
        else:
            user_query = user_query.order_by('username')

        users = user_query.distinct()

        # Generate report data
        for user in users:
            # Get attendance data for date range
            attendance_query = Attendance.objects.filter(
                user=user,
                date__gte=from_date,
                date__lte=to_date
            )

            # Apply status filter if provided
            if status_filter:
                attendance_query = attendance_query.filter(status=status_filter)

            attendance_data = attendance_query.order_by('date')

            # Skip users with no attendance data unless specifically requested
            if not attendance_data.exists() and report_type != 'user':
                continue

            # Calculate statistics
            total_days = (to_date - from_date).days + 1
            work_days = get_work_days(from_date, to_date)  # Exclude weekends/holidays

            present_count = attendance_data.filter(
                status__in=['Present', 'Present & Late', 'Work From Home']
            ).count()

            absent_count = attendance_data.filter(status='Absent').count()
            leave_count = attendance_data.filter(status__in=['On Leave', 'Comp Off']).count()
            late_count = attendance_data.filter(status__in=['Present & Late', 'Late']).count()
            wfh_count = attendance_data.filter(status='Work From Home').count()

            total_hours = attendance_data.aggregate(Sum('total_hours'))['total_hours__sum'] or Decimal('0')
            overtime_hours = attendance_data.aggregate(Sum('overtime_hours'))['overtime_hours__sum'] or Decimal('0')

            # Calculate average working hours (only consider days with hours)
            days_with_hours = attendance_data.exclude(total_hours__isnull=True).exclude(total_hours=0).count()
            avg_hours = Decimal('0')
            if days_with_hours > 0:
                avg_hours = total_hours / days_with_hours

            # Get groups/roles for this user
            roles = user.groups.values_list('name', flat=True)

            # Calculate attendance percentage based on work days
            attendance_percentage = 0
            if work_days > 0:
                attendance_percentage = (present_count / work_days) * 100

            # Calculate punctuality percentage
            punctuality = 0
            if present_count > 0:
                punctuality = ((present_count - late_count) / present_count) * 100

            # Get daily attendance details
            daily_records = []
            if report_type == 'user':  # Only generate detailed daily records for single user reports
                current_date = from_date
                while current_date <= to_date:
                    day_record = attendance_data.filter(date=current_date).first()
                    daily_entry = {
                        'date': current_date,
                        'day': current_date.strftime('%A'),
                        'status': day_record.status if day_record else 'Not Marked',
                        'check_in': day_record.clock_in_time.strftime('%H:%M') if day_record and day_record.clock_in_time else '-',
                        'check_out': day_record.clock_out_time.strftime('%H:%M') if day_record and day_record.clock_out_time else '-',
                        'total_hours': day_record.total_hours if day_record else Decimal('0'),
                        'overtime_hours': day_record.overtime_hours if day_record else Decimal('0'),
                    }
                    daily_records.append(daily_entry)
                    current_date += timedelta(days=1)

            user_data = {
                'user_id': user.id,
                'username': user.username,
                'full_name': f"{user.first_name} {user.last_name}".strip() or user.username,
                'email': user.email,
                'roles': ', '.join(roles),
                'total_days': total_days,
                'work_days': work_days,
                'present_days': present_count,
                'absent_days': absent_count,
                'leave_days': leave_count,
                'late_days': late_count,
                'wfh_days': wfh_count,
                'total_hours': total_hours,
                'avg_hours': avg_hours.quantize(Decimal('0.01')),
                'overtime_hours': overtime_hours,
                'attendance_percentage': round(attendance_percentage, 2),
                'punctuality_percentage': round(punctuality, 2),
                'daily_records': daily_records,
            }

            report_data.append(user_data)

        # Sort by attendance percentage if requested
        if sort_by == 'attendance' and report_data:
            report_data.sort(key=lambda x: x['attendance_percentage'], reverse=True)

        # Calculate summary
        if report_data:
            # Average attendance percentage
            avg_attendance = sum(d['attendance_percentage'] for d in report_data) / len(report_data)

            # Total hours worked
            total_all_hours = sum(d['total_hours'] for d in report_data)

            # Total overtime hours
            total_overtime = sum(d['overtime_hours'] for d in report_data)

            # Average hours per day
            avg_daily_hours = Decimal('0')
            total_present_days = sum(d['present_days'] for d in report_data)
            if total_present_days > 0:
                avg_daily_hours = total_all_hours / total_present_days

            # Most punctual users (top 3)
            most_punctual = sorted(report_data, key=lambda x: x['punctuality_percentage'], reverse=True)[:5]
            most_punctual = [{'name': d['full_name'], 'percentage': d['punctuality_percentage']} for d in most_punctual]

            # Users with most overtime (top 3)
            most_overtime = sorted(report_data, key=lambda x: x['overtime_hours'], reverse=True)[:5]
            most_overtime = [{'name': d['full_name'], 'hours': d['overtime_hours']} for d in most_overtime]

            date_range_str = get_date_range_display(date_filter_type, from_date, to_date, month, year, week)

            summary_data = {
                'user_count': len(report_data),
                'date_range': date_range_str,
                'date_type': date_filter_type,
                'avg_attendance': round(avg_attendance, 2),
                'total_hours': total_all_hours,
                'avg_daily_hours': avg_daily_hours.quantize(Decimal('0.01')),
                'total_overtime': total_overtime,
                'most_punctual': most_punctual,
                'most_overtime': most_overtime,
            }

        # Handle export
        if export_format and report_data:
            if export_format == 'csv':
                response = export_to_csv(report_data, from_date, to_date, report_type)
                return response
            elif export_format == 'excel':
                response = export_to_excel(report_data, from_date, to_date, summary_data, report_type)
                return response
            elif export_format == 'pdf':
                return generate_pdf_report(request, report_data, summary_data, from_date, to_date)

    except ValueError as e:
        messages.error(request, f"Error processing report: {str(e)}")

    # Get all available roles for filters
    roles = User.groups.through.objects.values_list('group__name', flat=True).distinct().order_by('group__name')

    # Get years and months for dropdowns
    current_year = datetime.now().year
    years = range(current_year - 3, current_year + 1)

    # Updated status choices to match the model definition (removed 'Half Day')
    status_choices = [
        'Present', 'Present & Late', 'Absent', 'Late', 'On Leave',
        'Work From Home', 'Weekend', 'Holiday', 'Comp Off', 'Not Marked'
    ]
    items = Attendance.objects.all()  # Your queryset
    paginator = Paginator(items, 10)  # Show 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    calander_report = hr_attendance_view(request)

    context = {
        'report_type': report_type,
        'date_filter_type': date_filter_type,
        'date_from': date_from,
        'page_obj': page_obj,
        'date_to': date_to,
        'single_date': single_date,
        'month': month,
        'year': year,
        'week': week,
        'user_id': user_id,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'sort_by': sort_by,
        'report_data': report_data,
        'summary_data': summary_data,
        'roles': roles,
        'status_choices': status_choices,
        'users': User.objects.filter(is_active=True).order_by('username'),
        'years': years,
        'months': range(1, 13),
        'weeks': range(1, 53),
        'calander_report': calander_report,
    }

    return render(request, 'components/hr/attendance/hr_generate_report.html', context)
def get_work_days(start_date, end_date):
    """
    Count the number of work days between two dates (excluding weekends)
    In a real app, this would also exclude holidays
    """
    work_days = 0
    current_date = start_date
    while current_date <= end_date:
        # If not weekend (Monday is 0, Sunday is 6)
        if current_date.weekday() < 5:
            work_days += 1
        current_date += timedelta(days=1)
    return work_days


def get_date_range_display(date_filter_type, from_date, to_date, month=None, year=None, week=None):
    """
    Generate a user-friendly date range display
    """
    if date_filter_type == 'daily':
        return f"Daily Report: {from_date.strftime('%B %d, %Y')}"
    elif date_filter_type == 'weekly':
        return f"Weekly Report: Week {week}, {year} ({from_date.strftime('%b %d')} - {to_date.strftime('%b %d, %Y')})"
    elif date_filter_type == 'monthly':
        return f"Monthly Report: {from_date.strftime('%B %Y')}"
    else:
        return f"Custom Period: {from_date.strftime('%b %d, %Y')} to {to_date.strftime('%b %d, %Y')}"


def export_to_csv(report_data, from_date, to_date, report_type):
    """
    Export report data to CSV format
    """
    response = HttpResponse(content_type='text/csv')
    filename = f"attendance_report_{from_date.strftime('%Y%m%d')}_to_{to_date.strftime('%Y%m%d')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # Write headers
    headers = [
        'Username', 'Full Name', 'Email', 'Roles',
        'Work Days', 'Present Days', 'Absent Days', 'Leave Days',
        'Late Days', 'Work From Home', 'Total Hours',
        'Average Hours/Day', 'Overtime Hours', 'Attendance %', 'Punctuality %'
    ]
    writer.writerow(headers)

    # Write data
    for data in report_data:
        row = [
            data['username'],
            data['full_name'],
            data['email'],
            data['roles'],
            data['work_days'],
            data['present_days'],
            data['absent_days'],
            data['leave_days'],
            data['late_days'],
            data['wfh_days'],
            data['total_hours'],
            data['avg_hours'],
            data['overtime_hours'],
            f"{data['attendance_percentage']:.2f}%",
            f"{data['punctuality_percentage']:.2f}%"
        ]
        writer.writerow(row)

        # Add daily records for individual reports
        if report_type == 'user' and data.get('daily_records'):
            writer.writerow(['Date', 'Day', 'Status', 'Check In', 'Check Out', 'Hours', 'Overtime'])
            for record in data['daily_records']:
                writer.writerow([
                    record['date'].strftime('%Y-%m-%d'),
                    record['day'],
                    record['status'],
                    record['check_in'],
                    record['check_out'],
                    record['total_hours'],
                    record['overtime_hours']
                ])
            writer.writerow([])  # Add blank row

    return response


def export_to_excel(report_data, from_date, to_date, summary_data, report_type):
    """
    Export report data to Excel format
    """
    import xlwt  # Import here since it's only needed for Excel export

    workbook = xlwt.Workbook(encoding='utf-8')

    # Summary sheet
    summary_sheet = workbook.add_sheet('Summary')

    # Styles
    header_style = xlwt.easyxf('font: bold on; align: wrap on, vert centre, horiz center; pattern: pattern solid, fore_color gray25;')
    date_style = xlwt.easyxf('font: bold on; align: wrap on, vert centre, horiz left;')
    percent_style = xlwt.easyxf('font: bold off; align: wrap on, vert centre, horiz right;', num_format_str='0.00%')
    decimal_style = xlwt.easyxf('font: bold off; align: wrap on, vert centre, horiz right;', num_format_str='0.00')

    # Write summary
    summary_sheet.write(0, 0, 'Attendance Report', header_style)
    summary_sheet.write(1, 0, summary_data['date_range'], date_style)
    summary_sheet.write(2, 0, 'Total Employees:', date_style)
    summary_sheet.write(2, 1, summary_data['user_count'])
    summary_sheet.write(3, 0, 'Average Attendance Rate:', date_style)
    summary_sheet.write(3, 1, summary_data['avg_attendance'] / 100, percent_style)
    summary_sheet.write(4, 0, 'Total Hours Worked:', date_style)
    summary_sheet.write(4, 1, float(summary_data['total_hours']), decimal_style)
    summary_sheet.write(5, 0, 'Average Daily Hours:', date_style)
    summary_sheet.write(5, 1, float(summary_data['avg_daily_hours']), decimal_style)
    summary_sheet.write(6, 0, 'Total Overtime Hours:', date_style)
    summary_sheet.write(6, 1, float(summary_data['total_overtime']), decimal_style)

    # Most punctual employees
    summary_sheet.write(8, 0, 'Most Punctual Employees', header_style)
    for i, emp in enumerate(summary_data['most_punctual']):
        summary_sheet.write(9 + i, 0, emp['name'])
        summary_sheet.write(9 + i, 1, emp['percentage'] / 100, percent_style)

    # Most overtime employees
    summary_sheet.write(8, 3, 'Most Overtime Employees', header_style)
    for i, emp in enumerate(summary_data['most_overtime']):
        summary_sheet.write(9 + i, 3, emp['name'])
        summary_sheet.write(9 + i, 4, float(emp['hours']), decimal_style)

    # Data sheet
    data_sheet = workbook.add_sheet('Attendance Data')

    # Write headers
    headers = [
        'Username', 'Full Name', 'Email', 'Roles',
        'Work Days', 'Present Days', 'Absent Days', 'Leave Days',
        'Late Days', 'Work From Home', 'Total Hours',
        'Average Hours/Day', 'Overtime Hours', 'Attendance %', 'Punctuality %'
    ]

    for col_idx, header in enumerate(headers):
        data_sheet.write(0, col_idx, header, header_style)
        data_sheet.col(col_idx).width = 256 * 15  # Set column width

    # Write data
    row_idx = 1
    for data in report_data:
        data_sheet.write(row_idx, 0, data['username'])
        data_sheet.write(row_idx, 1, data['full_name'])
        data_sheet.write(row_idx, 2, data['email'])
        data_sheet.write(row_idx, 3, data['roles'])
        data_sheet.write(row_idx, 4, data['work_days'])
        data_sheet.write(row_idx, 5, data['present_days'])
        data_sheet.write(row_idx, 6, data['absent_days'])
        data_sheet.write(row_idx, 7, data['leave_days'])
        data_sheet.write(row_idx, 8, data['late_days'])
        data_sheet.write(row_idx, 9, data['wfh_days'])
        data_sheet.write(row_idx, 10, float(data['total_hours']), decimal_style)
        data_sheet.write(row_idx, 11, float(data['avg_hours']), decimal_style)
        data_sheet.write(row_idx, 12, float(data['overtime_hours']), decimal_style)
        data_sheet.write(row_idx, 13, data['attendance_percentage'] / 100, percent_style)
        data_sheet.write(row_idx, 14, data['punctuality_percentage'] / 100, percent_style)
        row_idx += 1

    # Daily details for individual user
    if report_type == 'user' and len(report_data) == 1 and report_data[0].get('daily_records'):
        details_sheet = workbook.add_sheet('Daily Details')

        # Write headers
        detail_headers = ['Date', 'Day', 'Status', 'Check In', 'Check Out', 'Hours', 'Overtime']
        for col_idx, header in enumerate(detail_headers):
            details_sheet.write(0, col_idx, header, header_style)
            details_sheet.col(col_idx).width = 256 * 15

        # Write data
        for i, record in enumerate(report_data[0]['daily_records']):
            details_sheet.write(i + 1, 0, record['date'].strftime('%Y-%m-%d'))
            details_sheet.write(i + 1, 1, record['day'])
            details_sheet.write(i + 1, 2, record['status'])
            details_sheet.write(i + 1, 3, record['check_in'])
            details_sheet.write(i + 1, 4, record['check_out'])
            details_sheet.write(i + 1, 5, float(record['total_hours']), decimal_style)
            details_sheet.write(i + 1, 6, float(record['overtime_hours']), decimal_style)

    # Save to response
    response = HttpResponse(content_type='application/ms-excel')
    filename = f"attendance_report_{from_date.strftime('%Y%m%d')}_to_{to_date.strftime('%Y%m%d')}.xls"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def generate_pdf_report(request, report_data, summary_data, from_date, to_date):
    """
    Generate PDF report using reportlab or another PDF library
    This is a placeholder - in a real implementation, you would use reportlab or another
    PDF generation library to create a formatted PDF report
    """
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    # Create a file-like buffer to receive PDF data
    buffer = BytesIO()

    # Create the PDF object, using the BytesIO object as its "file"
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))

    # Create the stylesheet
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    subtitle_style = styles['Heading2']
    normal_style = styles['Normal']

    # Build the PDF content
    elements = []

    # Title
    elements.append(Paragraph("Attendance Report", title_style))
    elements.append(Paragraph(summary_data['date_range'], subtitle_style))
    elements.append(Spacer(1, 12))

    # Summary table
    summary_data_items = [
        ["Total Employees", str(summary_data['user_count'])],
        ["Average Attendance", f"{summary_data['avg_attendance']}%"],
        ["Total Hours", str(summary_data['total_hours'])],
        ["Average Daily Hours", str(summary_data['avg_daily_hours'])],
        ["Total Overtime", str(summary_data['total_overtime'])]
    ]

    summary_table = Table(summary_data_items, colWidths=[200, 100])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 24))

    # Main data table
    if report_data:
        # Headers
        headers = [
            'Name', 'Present', 'Absent', 'Leave', 'Late',
            'WFH', 'Hours', 'OT Hours', 'Attendance', 'Punctuality'
        ]

        # Data rows
        data_rows = [headers]
        for data in report_data:
            row = [
                data['full_name'],
                data['present_days'],
                data['absent_days'],
                data['leave_days'],
                data['late_days'],
                data['wfh_days'],
                str(data['total_hours']),
                str(data['overtime_hours']),
                f"{data['attendance_percentage']}%",
                f"{data['punctuality_percentage']}%"
            ]
            data_rows.append(row)

        # Create table
        table = Table(data_rows)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))

        elements.append(table)

    # Build the PDF
    doc.build(elements)

    # Get the value of the BytesIO buffer
    pdf = buffer.getvalue()
    buffer.close()

    # Create the HttpResponse with PDF headers
    response = HttpResponse(content_type='application/pdf')
    filename = f"attendance_report_{from_date.strftime('%Y%m%d')}_to_{to_date.strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Write the PDF to the response
    response.write(pdf)
    return response

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.core.exceptions import ValidationError
from django.forms import formset_factory, modelformset_factory
from django.db.models import Q
import json
import csv
import logging
from datetime import datetime, timedelta
from decimal import Decimal

from .models import Attendance, User, ShiftAssignment
from .forms import AttendanceForm, BulkAttendanceForm, AttendanceFilterForm

logger = logging.getLogger(__name__)

def is_hr_or_admin_check(user):
    """Check if user has HR or admin privileges"""
    return user.is_superuser or hasattr(user, 'is_hr') and user.is_hr

from django.db.models import Q

def can_manage_user(request_user, target_user):
    """
    Check if request_user can manage target_user based on group roles

    Args:
        request_user: The user attempting to manage another user
        target_user: The user being managed

    Returns:
        bool: True if request_user can manage target_user, False otherwise
    """
    # Superuser can manage everyone
    if request_user.is_superuser:
        return True

    # Can't manage yourself
    if request_user.id == target_user.id:
        return False

    # Check if request_user is in HR group
    if request_user.groups.filter(name='HR').exists():
        # HR can manage users in Management or Backoffice groups
        manageable_groups = target_user.groups.filter(
            Q(name='Management') |
            Q(name='Backoffice')
        ).exists()

        # HR cannot manage other HR
        is_hr = target_user.groups.filter(name='HR').exists()

        return manageable_groups and not is_hr

    return False

from .forms import ManualAttendanceForm

@login_required
@user_passes_test(is_hr_check)
def add_attendance(request):
    """
    View for HR to manually add attendance records
    """
    context = {
        'title': 'Add Attendance Records',
        'today': timezone.localdate(),
    }

    if request.method == 'POST':
        form = ManualAttendanceForm(request.POST)
        if form.is_valid():
            try:
                # Create attendance record
                attendance = form.save(commit=False)

                # Set additional default values
                attendance.modified_by = request.user
                attendance.location = 'Office'  # Default to Office

                # Additional default settings
                attendance.is_half_day = form.cleaned_data['status'] == 'Half Day'
                attendance.total_hours = 0  # No work hours tracked for manual entry

                # Save the attendance record
                attendance.save()

                messages.success(request, f'Attendance for {attendance.user.username} on {attendance.date} added successfully.')
                return redirect('attendance_list')  # Redirect to attendance list or dashboard

            except Exception as e:
                messages.error(request, f'Error adding attendance: {str(e)}')
    else:
        form = ManualAttendanceForm(initial={'date': timezone.localdate()})

    context['form'] = form
    return render(request, 'components/hr/attendance/hr_add_attendance.html', context)

# @login_required
# @user_passes_test(is_hr_check)
# def add_attendance(request):
#     """
#     View for HR to add/update attendance records (single or bulk)
#     """
#     context = {
#         'title': 'Add Attendance Records',
#         'today': timezone.localdate(),
#     }

#     # Get list of users that can be managed by current user
#     manageable_users = User.objects.filter(
#         Q(groups__name='Management') | Q(groups__name='Backoffice')
#     ).distinct().order_by('username')

#     # Prepare context for form fields
#     context.update({
#         'users': manageable_users,
#         'status_choices': Attendance.STATUS_CHOICES,
#         'location_choices': Attendance.LOCATION_CHOICES,
#     })

#     if request.method == 'POST':
#         try:
#             # Single Attendance Submission
#             if 'add_single' in request.POST:
#                 # Validate single attendance submission manually
#                 errors = validate_single_attendance(request.POST, manageable_users)
#                 if errors:
#                     for error in errors:
#                         messages.error(request, error)
#                 else:
#                     with transaction.atomic():
#                         attendance = process_single_attendance(request)
#                         messages.success(request, f"Attendance record added for {attendance.user.username} on {attendance.date}")
#                     return redirect('aps_attendance:add_attendance')

#             # Bulk Attendance Submission
#             elif 'add_bulk' in request.POST:
#                 # Process bulk attendance (you'll need to implement this method)
#                 with transaction.atomic():
#                     success_count = process_bulk_attendance(request, manageable_users)
#                     messages.success(request, f"Successfully added {success_count} attendance records")
#                     return redirect('aps_attendance:add_attendance')

#             # CSV Import
#             elif 'import_csv' in request.POST and request.FILES.get('csv_file'):
#                 with transaction.atomic():
#                     success_count = process_csv_import(request)
#                     messages.success(request, f"Successfully imported {success_count} attendance records")
#                     return redirect('aps_attendance:add_attendance')

#         except Exception as e:
#             logger.error(f"Error processing attendance: {e}")
#             messages.error(request, f"Error processing attendance: {str(e)}")

#     return render(request, 'components/hr/attendance/hr_add_attendance.html', context)

def validate_single_attendance(post_data, manageable_users):
    """
    Manually validate single attendance submission
    Returns a list of error messages
    """
    errors = []

    # User validation
    user_id = post_data.get('user')
    if not user_id:
        errors.append("User is required")
    else:
        try:
            user = manageable_users.get(id=user_id)
        except User.DoesNotExist:
            errors.append("Invalid user selection")

    # Date validation
    date_str = post_data.get('date')
    try:
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
        if date > timezone.localdate():
            errors.append("Date cannot be in the future")
    except (ValueError, TypeError):
        errors.append("Invalid date format")

    # Status validation
    status = post_data.get('status')
    if not status or status not in dict(Attendance.STATUS_CHOICES):
        errors.append("Invalid attendance status")

    # Status-specific validations
    if status == 'On Leave':
        if not post_data.get('leave_type'):
            errors.append("Leave type is required")

    elif status == 'Holiday':
        if not post_data.get('holiday_name'):
            errors.append("Holiday name is required")

    elif status in ['Present', 'Present & Late']:
        clock_in_time_str = post_data.get('clock_in_time')
        clock_out_time_str = post_data.get('clock_out_time')

        # Clock-in and clock-out time validation
        clock_in_time = None
        clock_out_time = None

        if clock_in_time_str:
            try:
                clock_in_time = datetime.strptime(clock_in_time_str, '%H:%M').time()
            except ValueError:
                errors.append("Invalid clock-in time format")

        if clock_out_time_str:
            try:
                clock_out_time = datetime.strptime(clock_out_time_str, '%H:%M').time()
            except ValueError:
                errors.append("Invalid clock-out time format")

        # Validate clock-in and clock-out times
        if clock_in_time and clock_out_time:
            if clock_in_time >= clock_out_time:
                errors.append("Clock-out time must be after clock-in time")

            # Validate against current time for today's entry
            if date == timezone.localdate():
                current_time = timezone.localtime().time()
                if clock_out_time > current_time:
                    errors.append("Clock-out can't be in the future")

        # Require clock-in time or regularization reason
        if not clock_in_time and not post_data.get('regularization_reason'):
            errors.append("Clock-in time or regularization reason is required")

    # Breaks validation
    breaks_str = post_data.get('breaks')
    if breaks_str:
        try:
            breaks_list = json.loads(breaks_str)
            if not isinstance(breaks_list, list):
                errors.append("Breaks must be a list of break periods")

            for break_item in breaks_list:
                if not isinstance(break_item, dict):
                    errors.append("Each break must be a dictionary")

                start_str = break_item.get('start', '')
                end_str = break_item.get('end', '')

                try:
                    start = datetime.strptime(start_str, '%H:%M').time()
                    end = datetime.strptime(end_str, '%H:%M').time()

                    if start >= end:
                        errors.append("Break end time must be after start time")
                except ValueError:
                    errors.append("Invalid break time format")

        except json.JSONDecodeError:
            errors.append("Invalid break format. Use [{'start':'HH:MM', 'end':'HH:MM'}]")

    return errors

def process_single_attendance(request):
    """Process a single attendance record directly from request"""
    post_data = request.POST

    # User selection
    user = User.objects.get(id=post_data.get('user'))

    # Validate user management permission
    if not can_manage_user(request.user, user):
        raise ValidationError(f"You don't have permission to manage {user.username}'s attendance")

    # Date parsing
    date = datetime.strptime(post_data.get('date'), '%Y-%m-%d').date()
    status = post_data.get('status')

    print(f"[DEBUG] Processing attendance for user: {user.username}")
    print(f"[DEBUG] Attendance Date: {date}, Status: {status}")

    # Get existing record or create new one
    try:
        attendance = Attendance.objects.get(user=user, date=date)
        print(f"[DEBUG] Existing attendance found for {user.username} on {date}")
        # Store original values to track changes
        attendance.original_status = attendance.status
        attendance.original_clock_in_time = attendance.clock_in_time
        attendance.original_clock_out_time = attendance.clock_out_time
    except Attendance.DoesNotExist:
        print(f"[DEBUG] No existing attendance. Creating new record.")
        attendance = Attendance(user=user, date=date)

    # Update attendance record
    attendance.status = status
    attendance.is_half_day = post_data.get('is_half_day') == 'on'
    attendance.location = post_data.get('location', 'Office')
    attendance.modified_by = request.user

    # Validate regularization reason is provided when needed
    if hasattr(attendance, 'original_status') and attendance.original_status != status:
        if not post_data.get('regularization_reason'):
            raise ValidationError("Regularization reason is required when changing attendance status")

    attendance.regularization_reason = post_data.get('regularization_reason', '')
    attendance.regularization_status = 'Approved'

    print(f"[DEBUG] Set location: {attendance.location}, is_half_day: {attendance.is_half_day}")

    # Clock-in with validation
    clock_in_time_str = post_data.get('clock_in_time')
    if clock_in_time_str:
        clock_in_time = datetime.combine(date, datetime.strptime(clock_in_time_str, '%H:%M').time())
        attendance.clock_in_time = timezone.make_aware(clock_in_time)
        print(f"[DEBUG] Clock-in time set to: {attendance.clock_in_time}")

    # Clock-out with validation
    clock_out_time_str = post_data.get('clock_out_time')
    if clock_out_time_str:
        clock_out_time = datetime.combine(date, datetime.strptime(clock_out_time_str, '%H:%M').time())
        attendance.clock_out_time = timezone.make_aware(clock_out_time)
        print(f"[DEBUG] Clock-out time set to: {attendance.clock_out_time}")

        # Validate clock-out is after clock-in
        if attendance.clock_in_time and attendance.clock_out_time <= attendance.clock_in_time:
            raise ValidationError("Clock out time must be after clock in time")

    # Calculate total hours (with improved break handling)
    if attendance.clock_in_time and attendance.clock_out_time:
        duration = attendance.clock_out_time - attendance.clock_in_time
        break_duration = timedelta(0)

        breaks_str = post_data.get('breaks')
        if breaks_str:
            try:
                breaks = json.loads(breaks_str) if isinstance(breaks_str, str) else breaks_str
                for break_item in breaks:
                    if break_item.get('start') and break_item.get('end'):
                        break_start = datetime.fromisoformat(break_item['start'])
                        break_end = datetime.fromisoformat(break_item['end'])

                        # Validate break times
                        if break_end <= break_start:
                            raise ValidationError("Break end time must be after break start time")

                        break_duration += break_end - break_start
                print(f"[DEBUG] Breaks total duration: {break_duration}")
            except json.JSONDecodeError:
                logger.error(f"Invalid break format: {breaks_str}")
                raise ValidationError("Invalid break format")
            except Exception as e:
                logger.error(f"Error calculating breaks: {e}")
                raise ValidationError(f"Error processing breaks: {str(e)}")

        total_hours = (duration - break_duration).total_seconds() / 3600
        attendance.total_hours = Decimal(str(round(total_hours, 2)))
        print(f"[DEBUG] Total hours worked: {attendance.total_hours}")

    # Shift information with better error handling
    is_late = False
    has_left_early = False

    try:
        current_shift = ShiftAssignment.get_user_current_shift(user, date)
        if current_shift:
            attendance.shift = current_shift
            attendance.expected_hours = Decimal(str(current_shift.shift_duration))
            print(f"[DEBUG] Shift assigned. Expected hours: {attendance.expected_hours}")

            # Late arrival
            if attendance.clock_in_time and current_shift.start_time:
                shift_start = timezone.make_aware(datetime.combine(date, current_shift.start_time))
                if attendance.clock_in_time > shift_start:
                    late_minutes = int((attendance.clock_in_time - shift_start).total_seconds() / 60)
                    attendance.late_minutes = late_minutes
                    print(f"[DEBUG] Late by {late_minutes} minutes")
                    if late_minutes > 0:
                        is_late = True

            # Early leave
            if attendance.clock_out_time and current_shift.end_time:
                shift_end = timezone.make_aware(datetime.combine(date, current_shift.end_time))
                if attendance.clock_out_time < shift_end:
                    early_minutes = int((shift_end - attendance.clock_out_time).total_seconds() / 60)
                    attendance.early_departure_minutes = early_minutes
                    print(f"[DEBUG] Left early by {early_minutes} minutes")
                    if early_minutes > 15:
                        attendance.left_early = True
                        has_left_early = True
                        print(f"[DEBUG] Marked as left early")
        else:
            logger.warning(f"No shift assignment found for user {user.username} on {date}")
    except Exception as e:
        logger.error(f"Error getting shift for user {user.username}: {e}")
        raise ValidationError(f"Error processing shift information: {str(e)}")

    # Update status based on late arrival and early departure
    if status == 'Present':
        if is_late and has_left_early:
            attendance.status = 'Present, Late & Early Exit'
        elif is_late:
            attendance.status = 'Present & Late'
        elif has_left_early:
            attendance.status = 'Present & Early Exit'

    # Special status
    if status == 'On Leave':
        leave_type = post_data.get('leave_type')
        if not leave_type:
            raise ValidationError("Leave type is required for 'On Leave' status")
        attendance.leave_type = leave_type
        print(f"[DEBUG] Leave type: {attendance.leave_type}")
    elif status == 'Holiday':
        attendance.is_holiday = True
        holiday_name = post_data.get('holiday_name', '')
        if not holiday_name:
            raise ValidationError("Holiday name is required for 'Holiday' status")
        attendance.holiday_name = holiday_name
        print(f"[DEBUG] Holiday marked: {attendance.holiday_name}")
    elif status == 'Weekend':
        attendance.is_weekend = True
        print(f"[DEBUG] Weekend marked")

    # Verify time data consistency
    if attendance.clock_in_time or attendance.clock_out_time:
        # Double-check that times in form match what's being saved
        if clock_in_time_str:
            form_clock_in = timezone.make_aware(datetime.combine(date, datetime.strptime(clock_in_time_str, '%H:%M').time()))
            if form_clock_in != attendance.clock_in_time:
                logger.error(f"Time inconsistency: Form clock-in {form_clock_in} != record clock-in {attendance.clock_in_time}")
                raise ValidationError("Clock-in time discrepancy detected")

        if clock_out_time_str:
            form_clock_out = timezone.make_aware(datetime.combine(date, datetime.strptime(clock_out_time_str, '%H:%M').time()))
            if form_clock_out != attendance.clock_out_time:
                logger.error(f"Time inconsistency: Form clock-out {form_clock_out} != record clock-out {attendance.clock_out_time}")
                raise ValidationError("Clock-out time discrepancy detected")

    # Overtime calculation
    if attendance.total_hours and attendance.expected_hours:
        overtime = float(attendance.total_hours) - float(attendance.expected_hours)
        if overtime > 0:
            attendance.overtime_hours = Decimal(str(round(overtime, 2)))
            print(f"[DEBUG] Overtime hours: {attendance.overtime_hours}")

    # Save with exception handling
    try:
        attendance.save()
        print(f"[DEBUG] Attendance record saved for {user.username} on {date}")
    except Exception as e:
        logger.error(f"Error saving attendance record: {e}")
        raise ValidationError(f"Failed to save attendance record: {str(e)}")

    logger.info(
        f"Attendance record modified by {request.user.username} for {user.username} "
        f"on {date}: status={attendance.status}, clock_in={attendance.clock_in_time}, "
        f"clock_out={attendance.clock_out_time}"
    )

    return attendance


def process_bulk_attendance(request, formset):
    """Process multiple attendance records from formset"""
    success_count = 0

    for form in formset:
        if form.is_valid() and not form.empty_permitted:
            try:
                attendance = process_single_attendance(request, form)
                success_count += 1
            except Exception as e:
                logger.error(f"Error processing bulk attendance record: {e}")
                raise

    return success_count

def process_csv_import(request):
    """Process attendance records from CSV upload"""
    csv_file = request.FILES['csv_file']
    if not csv_file.name.endswith('.csv'):
        raise ValidationError("File must be a CSV")

    decoded_file = csv_file.read().decode('utf-8').splitlines()
    reader = csv.DictReader(decoded_file)

    required_fields = ['username', 'date', 'status']
    for field in required_fields:
        if field not in reader.fieldnames:
            raise ValidationError(f"CSV missing required field: {field}")

    success_count = 0
    errors = []

    for row in reader:
        try:
            # Get user
            try:
                user = User.objects.get(username=row['username'])
                if not can_manage_user(request.user, user):
                    errors.append(f"No permission to manage {user.username}")
                    continue
            except User.DoesNotExist:
                errors.append(f"User not found: {row['username']}")
                continue

            # Parse date
            try:
                date = datetime.strptime(row['date'], '%Y-%m-%d').date()
            except ValueError:
                errors.append(f"Invalid date format for {user.username}: {row['date']}")
                continue

            # Validate status
            status = row['status']
            valid_statuses = [s[0] for s in Attendance.STATUS_CHOICES]
            if status not in valid_statuses:
                errors.append(f"Invalid status for {user.username}: {status}")
                continue

            # Get or create attendance record
            try:
                attendance = Attendance.objects.get(user=user, date=date)
                attendance.original_status = attendance.status
                attendance.original_clock_in_time = attendance.clock_in_time
                attendance.original_clock_out_time = attendance.clock_out_time
            except Attendance.DoesNotExist:
                attendance = Attendance(user=user, date=date)

            # Update attendance record
            attendance.status = status
            attendance.modified_by = request.user
            attendance.regularization_status = 'Approved'  # HR edits are pre-approved

            # Handle optional fields
            if 'clock_in_time' in row and row['clock_in_time']:
                try:
                    clock_in_time = datetime.strptime(row['clock_in_time'], '%H:%M')
                    attendance.clock_in_time = timezone.make_aware(datetime.combine(date, clock_in_time.time()))
                except ValueError:
                    errors.append(f"Invalid clock in time for {user.username}: {row['clock_in_time']}")

            if 'clock_out_time' in row and row['clock_out_time']:
                try:
                    clock_out_time = datetime.strptime(row['clock_out_time'], '%H:%M')
                    attendance.clock_out_time = timezone.make_aware(datetime.combine(date, clock_out_time.time()))
                except ValueError:
                    errors.append(f"Invalid clock out time for {user.username}: {row['clock_out_time']}")

            if 'location' in row and row['location']:
                valid_locations = [l[0] for l in Attendance.LOCATION_CHOICES]
                if row['location'] in valid_locations:
                    attendance.location = row['location']

            if 'is_half_day' in row:
                attendance.is_half_day = row['is_half_day'].lower() in ['true', 'yes', '1']

            if 'leave_type' in row and row['leave_type'] and status == 'On Leave':
                attendance.leave_type = row['leave_type']

            if 'remarks' in row and row['remarks']:
                attendance.remarks = row['remarks']

            # Calculate total hours if both clock times exist
            if attendance.clock_in_time and attendance.clock_out_time:
                total_hours = (attendance.clock_out_time - attendance.clock_in_time).total_seconds() / 3600
                attendance.total_hours = Decimal(str(round(total_hours, 2)))

            # Get shift information
            try:
                current_shift = ShiftAssignment.get_user_current_shift(user, date)
                if current_shift:
                    attendance.shift = current_shift
                    attendance.expected_hours = Decimal(str(current_shift.shift_duration))

                    # Calculate late minutes
                    if attendance.clock_in_time:
                        shift_start = timezone.make_aware(datetime.combine(date, current_shift.start_time))
                        if attendance.clock_in_time > shift_start:
                            late_minutes = int((attendance.clock_in_time - shift_start).total_seconds() / 60)
                            attendance.late_minutes = late_minutes
                            if late_minutes > 0 and status == 'Present':
                                attendance.status = 'Present & Late'

                    # Calculate early departure
                    if attendance.clock_out_time:
                        shift_end = timezone.make_aware(datetime.combine(date, current_shift.end_time))
                        if attendance.clock_out_time < shift_end:
                            early_minutes = int((shift_end - attendance.clock_out_time).total_seconds() / 60)
                            attendance.early_departure_minutes = early_minutes
                            if early_minutes > 15:  # Assuming 15 minutes grace period
                                attendance.left_early = True
            except Exception as e:
                logger.error(f"Error getting shift for user {user.username}: {e}")

            attendance.save()
            success_count += 1

        except Exception as e:
            logger.error(f"Error processing CSV row: {e}")
            errors.append(f"Error for {row.get('username', 'unknown')}: {str(e)}")

    if errors:
        logger.warning(f"CSV import completed with {len(errors)} errors: {errors[:5]}")

    return success_count

@login_required
@user_passes_test(is_hr_or_admin_check)
@require_POST
def mark_attendance_ajax(request):
    """AJAX endpoint for quickly marking attendance"""
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        date_str = data.get('date')
        status = data.get('status')

        if not user_id or not date_str or not status:
            return JsonResponse({'error': 'Missing required parameters'}, status=400)

        user = get_object_or_404(User, id=user_id)

        # Check permissions
        if not can_manage_user(request.user, user):
            return JsonResponse({'error': 'Permission denied'}, status=403)

        date = datetime.strptime(date_str, '%Y-%m-%d').date()

        # Get or create attendance record
        attendance, created = Attendance.objects.get_or_create(
            user=user,
            date=date,
            defaults={'status': status, 'modified_by': request.user}
        )

        if not created:
            attendance.original_status = attendance.status
            attendance.status = status
            attendance.modified_by = request.user
            attendance.regularization_status = 'Approved'
            attendance.save()

        return JsonResponse({
            'success': True,
            'message': f"Attendance marked as {status} for {user.username} on {date}"
        })

    except Exception as e:
        logger.error(f"Error in mark_attendance_ajax: {e}")
        return JsonResponse({'error': str(e)}, status=500)



def process_csv_import(request):
    """Process attendance records from CSV upload"""
    csv_file = request.FILES['csv_file']
    if not csv_file.name.endswith('.csv'):
        raise ValidationError("File must be a CSV")

    decoded_file = csv_file.read().decode('utf-8').splitlines()
    reader = csv.DictReader(decoded_file)

    required_fields = ['username', 'date', 'status']
    for field in required_fields:
        if field not in reader.fieldnames:
            raise ValidationError(f"CSV missing required field: {field}")

    success_count = 0
    errors = []

    for row in reader:
        try:
            # Get user
            try:
                user = User.objects.get(username=row['username'])
                if not can_manage_user(request.user, user):
                    errors.append(f"No permission to manage {user.username}")
                    continue
            except User.DoesNotExist:
                errors.append(f"User not found: {row['username']}")
                continue

            # Parse date
            try:
                date = datetime.strptime(row['date'], '%Y-%m-%d').date()
            except ValueError:
                errors.append(f"Invalid date format for {user.username}: {row['date']}")
                continue

            # Validate status
            status = row['status']
            valid_statuses = [s[0] for s in Attendance.STATUS_CHOICES]
            if status not in valid_statuses:
                errors.append(f"Invalid status for {user.username}: {status}")
                continue

            # Get or create attendance record
            try:
                attendance = Attendance.objects.get(user=user, date=date)
                attendance.original_status = attendance.status
                attendance.original_clock_in_time = attendance.clock_in_time
                attendance.original_clock_out_time = attendance.clock_out_time
            except Attendance.DoesNotExist:
                attendance = Attendance(user=user, date=date)

            # Update attendance record
            attendance.status = status
            attendance.modified_by = request.user
            attendance.regularization_status = 'Approved'  # HR edits are pre-approved

            # Handle optional fields
            if 'clock_in_time' in row and row['clock_in_time']:
                try:
                    clock_in_time = datetime.strptime(row['clock_in_time'], '%H:%M')
                    attendance.clock_in_time = timezone.make_aware(datetime.combine(date, clock_in_time.time()))
                except ValueError:
                    errors.append(f"Invalid clock in time for {user.username}: {row['clock_in_time']}")

            if 'clock_out_time' in row and row['clock_out_time']:
                try:
                    clock_out_time = datetime.strptime(row['clock_out_time'], '%H:%M')
                    attendance.clock_out_time = timezone.make_aware(datetime.combine(date, clock_out_time.time()))
                except ValueError:
                    errors.append(f"Invalid clock out time for {user.username}: {row['clock_out_time']}")

            if 'location' in row and row['location']:
                valid_locations = [l[0] for l in Attendance.LOCATION_CHOICES]
                if row['location'] in valid_locations:
                    attendance.location = row['location']

            if 'is_half_day' in row:
                attendance.is_half_day = row['is_half_day'].lower() in ['true', 'yes', '1']

            if 'leave_type' in row and row['leave_type'] and status == 'On Leave':
                attendance.leave_type = row['leave_type']

            if 'remarks' in row and row['remarks']:
                attendance.remarks = row['remarks']

            # Calculate total hours if both clock times exist
            if attendance.clock_in_time and attendance.clock_out_time:
                total_hours = (attendance.clock_out_time - attendance.clock_in_time).total_seconds() / 3600
                attendance.total_hours = Decimal(str(round(total_hours, 2)))

            # Get shift information
            try:
                current_shift = ShiftAssignment.get_user_current_shift(user, date)
                if current_shift:
                    attendance.shift = current_shift
                    attendance.expected_hours = Decimal(str(current_shift.shift_duration))

                    # Calculate late minutes
                    if attendance.clock_in_time:
                        shift_start = timezone.make_aware(datetime.combine(date, current_shift.start_time))
                        if attendance.clock_in_time > shift_start:
                            late_minutes = int((attendance.clock_in_time - shift_start).total_seconds() / 60)
                            attendance.late_minutes = late_minutes
                            if late_minutes > 0 and status == 'Present':
                                attendance.status = 'Present & Late'

                    # Calculate early departure
                    if attendance.clock_out_time:
                        shift_end = timezone.make_aware(datetime.combine(date, current_shift.end_time))
                        if attendance.clock_out_time < shift_end:
                            early_minutes = int((shift_end - attendance.clock_out_time).total_seconds() / 60)
                            attendance.early_departure_minutes = early_minutes
                            if early_minutes > 15:  # Assuming 15 minutes grace period
                                attendance.left_early = True
            except Exception as e:
                logger.error(f"Error getting shift for user {user.username}: {e}")

            attendance.save()
            success_count += 1

        except Exception as e:
            logger.error(f"Error processing CSV row: {e}")
            errors.append(f"Error for {row.get('username', 'unknown')}: {str(e)}")

    if errors:
        logger.warning(f"CSV import completed with {len(errors)} errors: {errors[:5]}")

    return success_count

@login_required
@user_passes_test(is_hr_or_admin_check)
@require_POST
def mark_attendance_ajax(request):
    """AJAX endpoint for quickly marking attendance"""
    try:
        data = json.loads(request.body)
        user_id = data.get('user_id')
        date_str = data.get('date')
        status = data.get('status')

        if not user_id or not date_str or not status:
            return JsonResponse({'error': 'Missing required parameters'}, status=400)

        user = get_object_or_404(User, id=user_id)

        # Check permissions
        if not can_manage_user(request.user, user):
            return JsonResponse({'error': 'Permission denied'}, status=403)

        date = datetime.strptime(date_str, '%Y-%m-%d').date()

        # Get or create attendance record
        attendance, created = Attendance.objects.get_or_create(
            user=user,
            date=date,
            defaults={'status': status, 'modified_by': request.user}
        )

        if not created:
            attendance.original_status = attendance.status
            attendance.status = status
            attendance.modified_by = request.user
            attendance.regularization_status = 'Approved'
            attendance.save()

        return JsonResponse({
            'success': True,
            'message': f"Attendance marked as {status} for {user.username} on {date}"
        })

    except Exception as e:
        logger.error(f"Error in mark_attendance_ajax: {e}")
        return JsonResponse({'error': str(e)}, status=500)



@login_required
@user_passes_test(is_hr_or_admin_check)
def bulk_update_attendance(request):
    """
    View for HR to bulk update attendance status
    """
    if request.method == 'POST':
        date_str = request.POST.get('date')
        status = request.POST.get('status')
        selected_users = request.POST.getlist('users')
        leave_type = request.POST.get('leave_type', '')
        remarks = request.POST.get('remarks', '')

        try:
            update_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            users = User.objects.filter(id__in=selected_users)

            updated_count = 0

            for user in users:
                # Get or create attendance record
                attendance, created = Attendance.objects.get_or_create(
                    user=user,
                    date=update_date,
                    defaults={'status': 'Not Marked'}
                )

                # Update status
                attendance.status = status
                attendance.modified_by = request.user
                attendance.regularization_reason = remarks
                attendance.regularization_status = 'Approved'  # Auto-approve HR updates

                # Update leave type if applicable
                if status in ['On Leave', 'Half Day'] and leave_type:
                    attendance.leave_type = leave_type
                    attendance.is_half_day = status == 'Half Day'

                attendance.save()
                updated_count += 1

            messages.success(request, f"Updated attendance for {updated_count} users.")
            return redirect('hr_attendance_list')

        except ValueError:
            messages.error(request, "Invalid date format.")

    # Get all active users
    users = User.objects.filter(is_active=True).order_by('username')

    # Get leave types
    leave_types = LeaveRequest.objects.values_list('leave_type__name', flat=True).distinct()

    context = {
        'users': users,
        'status_choices': Attendance.STATUS_CHOICES,
        'leave_types': leave_types
    }

    return render(request, 'components/hr/attendance/hr_bulk_update_attendance.html', context)

@login_required
@user_passes_test(is_hr_check)
def attendance_statistics(request):
    """
    View for HR to see attendance statistics and trends
    """
    # Get filter parameters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Default to last 30 days if no dates specified
    if not date_from or not date_to:
        today = timezone.localtime(timezone.now()).date()
        date_to = today.strftime('%Y-%m-%d')
        date_from = (today - timedelta(days=30)).strftime('%Y-%m-%d')

    try:
        from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        to_date = datetime.strptime(date_to, '%Y-%m-%d').date()

        # Calculate attendance statistics
        attendance_data = Attendance.objects.filter(
            date__gte=from_date,
            date__lte=to_date
        )

        # Status distribution
        status_distribution = attendance_data.values('status').annotate(
            count=Count('id')
        ).order_by('status')

        # Daily attendance counts
        daily_stats = {}
        current_date = from_date
        while current_date <= to_date:
            daily_data = attendance_data.filter(date=current_date)

            present_count = daily_data.filter(
                status__in=['Present', 'Present & Late', 'Work From Home']
            ).count()

            absent_count = daily_data.filter(status='Absent').count()
            leave_count = daily_data.filter(status__in=['On Leave', 'Half Day']).count()

            daily_stats[current_date.strftime('%Y-%m-%d')] = {
                'present': present_count,
                'absent': absent_count,
                'leave': leave_count,
                'total': present_count + absent_count + leave_count
            }

            current_date += timedelta(days=1)

        # Average time statistics
        avg_hours = attendance_data.exclude(total_hours__isnull=True).aggregate(Avg('total_hours'))['total_hours__avg'] or 0
        avg_late_mins = attendance_data.filter(late_minutes__gt=0).aggregate(Avg('late_minutes'))['late_minutes__avg'] or 0

        # Top 5 most absent users
        most_absent = User.objects.annotate(
            absent_count=Count(
                Case(
                    When(attendance__date__gte=from_date, attendance__date__lte=to_date, attendance__status='Absent', then=1),
                    output_field=IntegerField()
                )
            )
        ).filter(absent_count__gt=0).order_by('-absent_count')[:5]

        # Top 5 most late users
        most_late = User.objects.annotate(
            late_count=Count(
                Case(
                    When(attendance__date__gte=from_date, attendance__date__lte=to_date, attendance__status='Present & Late', then=1),
                    output_field=IntegerField()
                )
            )
        ).filter(late_count__gt=0).order_by('-late_count')[:5]

        # Prepare chart data
        status_chart = {
            'labels': [item['status'] for item in status_distribution],
            'data': [item['count'] for item in status_distribution],
        }

        daily_chart = {
            'labels': list(daily_stats.keys()),
            'present': [data['present'] for data in daily_stats.values()],
            'absent': [data['absent'] for data in daily_stats.values()],
            'leave': [data['leave'] for data in daily_stats.values()],
        }

        context = {
            'date_from': date_from,
            'date_to': date_to,
            'status_distribution': status_distribution,
            'daily_stats': daily_stats,
            'avg_hours': round(avg_hours, 2),
            'avg_late_mins': round(avg_late_mins, 2),
            'most_absent': most_absent,
            'most_late': most_late,
            'status_chart': json.dumps(status_chart),
            'daily_chart': json.dumps(daily_chart),
        }

    except ValueError:
        messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
        context = {
            'date_from': date_from,
            'date_to': date_to,
        }

    return render(request, 'components/hr/attendance/hr_attendance_statistics.html', context)

@login_required
def attendance_dashboard(request):
    """
    Personal attendance dashboard showing individual attendance records and statistics.
    """
    # Get current date and time in local timezone
    current_datetime = localtime(now())
    current_date = current_datetime.date()

    # Get filter parameters
    date_filter = request.GET.get('date_filter', 'current_month')
    date_range_start = request.GET.get('date_range_start', '')
    date_range_end = request.GET.get('date_range_end', '')

    # Set date range based on filter
    if date_filter == 'current_month':
        start_date = current_date.replace(day=1)
        end_date = (start_date.replace(month=start_date.month % 12 + 1, day=1) if start_date.month < 12
                   else start_date.replace(year=start_date.year + 1, month=1, day=1)) - timedelta(days=1)
    elif date_filter == 'previous_month':
        if current_date.month == 1:
            start_date = current_date.replace(year=current_date.year-1, month=12, day=1)
        else:
            start_date = current_date.replace(month=current_date.month-1, day=1)
        end_date = current_date.replace(day=1) - timedelta(days=1)
    elif date_filter == 'current_week':
        start_date = current_date - timedelta(days=current_date.weekday())
        end_date = start_date + timedelta(days=6)
    elif date_filter == 'previous_week':
        start_date = current_date - timedelta(days=current_date.weekday() + 7)
        end_date = start_date + timedelta(days=6)
    elif date_filter == 'custom_range' and date_range_start and date_range_end:
        try:
            start_date = datetime.strptime(date_range_start, '%Y-%m-%d').date()
            end_date = datetime.strptime(date_range_end, '%Y-%m-%d').date()
        except ValueError:
            start_date = current_date.replace(day=1)
            end_date = (current_date.replace(month=current_date.month % 12 + 1, day=1) if current_date.month < 12
                       else current_date.replace(year=current_date.year + 1, month=1, day=1)) - timedelta(days=1)
    else:
        # Default to current month
        start_date = current_date.replace(day=1)
        end_date = (current_date.replace(month=current_date.month % 12 + 1, day=1) if current_date.month < 12
                   else current_date.replace(year=current_date.year + 1, month=1, day=1)) - timedelta(days=1)

    # Get monthly attendance report
    month_report = Attendance.get_monthly_report(request.user, start_date.year, start_date.month)

    # Get today's attendance record if exists
    today_attendance = Attendance.objects.filter(user=request.user, date=current_date).first()

    absent_count = Attendance.objects.filter(user=request.user, date=current_date, status='Absent').count()

    # Check if user is currently in an active session
    is_active_session = False
    active_session = None

    active_session = UserSession.objects.filter(
        user=request.user,
        is_active=True
    ).first()

    if active_session:
        is_active_session = True

    # Convert dict_values to list for pagination
    attendance_records = list(month_report['days'].values())

    # Paginate attendance records
    paginator = Paginator(attendance_records, 10)
    page = request.GET.get('page', 1)

    try:
        paginated_records = paginator.page(page)
    except PageNotAnInteger:
        paginated_records = paginator.page(1)
    except EmptyPage:
        paginated_records = paginator.page(paginator.num_pages)

    # Get user's shift information
    user_shift = None
    try:
        shift_assignment = ShiftAssignment.get_user_current_shift(request.user, current_date)
        if shift_assignment:
            user_shift = shift_assignment
    except Exception as e:
        logger.error(f"Error getting shift for user {request.user.username}: {e}")

    # Calculate the correct present days (including late present days)
    total_present_days = month_report['summary']['present'] + month_report['summary']['present_late']

    context = {
        'attendance_records': paginated_records,
        'today_attendance': today_attendance,
        'is_active_session': is_active_session,
        'active_session': active_session,
        'current_date': current_date,
        'start_date': start_date,
        'end_date': end_date,
        'absent_count': absent_count,
        'total_days': len(month_report['days']),
        'present_days': total_present_days,  # Changed to include both present and present_late
        'present_late_days': month_report['summary']['present_late'],
        'absent_days': month_report['summary']['absent'],
        'late_days': month_report['summary']['late'],
        'leave_days': month_report['summary']['on_leave'],
        'wfh_days': month_report['summary']['work_from_home'],
        'total_hours': month_report['summary']['total_hours'],
        'overtime_hours': month_report['summary']['overtime_hours'],
        'date_filter': date_filter,
        'date_range_start': date_range_start,
        'date_range_end': date_range_end,
        'user_shift': user_shift,
    }

    return render(request, 'components/employee/attendance_dashboard.html', context)


@login_required
def attendance_calendar(request):
    """
    Monthly calendar view of personal attendance.
    """
    # Get current date
    current_date = localtime(now()).date()

    # Get month and year from request parameters or use current month/year
    month = int(request.GET.get('month', current_date.month))
    year = int(request.GET.get('year', current_date.year))

    # Calculate previous and next month
    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year

    if month == 12:
        next_month, next_year = 1, year + 1
    else:
        next_month, next_year = month + 1, year

    # Get all days in the month
    cal = calendar.monthcalendar(year, month)
    month_name = calendar.month_name[month]

    # Get the monthly attendance report
    month_report = Attendance.get_monthly_report(request.user, year, month)

    # Create a dictionary mapping days to attendance records
    attendance_by_day = {day: record for day, record in month_report['days'].items()}

    # Get user's shift information
    user_shift = None
    try:
        shift_assignment = ShiftAssignment.objects.filter(
            user=request.user,
            is_current=True
        ).select_related('shift').first()

        if shift_assignment:
            user_shift = shift_assignment.shift
    except:
        pass

    context = {
        'calendar': cal,
        'month': month,
        'year': year,
        'month_name': month_name,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
        'attendance_by_day': attendance_by_day,
        'current_date': current_date,
        'user_shift': user_shift,
    }

    return render(request, 'components/employee/attendance_calendar.html', context)

@login_required
def session_activity(request):
    """
    View and manage current session activity.
    """
    current_datetime = localtime(now())
    current_date = current_datetime.date()

    # Get today's attendance record
    today_attendance = Attendance.objects.filter(
        user=request.user,
        date=current_date
    ).first()

    # Get active session if any
    active_session = UserSession.objects.filter(
        user=request.user,
        is_active=True
    ).first()

    # Get all sessions for today
    today_sessions = UserSession.objects.filter(
        user=request.user,
        login_time__date=current_date
    ).order_by('login_time')

    # Calculate total session time and idle time
    total_session_time = timedelta(0)
    total_idle_time = timedelta(0)

    for session in today_sessions:
        if session.is_active:
            # For active sessions, calculate time until now
            session_duration = current_datetime - session.login_time
        else:
            # For completed sessions
            session_duration = session.logout_time - session.login_time

        total_session_time += session_duration
        if session.idle_time:
            total_idle_time += session.idle_time

    # Get user's shift information
    user_shift = None
    try:
        shift_assignment = ShiftAssignment.objects.filter(
            user=request.user,
            is_current=True
        ).select_related('shift').first()

        if shift_assignment:
            user_shift = shift_assignment.shift
    except:
        pass

    # Handle break actions
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'start_break':
            break_type = request.POST.get('break_type')
            if active_session:
                # Record break start in the active session
                if not active_session.breaks:
                    active_session.breaks = []

                active_session.breaks.append({
                    'type': break_type,
                    'start': current_datetime.isoformat(),
                    'end': None
                })
                active_session.save()

                messages.success(request, f"{break_type} break started.")
            else:
                messages.error(request, "No active session found to record break.")

        elif action == 'end_break':
            if active_session and active_session.breaks:
                # Find the last break without an end time
                for i in range(len(active_session.breaks) - 1, -1, -1):
                    if active_session.breaks[i].get('end') is None:
                        active_session.breaks[i]['end'] = current_datetime.isoformat()
                        active_session.save()
                        messages.success(request, "Break ended.")
                        break
                else:
                    messages.warning(request, "No active break found to end.")
            else:
                messages.error(request, "No active session or breaks found.")

        return redirect('aps_employee:session_activity')

    context = {
        'today_attendance': today_attendance,
        'active_session': active_session,
        'today_sessions': today_sessions,
        'total_session_time': total_session_time,
        'total_idle_time': total_idle_time,
        'current_datetime': current_datetime,
        'user_shift': user_shift,
    }

    return render(request, 'components/employee/session_activity.html', context)

def get_user_shift_for_date(user, date):
    """
    Get a user's assigned shift for a specific date.

    This function checks for:
    1. User-specific shift assignments for the given date
    2. Department-level default shifts
    3. User's default shift assignment

    Args:
        user: The User object
        date: datetime.date object for which to find the shift

    Returns:
        Shift object or None if no shift is found
    """
    from django.db.models import Q

    # Import here to avoid circular imports
    # These would be your actual model imports
    from .models import ShiftAssignment

    # Check if it's a weekend or holiday
    is_weekend = date.weekday() >= 5  # 5=Saturday, 6=Sunday

    # Try to find a date-specific shift assignment for this user
    # Check for specific date assignment first (highest priority)
    try:
        # Look for user-specific shift assignment for this date
        assignment = ShiftAssignment.objects.filter(
            user=user,
            date=date,
            is_active=True
        ).first()

        if assignment:
            return assignment.shift

        # If no specific assignment, check if user has a day-of-week assignment
        # E.g., "Every Monday" or "Every Weekend"
        day_of_week = date.strftime('%A').lower()  # 'monday', 'tuesday', etc.

        assignment = ShiftAssignment.objects.filter(
            Q(user=user) &
            Q(is_active=True) &
            (
                Q(day_of_week=day_of_week) |
                (Q(is_weekend_shift=True) & Q(is_weekend=is_weekend))
            )
        ).first()

        if assignment:
            return assignment.shift

        # Check if user has a default shift
        assignment = ShiftAssignment.objects.filter(
            user=user,
            is_default=True,
            is_active=True
        ).first()

        if assignment:
            return assignment.shift

        # If still no assignment, check if user's department has a default shift
        if hasattr(user, 'employee_profile') and user.employee_profile.department:
            department = user.employee_profile.department

            # Try department's day-specific shift
            dept_assignment = ShiftAssignment.objects.filter(
                department=department,
                day_of_week=day_of_week,
                is_active=True
            ).first()

            if dept_assignment:
                return dept_assignment.shift

            # Try department's default shift
            dept_assignment = ShiftAssignment.objects.filter(
                department=department,
                is_default=True,
                is_active=True
            ).first()

            if dept_assignment:
                return dept_assignment.shift

        # If we get here, try to return the organization's default shift
        default_shift = Shift.objects.filter(is_default=True).first()
        return default_shift

    except Exception as e:
        # Log the error
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error retrieving shift for user {user.username} on {date}: {str(e)}")
        return None


def combine_date_with_time(date, time):
    """
    Helper function to combine a date object with a time object into a datetime

    Args:
        date: datetime.date object
        time: datetime.time object

    Returns:
        datetime object with the date and time combined
    """
    from django.utils.timezone import make_aware
    from datetime import datetime

    dt = datetime.combine(date, time)
    return make_aware(dt)  # Make timezone-aware using default timezone


@login_required
def attendance_regularization(request):
    """
    Submit and manage attendance regularization requests.

    This view handles both:
    1. GET: Display form for regularization and show regularization history
    2. POST: Process the regularization submission
    """
    if request.method == 'POST':
        date_str = request.POST.get('date')
        clock_in_time = request.POST.get('clock_in_time')
        clock_out_time = request.POST.get('clock_out_time')
        reason = request.POST.get('reason')
        requested_status = request.POST.get('requested_status')
        location = request.POST.get('location', 'Office')

        try:
            # Parse the date and times
            regularization_date = datetime.strptime(date_str, '%Y-%m-%d').date()

            # Check if the date is valid for regularization (within policy limits)
            current_date = localtime(now()).date()
            days_difference = (current_date - regularization_date).days

            # Example policy: Can't regularize attendance older than 30 days
            if days_difference > 30:
                messages.error(request, "Regularization requests can only be submitted for dates within the last 30 days.")
                return redirect('aps_employee:attendance_regularization')

            # Check for future dates
            if regularization_date > current_date:
                messages.error(request, "Cannot regularize attendance for future dates.")
                return redirect('aps_employee:attendance_regularization')

            # Combine date and times
            clock_in_datetime = None
            clock_out_datetime = None

            if clock_in_time:
                clock_in_datetime = make_aware(datetime.combine(regularization_date,
                                     datetime.strptime(clock_in_time, '%H:%M').time()))

            if clock_out_time:
                clock_out_datetime = make_aware(datetime.combine(regularization_date,
                                     datetime.strptime(clock_out_time, '%H:%M').time()))

            # Time validation
            if clock_in_datetime and clock_out_datetime and clock_in_datetime >= clock_out_datetime:
                messages.error(request, "Clock-out time must be after clock-in time.")
                return redirect('aps_employee:attendance_regularization')

            # Get user's shift for the date to validate against shift timings
            user_shift = get_user_shift_for_date(request.user, regularization_date)

            if user_shift and clock_in_datetime and clock_out_datetime:
                # Example validation against shift times
                shift_start = combine_date_with_time(regularization_date, user_shift.start_time)
                shift_end = combine_date_with_time(regularization_date, user_shift.end_time)

                # Optional: Check if clock times make sense for the shift
                # This is a soft validation, just warning the user
                time_diff_start = abs((clock_in_datetime - shift_start).total_seconds() / 3600)
                time_diff_end = abs((clock_out_datetime - shift_end).total_seconds() / 3600)

                if time_diff_start > 2:  # More than 2 hours difference from shift start
                    messages.warning(request,
                        f"Your clock-in time is significantly different from your shift start time ({user_shift.start_time.strftime('%H:%M')})")

                if time_diff_end > 2:  # More than 2 hours difference from shift end
                    messages.warning(request,
                        f"Your clock-out time is significantly different from your shift end time ({user_shift.end_time.strftime('%H:%M')})")

            # Validate the request
            if not reason:
                messages.error(request, "Please provide a reason for the regularization request.")
                return redirect('aps_employee:attendance_regularization')

            if len(reason) < 10:
                messages.error(request, "Please provide a more detailed reason for the regularization request.")
                return redirect('aps_employee:attendance_regularization')

            # Get or create attendance record for the date
            attendance, created = Attendance.objects.get_or_create(
                user=request.user,
                date=regularization_date,
                defaults={
                    'status': 'Not Marked',
                }
            )

            # Check if there's already a pending regularization
            if attendance.regularization_status == 'Pending':
                messages.warning(request, "You already have a pending regularization request for this date.")
                return redirect('aps_employee:attendance_dashboard')

            # Store original values for audit purposes
            attendance.original_clock_in_time = attendance.clock_in_time
            attendance.original_clock_out_time = attendance.clock_out_time
            attendance.original_status = attendance.status

            # Update with regularization request
            attendance.regularization_status = 'Pending'
            attendance.regularization_reason = reason
            attendance.regularization_attempts += 1
            attendance.last_regularization_date = localtime(now())
            attendance.is_employee_notified = True
            attendance.is_hr_notified = False

            # Store the requested status change if provided
            if requested_status and requested_status in dict(Attendance.STATUS_CHOICES):
                attendance.requested_status = requested_status

            # Update location if provided
            if location and location in dict(Attendance.LOCATION_CHOICES):
                attendance.location = location

            if clock_in_datetime:
                attendance.clock_in_time = clock_in_datetime

            if clock_out_datetime:
                attendance.clock_out_time = clock_out_datetime

            # Calculate and update total hours if both clock times are provided
            if attendance.clock_in_time and attendance.clock_out_time:
                total_seconds = (attendance.clock_out_time - attendance.clock_in_time).total_seconds()
                attendance.total_hours = Decimal(total_seconds / 3600).quantize(Decimal('0.01'))

            attendance.save()

            # Notify HR about the regularization request
            notify_hr_about_regularization(attendance)

            messages.success(request, "Attendance regularization request submitted successfully.")
            return redirect('aps_employee:attendance_dashboard')

        except ValueError as e:
            messages.error(request, f"Invalid date or time format: {str(e)}")
            return redirect('aps_employee:attendance_regularization')
        except Exception as e:
            messages.error(request, f"Error submitting regularization request: {str(e)}")
            return redirect('aps_employee:attendance_regularization')

    # For GET requests, show regularization form
    # Get dates that might need regularization
    regularization_data = get_potential_regularization_dates(request.user)

    # Get pending regularization requests
    pending_requests = Attendance.objects.filter(
        user=request.user,
        regularization_status='Pending'
    ).order_by('-date')

    # Get recent regularization history
    regularization_history = Attendance.objects.filter(
        user=request.user,
        regularization_status__in=['Approved', 'Rejected']
    ).order_by('-last_regularization_date')[:10]  # Show last 10 regularizations

    context = {
        'records_needing_regularization': regularization_data['records_needing_regularization'],
        'pending_requests': pending_requests,
        'regularization_history': regularization_history,
        'current_date': regularization_data['current_date'],
        'status_choices': Attendance.STATUS_CHOICES,
        'location_choices': Attendance.LOCATION_CHOICES,
    }

    return render(request, 'components/employee/attendance_regularization.html', context)


@login_required
def view_regularization_history(request):
    """
    View for employees to see their regularization request history
    """
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Build query
    history_query = Attendance.objects.filter(
        user=request.user,
        regularization_status__isnull=False
    )

    # Apply filters
    if status_filter:
        history_query = history_query.filter(regularization_status=status_filter)

    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            history_query = history_query.filter(date__gte=date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            history_query = history_query.filter(date__lte=date_to_obj)
        except ValueError:
            pass

    # Order by date (newest first)
    history_list = history_query.order_by('-date')

    # Pagination
    paginator = Paginator(history_list, 10)  # 10 items per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
    }

    return render(request, 'components/employee/regularization_history.html', context)


def get_potential_regularization_dates(user):
    """
    Utility function to identify dates that may need regularization
    """
    current_date = localtime(now()).date()
    start_date = current_date - timedelta(days=30)  # Last 30 days

    # Get monthly report for regularization period
    month_report = Attendance.get_monthly_report(user, current_date.year, current_date.month)

    # Prepare records that might need regularization
    records_needing_regularization = []

    for day_record in month_report['days'].values():
        needs_regularization = False

        # Prepare record with consistent field names for template
        template_record = {
            'date': day_record['date'],
            'status': day_record['status'],
            'check_in': day_record.get('clock_in_time'),
            'check_out': day_record.get('clock_out_time'),
            'is_late': day_record.get('late_minutes', 0) > 0,
            'late_minutes': day_record.get('late_minutes', 0),
            'left_early': day_record.get('early_departure_minutes', 0) > 0,
            'early_departure_minutes': day_record.get('early_departure_minutes', 0)
        }

        # Missing clock-in or clock-out
        if (day_record['status'] == 'Present' and
            (not day_record.get('clock_in_time') or not day_record.get('clock_out_time'))):
            needs_regularization = True

        # Marked as absent but user might have been present
        if day_record['status'] == 'Absent':
            needs_regularization = True

        # Late arrival that might need justification
        if day_record.get('late_minutes', 0) > 15:
            needs_regularization = True

        # Early departure that might need justification
        if day_record.get('early_departure_minutes', 0) > 15:
            needs_regularization = True

        if needs_regularization:
            records_needing_regularization.append(template_record)

    return {
        'records_needing_regularization': records_needing_regularization,
        'current_date': current_date
    }

@login_required
def employee_attendance_view(request):
    # Current date in the local timezone
    current_date = localtime(now())
    print(f"Current date: {current_date}")

    # Get current month and year from query parameters or fallback to the current date
    current_month = int(request.GET.get('month', current_date.month))
    current_year = int(request.GET.get('year', current_date.year))
    current_month_name = calendar.month_name[current_month]
    print(f"Current month: {current_month}, Current year: {current_year}")

    # Calculate previous and next month and year
    prev_month = current_month - 1 if current_month > 1 else 12
    next_month = current_month + 1 if current_month < 12 else 1
    prev_year = current_year if current_month > 1 else current_year - 1
    next_year = current_year if current_month < 12 else current_year + 1
    print(f"Previous month: {prev_month}, Next month: {next_month}, Previous year: {prev_year}, Next year: {next_year}")

    # Generate the calendar for the current month
    cal = calendar.Calendar(firstweekday=6)  # Week starts on Sunday
    days_in_month = cal.monthdayscalendar(current_year, current_month)
    print(f"Days in month: {days_in_month}")

    # Query attendance and leave data for the current user
    user_attendance = Attendance.objects.filter(
        user=request.user,
        date__month=current_month,
        date__year=current_year
    ).select_related('user')
    print(f"User attendance records: {user_attendance.count()} found")

    # Get daily attendance aggregates (earliest clock in and latest clock out)
    daily_attendance = {}
    for attendance in user_attendance:
        date = attendance.date
        if date not in daily_attendance:
            daily_attendance[date] = {
                'first_clock_in': attendance.clock_in_time,
                'last_clock_out': attendance.clock_out_time,
                'total_hours': attendance.total_hours,
                'status': attendance.status,
                'location': attendance.location,
                'is_half_day': attendance.is_half_day,
                'regularization_status': attendance.regularization_status,
                'regularization_reason': attendance.regularization_reason,
                'breaks': attendance.breaks
            }
        else:
            # Update earliest clock in
            if attendance.clock_in_time and (not daily_attendance[date]['first_clock_in'] or
                attendance.clock_in_time < daily_attendance[date]['first_clock_in']):
                daily_attendance[date]['first_clock_in'] = attendance.clock_in_time

            # Update latest clock out
            if attendance.clock_out_time and (not daily_attendance[date]['last_clock_out'] or
                attendance.clock_out_time > daily_attendance[date]['last_clock_out']):
                daily_attendance[date]['last_clock_out'] = attendance.clock_out_time

            # Accumulate total hours
            if attendance.total_hours:
                daily_attendance[date]['total_hours'] = (daily_attendance[date]['total_hours'] or 0) + attendance.total_hours

    leaves = Leave.objects.filter(
        user=request.user,
        status='Approved',
        start_date__lte=datetime(current_year, current_month, calendar.monthrange(current_year, current_month)[1]),
        end_date__gte=datetime(current_year, current_month, 1)
    )
    print(f"Approved leaves: {leaves.count()} found")

    # Aggregate statistics including weekend work
    total_present = user_attendance.filter(status='Present').count()
    total_absent = user_attendance.filter(status='Absent').count()
    total_late = user_attendance.filter(status='Late').count()
    total_leave = user_attendance.filter(status='On Leave').count()
    total_wfh = user_attendance.filter(status='Work From Home').count()
    weekend_work = user_attendance.filter(is_weekend=True, status='Present').count()
    total_half_days = user_attendance.filter(is_half_day=True).count()
    print(f"Attendance stats - Present: {total_present}, Absent: {total_absent}, Late: {total_late}, Leave: {total_leave}, WFH: {total_wfh}, Weekend Work: {weekend_work}, Half Days: {total_half_days}")

    # Get average working hours
    avg_hours = user_attendance.exclude(total_hours__isnull=True).aggregate(
        avg_hours=Avg('total_hours')
    )['avg_hours'] or 0
    print(f"Average working hours: {avg_hours}")

    # Prepare calendar data with attendance and leave details
    calendar_data = []
    for week in days_in_month:
        week_data = []
        for day in week:
            if day == 0:
                week_data.append({'empty': True})
                print(f"Day {day} is empty")
            else:
                date = make_aware(datetime(current_year, current_month, day))
                leave_status = None
                leave_type = None

                # Check if leave exists for the day
                leave_on_day = leaves.filter(start_date__lte=date, end_date__gte=date).first()
                if leave_on_day:
                    leave_status = 'On Leave'
                    leave_type = leave_on_day.leave_type
                    print(f"Leave on day {day}: {leave_status}, Type: {leave_type}")

                # Get aggregated attendance data for the day
                attendance_date = date.date()
                attendance_data = daily_attendance.get(attendance_date, {})

                week_data.append({
                    'date': day,
                    'is_today': date.date() == current_date.date(),
                    'status': leave_status or attendance_data.get('status'),
                    'leave_type': leave_type,
                    'clock_in_time': attendance_data.get('first_clock_in'),
                    'clock_out_time': attendance_data.get('last_clock_out'),
                    'total_hours': attendance_data.get('total_hours'),
                    'breaks': attendance_data.get('breaks'),
                    'location': attendance_data.get('location'),
                    'is_half_day': attendance_data.get('is_half_day', False),
                    'is_sunday': date.weekday() == 6,
                    'is_weekend': date.weekday() >= 5,  # Saturday and Sunday
                    'regularization_status': attendance_data.get('regularization_status'),
                    'regularization_reason': attendance_data.get('regularization_reason'),
                    'empty': False
                })
        calendar_data.append(week_data)

    # Paginate the attendance records
    paginator = Paginator(user_attendance.order_by('-date'), 10)
    page = request.GET.get('page')
    print(f"Pagination page: {page}")
    try:
        records = paginator.get_page(page)
        print(f"Records on page: {len(records)}")
    except (EmptyPage, PageNotAnInteger):
        records = paginator.page(1)
        print("No valid page number provided, defaulting to page 1")

    return render(request, 'components/employee/calander.html', {
        'current_month': current_month_name,
        'current_year': current_year,
        'prev_month': prev_month,
        'next_month': next_month,
        'prev_year': prev_year,
        'next_year': next_year,
        'calendar_data': calendar_data,
        'total_present': total_present,
        'total_absent': total_absent,
        'total_late': total_late,
        'total_leave': total_leave,
        'total_wfh': total_wfh,
        'total_half_days': total_half_days,
        'weekend_work': weekend_work,
        'avg_hours': round(avg_hours, 2),
        'records': records,
    })

@login_required
@user_passes_test(is_manager)
def manager_attendance_view(request):
    # Prefetching related user manager data for efficiency
    team_attendance = Attendance.objects.filter(
        user__manager=request.user
    ).select_related('user').order_by('-date')

    # Pagination setup
    paginator = Paginator(team_attendance, 10)
    page = request.GET.get('page')

    try:
        team_records = paginator.get_page(page)
    except EmptyPage:
        team_records = paginator.page(paginator.num_pages)
    except PageNotAnInteger:
        team_records = paginator.page(1)

    return render(request, 'components/manager/manager_attendance.html', {
        'team_attendance': team_records
    })
@login_required
@user_passes_test(is_hr_check)
def hr_attendance_view(request):
    # Get the month and year from request params, default to current month
    today = timezone.now().date()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))

    # Get first and last day of selected month
    first_day = datetime(year, month, 1).date()
    last_day = datetime(year, month, calendar.monthrange(year, month)[1]).date()

    # Get all users with their details
    users = User.objects.select_related('profile').all().order_by('username')

    # Get all attendance records for the month with related data
    attendance_records = Attendance.objects.filter(
        date__range=[first_day, last_day]
    ).select_related('user', 'shift', 'modified_by')

    # Get leave records for the month
    leave_records = LeaveRequest.objects.filter(
        start_date__lte=last_day,
        end_date__gte=first_day,
        status='Approved'
    ).select_related('user', 'approver', 'leave_type')

    # Get current shift assignments for all users
    shift_assignments = ShiftAssignment.objects.filter(
        is_current=True
    ).select_related('user', 'shift')

    # Create a dictionary for quick access to shift assignments
    user_shifts = {assignment.user.id: assignment.shift for assignment in shift_assignments}

    # Create attendance matrix
    attendance_matrix = []
    days_in_month = calendar.monthrange(year, month)[1]

    for user in users:
        user_row = {
            'employee': user,
            'work_location': getattr(user.profile, 'work_location', 'Not set'),
            'attendance': {},
            'current_shift': user_shifts.get(user.id, None)
        }

        # Get the user's shift
        current_shift = user_shifts.get(user.id, None)

        # Initialize all days
        for day in range(1, days_in_month + 1):
            current_date = datetime(year, month, day).date()
            day_name = current_date.strftime('%a')

            # Determine if this is a working day based on shift settings
            is_weekend = is_weekend_for_user(current_date, current_shift)

            user_row['attendance'][current_date] = {
                'status': 'Weekend' if is_weekend else 'Not Marked',
                'working_hours': None,
                'day_name': day_name,
                'is_weekend': is_weekend,
                'is_holiday': False,
                'overtime_hours': 0,
                'late_minutes': 0,
                'breaks': [],
                'location': None,
                'regularization_status': None,
                'regularization_reason': None,
                'shift': current_shift,
                'modified_by': None,
                'remarks': None
            }

        # Fill in actual attendance records
        user_records = attendance_records.filter(user=user)
        for record in user_records:
            day_name = record.date.strftime('%a')
            working_hours = f"{record.total_hours:.1f}h" if record.total_hours else "-"

            # Determine if this is a working day based on the user's shift
            # Use the record's shift if available, otherwise fall back to current shift
            record_shift = record.shift if record.shift else current_shift
            is_weekend = is_weekend_for_user(record.date, record_shift)

            status = record.status
            # Only mark as Weekend Work if it's a weekend day AND they showed up as Present
            if is_weekend and status == 'Present':
                status = 'Weekend Work'

            user_row['attendance'][record.date] = {
                'status': status,
                'working_hours': working_hours,
                'day_name': day_name,
                'is_weekend': is_weekend,
                'is_holiday': record.is_holiday,
                'overtime_hours': record.overtime_hours,
                'late_minutes': record.late_minutes,
                'breaks': record.breaks,
                'location': record.location,
                'regularization_status': record.regularization_status,
                'regularization_reason': record.regularization_reason,
                'shift': record.shift.name if record.shift else 'No Shift',
                'shift_timing': f"{record.shift.start_time.strftime('%H:%M')} - {record.shift.end_time.strftime('%H:%M')}" if record.shift else None,
                'modified_by': record.modified_by.username if record.modified_by else None,
                'remarks': record.remarks,
                'clock_in': record.clock_in_time.strftime('%H:%M') if record.clock_in_time else None,
                'clock_out': record.clock_out_time.strftime('%H:%M') if record.clock_out_time else None
            }

        # Fill in leave records
        user_leaves = leave_records.filter(user=user)
        for leave in user_leaves:
            leave_dates = []
            current_date = max(leave.start_date, first_day)
            while current_date <= min(leave.end_date, last_day):
                leave_dates.append(current_date)
                current_date += timedelta(days=1)

            for date in leave_dates:
                if date in user_row['attendance']:
                    # Mark the day as a leave day
                    leave_status = 'On Leave'
                    if hasattr(leave, 'half_day') and leave.half_day:
                        leave_status = 'Half Day'

                    # Remove 'is_half_day' from the update dict
                    update_dict = {
                        'status': leave_status,
                        'leave_type': leave.leave_type.name,
                        'leave_reason': leave.reason,
                        'leave_approver': leave.approver.username if leave.approver else None
                    }
                    # If you want to show half day in the UI, you can add a custom key, but not 'is_half_day'
                    if hasattr(leave, 'half_day'):
                        update_dict['half_day'] = leave.half_day
                    user_row['attendance'][date].update(update_dict)

        attendance_matrix.append(user_row)

    # Calculate summary statistics
    summary = {
        'present_count': attendance_records.filter(status='Present').count(),
        'absent_count': attendance_records.filter(status='Absent').count(),
        'late_count': attendance_records.filter(status__in=['Late', 'Present & Late']).count(),
        'leave_count': attendance_records.filter(status='On Leave').count(),
        'wfh_count': attendance_records.filter(status='Work From Home').count(),
        # Removed 'half_day_count' as is_half_day is no longer in Attendance model
        # 'half_day_count': attendance_records.filter(is_half_day=True).count(),
        'weekend_work_count': attendance_records.filter(is_weekend=True, status='Present').count(),
        'not_marked_count': days_in_month * users.count() - attendance_records.count() - leave_records.count(),
        'total_overtime_hours': attendance_records.aggregate(Sum('overtime_hours'))['overtime_hours__sum'] or 0,
        'avg_working_hours': attendance_records.filter(total_hours__isnull=False).aggregate(Avg('total_hours'))['total_hours__avg'] or 0
    }

    # Get previous and next month links
    prev_month = 12 if month == 1 else month - 1
    prev_year = year - 1 if month == 1 else year
    next_month = 1 if month == 12 else month + 1
    next_year = year + 1 if month == 12 else year

    # Create days range for all days
    days_range = [datetime(year, month, day).date() for day in range(1, days_in_month + 1)]

    # Handle download requests
    if 'format' in request.GET:
        return handle_attendance_download(request, attendance_matrix, month, year)

    context = {
        'attendance_matrix': attendance_matrix,
        'days_range': days_range,
        'month': month,
        'year': year,
        'month_name': calendar.month_name[month],
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
        **summary
    }

    return render(request, 'components/hr/hr_admin_attendance.html', context)

def is_weekend_for_user(date, shift):
    """
    Determine if a date is a weekend day for a specific user based on their assigned shift.

    Args:
        date: The date to check
        shift: The ShiftMaster object representing the user's shift

    Returns:
        bool: True if it's a weekend day for this user, False otherwise
    """
    # Default behavior - Saturday and Sunday are weekend days
    if not shift:
        return date.weekday() >= 5  # Saturday (5) or Sunday (6)

    # Check based on shift's work_days setting
    if shift.work_days == 'All Days':
        return date.weekday() == 6  # Only Sunday is weekend
    elif shift.work_days == 'Weekdays':
        return date.weekday() >= 5  # Saturday and Sunday are weekend
    elif shift.work_days == 'Custom':
        # Get day names from custom_work_days
        if not shift.custom_work_days:
            return date.weekday() >= 5  # Default to weekend if no custom days

        custom_days = [day.strip() for day in shift.custom_work_days.split(',')]
        day_name = date.strftime('%A')  # Full day name (Monday, Tuesday, etc.)

        # If the day name is not in custom work days, it's a weekend
        return day_name not in custom_days

    # Default fallback
    return date.weekday() >= 5

def handle_attendance_download(request, attendance_matrix=None, month=None, year=None):
    """Handle attendance download requests for both direct and custom month downloads"""
    try:
        # Get format and date parameters
        export_format = request.GET.get('format', 'excel')

        # Check if this is a custom month request
        custom_month = request.GET.get('custom_month')
        if custom_month:
            # Parse custom month (format: YYYY-MM)
            year, month = map(int, custom_month.split('-'))
        else:
            # Use month and year from query parameters
            month = int(request.GET.get('month', datetime.now().month))
            year = int(request.GET.get('year', datetime.now().year))

        # If attendance_matrix wasn't provided, recreate it
        if not attendance_matrix:
            # Recreate attendance data
            first_day = datetime(year, month, 1).date()
            last_day = datetime(year, month, calendar.monthrange(year, month)[1]).date()

            # Get all users except clients
            employees = User.objects.select_related('profile').exclude(
                groups__name='Client'
            ).order_by('username')

            # Get all attendance records for the month
            attendance_records = Attendance.objects.filter(
                date__range=[first_day, last_day]
            ).select_related('user', 'shift')

            # Get leave records for the month
            leave_records = LeaveRequest.objects.filter(
                start_date__lte=last_day,
                end_date__gte=first_day,
                status='Approved'
            ).select_related('user', 'leave_type')

            # Get current shift assignments
            shift_assignments = ShiftAssignment.objects.filter(
                is_current=True
            ).select_related('user', 'shift')

            user_shifts = {assignment.user.id: assignment.shift for assignment in shift_assignments}

            # Create attendance matrix
            attendance_matrix = []
            days_in_month = calendar.monthrange(year, month)[1]

            for employee in employees:
                # Get the employee's shift
                current_shift = user_shifts.get(employee.id, None)

                employee_row = {
                    'employee': employee,
                    'work_location': getattr(employee.profile, 'work_location', 'Not set'),
                    'attendance': {},
                    'current_shift': current_shift
                }

                # Initialize all days
                for day in range(1, days_in_month + 1):
                    current_date = datetime(year, month, day).date()
                    day_name = current_date.strftime('%a')

                    # Determine if this is a working day based on shift settings
                    is_weekend = is_weekend_for_user(current_date, current_shift)

                    employee_row['attendance'][current_date] = {
                        'status': 'Weekend' if is_weekend else 'Not Marked',
                        'working_hours': None,
                        'day_name': day_name,
                        'is_weekend': is_weekend,
                        'shift': current_shift
                    }

                # Fill in actual attendance records
                employee_records = attendance_records.filter(user=employee)
                for record in employee_records:
                    # Determine if this is a working day based on the employee's shift
                    # Use the record's shift if available, otherwise fall back to current shift
                    record_shift = record.shift if record.shift else current_shift
                    is_weekend = is_weekend_for_user(record.date, record_shift)

                    status = record.status
                    if is_weekend and status == 'Present':
                        status = 'Weekend Work'

                    employee_row['attendance'][record.date]['status'] = status
                    employee_row['attendance'][record.date]['working_hours'] = record.total_hours
                    employee_row['attendance'][record.date]['shift'] = record.shift
                    employee_row['attendance'][record.date]['is_weekend'] = is_weekend

                # Fill in leave records
                employee_leaves = leave_records.filter(user=employee)
                for leave in employee_leaves:
                    leave_dates = []
                    current_date = max(leave.start_date, first_day)
                    while current_date <= min(leave.end_date, last_day):
                        leave_dates.append(current_date)
                        current_date += timedelta(days=1)

                    for date in leave_dates:
                        if date in employee_row['attendance']:
                            leave_status = 'On Leave'
                            if leave.half_day:
                                leave_status = 'Half Day'

                            employee_row['attendance'][date]['status'] = leave_status
                            employee_row['attendance'][date]['leave_type'] = leave.leave_type.name

                attendance_matrix.append(employee_row)

        # Export based on format
        if export_format == 'excel':
            return export_attendance_excel(attendance_matrix, month, year)
        elif export_format == 'csv':
            return export_attendance_csv(attendance_matrix, month, year)
        elif export_format == 'pdf':
            return export_attendance_pdf(attendance_matrix, month, year)
        else:
            raise Http404("Invalid export format")

    except Exception as e:
        # Log the error and return an error response
        print(f"Export error: {str(e)}")  # Replace with proper logging
        return HttpResponse(
            "Error generating report. Please try again.",
            status=500
        )

def export_attendance_excel(attendance_matrix, month, year):
    """Generate Excel version of attendance report with shift information"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import calendar
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance {calendar.month_name[month]} {year}"

    # Define styles
    header_fill = PatternFill(start_color='4B5563', end_color='4B5563', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    weekend_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
    present_fill = PatternFill(start_color='BBFFBB', end_color='BBFFBB', fill_type='solid')
    absent_fill = PatternFill(start_color='FFBBBB', end_color='FFBBBB', fill_type='solid')
    leave_fill = PatternFill(start_color='FFFFBB', end_color='FFFFBB', fill_type='solid')
    wfh_fill = PatternFill(start_color='BBBBFF', end_color='BBBBFF', fill_type='solid')
    late_fill = PatternFill(start_color='FFBBFF', end_color='FFBBFF', fill_type='solid')
    weekend_work_fill = PatternFill(start_color='FFD580', end_color='FFD580', fill_type='solid')  # Orange color for weekend work
    not_marked_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Get days in month
    days_in_month = calendar.monthrange(year, month)[1]
    dates = [datetime(year, month, day).date() for day in range(1, days_in_month + 1)]

    # Write headers
    headers = ['Employee', 'Username', 'Work Location', 'Shift', 'Shift Timing']
    # Add dates to headers
    for date in dates:
        headers.append(f"{date.day}\n{date.strftime('%a')}")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Write data
    row = 2
    for employee_data in attendance_matrix:
        employee = employee_data['employee']
        current_shift = employee_data['current_shift']

        ws.cell(row=row, column=1, value=f"{employee.first_name} {employee.last_name}").alignment = center_align
        ws.cell(row=row, column=1).border = border

        ws.cell(row=row, column=2, value=employee.username).alignment = center_align
        ws.cell(row=row, column=2).border = border

        ws.cell(row=row, column=3, value=employee_data['work_location']).alignment = center_align
        ws.cell(row=row, column=3).border = border

        # Shift info
        shift_name = 'No Shift'
        shift_timing = '-'
        if current_shift:
            shift_name = current_shift.name
            shift_timing = f"{current_shift.start_time.strftime('%H:%M')} - {current_shift.end_time.strftime('%H:%M')}"

        ws.cell(row=row, column=4, value=shift_name).alignment = center_align
        ws.cell(row=row, column=4).border = border

        ws.cell(row=row, column=5, value=shift_timing).alignment = center_align
        ws.cell(row=row, column=5).border = border

        # Attendance data
        col = 6
        for date in dates:
            attendance_data = employee_data['attendance'].get(date, {})
            status = attendance_data.get('status', 'Not Marked')

            cell = ws.cell(row=row, column=col, value=status)
            cell.alignment = center_align
            cell.border = border

            # Apply status-based formatting
            if status == 'Present':
                cell.fill = present_fill
            elif status == 'Absent':
                cell.fill = absent_fill
            elif status == 'On Leave':
                leave_type = attendance_data.get('leave_type', '')
                cell.value = f"{status}\n({leave_type})"
                cell.fill = leave_fill
            elif status == 'Half Day':
                leave_type = attendance_data.get('leave_type', '')
                if leave_type:
                    cell.value = f"{status}\n({leave_type})"
                cell.fill = leave_fill
            elif status == 'Weekend':
                cell.fill = weekend_fill
            elif status == 'Weekend Work':
                cell.fill = weekend_work_fill
            elif status == 'Holiday':
                cell.fill = weekend_fill
            elif status == 'Work From Home':
                cell.fill = wfh_fill
            elif 'Late' in status:
                cell.fill = late_fill
            elif status == 'Not Marked':
                cell.fill = not_marked_fill

            col += 1

        row += 1

    # Add summary row
    ws.cell(row=row + 1, column=1, value="SUMMARY").font = Font(bold=True)

    # Add a legend for status colors
    legend_row = row + 3
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)

    legend_items = [
        ("Present", present_fill),
        ("Absent", absent_fill),
        ("On Leave", leave_fill),
        ("Work From Home", wfh_fill),
        ("Late", late_fill),
        ("Weekend", weekend_fill),
        ("Weekend Work", weekend_work_fill),
        ("Holiday", weekend_fill),
        ("Not Marked", not_marked_fill)
    ]

    for i, (label, fill) in enumerate(legend_items):
        ws.cell(row=legend_row + i, column=2, value=label).alignment = center_align
        ws.cell(row=legend_row + i, column=2).border = border
        ws.cell(row=legend_row + i, column=2).fill = fill

    # Adjust column widths
    for col in range(1, 6):  # Employee info columns
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    for col in range(6, 6 + len(dates)):  # Date columns
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 12

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="attendance_{calendar.month_name[month]}_{year}.xlsx"'
    return response

def export_attendance_csv(attendance_matrix, month, year):
    """Generate CSV version of attendance report with shift information"""
    import csv
    from io import StringIO
    import calendar

    output = StringIO()
    writer = csv.writer(output)

    # Get days in month
    days_in_month = calendar.monthrange(year, month)[1]
    dates = [datetime(year, month, day).date() for day in range(1, days_in_month + 1)]

    # Write headers
    headers = ['Employee', 'Username', 'Work Location', 'Shift', 'Shift Timing']
    for date in dates:
        headers.append(f"{date.day} ({date.strftime('%a')})")

    writer.writerow(headers)

    # Write data
    for employee_data in attendance_matrix:
        employee = employee_data['employee']
        current_shift = employee_data['current_shift']

        # Employee info
        employee_info = [
            f"{employee.first_name} {employee.last_name}",
            employee.username,
            employee_data['work_location']
        ]

        # Shift info
        shift_name = 'No Shift'
        shift_timing = '-'
        if current_shift:
            shift_name = current_shift.name
            shift_timing = f"{current_shift.start_time.strftime('%H:%M')} - {current_shift.end_time.strftime('%H:%M')}"

        employee_info.extend([shift_name, shift_timing])

        # Attendance data
        for date in dates:
            attendance_data = employee_data['attendance'].get(date, {})
            status = attendance_data.get('status', 'Not Marked')

            if status == 'On Leave' or status == 'Half Day':
                leave_type = attendance_data.get('leave_type', '')
                if leave_type:
                    status = f"{status} ({leave_type})"

            employee_info.append(status)

        writer.writerow(employee_info)

    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="attendance_{calendar.month_name[month]}_{year}.csv"'
    return response


def export_attendance_pdf(attendance_matrix, month, year):
    """Generate PDF version of attendance report with shift information"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A3
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from io import BytesIO
    import calendar

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A3))
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title = Paragraph(f"Attendance Report - {calendar.month_name[month]} {year}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Get days in month
    days_in_month = calendar.monthrange(year, month)[1]
    dates = [datetime(year, month, day).date() for day in range(1, days_in_month + 1)]

    # Table data
    data = []

    # Headers
    headers = ['Employee', 'Username', 'Work Location', 'Shift', 'Shift Timing']
    for date in dates:
        headers.append(f"{date.day}\n({date.strftime('%a')})")

    data.append(headers)

    # Employee data
    for employee_data in attendance_matrix:
        employee = employee_data['employee']
        current_shift = employee_data['current_shift']

        row = [
            f"{employee.first_name} {employee.last_name}",
            employee.username,
            employee_data['work_location']
        ]

        # Shift info
        shift_name = 'No Shift'
        shift_timing = '-'
        if current_shift:
            shift_name = current_shift.name
            shift_timing = f"{current_shift.start_time.strftime('%H:%M')} - {current_shift.end_time.strftime('%H:%M')}"

        row.extend([shift_name, shift_timing])

        # Attendance data
        for date in dates:
            attendance_data = employee_data['attendance'].get(date, {})
            status = attendance_data.get('status', 'Not Marked')

            if status == 'On Leave' or status == 'Half Day':
                leave_type = attendance_data.get('leave_type', '')
                if leave_type:
                    status = f"{status}\n({leave_type})"

            row.append(status)

        data.append(row)

    # Create table and style
    table = Table(data)

    # Base style
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])

    # Add row colors for status
    for i in range(1, len(data)):
        for j in range(5, len(headers)):  # start from date columns
            status = data[i][j].split('\n')[0] if '\n' in data[i][j] else data[i][j]

            if status == 'Present':
                style.add('BACKGROUND', (j, i), (j, i), colors.lightgreen)
            elif status == 'Absent':
                style.add('BACKGROUND', (j, i), (j, i), colors.lightcoral)
            elif status == 'On Leave':
                style.add('BACKGROUND', (j, i), (j, i), colors.lightyellow)
            elif status == 'Half Day':
                style.add('BACKGROUND', (j, i), (j, i), colors.lightyellow)
            elif status == 'Weekend' or status == 'Holiday':
                style.add('BACKGROUND', (j, i), (j, i), colors.lightgrey)
            elif status == 'Work From Home':
                style.add('BACKGROUND', (j, i), (j, i), colors.lightblue)
            elif 'Late' in status:
                style.add('BACKGROUND', (j, i), (j, i), colors.plum)

    table.setStyle(style)
    elements.append(table)

    # Build the PDF
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="attendance_{calendar.month_name[month]}_{year}.pdf"'
    return response

'''------------------------------------------------ SUPPORT  AREA------------------------------------------------'''
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponseForbidden
from django.db.models import Q
from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponseForbidden
from django.db.models import Q
from django.core.paginator import Paginator
from .models import Support, TicketComment, TicketAttachment, TicketActivity
from .forms import TicketForm, CommentForm, TicketAttachmentForm
from django.db.models.functions import Trunc
from django.db.models.functions import TruncDay, TruncMonth, TruncWeek, TruncHour
from trueAlign.utils import send_ticket_notification


def get_user_roles(user):
    """Helper function to get user roles"""
    return {
        'is_admin': user.groups.filter(name='Admin').exists() or user.is_superuser,
        'is_hr': user.groups.filter(name='HR').exists(),
        'is_manager': user.groups.filter(name='Manager').exists(),
        'is_employee': user.groups.filter(name='Employee').exists()
    }


@login_required
def ticket_list(request):
    """View for listing tickets based on user role with enhanced filtering and sorting"""
    user = request.user
    user_roles = get_user_roles(user)

    # Determine which tickets to show based on user's role
    if user_roles['is_admin']:
        tickets = Support.objects.all()
    elif user_roles['is_hr']:
        tickets = Support.objects.filter(
            Q(assigned_group='HR') | Q(user=user)
        )
    elif user_roles['is_manager']:
        managed_users = User.objects.filter(department__manager=user)
        tickets = Support.objects.filter(
            Q(user=user) |
            Q(user__in=managed_users) |
            Q(assigned_to_user=user)
        ).distinct()
    else:
        tickets = Support.objects.filter(user=user)

    # Advanced filtering options
    status_filter = request.GET.get('status')
    if status_filter:
        tickets = tickets.filter(status=status_filter)

    priority_filter = request.GET.get('priority')
    if priority_filter:
        tickets = tickets.filter(priority=priority_filter)

    assigned_group_filter = request.GET.get('assigned_group')
    if assigned_group_filter and (user_roles['is_admin'] or user_roles['is_hr']):
        tickets = tickets.filter(assigned_group=assigned_group_filter)

    # Date range filtering
    date_from = request.GET.get('date_from')
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
            tickets = tickets.filter(created_at__gte=date_from)
        except ValueError:
            messages.error(request, 'Invalid date format for date_from. Use YYYY-MM-DD.')

    date_to = request.GET.get('date_to')
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d')
            date_to = date_to + timedelta(days=1)
            tickets = tickets.filter(created_at__lt=date_to)
        except ValueError:
            messages.error(request, 'Invalid date format for date_to. Use YYYY-MM-DD.')

    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        tickets = tickets.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(ticket_id__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )

    # Sorting
    sort_by = request.GET.get('sort', '-created_at')
    valid_sort_fields = ['created_at', '-created_at', 'priority', '-priority',
                         'status', '-status', 'user__username', '-user__username']
    if sort_by in valid_sort_fields:
        tickets = tickets.order_by(sort_by)
    else:
        tickets = tickets.order_by('-created_at')

    # Pagination with configurable page size
    page_size = int(request.GET.get('page_size', 10))
    paginator = Paginator(tickets, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'status_choices': Support.Status.choices,
        'priority_choices': Support.Priority.choices,
        'assigned_group_choices': Support.AssignedGroup.choices,
        'current_filters': {
            'status': status_filter,
            'priority': priority_filter,
            'assigned_group': assigned_group_filter,
            'date_from': date_from,
            'date_to': date_to,
            'search': search_query,
            'sort': sort_by,
            'page_size': page_size,
        },
        **user_roles,
    }
    return render(request, 'components/support/ticket_list.html', context)


@login_required
def create_ticket(request):
    from .utils import send_ticket_notification

    user = request.user
    user_roles = get_user_roles(user)

    if request.method == 'POST':
        form = TicketForm(request.POST, request.FILES)
        if user_roles['is_employee'] and not any([user_roles['is_admin'], user_roles['is_hr'], user_roles['is_manager']]):
            form.fields.pop('priority', None)
            form.fields.pop('assigned_group', None)
            form.fields.pop('assigned_to_user', None)

        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.user = request.user
            if not ticket.priority:
                if user_roles['is_manager']:
                    ticket.priority = Support.Priority.HIGH
                elif user_roles['is_employee']:
                    ticket.priority = Support.Priority.MEDIUM
            if not (user_roles['is_admin'] or user_roles['is_hr']) and not ticket.assigned_group:
                user_region = get_user_region(user)
                hr_issues = [Support.IssueType.HR, Support.IssueType.ACCESS]
                if ticket.issue_type in hr_issues:
                    ticket.assigned_group = Support.AssignedGroup.HR
                    regional_hr = get_regional_hr_contact(user_region)
                    if regional_hr:
                        ticket.assigned_to_user = regional_hr
                else:
                    ticket.assigned_group = Support.AssignedGroup.ADMIN
                    regional_it = get_regional_it_contact(user_region)
                    if regional_it:
                        ticket.assigned_to_user = regional_it
            ticket.sla_target_date = calculate_sla_target(ticket.priority)
            ticket.language = getattr(user, 'preferred_language', 'en')
            parent_ticket_id = request.POST.get('parent_ticket_id')
            if parent_ticket_id:
                try:
                    parent_ticket = Support.objects.get(ticket_id=parent_ticket_id)
                    ticket.parent_ticket = parent_ticket
                except Support.DoesNotExist:
                    messages.warning(request, f'Parent ticket {parent_ticket_id} not found. Creating as standalone ticket.')
            ticket.save(user=request.user)
            files = request.FILES.getlist('attachments')
            max_file_size = 5 * 1024 * 1024
            allowed_types = ['application/pdf', 'image/jpeg', 'image/png', 'image/gif',
                            'application/msword', 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                            'text/plain', 'application/zip']
            for file in files:
                if file.size > max_file_size:
                    messages.warning(request, f'File {file.name} exceeds the 5MB size limit and was not attached.')
                    continue
                file_type_valid = file.content_type in allowed_types
                if not file_type_valid:
                    file_extension = os.path.splitext(file.name)[1].lower()
                    allowed_extensions = ['.pdf', '.jpg', '.jpeg', '.png', '.gif', '.doc', '.docx', '.txt', '.zip']
                    if file_extension not in allowed_extensions:
                        messages.warning(request, f'File {file.name} type is not allowed and was not attached.')
                        continue
                TicketAttachment.objects.create(
                    ticket=ticket,
                    file=file,
                    uploaded_by=request.user,
                    file_size=file.size,
                    file_type=file.content_type
                )
            TicketActivity.objects.create(
                ticket=ticket,
                user=request.user,
                action=TicketActivity.Action.CREATED,
                details=f"Ticket created with priority {ticket.get_priority_display()}"
            )
            send_ticket_notification(ticket, 'created')
            messages.success(request, f'Ticket {ticket.ticket_id} created successfully. Your request will be processed according to its priority.')
            return redirect('aps_support:ticket_detail', pk=ticket.pk)
    else:
        form = TicketForm()
        if user_roles['is_employee'] and not any([user_roles['is_admin'], user_roles['is_hr'], user_roles['is_manager']]):
            form.fields.pop('priority', None)
            form.fields.pop('assigned_group', None)
            form.fields.pop('assigned_to_user', None)

    context = {
        'form': form,
        'title': 'Create New Support Ticket',
        'max_file_size': '5MB',
        'allowed_file_types': 'PDF, Word, Images, Text, ZIP',
        **user_roles,
    }
    return render(request, 'components/support/ticket_form.html', context)

def calculate_sla_target(priority):
    from django.utils import timezone
    now = timezone.now()
    SLA_HOURS = {
        'Critical': 4,
        'High': 24,
        'Medium': 72,
        'Low': 120,
    }
    if hasattr(priority, 'value'):
        priority_value = priority.value
    else:
        priority_value = str(priority)
    hours = SLA_HOURS.get(priority_value, 72)
    return now + timezone.timedelta(hours=hours)

def get_user_region(user):
    if hasattr(user, 'region') and user.region:
        return user.region
    if hasattr(user, 'profile') and hasattr(user.profile, 'region'):
        return user.profile.region
    return getattr(user, 'preferred_region', None) or 'default'

def get_regional_hr_contact(region):
    from django.contrib.auth import get_user_model
    User = get_user_model()
    try:
        return User.objects.filter(groups__name='HR', is_active=True, region=region).order_by('id').first()
    except Exception:
        return User.objects.filter(groups__name='HR', is_active=True).order_by('id').first()

def get_regional_it_contact(region):
    from django.contrib.auth import get_user_model
    User = get_user_model()
    try:
        return User.objects.filter(groups__name='IT', is_active=True, region=region).order_by('id').first()
    except Exception:
        return User.objects.filter(groups__name='IT', is_active=True).order_by('id').first()

def send_ticket_notification(ticket, *args, **kwargs):
    """
    Wrapper for send_ticket_notification that ignores unexpected keyword arguments.
    This prevents TypeError when extra kwargs (like old_status) are passed.
    """
    from .utils import send_ticket_notification as utils_send_ticket_notification

    # Only pass the arguments that utils_send_ticket_notification expects
    # (ticket, action, performed_by, extra_message)
    action = kwargs.get('action', None)
    performed_by = kwargs.get('user', None) or kwargs.get('performed_by', None)
    extra_message = kwargs.get('details', None) or kwargs.get('comment', None)

    return utils_send_ticket_notification(
        ticket,
        action=action,
        performed_by=performed_by,
        extra_message=extra_message
    )


@login_required
def ticket_detail(request, pk):
    """
    View for viewing a ticket's details with enhanced permissions and business logic.
    This version fixes status update issues by ensuring the status is always updated and saved.
    """
    ticket = get_object_or_404(Support, pk=pk)
    user = request.user
    user_roles = get_user_roles(user)
    is_ticket_owner = ticket.user == user

    # Enhanced permission check logic
    has_permission = (
        user_roles['is_admin'] or
        is_ticket_owner or
        (user_roles['is_hr'] and ticket.assigned_group == 'HR') or
        (ticket.assigned_to_user == user) or
        (user in ticket.cc_users.all())
    )

    # Add manager permission to view team tickets
    managed_users = None
    if user_roles['is_manager'] and not has_permission:
        # Remove department logic, just allow managers to view tickets of users they manage if such logic exists elsewhere
        # If you have a way to get managed users, use it; otherwise, skip this
        # For now, skip managed_users logic since department is not present
        pass

    if not has_permission:
        logger.warning(f"User {user.username} (ID: {user.id}) attempted unauthorized access to ticket {ticket.ticket_id}")
        return HttpResponseForbidden("You don't have permission to view this ticket.")

    # Check SLA status and update if needed
    current_time = timezone.now()
    if ticket.sla_target_date and ticket.status not in ['Resolved', 'Closed']:
        if current_time > ticket.sla_target_date:
            if ticket.sla_status != 'Breached':
                ticket.sla_status = 'Breached'
                ticket.save(update_fields=['sla_status'])
                if not TicketActivity.objects.filter(ticket=ticket, action='SLA_BREACHED').exists():
                    TicketActivity.objects.create(
                        ticket=ticket,
                        user=None,
                        action=TicketActivity.Action.SLA_BREACHED,
                        details=f"SLA breached. Target was {ticket.sla_target_date.strftime('%Y-%m-%d %H:%M')}"
                    )

    # --- STATUS UPDATE LOGIC: Allow status update via POST, even without a comment ---
    status_update_handled = False
    if request.method == 'POST':
        # Check if status update is requested
        new_status = request.POST.get('new_status')
        allowed_to_change_status = user_roles['is_admin'] or user_roles['is_hr']
        # Remove managed_users/department logic
        if user_roles['is_manager'] and is_ticket_owner:
            allowed_to_change_status = True
        if ticket.assigned_to_user == user:
            allowed_to_change_status = True
        if is_ticket_owner and new_status in ['Closed', 'Feedback']:
            allowed_to_change_status = True

        if new_status and allowed_to_change_status:
            old_status = ticket.status
            if new_status != old_status:
                ticket.status = new_status

                # Update resolution timestamps
                if new_status == 'Resolved' and not ticket.resolved_at:
                    ticket.resolved_at = timezone.now()
                    ticket.resolution_time = ticket.resolved_at - ticket.created_at
                    TicketActivity.objects.create(
                        ticket=ticket,
                        user=request.user,
                        action=TicketActivity.Action.RESOLVED,
                        details=f"Ticket resolved after {humanize_timedelta(ticket.resolution_time)}"
                    )

                # Handle feedback state for CSAT collection
                if new_status == 'Feedback' and old_status == 'Resolved':
                    ticket.satisfaction_survey_token = generate_unique_token()

                # Handle final closure
                if new_status == 'Closed':
                    ticket.closed_at = timezone.now()
                    if ticket.resolved_at:
                        ticket.time_to_close = ticket.closed_at - ticket.resolved_at
                    else:
                        ticket.resolved_at = ticket.closed_at
                        ticket.resolution_time = ticket.resolved_at - ticket.created_at
                    TicketActivity.objects.create(
                        ticket=ticket,
                        user=request.user,
                        action=TicketActivity.Action.CLOSED,
                        details=f"Ticket closed by {request.user.get_full_name() or request.user.username}"
                    )

                # Reset escalation level if status changes to In Progress
                if new_status == 'In Progress' and getattr(ticket, 'escalation_level', 0) > 0:
                    ticket.escalation_level = 0
                    messages.info(request, 'Escalation level has been reset as ticket is now In Progress.')

                # Always save the ticket after status change
                ticket.save(user=request.user)

                # Log status change activity
                TicketActivity.objects.create(
                    ticket=ticket,
                    user=request.user,
                    action='STATUS_CHANGED',
                    details=f"Status changed from {old_status} to {new_status}"
                )

                messages.info(request, f'Ticket status updated from {old_status} to {new_status}')
                send_ticket_notification(ticket, 'status_changed', old_status=old_status)
                status_update_handled = True

        # If only status was updated, redirect immediately
        if status_update_handled and not request.POST.get('comment') and not request.FILES.getlist('comment_attachments'):
            return redirect('aps_support:ticket_detail', pk=pk)

        # --- END STATUS UPDATE LOGIC ---

        # Comment form processing (if a comment is being added)
        comment_form = CommentForm(request.POST, request.FILES)
        if comment_form.is_valid():
            comment = comment_form.save(commit=False)
            comment.ticket = ticket
            comment.user = request.user

            # Only admins, HR and managers can make internal comments
            if comment_form.cleaned_data.get('is_internal') and not (user_roles['is_admin'] or user_roles['is_hr'] or user_roles['is_manager']):
                comment.is_internal = False

            comment.save()

            # Handle file attachments for comments
            comment_files = request.FILES.getlist('comment_attachments')
            for file in comment_files:
                CommentAttachment.objects.create(
                    comment=comment,
                    file=file,
                    uploaded_by=request.user
                )

            # If status was not already updated above, check again in case status is being changed with comment
            if not status_update_handled:
                new_status = request.POST.get('new_status')
                allowed_to_change_status = user_roles['is_admin'] or user_roles['is_hr']
                if user_roles['is_manager'] and is_ticket_owner:
                    allowed_to_change_status = True
                if ticket.assigned_to_user == user:
                    allowed_to_change_status = True
                if is_ticket_owner and new_status in ['Closed', 'Feedback']:
                    allowed_to_change_status = True

                if new_status and allowed_to_change_status:
                    old_status = ticket.status
                    if new_status != old_status:
                        ticket.status = new_status

                        # Update resolution timestamps
                        if new_status == 'Resolved' and not ticket.resolved_at:
                            ticket.resolved_at = timezone.now()
                            ticket.resolution_time = ticket.resolved_at - ticket.created_at
                            TicketActivity.objects.create(
                                ticket=ticket,
                                user=request.user,
                                action=TicketActivity.Action.RESOLVED,
                                details=f"Ticket resolved after {humanize_timedelta(ticket.resolution_time)}"
                            )

                        # Handle feedback state for CSAT collection
                        if new_status == 'Feedback' and old_status == 'Resolved':
                            ticket.satisfaction_survey_token = generate_unique_token()

                        # Handle final closure
                        if new_status == 'Closed':
                            ticket.closed_at = timezone.now()
                            if ticket.resolved_at:
                                ticket.time_to_close = ticket.closed_at - ticket.resolved_at
                            else:
                                ticket.resolved_at = ticket.closed_at
                                ticket.resolution_time = ticket.resolved_at - ticket.created_at
                            TicketActivity.objects.create(
                                ticket=ticket,
                                user=request.user,
                                action=TicketActivity.Action.CLOSED,
                                details=f"Ticket closed by {request.user.get_full_name() or request.user.username}"
                            )

                        # Reset escalation level if status changes to In Progress
                        if new_status == 'In Progress' and getattr(ticket, 'escalation_level', 0) > 0:
                            ticket.escalation_level = 0
                            messages.info(request, 'Escalation level has been reset as ticket is now In Progress.')

                        # Always save the ticket after status change
                        ticket.save(user=request.user)

                        # Log status change activity
                        TicketActivity.objects.create(
                            ticket=ticket,
                            user=request.user,
                            action='STATUS_CHANGED',
                            details=f"Status changed from {old_status} to {new_status}"
                        )

                        messages.info(request, f'Ticket status updated from {old_status} to {new_status}')
                        send_ticket_notification(ticket, 'status_changed', old_status=old_status)

            # Handle ticket escalation request
            if 'escalate_ticket' in request.POST and (is_ticket_owner or user_roles['is_manager']):
                if ticket.status not in ['Resolved', 'Closed']:
                    old_escalation = getattr(ticket, 'escalation_level', 0)
                    ticket.escalation_level = old_escalation + 1
                    ticket.last_escalated_at = timezone.now()
                    if ticket.escalation_level >= 2 and ticket.priority != 'Critical':
                        old_priority = ticket.priority
                        priority_order = ['Low', 'Medium', 'High', 'Critical']
                        current_index = priority_order.index(ticket.priority)
                        if current_index < len(priority_order) - 1:
                            ticket.priority = priority_order[current_index + 1]
                            messages.warning(request, f'Ticket priority automatically increased from {old_priority} to {ticket.priority} due to multiple escalations.')
                    # Remove department escalation logic
                    ticket.save(user=request.user)
                    TicketActivity.objects.create(
                        ticket=ticket,
                        user=request.user,
                        action=TicketActivity.Action.ESCALATED,
                        details=f"Ticket escalated from level {old_escalation} to level {ticket.escalation_level}"
                    )
                    send_ticket_notification(ticket, 'escalated')
                    messages.warning(request, f'Ticket has been escalated to level {ticket.escalation_level}.')
                else:
                    messages.error(request, 'Cannot escalate a resolved or closed ticket.')

            # Handle adding users to CC list
            cc_user_id = request.POST.get('add_cc_user')
            if cc_user_id and (user_roles['is_admin'] or user_roles['is_hr'] or user_roles['is_manager'] or is_ticket_owner):
                try:
                    cc_user = User.objects.get(id=cc_user_id)
                    ticket.cc_users.add(cc_user)
                    messages.success(request, f'{cc_user.get_full_name() or cc_user.username} added to CC list.')
                    send_cc_notification(ticket, cc_user)
                except User.DoesNotExist:
                    messages.error(request, 'User not found.')

            messages.success(request, 'Comment added successfully.')
            return redirect('aps_support:ticket_detail', pk=pk)
        else:
            # If comment form is invalid, fall through to render with errors
            comment_form = CommentForm(request.POST, request.FILES)
    else:
        comment_form = CommentForm()

    comments = ticket.comments.all()
    if not (user_roles['is_admin'] or user_roles['is_hr'] or user_roles['is_manager'] or ticket.assigned_to_user == user):
        comments = comments.filter(is_internal=False)

    activity_log = TicketActivity.objects.filter(ticket=ticket).order_by('-timestamp')

    sla_status = "N/A"
    sla_remaining = None
    if ticket.sla_target_date and ticket.status not in ['Resolved', 'Closed']:
        time_diff = ticket.sla_target_date - timezone.now()
        if time_diff.total_seconds() > 0:
            sla_status = "Within SLA"
            sla_remaining = humanize_timedelta(time_diff)
        else:
            sla_status = "SLA Breached"
            sla_remaining = f"Overdue by {humanize_timedelta(abs(time_diff))}"

    potential_duplicates = []
    if user_roles['is_admin'] or user_roles['is_hr'] or ticket.assigned_to_user == user:
        potential_duplicates = Support.objects.filter(
            Q(subject__icontains=ticket.subject[:30]) |
            Q(description__icontains=ticket.subject[:30])
        ).exclude(id=ticket.id)[:5]

    # Remove all logic that references user.department or ticket.user.department
    if user_roles['is_admin']:
        available_cc_users = User.objects.filter(is_active=True).exclude(id__in=ticket.cc_users.values_list('id', flat=True))
    elif user_roles['is_hr']:
        available_cc_users = User.objects.filter(
            Q(groups__name__in=['HR', 'Manager'])
        ).filter(is_active=True).exclude(id__in=ticket.cc_users.values_list('id', flat=True)).distinct()
    elif user_roles['is_manager'] or is_ticket_owner:
        available_cc_users = User.objects.filter(
            Q(id=ticket.assigned_to_user_id) if ticket.assigned_to_user_id else Q()
        ).filter(is_active=True).exclude(id__in=ticket.cc_users.values_list('id', flat=True)).distinct()
    else:
        available_cc_users = User.objects.none()

    context = {
        'ticket': ticket,
        'comments': comments,
        'comment_form': comment_form,
        'status_choices': Support.Status.choices,
        'activity_log': activity_log,
        'sla_status': sla_status,
        'sla_remaining': sla_remaining,
        'potential_duplicates': potential_duplicates,
        'available_cc_users': available_cc_users,
        'cc_users': ticket.cc_users.all(),
        **user_roles,
        'is_ticket_owner': is_ticket_owner,
        'is_assigned_to_user': ticket.assigned_to_user == user,
    }
    return render(request, 'components/support/ticket_detail.html', context )


@login_required
def update_ticket(request, pk):
    ticket = get_object_or_404(Support, pk=pk)
    user = request.user
    user_roles = get_user_roles(user)
    is_ticket_owner = ticket.user == user
    is_assigned_to_user = ticket.assigned_to_user == user

    has_permission = (
        user_roles['is_admin'] or
        is_ticket_owner or
        (user_roles['is_hr'] and ticket.assigned_group == 'HR') or
        is_assigned_to_user
    )

    if user_roles['is_manager'] and not has_permission:
        managed_users = User.objects.filter(department__manager=user)
        if ticket.user in managed_users:
            has_permission = True

    if not has_permission:
        logger.warning(f"User {user.username} (ID: {user.id}) attempted unauthorized ticket update for {ticket.ticket_id}")
        return HttpResponseForbidden("You don't have permission to update this ticket.")

    if ticket.status == 'Closed' and ticket.closed_at:
        days_since_closure = (timezone.now() - ticket.closed_at).days
        if days_since_closure > 30 and not user_roles['is_admin']:
            messages.error(request, "This ticket has been closed for more than 30 days and cannot be modified. Please create a new ticket if needed.")
            return redirect('aps_support:ticket_detail', pk=ticket.pk)

    if request.method == 'POST':
        form = TicketForm(request.POST, request.FILES, instance=ticket)
        if not user_roles['is_admin']:
            if not user_roles['is_hr'] or ticket.assigned_group != 'HR':
                form.fields.pop('assigned_group', None)
                form.fields.pop('assigned_to_user', None)
            if not (user_roles['is_hr'] or user_roles['is_manager'] or is_assigned_to_user):
                form.fields.pop('priority', None)
            if user_roles['is_employee'] and not is_ticket_owner:
                return HttpResponseForbidden("You don't have permission to update this ticket.")
            if ticket.status not in ['New', 'Open'] and 'issue_type' in form.fields:
                form.fields['issue_type'].disabled = True

        if form.is_valid():
            old_ticket = Support.objects.get(pk=ticket.pk)
            ticket = form.save(commit=False)
            changes = []
            for field in form.changed_data:
                if field not in ['attachments']:
                    old_value = getattr(old_ticket, field)
                    new_value = getattr(ticket, field)
                    if field == 'assigned_to_user' and old_value:
                        old_value = old_value.get_full_name() or old_value.username
                    if field == 'assigned_to_user' and new_value:
                        new_value = new_value.get_full_name() or new_value.username
                    display_method = f'get_{field}_display'
                    if hasattr(old_ticket, display_method):
                        try:
                            old_value = getattr(old_ticket, display_method)()
                            new_value = getattr(ticket, display_method)()
                        except:
                            pass
                    changes.append(f"{field.replace('_', ' ').title()}: {old_value} → {new_value}")

            if 'issue_type' in form.changed_data and not (user_roles['is_admin'] or user_roles['is_hr']):
                hr_issues = [Support.IssueType.HR, Support.IssueType.ACCESS]
                ticket.assigned_group = Support.AssignedGroup.HR if ticket.issue_type in hr_issues else Support.AssignedGroup.ADMIN
                changes.append(f"Assigned Group: Auto-assigned to {ticket.get_assigned_group_display()}")

            if 'issue_type' in form.changed_data and ticket.issue_type == Support.IssueType.OUTAGE and ticket.priority != 'Critical':
                old_priority = ticket.priority
                ticket.priority = Support.Priority.CRITICAL
                changes.append(f"Priority: {old_priority} → Critical (auto-updated for Outage)")

            if 'priority' in form.changed_data and ticket.status not in ['Resolved', 'Closed']:
                old_sla = ticket.sla_target_date
                ticket.sla_target_date = calculate_sla_target(ticket.priority)
                changes.append(f"SLA Target: {old_sla.strftime('%Y-%m-%d %H:%M') if old_sla else 'None'} → {ticket.sla_target_date.strftime('%Y-%m-%d %H:%M')}")

            if 'assigned_to_user' in form.changed_data and ticket.assigned_to_user:
                ticket.assigned_at = timezone.now()
                ticket.is_stale = False

            # --- STATUS UPDATE LOGIC FIX FOR UPDATE TICKET ---
            if 'status' in form.changed_data:
                old_status = old_ticket.status
                new_status = ticket.status
                if new_status == 'Resolved' and not ticket.resolved_at:
                    ticket.resolved_at = timezone.now()
                    ticket.resolution_time = ticket.resolved_at - ticket.created_at
                    TicketActivity.objects.create(
                        ticket=ticket,
                        user=request.user,
                        action=TicketActivity.Action.RESOLVED,
                        details=f"Ticket resolved after {humanize_timedelta(ticket.resolution_time)}"
                    )
                if new_status == 'Feedback' and old_status == 'Resolved':
                    ticket.satisfaction_survey_token = generate_unique_token()
                if new_status == 'Closed':
                    ticket.closed_at = timezone.now()
                    if ticket.resolved_at:
                        ticket.time_to_close = ticket.closed_at - ticket.resolved_at
                    else:
                        ticket.resolved_at = ticket.closed_at
                        ticket.resolution_time = ticket.resolved_at - ticket.created_at
                    TicketActivity.objects.create(
                        ticket=ticket,
                        user=request.user,
                        action=TicketActivity.Action.CLOSED,
                        details=f"Ticket closed by {request.user.get_full_name() or request.user.username}"
                    )
                if new_status == 'In Progress' and getattr(ticket, 'escalation_level', 0) > 0:
                    ticket.escalation_level = 0
                    messages.info(request, 'Escalation level has been reset as ticket is now In Progress.')
                # FIX: Use string for action, not enum attribute
                changes.append(f"Status: {old_status} → {new_status}")
            # --- END STATUS UPDATE LOGIC FIX ---

            ticket.save(user=request.user)

            if changes:
                TicketActivity.objects.create(
                    ticket=ticket,
                    user=request.user,
                    action=TicketActivity.Action.UPDATED,
                    details="Updated ticket fields:\n• " + "\n• ".join(changes)
                )

            files = request.FILES.getlist('attachments')
            for file in files:
                max_file_size = 5 * 1024 * 1024
                if file.size > max_file_size:
                    messages.warning(request, f'File {file.name} exceeds the 5MB size limit and was not attached.')
                    continue
                TicketAttachment.objects.create(
                    ticket=ticket,
                    file=file,
                    uploaded_by=request.user,
                    file_size=file.size,
                    file_type=file.content_type
                )

            significant_fields = ['status', 'priority', 'assigned_to_user', 'assigned_group']
            if any(field in form.changed_data for field in significant_fields):
                send_ticket_notification(ticket, 'updated', changed_fields=form.changed_data)

            messages.success(request, 'Ticket updated successfully.')
            return redirect('aps_support:ticket_detail', pk=ticket.pk)
    else:
        form = TicketForm(instance=ticket)
        if not user_roles['is_admin']:
            if not user_roles['is_hr'] or ticket.assigned_group != 'HR':
                form.fields.pop('assigned_group', None)
                form.fields.pop('assigned_to_user', None)
            if not (user_roles['is_hr'] or user_roles['is_manager'] or is_assigned_to_user):
                form.fields.pop('priority', None)
            if ticket.status not in ['New', 'Open'] and 'issue_type' in form.fields:
                form.fields['issue_type'].disabled = True

    context = {
        'form': form,
        'ticket': ticket,
        'title': f'Update Ticket {ticket.ticket_id}',
        'max_file_size': '5MB',
        'allowed_file_types': 'PDF, Word, Images, Text, ZIP',
        **user_roles,
        'is_ticket_owner': is_ticket_owner,
    }
    return render(request, 'components/support/ticket_form.html', context)


@login_required
def assign_ticket(request, pk):
    ticket = get_object_or_404(Support, pk=pk)
    user = request.user
    user_roles = get_user_roles(user)
    has_permission = user_roles['is_admin'] or (user_roles['is_hr'] and ticket.assigned_group == 'HR')
    if user_roles['is_manager'] and not has_permission:
        managed_users = User.objects.filter(department__manager=user)
        if ticket.user in managed_users:
            has_permission = True
    if not has_permission:
        logger.warning(f"User {user.username} (ID: {user.id}) attempted unauthorized ticket assignment for {ticket.ticket_id}")
        return HttpResponseForbidden("You don't have permission to assign this ticket.")
    if ticket.status in ['Resolved', 'Closed']:
        messages.warning(request, "This ticket is already resolved or closed. Reopening it for reassignment.")
    old_status = ticket.status
    old_assigned_to = ticket.assigned_to_user
    old_assigned_group = ticket.assigned_group
    if request.method == 'POST':
        new_status = request.POST.get('status')
        assigned_to_id = request.POST.get('assigned_to_user')
        new_assigned_group = request.POST.get('assigned_group')
        changes = []
        # --- STATUS UPDATE LOGIC FIX FOR ASSIGN TICKET ---
        if new_status and new_status != old_status:
            ticket.status = new_status
            changes.append(f"Status: {old_status} → {new_status}")
            if new_status == 'In Progress' and not ticket.in_progress_at:
                ticket.in_progress_at = timezone.now()
                ticket.response_time = ticket.in_progress_at - ticket.created_at
                changes.append(f"First response time: {humanize_timedelta(ticket.response_time)}")
            if new_status == 'Resolved' and not ticket.resolved_at:
                ticket.resolved_at = timezone.now()
                ticket.resolution_time = ticket.resolved_at - ticket.created_at
                TicketActivity.objects.create(
                    ticket=ticket,
                    user=request.user,
                    action=TicketActivity.Action.RESOLVED,
                    description=f"Ticket resolved after {humanize_timedelta(ticket.resolution_time)}"
                )
            if new_status == 'Closed':
                ticket.closed_at = timezone.now()
                if ticket.resolved_at:
                    ticket.time_to_close = ticket.closed_at - ticket.resolved_at
                else:
                    ticket.resolved_at = ticket.closed_at
                    ticket.resolution_time = ticket.resolved_at - ticket.created_at
                TicketActivity.objects.create(
                    ticket=ticket,
                    user=request.user,
                    action=TicketActivity.Action.CLOSED,
                    description=f"Ticket closed by {request.user.get_full_name() or request.user.username}"
                )
            if new_status == 'In Progress' and getattr(ticket, 'escalation_level', 0) > 0:
                ticket.escalation_level = 0
                messages.info(request, 'Escalation level has been reset as ticket is now In Progress.')
        # --- END STATUS UPDATE LOGIC FIX FOR ASSIGN TICKET ---
        if assigned_to_id:
            try:
                new_assigned_user = User.objects.get(id=assigned_to_id)
                if old_assigned_to != new_assigned_user:
                    ticket.assigned_to_user = new_assigned_user
                    ticket.assigned_at = timezone.now()
                    ticket.is_stale = False
                    old_assigned_name = old_assigned_to.get_full_name() if old_assigned_to else "None"
                    new_assigned_name = new_assigned_user.get_full_name() or new_assigned_user.username
                    changes.append(f"Assigned to: {old_assigned_name} → {new_assigned_name}")
            except User.DoesNotExist:
                messages.error(request, "Selected user does not exist.")
        if (user_roles['is_admin'] and new_assigned_group) or \
           (user_roles['is_hr'] and ticket.assigned_group == 'HR' and new_assigned_group == 'HR'):
            if old_assigned_group != new_assigned_group:
                ticket.assigned_group = new_assigned_group
                changes.append(f"Assigned group: {old_assigned_group} → {new_assigned_group}")
        new_priority = request.POST.get('priority')
        if new_priority and user_roles['is_admin'] and ticket.priority != new_priority:
            old_priority = ticket.priority
            ticket.priority = new_priority
            changes.append(f"Priority: {old_priority} → {new_priority}")
            if ticket.status not in ['Resolved', 'Closed']:
                old_sla = ticket.sla_target_date
                ticket.sla_target_date = calculate_sla_target(new_priority)
                changes.append(f"SLA Target: {old_sla.strftime('%Y-%m-%d %H:%M') if old_sla else 'None'} → {ticket.sla_target_date.strftime('%Y-%m-%d %H:%M')}")
        assignment_note = request.POST.get('assignment_note', '').strip()
        ticket.save(user=request.user)
        if changes:
            TicketActivity.objects.create(
                ticket=ticket,
                user=request.user,
                action=TicketActivity.Action.ASSIGNED,
                description="Updated ticket assignment:\n• " + "\n• ".join(changes)
            )
        if assignment_note:
            comment = Comment.objects.create(
                ticket=ticket,
                user=request.user,
                content=assignment_note,
                is_internal=True
            )
            TicketActivity.objects.create(
                ticket=ticket,
                user=request.user,
                action=TicketActivity.Action.COMMENTED,
                description="Added assignment note"
            )
        send_ticket_notification(ticket, 'assigned', old_assigned_to=old_assigned_to, assignment_note=assignment_note if assignment_note else None)
        messages.success(request, 'Ticket assignment updated successfully.')
        return redirect('aps_support:ticket_detail', pk=pk)
    if user_roles['is_admin']:
        assignable_users = User.objects.filter(is_active=True)
    elif user_roles['is_hr'] and ticket.assigned_group == 'HR':
        assignable_users = User.objects.filter(groups__name='HR', is_active=True)
    elif user_roles['is_manager']:
        assignable_users = User.objects.filter(
            Q(department__manager=user) | Q(id=user.id),
            is_active=True
        )
    else:
        assignable_users = User.objects.none()
    recently_active_users = User.objects.filter(
        last_login__gte=timezone.now() - timedelta(days=7),
        is_active=True
    ).intersection(assignable_users)[:5]
    related_ticket_handlers = User.objects.filter(
        assigned_tickets__issue_type=ticket.issue_type,
        is_active=True
    ).intersection(assignable_users).distinct()[:5]
    context = {
        'ticket': ticket,
        'status_choices': Support.Status.choices,
        'priority_choices': Support.Priority.choices,
        'assigned_group_choices': Support.AssignedGroup.choices,
        'assignable_users': assignable_users,
        'recently_active_users': recently_active_users,
        'related_ticket_handlers': related_ticket_handlers,
        **user_roles,
    }
    return render(request, 'components/support/assign_ticket.html', context)

@login_required
def support_dashboard(request):
    from datetime import timedelta, datetime
    import pytz

    def humanize_timedelta(td):
        if td is None:
            return None
        if isinstance(td, (int, float)):
            td = timedelta(seconds=td)
        total_seconds = int(td.total_seconds())
        if total_seconds < 60:
            return f"{total_seconds}s"
        minutes, seconds = divmod(total_seconds, 60)
        if minutes < 60:
            return f"{minutes}m {seconds}s"
        hours, minutes = divmod(minutes, 60)
        if hours < 24:
            return f"{hours}h {minutes}m"
        days, hours = divmod(hours, 24)
        return f"{days}d {hours}h"

    user = request.user
    user_roles = get_user_roles(user)
    from django.conf import settings
    tz = pytz.timezone(getattr(settings, "TIME_ZONE", "UTC"))
    date_range = request.GET.get('date_range', 'last_30_days')
    from_date = None
    to_date = timezone.now().astimezone(tz)

    def make_aware_if_naive(dt):
        if dt is not None:
            if timezone.is_naive(dt):
                return tz.localize(dt)
            return dt.astimezone(tz)
        return dt

    if date_range == 'last_7_days':
        from_date = to_date - timedelta(days=7)
    elif date_range == 'last_30_days':
        from_date = to_date - timedelta(days=30)
    elif date_range == 'last_90_days':
        from_date = to_date - timedelta(days=90)
    elif date_range == 'year_to_date':
        from_date = datetime(to_date.year, 1, 1, tzinfo=tz)
    elif date_range == 'custom':
        try:
            from_date_str = request.GET.get('from_date')
            to_date_str = request.GET.get('to_date')
            if from_date_str:
                from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
                from_date = make_aware_if_naive(from_date)
            if to_date_str:
                to_date = datetime.strptime(to_date_str, '%Y-%m-%d')
                to_date = make_aware_if_naive(to_date)
                to_date = to_date + timedelta(days=1)
        except ValueError:
            messages.error(request, 'Invalid date format. Using default 30 days range.')
            from_date = to_date - timedelta(days=30)

    from_date = make_aware_if_naive(from_date)
    to_date = make_aware_if_naive(to_date)

    if user_roles['is_admin']:
        tickets = Support.objects.all()
    elif user_roles['is_hr']:
        tickets = Support.objects.filter(
            Q(assigned_group='HR') | Q(user=user)
        )
    elif user_roles['is_manager']:
        managed_users = User.objects.filter(department__manager=user)
        tickets = Support.objects.filter(
            Q(user=user) |
            Q(user__in=managed_users) |
            Q(assigned_to_user=user)
        ).distinct()
    else:
        tickets = Support.objects.filter(user=user)

    if from_date:
        tickets_in_period = tickets.filter(created_at__gte=from_date, created_at__lt=to_date)
    else:
        tickets_in_period = tickets

    total_tickets = tickets_in_period.count()
    open_tickets = tickets_in_period.filter(status__in=['New', 'Open', 'In Progress']).count()
    in_progress_tickets = tickets_in_period.filter(status='In Progress').count()
    resolved_tickets = tickets_in_period.filter(status__in=['Resolved', 'Closed']).count()
    closed_tickets = tickets_in_period.filter(status='Closed').count()

    if from_date:
        period_length = (to_date - from_date).days
        previous_from_date = from_date - timedelta(days=period_length)
        previous_to_date = from_date
        previous_from_date = make_aware_if_naive(previous_from_date)
        previous_to_date = make_aware_if_naive(previous_to_date)
        previous_tickets = tickets.filter(created_at__gte=previous_from_date, created_at__lt=previous_to_date)
        previous_total = previous_tickets.count()
        if previous_total > 0:
            ticket_growth_rate = ((total_tickets - previous_total) / previous_total) * 100
        else:
            ticket_growth_rate = 100 if total_tickets > 0 else 0
    else:
        ticket_growth_rate = 0

    priority_distribution = tickets_in_period.values('priority').annotate(count=Count('id'))
    priority_counts = {
        'Critical': tickets_in_period.filter(priority='Critical').count(),
        'High': tickets_in_period.filter(priority='High').count(),
        'Medium': tickets_in_period.filter(priority='Medium').count(),
        'Low': tickets_in_period.filter(priority='Low').count(),
    }
    issue_distribution = tickets_in_period.values('issue_type').annotate(count=Count('id'))
    recent_tickets = tickets.prefetch_related(
        'user', 'assigned_to_user', 'comments'
    ).order_by('-created_at')[:10]
    time_analytics = {
        'avg_resolution_time': None,
        'avg_response_time': None,
        'sla_compliance': 0,
        'avg_time_to_close': None,
        'avg_resolution_time_human': None,
        'avg_response_time_human': None,
        'avg_time_to_close_human': None,
    }
    resolved_with_time = tickets_in_period.filter(resolution_time__isnull=False)
    responded_tickets = tickets_in_period.filter(response_time__isnull=False)
    closed_with_time = tickets_in_period.filter(
        Q(time_to_close__isnull=False) |
        Q(resolution_time__isnull=False, status__in=['Resolved', 'Closed'])
    )
    try:
        if resolved_with_time.exists():
            avg_res_time = resolved_with_time.aggregate(avg_time=Avg('resolution_time'))['avg_time']
            time_analytics['avg_resolution_time'] = avg_res_time
            time_analytics['avg_resolution_time_human'] = humanize_timedelta(avg_res_time)
    except ValueError:
        time_analytics['avg_resolution_time'] = None
        time_analytics['avg_resolution_time_human'] = "Invalid data"
    try:
        if responded_tickets.exists():
            avg_resp_time = responded_tickets.aggregate(avg_time=Avg('response_time'))['avg_time']
            time_analytics['avg_response_time'] = avg_resp_time
            time_analytics['avg_response_time_human'] = humanize_timedelta(avg_resp_time)
    except ValueError:
        time_analytics['avg_response_time'] = None
        time_analytics['avg_response_time_human'] = "Invalid data"
    try:
        if closed_with_time.exists():
            if hasattr(closed_with_time.model, 'time_to_close'):
                avg_close_time = closed_with_time.aggregate(avg_time=Avg('time_to_close'))['avg_time']
            else:
                avg_close_time = closed_with_time.aggregate(avg_time=Avg('resolution_time'))['avg_time']
            time_analytics['avg_time_to_close'] = avg_close_time
            time_analytics['avg_time_to_close_human'] = humanize_timedelta(avg_close_time)
    except ValueError:
        time_analytics['avg_time_to_close'] = None
        time_analytics['avg_time_to_close_human'] = "Invalid data"
    try:
        tickets_with_sla = tickets_in_period.filter(sla_target_date__isnull=False)
        if tickets_with_sla.exists():
            compliant_tickets = tickets_with_sla.filter(
                Q(
                    Q(resolved_at__isnull=False) & Q(resolved_at__lte=F('sla_target_date'))
                ) |
                Q(
                    Q(status__in=['New', 'Open', 'In Progress']) &
                    Q(sla_target_date__gt=timezone.now().astimezone(tz))
                )
            ).count()
            time_analytics['sla_compliance'] = (compliant_tickets / tickets_with_sla.count()) * 100 if tickets_with_sla.count() > 0 else 100
    except ValueError:
        time_analytics['sla_compliance'] = None
    workload_distribution = None
    bottleneck_analysis = None
    performance_by_assignee = None
    if user_roles['is_admin'] or user_roles['is_hr'] or user_roles['is_manager']:
        try:
            unassigned_tickets = tickets_in_period.filter(assigned_to_user__isnull=True).count()
            assigned_tickets = total_tickets - unassigned_tickets
        except ValueError:
            unassigned_tickets = None
            assigned_tickets = None
        if assigned_tickets and assigned_tickets > 0:
            try:
                workload_distribution = tickets_in_period.filter(
                    assigned_to_user__isnull=False
                ).values(
                    'assigned_to_user__username',
                    'assigned_to_user__first_name',
                    'assigned_to_user__last_name'
                ).annotate(
                    open_count=Count('id', filter=Q(status__in=['New', 'Open', 'In Progress'])),
                    total_count=Count('id'),
                    avg_resolution=Avg('resolution_time', filter=Q(resolution_time__isnull=False))
                ).order_by('-open_count')[:10]
            except ValueError:
                workload_distribution = None
        current_time = timezone.now().astimezone(tz)
        try:
            bottleneck_tickets = tickets.filter(
                Q(status='New', created_at__lt=current_time - timedelta(days=2)) |
                Q(status='Open', created_at__lt=current_time - timedelta(days=3)) |
                Q(status='In Progress', created_at__lt=current_time - timedelta(days=5))
            )
            bottleneck_analysis = bottleneck_tickets.values('status').annotate(count=Count('id'))
        except ValueError:
            bottleneck_analysis = None
        try:
            performance_by_assignee = tickets_in_period.filter(
                assigned_to_user__isnull=False,
                resolution_time__isnull=False
            ).values(
                'assigned_to_user__username',
                'assigned_to_user__first_name',
                'assigned_to_user__last_name'
            ).annotate(
                resolved_count=Count('id'),
                avg_resolution_time=Avg('resolution_time'),
                sla_breached=Count('id', filter=Q(sla_status='Breached')),
                sla_met=Count('id', filter=Q(sla_status__in=['Within SLA', '']) | Q(sla_status__isnull=True))
            ).filter(resolved_count__gt=0).order_by('avg_resolution_time')[:10]
        except ValueError:
            performance_by_assignee = None
    if from_date and to_date:
        try:
            time_series_data = generate_time_series_data(tickets, from_date, to_date)
        except ValueError:
            fallback_from = timezone.now().astimezone(tz) - timedelta(days=30)
            fallback_to = timezone.now().astimezone(tz)
            try:
                time_series_data = generate_time_series_data(tickets, fallback_from, fallback_to)
            except Exception:
                time_series_data = []
    else:
        default_from = timezone.now().astimezone(tz) - timedelta(days=30)
        default_to = timezone.now().astimezone(tz)
        try:
            time_series_data = generate_time_series_data(tickets, default_from, default_to)
        except Exception:
            time_series_data = []
    recommendations = generate_recommendations(
        tickets_in_period,
        time_analytics,
        bottleneck_analysis,
        performance_by_assignee,
        priority_counts,
        user_roles
    )
    first_response_breaches = None
    if user_roles['is_admin'] or user_roles['is_hr'] or user_roles['is_manager']:
        try:
            first_response_breaches = analyze_first_response_times(tickets_in_period)
        except ValueError:
            first_response_breaches = None
    user_satisfaction = None
    if hasattr(Support, 'satisfaction_rating') and (user_roles['is_admin'] or user_roles['is_hr']):
        try:
            tickets_with_ratings = tickets_in_period.exclude(
                Q(satisfaction_rating__isnull=True) | Q(satisfaction_rating=0)
            )
            if tickets_with_ratings.exists():
                user_satisfaction = {
                    'avg_rating': tickets_with_ratings.aggregate(avg=Avg('satisfaction_rating'))['avg'],
                    'rating_counts': tickets_with_ratings.values('satisfaction_rating').annotate(count=Count('id')),
                    'total_rated': tickets_with_ratings.count(),
                    'pct_of_total': (tickets_with_ratings.count() / total_tickets * 100) if total_tickets > 0 else 0
                }
        except ValueError:
            user_satisfaction = None
    escalation_data = None
    if user_roles['is_admin'] or user_roles['is_hr'] or user_roles['is_manager']:
        try:
            escalated_tickets = tickets_in_period.filter(escalation_level__gt=0)
            if escalated_tickets.exists():
                escalation_data = {
                    'total_escalated': escalated_tickets.count(),
                    'pct_escalated': (escalated_tickets.count() / total_tickets * 100) if total_tickets > 0 else 0,
                    'avg_escalation_level': escalated_tickets.aggregate(avg=Avg('escalation_level'))['avg'],
                    'by_priority': escalated_tickets.values('priority').annotate(count=Count('id')),
                    'by_issue_type': escalated_tickets.values('issue_type').annotate(count=Count('id')),
                }
        except ValueError:
            escalation_data = None
    context = {
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'in_progress_tickets': in_progress_tickets,
        'resolved_tickets': resolved_tickets,
        'closed_tickets': closed_tickets,
        'ticket_growth_rate': ticket_growth_rate,
        'recent_tickets': recent_tickets,
        'critical_tickets': priority_counts['Critical'],
        'high_tickets': priority_counts['High'],
        'medium_tickets': priority_counts['Medium'],
        'low_tickets': priority_counts['Low'],
        'priority_distribution': priority_distribution,
        'issue_distribution': issue_distribution,
        'time_analytics': time_analytics,
        'time_series_data': time_series_data,
        'workload_distribution': workload_distribution,
        'bottleneck_analysis': bottleneck_analysis,
        'performance_by_assignee': performance_by_assignee,
        'first_response_breaches': first_response_breaches,
        'user_satisfaction': user_satisfaction,
        'escalation_data': escalation_data,
        'recommendations': recommendations,
        'unassigned_tickets': unassigned_tickets if 'unassigned_tickets' in locals() else None,
        'assigned_tickets': assigned_tickets if 'assigned_tickets' in locals() else None,
        'critical_percentage': (priority_counts['Critical'] / total_tickets * 100) if total_tickets > 0 else 0,
        'high_percentage': (priority_counts['High'] / total_tickets * 100) if total_tickets > 0 else 0,
        'medium_percentage': (priority_counts['Medium'] / total_tickets * 100) if total_tickets > 0 else 0,
        'low_percentage': (priority_counts['Low'] / total_tickets * 100) if total_tickets > 0 else 0,
        'date_range': date_range,
        'from_date': from_date,
        'to_date': to_date - timedelta(days=1) if to_date else None,
        **user_roles,
    }
    return render(request, 'components/support/dashboard.html', context)


def generate_time_series_data(tickets_queryset, from_date, to_date):
    """Generate time series data for ticket trends"""
    from datetime import timedelta
    date_range = (to_date - from_date).days

    # Adjust grouping based on the date range
    if date_range <= 7:
        # Group by hour if range is very short
        trunc_function = TruncHour('created_at')
        date_format = '%Y-%m-%d %H:%M'
    elif date_range <= 31:
        # Group by day for up to a month
        trunc_function = TruncDay('created_at')
        date_format = '%Y-%m-%d'
    elif date_range <= 90:
        # Group by week for up to 3 months
        trunc_function = TruncWeek('created_at')
        date_format = 'Week %W, %Y'
    else:
        # Group by month for longer periods
        trunc_function = TruncMonth('created_at')
        date_format = '%b %Y'

    # Defensive: handle ValueError from invalid datetimes in DB
    try:
        # Get created ticket counts
        created_tickets = tickets_queryset.filter(
            created_at__gte=from_date,
            created_at__lt=to_date
        ).annotate(
            date=trunc_function
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
    except ValueError:
        created_tickets = []

    try:
        # Get resolved ticket counts
        resolved_tickets = tickets_queryset.filter(
            resolved_at__isnull=False,
            resolved_at__gte=from_date,
            resolved_at__lt=to_date
        ).annotate(
            date=Trunc('resolved_at', kind='day')  # Or 'month', 'year', etc. depending on your needs
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')
    except ValueError:
        resolved_tickets = []

    # Convert querysets to dictionaries for easier manipulation
    created_dict = {}
    for item in created_tickets:
        try:
            created_dict[item['date'].strftime(date_format)] = item['count']
        except Exception:
            continue
    resolved_dict = {}
    for item in resolved_tickets:
        try:
            resolved_dict[item['date'].strftime(date_format)] = item['count']
        except Exception:
            continue

    # Generate a complete date series
    current = from_date
    all_dates = []

    # Function to increment date based on grouping
    def get_increment_function(trunc_function):
        trunc_class_name = trunc_function.__class__.__name__
        if trunc_class_name == 'TruncHour':
            return lambda d: d + timedelta(hours=1)
        elif trunc_class_name == 'TruncDay':
            return lambda d: d + timedelta(days=1)
        elif trunc_class_name == 'TruncWeek':
            return lambda d: d + timedelta(weeks=1)
        elif trunc_class_name == 'TruncMonth':
            def increment_month(d):
                month = d.month + 1
                year = d.year
                if month > 12:
                    month = 1
                    year += 1
                return d.replace(year=year, month=month, day=1)
            return increment_month
        else:  # Default case
            return lambda d: d + timedelta(days=1)

    increment = get_increment_function(trunc_function)

    while current < to_date:
        try:
            date_str = current.strftime(date_format)
            all_dates.append(date_str)
            current = increment(current)
        except Exception:
            break

    # Combine the data
    result = []
    for date_str in all_dates:
        result.append({
            'date': date_str,
            'created': created_dict.get(date_str, 0),
            'resolved': resolved_dict.get(date_str, 0)
        })

    return result


def analyze_first_response_times(tickets):
    """Analyze first response times for tickets"""
    from datetime import timedelta
    # Consider tickets that have a recorded first response time
    try:
        tickets_with_response = tickets.filter(response_time__isnull=False)
    except ValueError:
        return None

    if not tickets_with_response.exists():
        return None

    # Get SLA targets for first response based on priority
    priority_sla_targets = {
        'Critical': timedelta(hours=1),
        'High': timedelta(hours=4),
        'Medium': timedelta(hours=8),
        'Low': timedelta(hours=24)
    }

    # Analyze breaches by priority
    breaches_by_priority = {}
    for priority, target in priority_sla_targets.items():
        try:
            priority_tickets = tickets_with_response.filter(priority=priority)
            if priority_tickets.exists():
                breached = priority_tickets.filter(response_time__gt=target).count()
                total = priority_tickets.count()
                breaches_by_priority[priority] = {
                    'breached': breached,
                    'total': total,
                    'percentage': (breached / total * 100) if total > 0 else 0,
                    'target_hours': target.total_seconds() / 3600
                }
        except ValueError:
            breaches_by_priority[priority] = {
                'breached': 0,
                'total': 0,
                'percentage': 0,
                'target_hours': target.total_seconds() / 3600
            }

    try:
        overall_avg = tickets_with_response.aggregate(avg=Avg('response_time'))['avg']
    except ValueError:
        overall_avg = None

    return {
        'overall_avg_response': overall_avg,
        'breaches_by_priority': breaches_by_priority
    }


def generate_recommendations(tickets, time_analytics, bottleneck_analysis, performance_by_assignee, priority_counts, user_roles):
    """Generate data-driven recommendations based on ticket analytics"""
    recommendations = []

    # Only generate recommendations for admins, HR, and managers
    if not (user_roles['is_admin'] or user_roles['is_hr'] or user_roles['is_manager']):
        return []

    # Check total ticket volume
    total_tickets = sum(priority_counts.values())

    # Check SLA compliance
    if time_analytics['sla_compliance'] is not None and time_analytics['sla_compliance'] < 80:
        recommendations.append({
            'title': 'SLA Compliance Needs Attention',
            'description': f"Current SLA compliance rate is {time_analytics['sla_compliance']:.1f}%. Consider revising your SLA targets or optimizing response workflow.",
            'severity': 'high',
            'action_items': [
                'Review SLA targets for different priorities',
                'Consider adding more resources to high-priority tickets',
                'Implement automated notifications for approaching SLA deadlines'
            ]
        })

    # Check for bottlenecks
    if bottleneck_analysis and len(bottleneck_analysis) > 0:
        total_bottleneck_tickets = sum(item['count'] for item in bottleneck_analysis)
        if total_bottleneck_tickets > 0:
            status_list = ", ".join([f"{item['status']} ({item['count']})" for item in bottleneck_analysis])
            recommendations.append({
                'title': 'Workflow Bottlenecks Detected',
                'description': f"Found {total_bottleneck_tickets} tickets stuck in the following statuses: {status_list}",
                'severity': 'medium',
                'action_items': [
                    'Review tickets that have been in their current status for too long',
                    'Consider implementing maximum time limits for each status',
                    'Analyze if additional resources or process changes are needed'
                ]
            })

    # Check high critical ticket volume
    if priority_counts['Critical'] > 0 and (priority_counts['Critical'] / total_tickets) > 0.15:
        recommendations.append({
            'title': 'High Volume of Critical Tickets',
            'description': f"Critical tickets make up {(priority_counts['Critical'] / total_tickets * 100):.1f}% of all tickets. Review priority assignment criteria.",
            'severity': 'high',
            'action_items': [
                'Audit critical ticket classifications',
                'Consider refining the criteria for critical priority',
                'Create a dedicated team or process for handling critical issues'
            ]
        })

    # Check unassigned tickets
    try:
        unassigned_count = tickets.filter(assigned_to_user__isnull=True).count()
    except ValueError:
        unassigned_count = 0
    if unassigned_count > 10 or (total_tickets > 0 and unassigned_count / total_tickets > 0.2):
        recommendations.append({
            'title': 'High Number of Unassigned Tickets',
            'description': f"{unassigned_count} tickets ({((unassigned_count / total_tickets) * 100):.1f}% of total) are currently unassigned.",
            'severity': 'medium',
            'action_items': [
                'Implement auto-assignment based on ticket type and available resources',
                'Review ticket triage process',
                'Set up alerts for tickets unassigned after a certain period'
            ]
        })

    # Check for workload imbalance
    if performance_by_assignee and len(performance_by_assignee) >= 2:
        assigned_counts = [item['resolved_count'] for item in performance_by_assignee]
        min_assigned = min(assigned_counts)
        max_assigned = max(assigned_counts)

        if max_assigned > min_assigned * 3:  # Significant imbalance
            recommendations.append({
                'title': 'Workload Imbalance Detected',
                'description': 'There is a significant imbalance in ticket assignments among team members.',
                'severity': 'medium',
                'action_items': [
                    'Review ticket assignment process',
                    'Consider load balancing techniques',
                    'Evaluate team capacity and specialization'
                ]
            })

    # Check for response time issues
    if time_analytics['avg_response_time'] and hasattr(time_analytics['avg_response_time'], 'total_seconds') and time_analytics['avg_response_time'].total_seconds() > 24*3600:  # More than 24 hours
        recommendations.append({
            'title': 'Slow Initial Response Times',
            'description': f"Average first response time is {humanize_timedelta(time_analytics['avg_response_time'])}. Consider improving initial triage.",
            'severity': 'high',
            'action_items': [
                'Implement auto-acknowledgment for new tickets',
                'Review ticket monitoring process',
                'Train team on faster initial assessment techniques'
            ]
        })

    # Check reopened tickets
    try:
        reopened_count = tickets.filter(
            Q(status='Resolved') | Q(status='Closed'),
            ticket_activity__action='REOPENED'
        ).distinct().count()
    except ValueError:
        reopened_count = 0

    if reopened_count > 0 and total_tickets > 0 and (reopened_count / total_tickets) > 0.1:
        recommendations.append({
            'title': 'High Rate of Reopened Tickets',
            'description': f"{reopened_count} tickets ({((reopened_count / total_tickets) * 100):.1f}% of total) were reopened after resolution.",
            'severity': 'medium',
            'action_items': [
                'Review resolution quality checks',
                'Implement follow-up surveys after resolution',
                'Analyze common reasons for ticket reopening'
            ]
        })

    # Check ticket categorization
    try:
        uncategorized = tickets.filter(
            Q(issue_type__isnull=True) | Q(issue_type='')
        ).count()
    except ValueError:
        uncategorized = 0

    if uncategorized > 0 and (uncategorized / total_tickets) > 0.05:
        recommendations.append({
            'title': 'Improve Ticket Categorization',
            'description': f"{uncategorized} tickets ({((uncategorized / total_tickets) * 100):.1f}% of total) are not properly categorized.",
            'severity': 'low',
            'action_items': [
                'Review and update ticket categorization guidelines',
                'Implement mandatory categorization in the ticket form',
                'Provide better category descriptions for submitters'
            ]
        })

    # Limit to top 5 recommendations
    return sorted(recommendations, key=lambda x: {'high': 0, 'medium': 1, 'low': 2}[x['severity']])[:5]


def humanize_timedelta(td):
    """Convert a timedelta into a human-readable format"""
    if not td:
        return "N/A"

    total_seconds = int(td.total_seconds())
    days, remainder = divmod(total_seconds, 86400)
    hours, remainder = divmod(remainder, 3600)
    minutes, seconds = divmod(remainder, 60)

    result = ""
    if days > 0:
        result += f"{days} day{'s' if days != 1 else ''} "
    if hours > 0 or days > 0:
        result += f"{hours} hour{'s' if hours != 1 else ''} "
    if minutes > 0 or hours > 0 or days > 0:
        result += f"{minutes} minute{'s' if minutes != 1 else ''}"
    else:
        result += f"{seconds} second{'s' if seconds != 1 else ''}"

    return result.strip()


'''---------------------------------------- HOLIDAY AREA ----------------------------------'''

from .models import Holiday
from .forms import HolidayForm

# Permission check function

# Helper function to check if user is HR
def is_hr(user):
    return user.groups.filter(name='HR').exists()

@login_required
def holiday_lists(request):
    """List holidays - all users can view, HR can perform CRUD"""
    holidays = Holiday.objects.all()
    user_is_hr = is_hr(request.user)
    today = timezone.now().date()

    # Get upcoming holidays (next 30 days)
    upcoming_holidays = []
    for i in range(30):
        check_date = today + timedelta(days=i)
        if Holiday.is_holiday(check_date):
            matches = list(Holiday.objects.filter(date=check_date)) + list(
                Holiday.objects.filter(
                    recurring_yearly=True,
                    date__month=check_date.month,
                    date__day=check_date.day
                )
            )
            for holiday in matches:
                upcoming_holidays.append({
                    'name': holiday.name,
                    'date': check_date,
                    'days_away': i,
                    'weekday': check_date.strftime('%A')
                })

    # Calendar for current month
    year = today.year
    month = today.month
    month_name = today.strftime('%B')
    _, days_in_month = monthrange(year, month)

    weekdays = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

    calendar_days = []
    for day in range(1, days_in_month + 1):
        check_date = datetime(year, month, day).date()
        calendar_days.append({
            'date': check_date,
            'is_holiday': Holiday.is_holiday(check_date),
            'is_today': check_date == today,
            'weekday': check_date.strftime('%a')
        })

    context = {
        'holidays': holidays,
        'upcoming_holidays': upcoming_holidays,
        'month_name': month_name,
        'year': year,
        'calendar_days': calendar_days,
        'is_hr': user_is_hr,
        'weekdays': weekdays
    }
    return render(request, 'components/holidays/holiday_lists.html', context)

@login_required
def holiday_dashboard(request):
    """Smart dashboard view for employees and managers to see holidays"""
    today = timezone.now().date()
    is_hr = User.groups.filter(name='HR').exists()


    # Next holiday
    next_holiday = None
    days_to_next = 0

    # Check next 60 days
    for i in range(1, 60):
        check_date = today + timedelta(days=i)
        if Holiday.is_holiday(check_date):
            # Find which holiday it is
            exact_match = Holiday.objects.filter(date=check_date).first()
            recurring_match = Holiday.objects.filter(
                recurring_yearly=True,
                date__month=check_date.month,
                date__day=check_date.day
            ).first()

            holiday = exact_match or recurring_match
            if holiday:
                next_holiday = {
                    'name': holiday.name,
                    'date': check_date,
                    'days_away': i,
                    'weekday': check_date.strftime('%A')
                }
                days_to_next = i
                break

    # Get all holidays in the current year
    current_year = today.year
    start_of_year = datetime(current_year, 1, 1).date()
    end_of_year = datetime(current_year, 12, 31).date()

    yearly_holidays = []

    # Add exact date holidays in this year
    exact_holidays = Holiday.objects.filter(
        date__gte=start_of_year,
        date__lte=end_of_year
    )

    for holiday in exact_holidays:
        yearly_holidays.append({
            'name': holiday.name,
            'date': holiday.date,
            'passed': holiday.date < today,
            'weekday': holiday.date.strftime('%A')
        })

    # Add recurring holidays for this year
    recurring_holidays = Holiday.objects.filter(recurring_yearly=True)

    for holiday in recurring_holidays:
        # Create date for this year
        try:
            holiday_date = datetime(current_year, holiday.date.month, holiday.date.day).date()

            # Check if this recurring holiday is already in the list from exact matches
            if not any(h['date'] == holiday_date for h in yearly_holidays):
                yearly_holidays.append({
                    'name': holiday.name,
                    'date': holiday_date,
                    'passed': holiday_date < today,
                    'weekday': holiday_date.strftime('%A')
                })
        except ValueError:
            # Handle Feb 29 in non-leap years
            continue

    # Sort holidays by date
    yearly_holidays.sort(key=lambda x: x['date'])

    # Get holidays by month for easy viewing
    holidays_by_month = {}
    for i in range(1, 13):
        month_name = datetime(2000, i, 1).strftime('%B')
        month_holidays = [h for h in yearly_holidays if h['date'].month == i]
        holidays_by_month[month_name] = month_holidays

    context = {
        'next_holiday': next_holiday,
        'days_to_next': days_to_next,
        'yearly_holidays': yearly_holidays,
        'holidays_by_month': holidays_by_month,
        'year': current_year,
        'now': today , # For timeline calculation
        'is_hr':is_hr
    }

    return render(request, 'components/holidays/holiday_dashboard.html', context)

# ===============================
# HR-ONLY VIEWS (CRUD)
# ===============================

@login_required
@user_passes_test(is_hr)
def holidays_create(request):
    """HR: Create new holiday"""
    if request.method == 'POST':
        form = HolidayForm(request.POST)
        if form.is_valid():
            holiday = form.save()
            messages.success(request, f"Holiday '{holiday.name}' created successfully!")
            return redirect('aps_holiday:holiday_lists')
    else:
        form = HolidayForm()
    return render(request, 'components/holidays/holiday_form.html', {'form': form})


@login_required
@user_passes_test(is_hr)
def holidays_update(request, pk):
    """HR: Update holiday"""
    holiday = get_object_or_404(Holiday, pk=pk)
    if request.method == 'POST':
        form = HolidayForm(request.POST, instance=holiday)
        if form.is_valid():
            holiday = form.save()
            messages.success(request, f"Holiday '{holiday.name}' updated successfully!")
            return redirect('aps_holiday:holiday_lists')
    else:
        form = HolidayForm(instance=holiday)
    return render(request, 'components/holidays/holiday_form.html', {'form': form})


@login_required
@user_passes_test(is_hr)
def holidays_delete(request, pk):
    """HR: Delete holiday"""
    holiday = get_object_or_404(Holiday, pk=pk)
    if request.method == 'POST':
        name = holiday.name
        holiday.delete()
        messages.success(request, f"Holiday '{name}' deleted successfully!")
        return redirect('aps_holiday:holiday_lists')
    return render(request, 'components/holidays/holiday_confirm_delete.html', {'object': holiday})





'''--------------------------------------------------------------------------------'''

@login_required
@user_passes_test(is_employee)
def employee_support(request):
    """Employee's Support Home with the ability to create a ticket."""

    if request.method == 'POST':
        issue_type = request.POST.get('issue_type')
        description = request.POST.get('description', '').strip()
        subject = request.POST.get('subject', 'No subject').strip()

        # Validate required fields
        if not issue_type or not description:
            messages.error(request, "Issue Type and Description are required.")
            return redirect('aps_employee:employee_support')

        # Assign responsible department (HR or Admin)
        assigned_to = (
            Support.AssignedTo.HR if issue_type == Support.IssueType.HR else Support.AssignedTo.ADMIN
        )

        # Assign priority dynamically
        priority_mapping = {
            Support.IssueType.HARDWARE: Support.Priority.HIGH,
            Support.IssueType.SOFTWARE: Support.Priority.MEDIUM,
            Support.IssueType.NETWORK: Support.Priority.CRITICAL,
            Support.IssueType.INTERNET: Support.Priority.CRITICAL,
            Support.IssueType.APPLICATION: Support.Priority.MEDIUM,
            Support.IssueType.HR: Support.Priority.LOW,
            Support.IssueType.ACCESS: Support.Priority.HIGH,
            Support.IssueType.SECURITY: Support.Priority.CRITICAL,
            Support.IssueType.SERVICE: Support.Priority.MEDIUM,
        }
        priority = priority_mapping.get(issue_type, Support.Priority.MEDIUM)

        # Set due date based on priority
        due_date_mapping = {
            Support.Priority.CRITICAL: now() + timedelta(hours=4),
            Support.Priority.HIGH: now() + timedelta(days=1),
            Support.Priority.MEDIUM: now() + timedelta(days=3),
            Support.Priority.LOW: now() + timedelta(days=5),
        }
        due_date = due_date_mapping.get(priority)

        try:
            with transaction.atomic():
                ticket = Support.objects.create(
                    user=request.user,
                    issue_type=issue_type,
                    description=description,
                    subject=subject,
                    status=Support.Status.NEW,
                    priority=priority,
                    assigned_to=assigned_to,
                    due_date=due_date,
                )
            messages.success(request, f"Ticket #{ticket.ticket_id} created successfully.")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            return redirect('aps_employee:employee_support')

    # Fetch tickets raised by the logged-in employee
    tickets = Support.objects.filter(user=request.user).order_by('-created_at')

    # Fetch issue type choices dynamically
    issue_type_choices = [choice[0] for choice in Support.IssueType.choices]

    return render(request, 'components/employee/support.html', {
        'tickets': tickets,
        'issue_type_choices': issue_type_choices
    })


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q
from .models import Support, StatusLog
from django.utils.timezone import now

def is_admin(user):
    return user.groups.filter(name='Admin').exists()

@login_required
@user_passes_test(is_admin)
def admin_support(request, ticket_id=None):
    try:
        if ticket_id:
            ticket = get_object_or_404(Support, ticket_id=ticket_id)

            if request.method == 'POST':
                new_status = request.POST.get('status')
                if new_status in dict(Support.Status.choices):
                    old_status = ticket.status
                    ticket.status = new_status

                    # Update resolved_at if status changed to Resolved
                    if new_status == Support.Status.RESOLVED and old_status != Support.Status.RESOLVED:
                        ticket.resolved_at = now()
                    elif new_status != Support.Status.RESOLVED:
                        ticket.resolved_at = None

                    ticket.save()

                    # Log status change
                    StatusLog.objects.create(
                        ticket=ticket,
                        old_status=old_status,
                        new_status=new_status,
                        changed_by=request.user
                    )

                    messages.success(request, f"Ticket {ticket.ticket_id} updated to {new_status}.")
                    return redirect('aps_admin:admin_support')
                else:
                    messages.error(request, "Invalid status selected.")

            return render(request, 'components/admin/support_admin.html', {
                'ticket': ticket,
                'is_admin': True
            })

        # List view with filters
        tickets = Support.objects.all().select_related('user')

        # Apply filters
        status_filter = request.GET.get('status')
        issue_type_filter = request.GET.get('issue_type')

        if status_filter:
            tickets = tickets.filter(status=status_filter)
        if issue_type_filter:
            tickets = tickets.filter(issue_type=issue_type_filter)

        context = {
            'tickets': tickets,
            'open_tickets': Support.objects.filter(status=Support.Status.OPEN).count(),
            'in_progress_tickets': Support.objects.filter(status=Support.Status.IN_PROGRESS).count(),
            'resolved_tickets': Support.objects.filter(status=Support.Status.RESOLVED).count(),
            'is_admin': True
        }
        return render(request, 'components/admin/support_admin.html', context)

    except Exception as e:
        messages.error(request, f"An error occurred while managing tickets: {str(e)}")
        return redirect('aps_admin:admin_support')

def is_hr(user):
    return user.groups.filter(name='HR').exists()

@login_required
@user_passes_test(is_hr)
def hr_support(request, ticket_id=None):
    """HR view to manage tickets and see ticket details."""
    try:
        if ticket_id:
            ticket = get_object_or_404(Support, ticket_id=ticket_id)

            if request.method == 'POST':
                new_status = request.POST.get('status')
                if new_status in dict(Support.Status.choices):
                    if ticket.assigned_to == Support.AssignedTo.HR:
                        old_status = ticket.status
                        ticket.status = new_status

                        # Update resolved_at if status changed to Resolved
                        if new_status == Support.Status.RESOLVED and old_status != Support.Status.RESOLVED:
                            ticket.resolved_at = now()
                        elif new_status != Support.Status.RESOLVED:
                            ticket.resolved_at = None

                        ticket.save()

                        # Log status change
                        StatusLog.objects.create(
                            ticket=ticket,
                            old_status=old_status,
                            new_status=new_status,
                            changed_by=request.user
                        )

                        messages.success(request, f"Ticket {ticket.ticket_id} updated to {new_status}.")
                        return redirect('aps_hr:hr_support')
                    else:
                        messages.error(request, "You can only update HR-assigned tickets.")
                else:
                    messages.error(request, "Invalid status selected.")

            return render(request, 'components/hr/support_hr.html', {
                'ticket': ticket,
                'is_hr': True,
                'can_update': ticket.assigned_to == Support.AssignedTo.HR
            })

        # List view with filters
        tickets = Support.objects.filter(
            Q(assigned_to=Support.AssignedTo.HR) |
            Q(issue_type=Support.IssueType.HR)
        ).select_related('user').order_by('-created_at')

        # Apply filters
        status_filter = request.GET.get('status')
        if status_filter:
            tickets = tickets.filter(status=status_filter)

        context = {
            'tickets': tickets,
            'open_tickets': tickets.filter(status=Support.Status.OPEN).count(),
            'in_progress_tickets': tickets.filter(status=Support.Status.IN_PROGRESS).count(),
            'resolved_tickets': tickets.filter(status=Support.Status.RESOLVED).count(),
            'is_hr': True,
            'total_tickets': tickets.count()
        }

        return render(request, 'components/hr/support_hr.html', context)

    except Exception as e:
        print(f"HR Support Error: {str(e)}")  # For debugging
        messages.error(request, f"An error occurred while managing tickets: {str(e)}")
        return redirect('aps_hr:hr_support')

'''----- Temeporray views -----'''

# Assign Tasks View
@login_required
def assign_tasks(request):
    # Placeholder context data
    context = {
        'title': 'Assign Tasks',
        'tasks': [],  # Example data (you can replace this with actual task data)
    }
    return render(request, 'components/manager/assign_tasks.html', context)

# Approve Leaves View
@login_required
def approve_leave(request):
    # Placeholder context data
    context = {
        'title': 'Approve Leaves',
        'leave_requests': [],  # Example data (you can replace this with actual leave request data)
    }
    return render(request, 'components/manager/approve_leave.html', context)

'''-------------------------- CHAT SYSTEM --------------------------------'''
# from django.shortcuts import render, redirect, get_object_or_404
# from django.contrib.auth.decorators import login_required, user_passes_test
# from django.contrib.auth.models import User, Group
# from django.contrib import messages
# from django.db.models import Q, Count, OuterRef, Subquery, F
# from django.core.exceptions import PermissionDenied
# from django.utils import timezone
# from django.http import JsonResponse
# from functools import wraps
# import re

# from .models import ChatGroup, GroupMember, DirectMessage, Message, MessageRead
# from .services import get_chat_history, mark_messages_as_read, get_unread_counts, create_group
# from .utils import validate_user_in_chat, send_notification

# @login_required
# def chat_home(request, chat_type=None, chat_id=None):
#     """Main chat view that renders the chat interface and handles all chat functionality"""
#     try:
#         # Get available users based on role
#         is_admin = request.user.groups.filter(name='Admin').exists()
#         is_manager = request.user.groups.filter(name='Manager').exists()

#         available_users = User.objects.exclude(id=request.user.id)
#         if not is_admin:
#             if is_manager:
#                 available_users = available_users.filter(groups__name='Employee')
#             else:
#                 available_users = available_users.filter(
#                     Q(groups__name='Manager') | Q(groups__name='HR')
#                 )

#         # Get chat lists for sidebar - Updated query
#         group_chats = ChatGroup.objects.filter(
#             memberships__user=request.user,
#             memberships__is_active=True,
#             is_active=True
#         ).annotate(
#             unread_count=Count(
#                 'messages',
#                 filter=Q(
#                     messages__read_receipts__user=request.user,
#                     messages__read_receipts__read_at__isnull=True,
#                     messages__is_deleted=False
#                 )
#             ),
#             latest_message=Subquery(
#                 Message.objects.filter(
#                     group=OuterRef('pk'),
#                     is_deleted=False
#                 ).order_by('-sent_at').values('content')[:1]
#             )
#         ).prefetch_related('memberships', 'messages')

#         direct_messages = DirectMessage.objects.filter(
#             participants=request.user,
#             is_active=True
#         ).annotate(
#             unread_count=Count(
#                 'messages',
#                 filter=Q(
#                     messages__read_receipts__user=request.user,
#                     messages__read_receipts__read_at__isnull=True,
#                     messages__is_deleted=False
#                 )
#             ),
#             latest_message=Subquery(
#                 Message.objects.filter(
#                     direct_message=OuterRef('pk'),
#                     is_deleted=False
#                 ).order_by('-sent_at').values('content')[:1]
#             )
#         ).prefetch_related('participants', 'messages')

#         # Add other participant info for direct messages
#         for dm in direct_messages:
#             dm.other_user = dm.participants.exclude(id=request.user.id).first()

#         context = {
#             'group_chats': group_chats,
#             'direct_messages': direct_messages,
#             'available_users': available_users,
#             'is_admin': is_admin,
#             'is_manager': is_manager,
#             'unread_counts': get_unread_counts(request.user)
#         }

#         # Handle chat detail view
#         if chat_type and chat_id:
#             try:
#                 validate_user_in_chat(request.user, chat_id)

#                 if chat_type == 'group':
#                     chat = get_object_or_404(ChatGroup, id=chat_id, is_active=True)
#                     other_participant = None
#                 else:
#                     chat = get_object_or_404(DirectMessage, id=chat_id, is_active=True)
#                     other_participant = chat.participants.exclude(id=request.user.id).first()

#                 messages_list = get_chat_history(chat_id, request.user, chat_type)

#                 # Mark messages as read and only send notification if messages were actually read
#                 unread_count = mark_messages_as_read(chat_id, request.user, chat_type)

#                 # Only send notification if there were unread messages
#                 if unread_count > 0:
#                     try:
#                         send_notification(
#                             request.user.id,
#                             "Messages marked as read",
#                             "read_status",
#                             chat_id
#                         )
#                     except Exception as notify_error:
#                         # Log notification errors without breaking the app flow
#                         import logging
#                         logger = logging.getLogger(__name__)
#                         logger.error(f"Notification error: {str(notify_error)}")

#                 context.update({
#                     'chat': chat,
#                     'chat_type': chat_type,
#                     'messages': messages_list,
#                     'other_participant': other_participant,
#                     'can_manage': request.user.groups.filter(name__in=['Admin', 'Manager']).exists(),
#                     'chat_detail_view': True
#                 })

#             except Exception as e:
#                 messages.error(request, f'Error loading chat: {str(e)}')
#                 if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
#                     return JsonResponse({'error': str(e)}, status=400)
#                 return redirect('dashboard')


#         # Handle create group chat
#         if request.method == 'POST' and request.POST.get('action') == 'create_group':
#             if not request.user.groups.filter(name__in=['Admin', 'Manager']).exists():
#                 raise PermissionDenied("You don't have permission to create groups")

#             try:
#                 name = request.POST.get('name')
#                 description = request.POST.get('description', '')
#                 member_ids = request.POST.getlist('members')

#                 chat = create_group(name, request.user, description)
#                 GroupMember.objects.bulk_create([
#                     GroupMember(group=chat, user_id=member_id, role='member', is_active=True)
#                     for member_id in member_ids
#                 ])

#                 for member_id in member_ids:
#                     send_notification(
#                         member_id,
#                         f"You've been added to group chat: {name}",
#                         "group_add",
#                         chat.id,
#                         request.user.username
#                     )

#                 messages.success(request, 'Group chat created successfully')
#                 return redirect('chat:detail', chat_type='group', chat_id=chat.id)

#             except Exception as e:
#                 messages.error(request, f'Error creating group: {str(e)}')
#                 return redirect('dashboard')

#         # Handle create direct message
#         if request.method == 'POST' and request.POST.get('action') == 'create_direct':
#             try:
#                 user_id = request.POST.get('user_id')
#                 if not user_id:
#                     raise ValueError("No user_id provided")

#                 other_user = get_object_or_404(User, id=user_id)

#                 existing_chat = DirectMessage.objects.filter(
#                     participants=request.user
#                 ).filter(
#                     participants=other_user,
#                     is_active=True
#                 ).first()

#                 if existing_chat:
#                     if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
#                         return JsonResponse({'chat_id': existing_chat.id})
#                     return redirect('chat:detail', chat_type='direct', chat_id=existing_chat.id)

#                 chat = DirectMessage.objects.create(is_active=True)
#                 chat.participants.add(request.user)
#                 chat.participants.add(other_user)
#                 chat.save()

#                 send_notification(
#                     other_user.id,
#                     f"New message from {request.user.get_full_name() or request.user.username}",
#                     "direct_message",
#                     chat.id,
#                     request.user.username
#                 )

#                 if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
#                     return JsonResponse({'chat_id': chat.id})
#                 return redirect('chat:detail', chat_type='direct', chat_id=chat.id)

#             except Exception as e:
#                 if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
#                     return JsonResponse({'error': str(e)}, status=400)
#                 messages.error(request, f'Error creating chat: {str(e)}')
#                 return redirect('dashboard')

#         # Handle message sending
#         if request.method == 'POST' and request.POST.get('message'):
#             try:
#                 content = request.POST.get('message')
#                 message_type = request.POST.get('message_type', 'text')
#                 file_attachment = request.FILES.get('file_attachment')  # Get the file attachment

#                 # Start a database transaction to ensure message and read receipts are created atomically
#                 from django.db import transaction

#                 with transaction.atomic():
#                     if chat_type == 'group':
#                         chat = get_object_or_404(ChatGroup, id=chat_id)
#                         message = Message.objects.create(
#                             group=chat,
#                             sender=request.user,
#                             content=content,
#                             message_type=message_type,
#                             file_attachment=file_attachment  # Attach the file
#                         )

#                         # Get all active members for this group
#                         participants = User.objects.filter(
#                             group_memberships__group=chat,
#                             group_memberships__is_active=True
#                         )

#                     else:
#                         chat = get_object_or_404(DirectMessage, id=chat_id)
#                         message = Message.objects.create(
#                             direct_message=chat,
#                             sender=request.user,
#                             content=content,
#                             message_type=message_type,
#                             file_attachment=file_attachment  # Attach the file
#                         )

#                         # Get all participants for this direct message
#                         participants = chat.participants.all()

#                     # Create read receipt for sender (already read)
#                     MessageRead.objects.create(
#                         message=message,
#                         user=request.user,
#                         read_at=timezone.now()
#                     )

#                     # Create read receipts for other participants (unread)
#                     other_participants = participants.exclude(id=request.user.id)
#                     read_receipts = [
#                         MessageRead(message=message, user=participant)
#                         for participant in other_participants
#                     ]

#                     if read_receipts:
#                         MessageRead.objects.bulk_create(read_receipts)

#                     # Notify other participants about the new message
#                     for participant in other_participants:
#                         try:
#                             sender_name = request.user.get_full_name() or request.user.username
#                             send_notification(
#                                 participant.id,
#                                 f"New message from {sender_name}",
#                                 "new_message",
#                                 chat_id,
#                                 sender_name
#                             )
#                         except Exception as notify_error:
#                             # Log notification errors without breaking the message flow
#                             import logging
#                             logger = logging.getLogger(__name__)
#                             logger.error(f"Notification error: {str(notify_error)}")

#             except Exception as e:
#                 if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
#                     return JsonResponse({'status': 'error', 'error': str(e)}, status=400)
#                 messages.error(request, f'Error sending message: {str(e)}')
#                 return redirect('chat:detail', chat_type=chat_type, chat_id=chat_id)

#             if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
#                 return JsonResponse({'status': 'success', 'message_id': message.id})
#             return redirect('chat:detail', chat_type=chat_type, chat_id=chat_id)

#         # Default return for GET requests when no other actions are taken
#         if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
#             return render(request, 'chat/chat_content.html', context)
#         return render(request, 'chat/chat_home.html', context)

#     except Exception as e:
#         import logging
#         logger = logging.getLogger(__name__)
#         logger.error(f"Error in chat_home view: {str(e)}")

#         if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
#             return JsonResponse({'status': 'error', 'error': str(e)}, status=400)
#         messages.error(request, f'Error: {str(e)}')
#         return redirect('dashboard')





'''-------------------------- MANUAL ATTENDACE BY HR  --------------------------------'''
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils.timezone import now
from django.contrib.auth.models import User, Group
from .models import Presence, PresenceStatus
from datetime import datetime
from django.db.models import Q
from trueAlign.context_processors import is_management


def is_hr(user):
    """Check if user is HR"""
    return user.groups.filter(name="HR").exists()

def is_management(request):
    """Check if the user belongs to the 'Management' group."""
    return {'is_management': request.user.groups.filter(name="Management").exists()} if request.user.is_authenticated else {'is_management': False}

@login_required
@user_passes_test(is_hr)
def manual_attendance(request):
    try:
        # Get date from query parameters or use current date
        date_str = request.GET.get('date')
        attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else now().date()

        # Get manageable users (Management and Backoffice users)
        manageable_users = User.objects.filter(is_active=True).filter(
            Q(groups__name="Management") | Q(groups__name="Backoffice")
        ).distinct()

        if request.method == "POST":
            attendance_date = datetime.strptime(request.POST.get('date'), '%Y-%m-%d').date()

            # Process each user's attendance data
            for user in manageable_users:
                user_id = str(user.id)
                status_key = f'attendance[{user_id}][status]'
                notes_key = f'attendance[{user_id}][notes]'

                if status_key in request.POST:
                    status = request.POST.get(status_key)
                    notes = request.POST.get(notes_key, '').strip()

                    # Update or create attendance record
                    Presence.objects.update_or_create(
                        user=user,
                        date=attendance_date,
                        defaults={
                            'status': status,
                            'notes': notes,
                            'marked_by': request.user,
                            'marked_at': now()
                        }
                    )

            messages.success(request, "Attendance marked successfully")
            return redirect('aps_hr:manual_attendance')

        # Get existing presence records for the selected date
        presences = Presence.objects.filter(date=attendance_date, user__in=manageable_users)

        # Check if user is in Management group directly
        is_management_user = request.user.groups.filter(name="Management").exists()

        # Prepare context for template rendering
        context = {
            'presences': presences,
            'employees': manageable_users,
            'statuses': PresenceStatus.choices,
            'date_filter': attendance_date,
            'today': now().date(),
            'is_hr': True,
            'is_management': is_management(request)  # Add this directly
        }

        # Debug line to verify the context
        print(f"Context contains is_management: {context}")

        return render(request, 'components/hr/markAttendace.html', context)

    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")
        return redirect('aps_hr:manual_attendance')


'''----------------------------------- APPLICATION FOR EMPLOYEE --------------------------------'''

def is_employee(user):
    """
    Check if user is an employee (not admin/management).
    Returns True if user belongs to Employee group.
    """
    return user.groups.filter(name='Employee').exists()
@login_required
@user_passes_test(is_employee)
def application_for_user(request):
    """
    Placeholder view that renders the application dashboard template.
    This is currently a dummy view that will be implemented later.
    """
    try:
        # Dummy data for now
        context = {
            'applications': {},
            'user': request.user,
            'page_title': 'Application Dashboard'
        }
        return render(request, 'components/employee/application_dashboard.html', context)

    except Exception as e:
        messages.error(request, f"Error loading application dashboard: {str(e)}")
        return redirect('dashboard')



'''------------------------------------- SHIFTS ----------------------------------'''
"""
SHIFT MASTER VIEWS
This module provides views for managing shifts, assignments and holidays.
Access is restricted to users in the 'Manager' or 'Admin' groups.
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count, Prefetch
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponseBadRequest
from django.contrib.auth.models import User
from django.db import transaction
from datetime import timedelta, datetime

from .models import ShiftMaster, ShiftAssignment, Holiday

def is_manager_or_admin(user):
    return user.groups.filter(name__in=['Manager', 'Admin']).exists() or user.is_superuser
@login_required
@user_passes_test(is_manager_or_admin)
def shift_dashboard(request):
    # Get shift statistics
    shift_stats = ShiftMaster.objects.aggregate(
        total_shifts=Count('id'),
        active_shifts=Count('id', filter=Q(is_active=True))
    )

    # Get upcoming holidays
    today = timezone.now().date()
    upcoming_holidays = Holiday.objects.filter(
        date__gte=today
    ).order_by('date')[:5]

    # Get current user's shift
    user_shift = ShiftAssignment.get_user_current_shift(request.user)

    # Get recent shift assignments with user and shift details
    recent_assignments = ShiftAssignment.objects.filter(
        is_current=True
    ).select_related(
        'user',
        'shift'
    ).order_by(
        '-created_at'
    )[:10]

    context = {
        'total_shifts': shift_stats['total_shifts'],
        'active_shifts': shift_stats['active_shifts'],
        'total_holidays': Holiday.objects.count(),
        'upcoming_holidays': upcoming_holidays,
        'user_shift': user_shift,
        'recent_assignments': recent_assignments,
    }

    return render(request, 'components/manager/shifts/dashboard.html', context)

@login_required
@user_passes_test(is_manager_or_admin)
def shift_list(request):
    try:
        is_active = request.GET.get('is_active')
        search_query = request.GET.get('search', '')
        shifts = ShiftMaster.objects.all()
        if is_active in ['true', 'false']:
            shifts = shifts.filter(is_active=(is_active == 'true'))
        if search_query:
            shifts = shifts.filter(Q(name__icontains=search_query))
        shifts = shifts.prefetch_related(
            Prefetch(
                'shiftassignment_set',
                queryset=ShiftAssignment.objects.filter(is_current=True),
                to_attr='current_assignments'
            )
        ).order_by('-created_at')
        paginator = Paginator(shifts, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Get shift detail if requested
        shift_detail = None
        shift_pk = request.GET.get('shift_detail')
        assignments = None
        total_assigned = None
        if shift_pk:
            try:
                shift_detail = get_object_or_404(ShiftMaster, pk=shift_pk)
                assignments = ShiftAssignment.objects.filter(
                    shift=shift_detail,
                    is_current=True
                ).select_related('user')
                total_assigned = assignments.count()
            except ShiftMaster.DoesNotExist:
                messages.error(request, "Requested shift does not exist.")
                return redirect('aps_manager:shift_list')

        context = {
            'page_obj': page_obj,
            'search_query': search_query,
            'is_active': is_active,
            'weekdays': ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'],
            'shift_detail': shift_detail,
            'assignments': assignments,
            'total_assigned': total_assigned,
        }

        # If editing a shift, add it to context
        shift_edit_pk = request.GET.get('shift_edit')
        if shift_edit_pk:
            try:
                shift = get_object_or_404(ShiftMaster, pk=shift_edit_pk)
                context['shift'] = shift
                context['selected_days'] = shift.custom_work_days.split(',') if shift.custom_work_days else []
            except ShiftMaster.DoesNotExist:
                messages.error(request, "Shift to edit does not exist.")
                return redirect('aps_manager:shift_list')

        # If deleting a shift, add delete context
        shift_delete_pk = request.GET.get('shift_delete')
        if shift_delete_pk:
            try:
                shift = get_object_or_404(ShiftMaster, pk=shift_delete_pk)
                active_assignments = ShiftAssignment.objects.filter(shift=shift, is_current=True)
                has_active_assignments = active_assignments.exists()
                all_assignments = ShiftAssignment.objects.filter(shift=shift)
                has_any_assignments = all_assignments.exists()

                context['shift'] = shift
                context['has_active_assignments'] = has_active_assignments
                context['has_any_assignments'] = has_any_assignments
                context['assignment_count'] = all_assignments.count()
            except ShiftMaster.DoesNotExist:
                messages.error(request, "Shift to delete does not exist.")
                return redirect('aps_manager:shift_list')

        return render(request, 'components/manager/shifts/shift_list.html', context)

    except Exception as e:
        messages.error(request, f"An error occurred while loading shifts: {str(e)}")
        return redirect('aps_manager:shift_list')

@login_required
@user_passes_test(is_manager_or_admin)
def shift_detail(request, pk):
    try:
        # Redirect to shift_list with shift_detail parameter
        return redirect(f'aps_manager:shift_list')
    except Exception as e:
        messages.error(request, f"Error accessing shift detail: {str(e)}")
        return redirect('aps_manager:shift_list')

@login_required
@user_passes_test(is_manager_or_admin)
def shift_create(request):
    try:
        if request.method == 'POST':
            name = request.POST.get('name')
            start_time = request.POST.get('start_time')
            end_time = request.POST.get('end_time')
            is_active = request.POST.get('is_active') == 'on'
            work_days = request.POST.get('work_days')

            if not name or not start_time or not end_time or not work_days:
                messages.error(request, "All fields are required.")
                return redirect('aps_manager:shift_list')

            # Handle custom work days
            custom_work_days = None
            if work_days == 'Custom':
                custom_days = request.POST.getlist('custom_days')
                if not custom_days:
                    messages.error(request, "Please select at least one day for custom schedule.")
                    return redirect('aps_manager:shift_list')
                custom_work_days = ','.join(custom_days)

            try:
                shift = ShiftMaster.objects.create(
                    name=name,
                    start_time=start_time,
                    end_time=end_time,
                    is_active=is_active,
                    work_days=work_days,
                    custom_work_days=custom_work_days
                )
                messages.success(request, f"Shift '{shift.name}' created successfully!")
                return redirect('aps_manager:shift_list')
            except Exception as e:
                messages.error(request, f"Error creating shift: {str(e)}")
                return redirect('aps_manager:shift_list')

        # For GET requests, redirect to shift_list with shift_create parameter
        return redirect('aps_manager:shift_list')
    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        return redirect('aps_manager:shift_list')

@login_required
@user_passes_test(is_manager_or_admin)
def shift_update(request, pk):
    try:
        shift = get_object_or_404(ShiftMaster, pk=pk)
        if request.method == 'POST':
            try:
                name = request.POST.get('name')
                start_time = request.POST.get('start_time')
                end_time = request.POST.get('end_time')
                is_active = request.POST.get('is_active') == 'on'
                work_days = request.POST.get('work_days')

                if not name or not start_time or not end_time or not work_days:
                    messages.error(request, "All fields are required.")
                    return redirect('aps_manager:shift_list')

                # Handle custom work days
                custom_work_days = None
                if work_days == 'Custom':
                    custom_days = request.POST.getlist('custom_days')
                    if not custom_days:
                        messages.error(request, "Please select at least one day for custom schedule.")
                        return redirect('aps_manager:shift_list')
                    custom_work_days = ','.join(custom_days)

                shift.name = name
                shift.start_time = start_time
                shift.end_time = end_time
                shift.is_active = is_active
                shift.work_days = work_days
                shift.custom_work_days = custom_work_days
                shift.save()
                messages.success(request, f"Shift '{shift.name}' updated successfully!")
                return redirect('aps_manager:shift_list')
            except Exception as e:
                messages.error(request, f"Error updating shift: {str(e)}")
                return redirect('aps_manager:shift_list')
        else:
            # Redirect to shift_list with shift_edit parameter
            return redirect(f'aps_manager:shift_list')
    except ShiftMaster.DoesNotExist:
        messages.error(request, "Shift not found.")
        return redirect('aps_manager:shift_list')
    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        return redirect('aps_manager:shift_list')

@login_required
@user_passes_test(is_manager_or_admin)
def shift_delete(request, pk):
    try:
        shift = get_object_or_404(ShiftMaster, pk=pk)

        if request.method == 'POST':
            try:
                # Check for active assignments
                active_assignments = ShiftAssignment.objects.filter(shift=shift, is_current=True)
                has_active_assignments = active_assignments.exists()
                all_assignments = ShiftAssignment.objects.filter(shift=shift)

                # Check if force delete option was selected
                force_delete = request.POST.get('force_delete') == 'true'

                if has_active_assignments and not force_delete:
                    # Only show warning if not forcing deletion
                    messages.warning(
                        request,
                        f"Shift '{shift.name}' has active assignments. Use the force delete option to delete anyway."
                    )
                    return redirect('aps_manager:shift_list')

                with transaction.atomic():
                    shift_name = shift.name
                    # Delete all assignments related to this shift
                    all_assignments.delete()
                    shift.delete()
                    messages.success(request, f"Shift '{shift_name}' deleted successfully!")

                return redirect('aps_manager:shift_list')
            except Exception as e:
                messages.error(request, f"Error during shift deletion: {str(e)}")
                return redirect('aps_manager:shift_list')
        else:
            # Redirect to shift_list with shift_delete parameter
            return redirect(f'aps_manager:shift_list')
    except ShiftMaster.DoesNotExist:
        messages.error(request, "Shift not found.")
        return redirect('aps_manager:shift_list')
    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        return redirect('aps_manager:shift_list')

@login_required
@user_passes_test(is_manager_or_admin)
def holiday_list(request):
    year = request.GET.get('year', timezone.now().year)
    try:
        year = int(year)
    except (ValueError, TypeError):
        year = timezone.now().year
    recurring_only = request.GET.get('recurring_only') == 'true'

    # Fix: Initialize holidays query properly
    if recurring_only:
        holidays = Holiday.objects.filter(recurring_yearly=True)
    else:
        start_date = datetime(year, 1, 1).date()
        end_date = datetime(year, 12, 31).date()
        holidays = Holiday.objects.filter(
            Q(recurring_yearly=True) |
            Q(date__range=(start_date, end_date))
        )

    # Use database-agnostic way to extract month and day
    holidays = holidays.order_by('date__month', 'date__day')

    current_year = timezone.now().year
    year_range = range(current_year - 2, current_year + 3)
    paginator = Paginator(holidays, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'page_obj': page_obj,
        'year': year,
        'year_range': year_range,
        'recurring_only': recurring_only,
    }
    return render(request, 'components/manager/shifts/holiday_list.html', context)

@login_required
@user_passes_test(is_manager_or_admin)
def holiday_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        date = request.POST.get('date')
        recurring_yearly = request.POST.get('recurring_yearly') == 'on'
        if not name or not date:
            messages.error(request, "Name and date are required.")
            return redirect('aps_manager:holiday_list')
        try:
            # Validate date format before creating
            date_obj = datetime.strptime(date, '%Y-%m-%d').date()
            holiday = Holiday.objects.create(
                name=name,
                date=date_obj,
                recurring_yearly=recurring_yearly
            )
            messages.success(request, f"Holiday '{holiday.name}' created successfully!")
            return redirect('aps_manager:holiday_list')
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
            return redirect('aps_manager:holiday_list')
        except Exception as e:
            messages.error(request, f"Error creating holiday: {str(e)}")
            return redirect('aps_manager:holiday_list')
    else:
        return redirect('aps_manager:holiday_list')

@login_required
@user_passes_test(is_manager_or_admin)
def holiday_update(request, pk):
    holiday = get_object_or_404(Holiday, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name')
        date = request.POST.get('date')
        recurring_yearly = request.POST.get('recurring_yearly') == 'on'
        if not name or not date:
            messages.error(request, "Name and date are required.")
            return redirect('aps_manager:holiday_list')
        try:
            # Validate date format before updating
            date_obj = datetime.strptime(date, '%Y-%m-%d').date()
            holiday.name = name
            holiday.date = date_obj
            holiday.recurring_yearly = recurring_yearly
            holiday.save()
            messages.success(request, f"Holiday '{holiday.name}' updated successfully!")
            return redirect('aps_manager:holiday_list')
        except ValueError:
            messages.error(request, "Invalid date format. Please use YYYY-MM-DD.")
            return redirect('aps_manager:holiday_list')
        except Exception as e:
            messages.error(request, f"Error updating holiday: {str(e)}")
            return redirect('aps_manager:holiday_list')
    else:
        return redirect('aps_manager:holiday_list')

@login_required
@user_passes_test(is_manager_or_admin)
def holiday_delete(request, pk):
    holiday = get_object_or_404(Holiday, pk=pk)
    if request.method == 'POST':
        with transaction.atomic():
            holiday_name = holiday.name
            holiday.delete()
            messages.success(request, f"Holiday '{holiday_name}' deleted successfully!")
        return redirect('aps_manager:holiday_list')
    else:
        return redirect('aps_manager:holiday_list')

@login_required
@user_passes_test(is_manager_or_admin)
def assignment_list(request):
    # Base queryset with select_related for better performance
    assignments = ShiftAssignment.objects.all().select_related('user', 'shift')

    # Get filter parameters from request
    user_id = request.GET.get('user_id', '')
    shift_id = request.GET.get('shift_id', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    status = request.GET.get('status', '')
    current_only = request.GET.get('current_only', False)
    overdue = request.GET.get('overdue', False)
    unassigned = request.GET.get('unassigned', False)

    # Apply filters
    if user_id and user_id.isdigit():
        assignments = assignments.filter(user_id=int(user_id))

    if shift_id and shift_id.isdigit():
        assignments = assignments.filter(shift_id=int(shift_id))


    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            assignments = assignments.filter(effective_from__gte=start_date_obj)
        except ValueError:
            pass

    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            assignments = assignments.filter(effective_to__lte=end_date_obj)
        except ValueError:
            pass

    if status:
        today = datetime.now().date()
        if status == 'active':
            # Active: current date is between effective_from and effective_to, and has shift assigned
            assignments = assignments.filter(
                effective_from__lte=today,
                effective_to__gte=today,
                shift__isnull=False
            )
        elif status == 'pending':
            # Pending: effective_from is in the future, or shift not yet assigned
            assignments = assignments.filter(
                Q(effective_from__gt=today) | Q(shift__isnull=True)
            )
        elif status == 'completed':
            # Completed: effective_to date has passed
            assignments = assignments.filter(
                effective_to__lt=today,
                shift__isnull=False
            )

    if current_only == 'true':
        today = datetime.now().date()
        assignments = assignments.filter(
            Q(effective_from__lte=today) & (Q(effective_to__gte=today) | Q(effective_to__isnull=True))
        )

    if overdue == 'true':
        today = datetime.now().date()
        assignments = assignments.filter(effective_to__lt=today, status__in=['active', 'pending'])

    if unassigned == 'true':
        assignments = assignments.filter(user__isnull=True)

    # Order the results
    assignments = assignments.order_by('-created_at')

    # Pagination
    paginator = Paginator(assignments, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Context data for the template
    context = {
        'page_obj': page_obj,
        'users': User.objects.filter(is_active=True).order_by('username'),
        'shifts': ShiftMaster.objects.filter(is_active=True).order_by('name'),

        # Filter states for form repopulation
        'current_only': current_only == 'true',
        'overdue': overdue == 'true',
        'unassigned': unassigned == 'true',
        'selected_user': int(user_id) if user_id and user_id.isdigit() else None,
        'selected_shift': int(shift_id) if shift_id and shift_id.isdigit() else None,
        'start_date': datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None,
        'end_date': datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None,
        'selected_status': status,

        # Flag to indicate if any filter has been applied
        'filter_applied': any([user_id, shift_id, start_date, end_date,
                              status, current_only, overdue, unassigned]),
    }

    return render(request, 'components/manager/shifts/assignment_list.html', context)

@login_required
@user_passes_test(is_manager_or_admin)
def assignment_create(request):
    if request.method != 'POST':
        return HttpResponseBadRequest("Only POST requests are allowed")
    user_id = request.POST.get('user')
    shift_id = request.POST.get('shift')
    effective_from = request.POST.get('effective_from')
    effective_to = request.POST.get('effective_to')
    is_current = request.POST.get('is_current') == 'on'
    # Validate required fields
    if not user_id or not shift_id or not effective_from:
        messages.error(request, "User, shift, and effective from date are required.")
        return redirect('aps_manager:assignment_list')
    # Validate date format
    try:
        effective_from_date = datetime.strptime(effective_from, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        messages.error(request, "Effective from date must be in YYYY-MM-DD format.")
        return redirect('aps_manager:assignment_list')
    effective_to_date = None
    if effective_to:
        try:
            effective_to_date = datetime.strptime(effective_to, "%Y-%m-%d").date()
            # Validate date range
            if effective_to_date < effective_from_date:
                messages.error(request, "Effective to date cannot be earlier than effective from date.")
                return redirect('aps_manager:assignment_list')
        except (ValueError, TypeError):
            messages.error(request, "Effective to date must be in YYYY-MM-DD format.")
            return redirect('aps_manager:assignment_list')
    try:
        user = User.objects.get(pk=user_id)
        shift = ShiftMaster.objects.get(pk=shift_id)

        # Check if shift is active
        if not shift.is_active:
            messages.error(request, f"Cannot assign inactive shift '{shift.name}'.")
            return redirect('aps_manager:assignment_list')

        assignment = ShiftAssignment.objects.create(
            user=user,
            shift=shift,
            effective_from=effective_from_date,
            effective_to=effective_to_date,
            is_current=is_current
        )
        messages.success(request, f"Shift assignment for {assignment.user.username} created successfully!")
    except User.DoesNotExist:
        messages.error(request, "Selected user does not exist.")
    except ShiftMaster.DoesNotExist:
        messages.error(request, "Selected shift does not exist.")
    except Exception as e:
        messages.error(request, f"Failed to create shift assignment: {str(e)}")
    return redirect('aps_manager:assignment_list')

@login_required
@user_passes_test(is_manager_or_admin)
def bulk_assignment(request):
    if request.method != 'POST':
        return HttpResponseBadRequest("Only POST requests are allowed")
    user_ids = request.POST.getlist('user_ids')
    if not user_ids:
        messages.error(request, "No users selected for bulk assignment")
        return redirect('aps_manager:assignment_list')
    shift_id = request.POST.get('shift_id')
    effective_from = request.POST.get('effective_from')
    effective_to = request.POST.get('effective_to')
    is_current = request.POST.get('is_current') == 'on'
    # Validate required fields
    if not shift_id or not effective_from:
        messages.error(request, "Shift and effective from date are required for adding users")
        return redirect('aps_manager:assignment_list')
    # Validate date format
    try:
        effective_from_date = datetime.strptime(effective_from, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        messages.error(request, "Effective from date must be in YYYY-MM-DD format.")
        return redirect('aps_manager:assignment_list')
    effective_to_date = None
    if effective_to:
        try:
            effective_to_date = datetime.strptime(effective_to, "%Y-%m-%d").date()
            # Validate date range
            if effective_to_date < effective_from_date:
                messages.error(request, "Effective to date cannot be earlier than effective from date.")
                return redirect('aps_manager:assignment_list')
        except (ValueError, TypeError):
            messages.error(request, "Effective to date must be in YYYY-MM-DD format.")
            return redirect('aps_manager:assignment_list')
    try:
        shift = ShiftMaster.objects.get(pk=shift_id)
        # Check if shift is active
        if not shift.is_active:
            messages.error(request, f"Cannot assign inactive shift '{shift.name}'.")
            return redirect('aps_manager:assignment_list')
    except ShiftMaster.DoesNotExist:
        messages.error(request, "Selected shift does not exist")
        return redirect('aps_manager:assignment_list')
    created_count = 0
    errors = []
    with transaction.atomic():
        for user_id in user_ids:
            try:
                user = User.objects.get(pk=user_id)
                exists = ShiftAssignment.objects.filter(
                    user=user,
                    shift=shift,
                    effective_from=effective_from_date,
                    effective_to=effective_to_date,
                ).exists()
                if not exists:
                    ShiftAssignment.objects.create(
                        user=user,
                        shift=shift,
                        effective_from=effective_from_date,
                        effective_to=effective_to_date,
                        is_current=is_current,
                    )
                    created_count += 1
            except User.DoesNotExist:
                errors.append(f"User ID {user_id}: User does not exist")
            except ValueError as ve:
                errors.append(f"User ID {user_id}: Invalid date format. {str(ve)}")
            except Exception as e:
                errors.append(f"User ID {user_id}: {str(e)}")
    if created_count:
        messages.success(request, f"Shift assigned to {created_count} employee(s) successfully!")
    if errors:
        messages.error(request, "Some assignments failed: " + "; ".join(errors))
    return redirect('aps_manager:assignment_list')

@login_required
@user_passes_test(is_manager_or_admin)
def assignment_update(request, pk):
    if request.method != 'POST':
        return HttpResponseBadRequest("Only POST requests are allowed")
    assignment = get_object_or_404(ShiftAssignment, pk=pk)
    user_id = request.POST.get('user')
    shift_id = request.POST.get('shift')
    effective_from = request.POST.get('effective_from')
    effective_to = request.POST.get('effective_to')
    is_current = request.POST.get('is_current') == 'on'
    # Validate required fields
    if not user_id or not shift_id or not effective_from:
        messages.error(request, "User, shift, and effective from date are required.")
        return redirect('aps_manager:assignment_list')
    # Validate date format
    try:
        effective_from_date = datetime.strptime(effective_from, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        messages.error(request, "Effective from date must be in YYYY-MM-DD format.")
        return redirect('aps_manager:assignment_list')
    effective_to_date = None
    if effective_to:
        try:
            effective_to_date = datetime.strptime(effective_to, "%Y-%m-%d").date()
            # Validate date range
            if effective_to_date < effective_from_date:
                messages.error(request, "Effective to date cannot be earlier than effective from date.")
                return redirect('aps_manager:assignment_list')
        except (ValueError, TypeError):
            messages.error(request, "Effective to date must be in YYYY-MM-DD format.")
            return redirect('aps_manager:assignment_list')
    try:
        user = User.objects.get(pk=user_id)
        shift = ShiftMaster.objects.get(pk=shift_id)

        # Check if shift is active
        if not shift.is_active:
            messages.error(request, f"Cannot assign inactive shift '{shift.name}'.")
            return redirect('aps_manager:assignment_list')

        assignment.user = user
        assignment.shift = shift
        assignment.effective_from = effective_from_date
        assignment.effective_to = effective_to_date
        assignment.is_current = is_current
        assignment.save()
        messages.success(request, f"Shift assignment for {assignment.user.username} updated successfully!")
    except User.DoesNotExist:
        messages.error(request, "Selected user does not exist.")
    except ShiftMaster.DoesNotExist:
        messages.error(request, "Selected shift does not exist.")
    except Exception as e:
        messages.error(request, f"Failed to update shift assignment: {str(e)}")
    return redirect('aps_manager:assignment_list')

@login_required
@user_passes_test(is_manager_or_admin)
def assignment_delete(request, pk):
    if request.method != 'POST':
        return HttpResponseBadRequest("Only POST requests are allowed")
    assignment = get_object_or_404(ShiftAssignment, pk=pk)
    try:
        with transaction.atomic():
            user_name = assignment.user.username
            assignment.delete()
            messages.success(request, f"Shift assignment for {user_name} deleted successfully!")
    except Exception as e:
        messages.error(request, f"Failed to delete shift assignment: {str(e)}")
    return redirect('aps_manager:assignment_list')

@login_required
@user_passes_test(is_manager_or_admin)
def user_shift_info(request, user_id=None):
    try:
        if user_id:
            user = get_object_or_404(User, pk=user_id)
        else:
            user = request.user
        current_date = timezone.now().date()
        shift = ShiftAssignment.get_user_current_shift(user, current_date)
        if not shift:
            context = {
                'user': user,
                'shift': None,
                'error': "No active shift assignment found for this user"
            }
            return render(request, 'components/manager/shifts/user_shift_info.html', context)
        is_holiday = Holiday.is_holiday(current_date)
        is_working_day = shift.is_working_day(current_date) and not is_holiday

        # Get holidays for the next 7 days
        date_range = [current_date + timedelta(days=i) for i in range(7)]
        holidays = Holiday.objects.filter(date__in=date_range)
        recurring_holidays = Holiday.objects.filter(recurring_yearly=True)

        schedule = []
        for i in range(7):
            date = current_date + timedelta(days=i)

            # Check if date is a holiday (exact match or recurring)
            is_day_holiday = False
            holiday_name = None

            # Check exact date match
            exact_holiday = holidays.filter(date=date).first()
            if exact_holiday:
                is_day_holiday = True
                holiday_name = exact_holiday.name
            else:
                # Check recurring holidays
                for h in recurring_holidays:
                    if h.date.day == date.day and h.date.month == date.month:
                        is_day_holiday = True
                        holiday_name = h.name
                        break

            is_day_working = shift.is_working_day(date) and not is_day_holiday
            schedule.append({
                'date': date.strftime('%Y-%m-%d'),
                'day_name': date.strftime('%A'),
                'is_working_day': is_day_working,
                'is_holiday': is_day_holiday,
                'holiday_name': holiday_name,
                'shift_start': shift.start_time.strftime('%H:%M') if is_day_working else 'Off',
                'shift_end': shift.end_time.strftime('%H:%M') if is_day_working else 'Off',
            })
        context = {
            'user': user,
            'shift': shift,
            'is_holiday': is_holiday,
            'is_working_day': is_working_day,
            'schedule': schedule,
        }
        return render(request, 'components/manager/shifts/user_shift_info.html', context)
    except Exception as e:
        context = {
            'user': user if 'user' in locals() else request.user,
            'error': f"Error retrieving shift information: {str(e)}"
        }
        return render(request, 'components/manager/shifts/user_shift_info.html', context)

@login_required
@user_passes_test(is_manager_or_admin)
def shift_calendar(request):
    today = timezone.now().date()
    try:
        start_date_str = request.GET.get('start_date')
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = today.replace(day=1)
    except ValueError:
        start_date = today.replace(day=1)
    try:
        end_date_str = request.GET.get('end_date')
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            if start_date.month == 12:
                end_date = start_date.replace(year=start_date.year+1, month=1, day=1) - timedelta(days=1)
            else:
                end_date = start_date.replace(month=start_date.month+1, day=1) - timedelta(days=1)
    except ValueError:
        if start_date.month == 12:
            end_date = start_date.replace(year=start_date.year+1, month=1, day=1) - timedelta(days=1)
        else:
            end_date = start_date.replace(month=start_date.month+1, day=1) - timedelta(days=1)

    # Get all holidays that might apply to the date range
    date_specific_holidays = Holiday.objects.filter(
        date__range=(start_date, end_date),
        recurring_yearly=False
    )
    recurring_holidays = Holiday.objects.filter(recurring_yearly=True)

    calendar_data = []
    current_date = start_date
    while current_date <= end_date:
        is_holiday = False
        holiday_name = None

        # Check for date-specific holiday
        date_holiday = date_specific_holidays.filter(date=current_date).first()
        if date_holiday:
            is_holiday = True
            holiday_name = date_holiday.name
        else:
            # Check for recurring holiday
            for h in recurring_holidays:
                if h.date.day == current_date.day and h.date.month == current_date.month:
                    is_holiday = True
                    holiday_name = h.name
                    break

        shift_assignments = []
        if request.GET.get('show_shifts') == 'true':
            active_assignments = ShiftAssignment.objects.filter(
                is_current=True,
                effective_from__lte=current_date,
            ).filter(
                Q(effective_to__isnull=True) | Q(effective_to__gte=current_date)
            ).select_related('user', 'shift')[:5]
            for assignment in active_assignments:
                if assignment.shift.is_working_day(current_date) and not is_holiday:
                    shift_assignments.append({
                        'user': assignment.user.get_full_name() or assignment.user.username,
                        'shift_name': assignment.shift.name,
                        'start_time': assignment.shift.start_time.strftime('%H:%M'),
                        'end_time': assignment.shift.end_time.strftime('%H:%M'),
                    })
        calendar_data.append({
            'date': current_date,
            'day_name': current_date.strftime('%a'),
            'is_holiday': is_holiday,
            'holiday_name': holiday_name,
            'is_weekend': current_date.weekday() >= 5,
            'shift_assignments': shift_assignments,
            'is_today': current_date == today,
        })
        current_date += timedelta(days=1)
    if start_date.day == 1:
        prev_month = (start_date - timedelta(days=1)).replace(day=1)
        if start_date.month == 12:
            next_month = start_date.replace(year=start_date.year+1, month=1, day=1)
        else:
            next_month = start_date.replace(month=start_date.month+1, day=1)
    else:
        days_displayed = (end_date - start_date).days + 1
        prev_month = start_date - timedelta(days=days_displayed)
        next_month = end_date + timedelta(days=1)
    context = {
        'calendar_data': calendar_data,
        'start_date': start_date,
        'end_date': end_date,
        'prev_month': prev_month,
        'next_month': next_month,
        'show_shifts': request.GET.get('show_shifts') == 'true',
        'today': today,
    }
    return render(request, 'components/manager/shifts/shift_calendar.html', context)

@login_required
@user_passes_test(is_manager_or_admin)
def api_get_shift_users(request, shift_id):
    shift = get_object_or_404(ShiftMaster, pk=shift_id)
    assignments = ShiftAssignment.objects.filter(
        shift=shift,
        is_current=True
    ).select_related('user')
    users = [{
        'id': a.user.id,
        'username': a.user.username,
        'full_name': a.user.get_full_name(),
        'email': a.user.email,
        'assignment_id': a.id,
        'effective_from': a.effective_from.strftime('%Y-%m-%d'),
        'effective_to': a.effective_to.strftime('%Y-%m-%d') if a.effective_to else None,
    } for a in assignments]
    return JsonResponse({'users': users})

@login_required
@user_passes_test(is_manager_or_admin)
def api_get_user_shift(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    current_date = timezone.now().date()
    shift = ShiftAssignment.get_user_current_shift(user, current_date)
    if not shift:
        return JsonResponse({'success': False, 'error': 'No active shift found for this user'})
    is_holiday = Holiday.is_holiday(current_date)
    is_working_day = shift.is_working_day(current_date) and not is_holiday
    shift_data = {
        'id': shift.id,
        'name': shift.name,
        'description': getattr(shift, 'description', ''),
        'start_time': shift.start_time.strftime('%H:%M'),
        'end_time': shift.end_time.strftime('%H:%M'),
        'is_active': shift.is_active,
        'working_days': shift.working_days,
        'is_working_today': is_working_day,
        'is_holiday_today': is_holiday,
    }
    return JsonResponse({'success': True, 'shift': shift_data})

@login_required
@user_passes_test(is_manager_or_admin)
def api_toggle_shift_active(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST requests are allowed'})
    shift = get_object_or_404(ShiftMaster, pk=pk)
    try:
        with transaction.atomic():
            shift.is_active = not shift.is_active
            shift.save()
        return JsonResponse({'success': True, 'is_active': shift.is_active})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


'''---------------------------------- APPRAISAL --------------------------------'''
from .models import Appraisal, AppraisalItem, AppraisalAttachment, AppraisalWorkflow


def is_manager_or_admin(user):
    return user.groups.filter(name="Manager").exists() or user.is_superuser

def is_hr_or_admin(user):
    return user.groups.filter(name="HR").exists() or user.is_superuser

def is_finance_or_admin(user):
    return user.groups.filter(name="Finance").exists() or user.is_superuser

def is_management_or_admin(user):
    return user.groups.filter(name="Management").exists() or user.is_superuser

@login_required
def appraisal_list(request):
    """View for listing appraisals based on user role"""
    print(f"[DEBUG] Accessing appraisal_list view for user: {request.user}")
    context = {}
    user = request.user

    # Check if user is in any of the special groups
    special_groups = ["Manager", "HR", "Finance", "Management"]
    is_special_user = user.groups.filter(name__in=special_groups).exists()
    context['is_special_user'] = is_special_user
    print(f"[DEBUG] User {user} is_special_user: {is_special_user}")

    if user.groups.filter(name="Manager").exists():
        print("[DEBUG] User is in Manager group")
        # Managers see appraisals where they are assigned as manager and status is submitted
        appraisals = Appraisal.objects.filter(
            manager=user,
            status='submitted'  # Only show submitted appraisals to manager
        )
        pending_reviews = appraisals.count()
        print(f"[DEBUG] Manager pending reviews: {pending_reviews}")
        context['pending_reviews'] = pending_reviews

    elif user.groups.filter(name="HR").exists():
        print("[DEBUG] User is in HR group")
        # HR sees appraisals in HR review status
        appraisals = Appraisal.objects.filter(status='hr_review')
        pending_reviews = appraisals.count()
        print(f"[DEBUG] HR pending reviews: {pending_reviews}")
        context['pending_reviews'] = pending_reviews

    elif user.groups.filter(name="Finance").exists():
        print("[DEBUG] User is in Finance group")
        # Finance sees appraisals in finance review status
        appraisals = Appraisal.objects.filter(status='finance_review')
        pending_reviews = appraisals.count()
        print(f"[DEBUG] Finance pending reviews: {pending_reviews}")
        context['pending_reviews'] = pending_reviews

    else:
        print("[DEBUG] User is regular employee")
        # Regular employees see their own appraisals
        appraisals = Appraisal.objects.filter(user=user)

    # Add special group flags to context
    group_flags = {
        'is_manager': user.groups.filter(name="Manager").exists(),
        'is_hr': user.groups.filter(name="HR").exists(),
        'is_finance': user.groups.filter(name="Finance").exists(),
        'is_management': user.groups.filter(name="Management").exists()
    }
    print(f"[DEBUG] User group flags: {group_flags}")
    context.update(group_flags)

    # Add stats for management view
    if group_flags['is_management']:
        all_appraisals = Appraisal.objects.all()
        context.update({
            'total_appraisals': all_appraisals.count(),
            'pending_reviews': all_appraisals.filter(status='submitted').count(),
            'approved_appraisals': all_appraisals.filter(status='approved').count(),
            'rejected_appraisals': all_appraisals.filter(status='rejected').count()
        })

    context['appraisals'] = appraisals
    print(f"[DEBUG] Total appraisals in context: {appraisals.count()}")
    return render(request, 'components/appraisal/appraisal_list.html', context)

@login_required
def appraisal_detail(request, pk):
    """View for displaying appraisal details"""
    print(f"[DEBUG] Accessing appraisal_detail view for pk: {pk}")
    appraisal = get_object_or_404(Appraisal, pk=pk)
    user = request.user

    print(f"[DEBUG] User {user} accessing appraisal for user {appraisal.user}")

    # Check permissions based on role and status
    if not (user == appraisal.user or
            user == appraisal.manager or
            user.groups.filter(name__in=['HR', 'Finance', 'Management']).exists()):
        print(f"[DEBUG] Permission denied for user {user}")
        raise PermissionDenied

    context = {
        'appraisal': appraisal,
        'items': appraisal.items.all(),
        'attachments': appraisal.attachments.all(),
        'workflow_history': appraisal.workflow_history.all()
    }

    # Add ability to change status from draft to submit
    if appraisal.status == 'draft' and user == appraisal.user:
        context['can_submit'] = True

    print(f"[DEBUG] Appraisal details - Items: {context['items'].count()}, Attachments: {context['attachments'].count()}")
    return render(request, 'components/appraisal/appraisal_detail.html', context)
def is_valid_status_transition(from_status, to_status):
    """
    Validates that status transitions follow the correct workflow

    Valid transitions:
    - draft -> submitted
    - submitted -> hr_review (if approved by manager)
    - submitted -> rejected (if rejected by manager)
    - hr_review -> finance_review (if approved by HR)
    - hr_review -> rejected (if rejected by HR)
    - finance_review -> approved (if approved by Finance)
    - finance_review -> rejected (if rejected by Finance)
    """
    valid_transitions = {
        'draft': ['submitted'],
        'submitted': ['hr_review', 'rejected'],
        'hr_review': ['finance_review', 'rejected'],
        'finance_review': ['approved', 'rejected']
    }

    return to_status in valid_transitions.get(from_status, [])

@login_required
def appraisal_create(request):
    """View for creating new appraisal"""
    print(f"[DEBUG] Accessing appraisal_create view for user: {request.user}")

    if request.method == 'POST':
        print("[DEBUG] Processing POST request for appraisal creation")
        try:
            with transaction.atomic():
                # Get manager user object
                manager_id = request.POST.get('manager')
                print(f"[DEBUG] Selected manager ID: {manager_id}")

                if not manager_id:
                    raise ValueError("Manager selection is required")

                manager = User.objects.get(id=manager_id)

                # Validate required fields
                required_fields = ['title', 'overview', 'period_start', 'period_end']
                for field in required_fields:
                    if not request.POST.get(field):
                        raise ValueError(f"{field} is required")

                # Create main appraisal
                appraisal_data = {
                    'user': request.user,
                    'manager': manager,
                    'title': request.POST['title'],
                    'overview': request.POST['overview'],
                    'period_start': request.POST['period_start'],
                    'period_end': request.POST['period_end'],
                    'status': 'draft',
                    'submitted_at': None
                }
                print(f"[DEBUG] Creating appraisal with data: {appraisal_data}")
                appraisal = Appraisal.objects.create(**appraisal_data)

                # Create initial workflow history entry
                workflow_data = {
                    'appraisal': appraisal,
                    'from_status': None,
                    'to_status': 'draft',
                    'action_by': request.user,
                    'comments': 'Initial appraisal creation'
                }
                print(f"[DEBUG] Creating initial workflow history: {workflow_data}")
                AppraisalWorkflow.objects.create(**workflow_data)

                # Handle appraisal items
                items_data = json.loads(request.POST.get('items', '[]'))
                print(f"[DEBUG] Processing {len(items_data)} appraisal items")

                # Require at least one item
                if not items_data:
                    raise ValueError("At least one appraisal item is required")

                # Additional validation for items - ensure required fields exist
                for idx, item in enumerate(items_data):
                    if not item.get('category'):
                        raise ValueError(f"Category is required for item #{idx+1}")
                    if not item.get('title'):
                        raise ValueError(f"Title is required for item #{idx+1}")
                    if not item.get('description'):
                        raise ValueError(f"Description is required for item #{idx+1}")

                for item in items_data:
                    AppraisalItem.objects.create(
                        appraisal=appraisal,
                        category=item['category'],
                        title=item['title'],
                        description=item['description'],
                        date=item.get('date'),
                        employee_rating=item.get('employee_rating')
                    )

                # Handle file attachments
                attachments = request.FILES.getlist('attachments')
                print(f"[DEBUG] Processing {len(attachments)} file attachments")
                for file in attachments:
                    AppraisalAttachment.objects.create(
                        appraisal=appraisal,
                        file=file,
                        title=file.name,
                        uploaded_by=request.user
                    )

                # Notify manager
                print(f"[DEBUG] Notifying manager {manager.email} about new appraisal")

                print(f"[DEBUG] Successfully created appraisal with ID: {appraisal.id}")
                return JsonResponse({'success': True, 'appraisal_id': appraisal.id})
        except ValueError as e:
            print(f"[DEBUG] Validation error: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
        except Exception as e:
            print(f"[DEBUG] Error creating appraisal: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})

    # Get managers for select field
    managers = User.objects.filter(groups__name='Manager')
    print(f"[DEBUG] Available managers: {managers.count()}")

    # Get category choices from model
    categories = dict(AppraisalItem.CATEGORY_CHOICES)
    print(f"[DEBUG] Available categories: {categories}")

    context = {
        'managers': managers,
        'categories': categories
    }
    return render(request, 'components/appraisal/appraisal_form.html', context)

@login_required
def appraisal_update(request, pk):
    """View for updating appraisal"""
    print(f"[DEBUG] Accessing appraisal_update view for pk: {pk}")
    appraisal = get_object_or_404(Appraisal, pk=pk)

    # Only allow updates if user owns the appraisal and it's in draft
    if not (request.user == appraisal.user and appraisal.status == 'draft'):
        print(f"[DEBUG] Permission denied for user {request.user}")
        raise PermissionDenied

    if request.method == 'POST':
        print("[DEBUG] Processing POST request for appraisal update")
        try:
            with transaction.atomic():
                # Update main appraisal
                appraisal.title = request.POST['title']
                appraisal.overview = request.POST['overview']
                appraisal.save()
                print(f"[DEBUG] Updated appraisal main details for ID: {appraisal.id}")

                # Update items
                items_data = json.loads(request.POST.get('items', '[]'))
                print(f"[DEBUG] Deleting existing items and creating {len(items_data)} new items")

                # Require at least one item
                if not items_data:
                    raise ValueError("At least one appraisal item is required")

                appraisal.items.all().delete()  # Remove existing items
                for item in items_data:
                    AppraisalItem.objects.create(
                        appraisal=appraisal,
                        category=item['category'],
                        title=item['title'],
                        description=item['description'],
                        date=item.get('date'),
                        employee_rating=item.get('employee_rating')
                    )

                # Handle new attachments
                new_attachments = request.FILES.getlist('attachments')
                print(f"[DEBUG] Processing {len(new_attachments)} new attachments")
                for file in new_attachments:
                    AppraisalAttachment.objects.create(
                        appraisal=appraisal,
                        file=file,
                        title=file.name,
                        uploaded_by=request.user
                    )

                print("[DEBUG] Successfully updated appraisal")
                return JsonResponse({'success': True})
        except ValueError as e:
            print(f"[DEBUG] Validation error: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
        except Exception as e:
            print(f"[DEBUG] Error updating appraisal: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})

    return render(request, 'components/appraisal/appraisal_form.html', {'appraisal': appraisal})

@login_required
def appraisal_submit(request, pk):
    """View for submitting appraisal for review"""
    print(f"[DEBUG] Accessing appraisal_submit view for pk: {pk}")
    appraisal = get_object_or_404(Appraisal, pk=pk)

    if not (request.user == appraisal.user and appraisal.status == 'draft'):
        print(f"[DEBUG] Permission denied for user {request.user}")
        raise PermissionDenied

    try:
        with transaction.atomic():
            # Verify manager is assigned
            if not appraisal.manager:
                raise ValueError("Cannot submit appraisal without assigned manager")

            # Enhanced verification that appraisal has valid items
            items = appraisal.items.all()
            if not items.exists():
                raise ValueError("Cannot submit appraisal without any items")

            # Check that all required item fields are filled
            for item in items:
                if not item.category:
                    raise ValueError(f"Category is required for item: {item.title}")
                if not item.title:
                    raise ValueError("Title is required for all items")
                if not item.description:
                    raise ValueError(f"Description is required for item: {item.title}")
                if not item.employee_rating:
                    raise ValueError(f"Self-rating is required for item: {item.title}")

            # Validate status transition
            if not is_valid_status_transition('draft', 'submitted'):
                raise ValueError("Invalid status transition from draft")

            appraisal.status = 'submitted'
            appraisal.submitted_at = timezone.now()
            appraisal.save()
            print(f"[DEBUG] Updated appraisal status to submitted for ID: {appraisal.id}")

            # Log workflow transition
            workflow_data = {
                'appraisal': appraisal,
                'from_status': 'draft',
                'to_status': 'submitted',
                'action_by': request.user,
                'comments': request.POST.get('comments', '')
            }
            print(f"[DEBUG] Creating workflow history entry: {workflow_data}")
            AppraisalWorkflow.objects.create(**workflow_data)
            print("[DEBUG] Successfully submitted appraisal")
            return redirect('appraisal:appraisal_detail', pk=appraisal.id)
    except ValueError as e:
        print(f"[DEBUG] Validation error: {str(e)}")
        messages.error(request, str(e))
        return redirect('appraisal:appraisal_detail', pk=appraisal.id)
    except Exception as e:
        print(f"[DEBUG] Error submitting appraisal: {str(e)}")
        messages.error(request, "An error occurred while submitting the appraisal")
        return redirect('appraisal:appraisal_detail', pk=appraisal.id)

@login_required
@user_passes_test(lambda u: is_manager_or_admin(u) or is_hr_or_admin(u) or is_finance_or_admin(u))
def appraisal_review(request, pk):
    """View for reviewing appraisals (Manager/HR/Finance)"""
    print(f"[DEBUG] Accessing appraisal_review view for pk: {pk}")
    appraisal = get_object_or_404(Appraisal, pk=pk)
    user = request.user

    if request.method == 'POST':
        print("[DEBUG] Processing POST request for appraisal review")
        try:
            with transaction.atomic():
                action = request.POST.get('action')
                from_status = appraisal.status
                comments = request.POST.get('comments', '')
                print(f"[DEBUG] Review action: {action}, Current status: {from_status}")

                # Manager Review
                if is_manager_or_admin(user) and appraisal.status == 'submitted':
                    print(f"[DEBUG] Processing manager review for user {user} on appraisal {appraisal.id}")

                    if user != appraisal.manager:
                        print(f"[DEBUG] Permission denied for manager {user} - not assigned manager {appraisal.manager}")
                        raise PermissionDenied

                    if appraisal.items.count() == 0:
                        print(f"[DEBUG] Appraisal {appraisal.id} has no items, cannot be reviewed")
                        raise ValueError("Cannot review an appraisal with no items")

                    # Get all appraisal items that need review
                    appraisal_items = appraisal.items.all()

                    # Process form data - NEW CODE STARTS HERE
                    print("[DEBUG] Extracting item data from form")
                    items_data = []

                    # Extract data from POST dictionary
                    for item in appraisal_items:
                        item_id = str(item.id)
                        manager_rating = request.POST.get(f'items[{item_id}][manager_rating]')
                        manager_comments = request.POST.get(f'items[{item_id}][manager_comments]', '')

                        print(f"[DEBUG] Extracted data for item {item_id}: Rating={manager_rating}, Comments={manager_comments}")

                        if manager_rating:
                            items_data.append({
                                'id': item_id,
                                'manager_rating': manager_rating,
                                'manager_comments': manager_comments
                            })
                    # NEW CODE ENDS HERE

                    print(f"[DEBUG] Items data processed: {items_data}")

                    # Validate that all items have ratings
                    if not items_data:
                        print("[DEBUG] No items data received")
                        raise ValueError("Manager must provide ratings for all items")

                    if len(items_data) != appraisal_items.count():
                        print(f"[DEBUG] Mismatch in items count - Expected: {appraisal_items.count()}, Received: {len(items_data)}")
                        raise ValueError(f"Manager must rate all appraisal items. Expected {appraisal_items.count()} items, received {len(items_data)}.")

                    # Validate each item has a rating
                    for item_data in items_data:
                        if not item_data.get('manager_rating'):
                            print(f"[DEBUG] Missing manager rating for item: {item_data.get('id')}")
                            raise ValueError("Manager rating is required for all items")

                    print(f"[DEBUG] Updating {len(items_data)} items with manager ratings")

                    # Process item updates only after all validations pass
                    for item_data in items_data:
                        print(f"[DEBUG] Processing item data: {item_data}")
                        item = get_object_or_404(AppraisalItem, pk=item_data['id'])
                        print(f"[DEBUG] Found appraisal item {item.id}")

                        old_rating = item.manager_rating
                        old_comments = item.manager_comments

                        item.manager_rating = item_data.get('manager_rating')
                        item.manager_comments = item_data.get('manager_comments', '')
                        item.save()

                        print(f"[DEBUG] Updated item {item.id}:")
                        print(f"  - Rating changed from {old_rating} to {item.manager_rating}")
                        print(f"  - Comments changed from '{old_comments}' to '{item.manager_comments}'")

                    to_status = 'hr_review' if action == 'approve' else 'rejected'
                    if not is_valid_status_transition(from_status, to_status):
                        print(f"[DEBUG] Invalid status transition from {from_status} to {to_status}")
                        raise ValueError(f"Invalid status transition")
                    print(f"[DEBUG] Setting new status to: {to_status} based on action: {action}")

                # HR Review
                elif is_hr_or_admin(user) and appraisal.status == 'hr_review':
                    print("[DEBUG] Processing HR review")
                    to_status = 'finance_review' if action == 'approve' else 'rejected'
                    if not is_valid_status_transition(from_status, to_status):
                        print(f"[DEBUG] Invalid status transition from {from_status} to {to_status}")
                        raise ValueError(f"Invalid status transition")

                # Finance Review
                elif is_finance_or_admin(user) and appraisal.status == 'finance_review':
                    print("[DEBUG] Processing Finance review")
                    to_status = 'approved' if action == 'approve' else 'rejected'
                    if not is_valid_status_transition(from_status, to_status):
                        print(f"[DEBUG] Invalid status transition from {from_status} to {to_status}")
                        raise ValueError(f"Invalid status transition")
                    if to_status == 'approved':
                        appraisal.approved_at = timezone.now()
                else:
                    print(f"[DEBUG] Invalid review state for user {user}")
                    raise PermissionDenied

                appraisal.status = to_status
                appraisal.save()
                print(f"[DEBUG] Updated appraisal status to: {to_status}")

                # Log workflow transition
                workflow_data = {
                    'appraisal': appraisal,
                    'from_status': from_status,
                    'to_status': to_status,
                    'action_by': user,
                    'comments': comments
                }
                print(f"[DEBUG] Creating workflow history entry: {workflow_data}")
                AppraisalWorkflow.objects.create(**workflow_data)
                print("[DEBUG] Successfully completed review")
                return redirect('appraisal:appraisal_list')
        except ValueError as e:
            print(f"[DEBUG] Validation error: {str(e)}")
            messages.error(request, str(e))
            return redirect('appraisal:appraisal_review', pk=appraisal.pk)
        except Exception as e:
            print(f"[DEBUG] Error during review: {str(e)}")
            messages.error(request, f"Error during review: {str(e)}")
            return redirect('appraisal:appraisal_review', pk=appraisal.pk)

    context = {
        'appraisal': appraisal,
        'items': appraisal.items.all(),
        'workflow_history': appraisal.workflow_history.all(),
        'is_manager': is_manager_or_admin(user),
        'is_hr': is_hr_or_admin(user),
        'is_finance': is_finance_or_admin(user)
    }
    print(f"[DEBUG] Rendering review template with {context['items'].count()} items")
    return render(request, 'components/appraisal/appraisal_review.html', context)

from django.db.models.functions import TruncMonth
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Count, Avg, Min, Max, F, Q, FloatField, DurationField
from django.db.models.functions import Cast, TruncMonth
from django.utils import timezone
from datetime import timedelta

@login_required
@user_passes_test(lambda u: is_management_or_admin(u) or is_hr_or_admin(u) or is_finance_or_admin(u))
def appraisal_dashboard(request):
    """Dashboard view for management and HR"""
    print("[DEBUG] Accessing appraisal dashboard view")

    # Base queryset with related fields
    appraisals = Appraisal.objects.select_related('user', 'manager').all()

    # Calculate overall statistics
    total_appraisals = appraisals.count()
    completed_appraisals = appraisals.filter(status__in=['approved', 'rejected']).count()
    in_progress = appraisals.exclude(status__in=['approved', 'rejected', 'draft']).count()
    pending_reviews = appraisals.filter(status='submitted').count()

    # Calculate status distribution
    status_counts = appraisals.values('status').annotate(count=Count('id'))
    status_distribution = {
        status: count for status_obj in status_counts
        for status, count in [(status_obj['status'], status_obj['count'])]
    }

    # Calculate monthly trends - using a safer approach
    # Get time range for trends
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=365)  # Last 12 months

    # Get all appraisals within date range
    period_appraisals = appraisals.filter(
        created_at__date__range=(start_date, end_date)
    )

    # Manual monthly aggregation
    monthly_trends = []
    current_date = start_date.replace(day=1)
    end_month = end_date.replace(day=1)

    while current_date <= end_month:
        next_month = (current_date.replace(day=28) + timedelta(days=4)).replace(day=1)
        month_appraisals = period_appraisals.filter(
            created_at__date__gte=current_date,
            created_at__date__lt=next_month
        )

        monthly_data = {
            'month': current_date,
            'total': month_appraisals.count(),
            'approved': month_appraisals.filter(status='approved').count(),
            'rejected': month_appraisals.filter(status='rejected').count()
        }
        monthly_trends.append(monthly_data)
        current_date = next_month

    # Calculate completion time statistics without relying on database duration calculations
    completed_appraisals_data = appraisals.filter(
        status__in=['approved', 'rejected'],
        submitted_at__isnull=False,
        approved_at__isnull=False
    )

    # Calculate durations in Python rather than at the database level
    completion_durations = []
    for appraisal in completed_appraisals_data:
        if appraisal.submitted_at and appraisal.approved_at:
            try:
                duration = appraisal.approved_at - appraisal.submitted_at
                completion_durations.append(duration.total_seconds())
            except (TypeError, ValueError):
                continue

    # Convert durations to human readable strings
    completion_times = {}
    if completion_durations:
        avg_seconds = sum(completion_durations)/len(completion_durations)
        min_seconds = min(completion_durations)
        max_seconds = max(completion_durations)

        completion_times = {
            'avg_time': f"{int(avg_seconds/86400)} days {int((avg_seconds%86400)/3600)} hours",
            'min_time': f"{int(min_seconds/86400)} days {int((min_seconds%86400)/3600)} hours",
            'max_time': f"{int(max_seconds/86400)} days {int((max_seconds%86400)/3600)} hours",
        }
    else:
        completion_times = {
            'avg_time': None,
            'min_time': None,
            'max_time': None
        }

    # Calculate rating statistics
    employee_ratings = AppraisalItem.objects.filter(
        appraisal__in=appraisals,
        employee_rating__isnull=False
    ).values_list('employee_rating', flat=True)

    manager_ratings = AppraisalItem.objects.filter(
        appraisal__in=appraisals,
        manager_rating__isnull=False
    ).values_list('manager_rating', flat=True)

    employee_ratings_list = list(employee_ratings)
    manager_ratings_list = list(manager_ratings)

    rating_stats = {
        'employee': {
            'avg': sum(employee_ratings_list) / len(employee_ratings_list) if employee_ratings_list else None,
            'min': min(employee_ratings_list) if employee_ratings_list else None,
            'max': max(employee_ratings_list) if employee_ratings_list else None,
        },
        'manager': {
            'avg': sum(manager_ratings_list) / len(manager_ratings_list) if manager_ratings_list else None,
            'min': min(manager_ratings_list) if manager_ratings_list else None,
            'max': max(manager_ratings_list) if manager_ratings_list else None,
        }
    }

    context = {
        'overview_stats': {
            'total_appraisals': total_appraisals,
            'completed_appraisals': completed_appraisals,
            'in_progress': in_progress,
            'pending_reviews': pending_reviews,
        },
        'status_distribution': status_distribution,
        'monthly_trends': monthly_trends,
        'completion_times': completion_times,
        'rating_stats': rating_stats,
        'filter_options': {
            'managers': User.objects.filter(groups__name='Manager'),
            'statuses': dict(Appraisal.STATUS_CHOICES),
            'date_range': {
                'start': start_date,
                'end': end_date
            }
        }
    }

    print(f"[DEBUG] Dashboard statistics prepared: {context}")
    return render(request, 'components/appraisal/dashboard.html', context)

'''--------------------------------- FINANCE --------------------------'''
# Finance Views

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, FileResponse
from django.db.models import Sum, Count, Q, F, Value, Case, When
from django.db.models.functions import Coalesce, ExtractMonth, ExtractYear
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import json
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle

from .models import (
    DailyExpense, Voucher, VoucherDetail, BankAccount,
    BankPayment, Subscription, ClientInvoice, ChartOfAccount
)
from django.contrib.auth.decorators import user_passes_test
from django.db.models import DecimalField

def is_finance(user):
    return user.groups.filter(name='Finance').exists()


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.utils import timezone
from django.db.models import Q
from django.contrib.contenttypes.models import ContentType

from .models import FinancialParameter



@login_required
@user_passes_test(is_finance)
def financial_parameter_list(request):
    """List view with extensive filtering options"""
    # Base queryset
    queryset = FinancialParameter.objects.all().order_by('-updated_at')

    # Filter options
    key_filter = request.GET.get('key', '')
    name_filter = request.GET.get('name', '')
    category_filter = request.GET.get('category', '')
    value_type_filter = request.GET.get('value_type', '')
    is_global_filter = request.GET.get('is_global', '')
    fiscal_year_filter = request.GET.get('fiscal_year', '')
    is_approved_filter = request.GET.get('is_approved', '')

    # Apply filters
    if key_filter:
        queryset = queryset.filter(key__icontains=key_filter)
    if name_filter:
        queryset = queryset.filter(name__icontains=name_filter)
    if category_filter:
        queryset = queryset.filter(category=category_filter)
    if value_type_filter:
        queryset = queryset.filter(value_type=value_type_filter)
    if is_global_filter:
        is_global_value = is_global_filter == 'true'
        queryset = queryset.filter(is_global=is_global_value)
    if fiscal_year_filter:
        queryset = queryset.filter(fiscal_year=fiscal_year_filter)
    if is_approved_filter:
        is_approved_value = is_approved_filter == 'true'
        queryset = queryset.filter(is_approved=is_approved_value)

    # Get unique values for dropdowns
    fiscal_years = FinancialParameter.objects.values_list('fiscal_year', flat=True).distinct()
    fiscal_years = [fy for fy in fiscal_years if fy]  # Remove None values

    # Pagination
    paginator = Paginator(queryset, 15)
    page = request.GET.get('page')
    parameters = paginator.get_page(page)

    context = {
        'parameters': parameters,
        'value_type_choices': FinancialParameter.VALUE_TYPE_CHOICES,
        'category_choices': FinancialParameter.CATEGORY_CHOICES,
        'fiscal_years': sorted(fiscal_years, reverse=True),
        'filters': {
            'key': key_filter,
            'name': name_filter,
            'category': category_filter,
            'value_type': value_type_filter,
            'is_global': is_global_filter,
            'fiscal_year': fiscal_year_filter,
            'is_approved': is_approved_filter,
        }
    }

    return render(request, 'components/finance/financial_parameter_list.html', context)


@login_required
@user_passes_test(is_finance)
def financial_parameter_create(request):
    """Create a new financial parameter"""
    if request.method == 'POST':
        data = request.POST
        try:
            # Create parameter with basic fields
            parameter = FinancialParameter(
                key=data.get('key'),
                name=data.get('name'),
                description=data.get('description'),
                category=data.get('category'),
                value_type=data.get('value_type'),
                is_global=data.get('is_global') == 'on',
                valid_from=data.get('valid_from'),
                valid_to=data.get('valid_to') or None,
                fiscal_year=data.get('fiscal_year') or None,
                fiscal_quarter=data.get('fiscal_quarter') or None,
                created_by=request.user,
                updated_by=request.user
            )

            # Handle entity relation if not global
            if not parameter.is_global and data.get('content_type_id') and data.get('object_id'):
                parameter.content_type_id = data.get('content_type_id')
                parameter.object_id = data.get('object_id')

            # Set the value with proper type conversion
            parameter.set_value(data.get('value'))

            # Save the parameter
            parameter.save()

            # Auto-approve if user has permission
            if request.user.has_perm('finance.approve_financialparameter'):
                parameter.approve(request.user)
                messages.success(request, 'Parameter created and approved successfully.')
            else:
                messages.success(request, 'Parameter created successfully. Awaiting approval.')

            return redirect('finance:financial_parameter_detail', pk=parameter.pk)
        except Exception as e:
            messages.error(request, f'Error creating parameter: {str(e)}')

    # Get content types for entity selection
    content_types = ContentType.objects.all().order_by('app_label', 'model')

    context = {
        'title': 'Create Financial Parameter',
        'value_type_choices': FinancialParameter.VALUE_TYPE_CHOICES,
        'category_choices': FinancialParameter.CATEGORY_CHOICES,
        'content_types': content_types,
        'can_approve': request.user.has_perm('finance.approve_financialparameter'),
    }
    return render(request, 'components/finance/financial_parameter_form.html', context)


@login_required
@user_passes_test(is_finance)
def financial_parameter_detail(request, pk):
    """View a single parameter's details"""
    parameter = get_object_or_404(FinancialParameter, pk=pk)

    # Get entity details if not global
    entity_details = None
    if not parameter.is_global and parameter.content_type and parameter.object_id:
        try:
            entity_model = parameter.content_type.model_class()
            entity = entity_model.objects.get(pk=parameter.object_id)
            entity_details = {
                'type': parameter.content_type.model,
                'id': parameter.object_id,
                'name': str(entity)
            }
        except:
            entity_details = {
                'type': parameter.content_type.model,
                'id': parameter.object_id,
                'name': 'Unknown or deleted entity'
            }

    # Check if user can approve
    can_approve = request.user.has_perm('finance.approve_financialparameter')

    # Handle approval action
    if request.method == 'POST' and 'approve' in request.POST:
        if can_approve:
            parameter.approve(request.user)
            messages.success(request, 'Parameter approved successfully.')
            return redirect('aps_finance:financial_parameter_detail', pk=parameter.pk)
        else:
            messages.error(request, 'You do not have permission to approve parameters.')

    context = {
        'parameter': parameter,
        'entity_details': entity_details,
        'typed_value': parameter.get_typed_value(),
        'can_approve': can_approve and not parameter.is_approved,
    }
    return render(request, 'components/finance/financial_parameter_detail.html', context)


@login_required
@user_passes_test(is_finance)
def financial_parameter_update(request, pk):
    """Update an existing parameter"""
    parameter = get_object_or_404(FinancialParameter, pk=pk)

    # Check if user has permission to edit
    if parameter.is_approved and not request.user.has_perm('finance.change_approved_financialparameter'):
        messages.error(request, 'Cannot edit an approved parameter. Create a new version instead.')
        return redirect('aps_finance:financial_parameter_detail', pk=parameter.pk)

    if request.method == 'POST':
        data = request.POST
        try:
            # Update basic fields
            parameter.key = data.get('key')
            parameter.name = data.get('name')
            parameter.description = data.get('description')
            parameter.category = data.get('category')
            parameter.value_type = data.get('value_type')
            parameter.is_global = data.get('is_global') == 'on'
            parameter.valid_from = data.get('valid_from')
            parameter.valid_to = data.get('valid_to') or None
            parameter.fiscal_year = data.get('fiscal_year') or None
            parameter.fiscal_quarter = data.get('fiscal_quarter') or None
            parameter.updated_by = request.user

            # Handle entity relation
            if not parameter.is_global and data.get('content_type_id') and data.get('object_id'):
                parameter.content_type_id = data.get('content_type_id')
                parameter.object_id = data.get('object_id')
            elif parameter.is_global:
                parameter.content_type = None
                parameter.object_id = None

            # Set the value with proper type conversion
            parameter.set_value(data.get('value'))

            # Save changes
            parameter.save()

            # Auto-approve if user has permission
            if not parameter.is_approved and request.user.has_perm('finance.approve_financialparameter'):
                if data.get('approve') == 'on':
                    parameter.approve(request.user)
                    messages.success(request, 'Parameter updated and approved successfully.')
                else:
                    messages.success(request, 'Parameter updated successfully. Not approved.')
            else:
                messages.success(request, 'Parameter updated successfully.')

            return redirect('aps_finance:financial_parameter_detail', pk=parameter.pk)
        except Exception as e:
            messages.error(request, f'Error updating parameter: {str(e)}')

    # Get content types for entity selection
    content_types = ContentType.objects.all().order_by('app_label', 'model')

    # Get entity details if not global
    selected_entity = None
    if not parameter.is_global and parameter.content_type and parameter.object_id:
        try:
            entity_model = parameter.content_type.model_class()
            entity = entity_model.objects.get(pk=parameter.object_id)
            selected_entity = {
                'content_type_id': parameter.content_type.id,
                'object_id': parameter.object_id,
                'name': str(entity)
            }
        except:
            selected_entity = {
                'content_type_id': parameter.content_type.id,
                'object_id': parameter.object_id,
                'name': 'Unknown or deleted entity'
            }

    context = {
        'parameter': parameter,
        'title': 'Update Financial Parameter',
        'value_type_choices': FinancialParameter.VALUE_TYPE_CHOICES,
        'category_choices': FinancialParameter.CATEGORY_CHOICES,
        'content_types': content_types,
        'selected_entity': selected_entity,
        'can_approve': not parameter.is_approved and request.user.has_perm('finance.approve_financialparameter'),
    }
    return render(request, 'components/finance/financial_parameter_form.html', context)


@login_required
@user_passes_test(is_finance)
def financial_parameter_delete(request, pk):
    """Delete a parameter"""
    parameter = get_object_or_404(FinancialParameter, pk=pk)

    # Prevent deletion of approved parameters without special permission
    if parameter.is_approved and not request.user.has_perm('finance.delete_approved_financialparameter'):
        messages.error(request, 'Cannot delete an approved parameter.')
        return redirect('aps_finance:financial_parameter_detail', pk=parameter.pk)

    if request.method == 'POST':
        parameter.delete()
        messages.success(request, 'Parameter deleted successfully.')
        return redirect('aps_finance:financial_parameter_list')

    return render(request, 'components/finance/financial_parameter_confirm_delete.html', {
        'parameter': parameter
    })


@login_required
@user_passes_test(is_finance)
def financial_parameter_duplicate(request, pk):
    """Create a new parameter based on an existing one"""
    source_parameter = get_object_or_404(FinancialParameter, pk=pk)

    # Create a new parameter as a copy but with unique identifiers
    new_parameter = FinancialParameter(
        key=source_parameter.key,
        name=f"Copy of {source_parameter.name}",
        description=source_parameter.description,
        category=source_parameter.category,
        value=source_parameter.value,
        value_type=source_parameter.value_type,
        is_global=source_parameter.is_global,
        content_type=source_parameter.content_type,
        object_id=source_parameter.object_id,
        valid_from=timezone.now().date(),  # Start from today
        fiscal_year=source_parameter.fiscal_year,
        fiscal_quarter=source_parameter.fiscal_quarter,
        created_by=request.user,
        updated_by=request.user
    )

    # Save without validating unique constraints yet - will be changed in form
    new_parameter.save(force_insert=True)

    messages.info(request, 'Created a duplicate. Please modify as needed and save.')
    return redirect('aps_finance:financial_parameter_update', pk=new_parameter.pk)


@login_required
@user_passes_test(is_finance)
def entity_parameter_list(request, content_type_id, object_id):
    """View all parameters for a specific entity"""
    content_type = get_object_or_404(ContentType, pk=content_type_id)

    try:
        # Try to get the actual entity
        model_class = content_type.model_class()
        entity = model_class.objects.get(pk=object_id)
        entity_name = str(entity)
    except:
        entity = None
        entity_name = f"{content_type.model} #{object_id} (not found)"

    # Get parameters specific to this entity
    entity_parameters = FinancialParameter.objects.filter(
        content_type_id=content_type_id,
        object_id=object_id,
        is_global=False
    ).order_by('key', '-valid_from')

    # Get today's date for parameter applicability
    today = timezone.now().date()

    # Get all current applicable parameters (both entity-specific and global)
    current_parameters = {}
    if entity:
        # Get all parameter keys first
        keys = FinancialParameter.objects.filter(
            Q(is_global=True) |
            Q(content_type_id=content_type_id, object_id=object_id, is_global=False)
        ).values_list('key', flat=True).distinct()

        # For each key, get the applicable value
        for key in keys:
            value = FinancialParameter.get_param(key, entity=entity)
            if value is not None:
                current_parameters[key] = value

    context = {
        'entity_type': content_type.model,
        'entity_id': object_id,
        'entity_name': entity_name,
        'entity_parameters': entity_parameters,
        'current_parameters': current_parameters,
        'today': today,
    }

    return render(request, 'components/finance/entity_parameter_list.html', context)


@login_required
@user_passes_test(is_finance)
def parameter_history(request, key):
    """View historical values of a parameter"""
    # Get all versions of this parameter, ordered by validity date
    parameters = FinancialParameter.objects.filter(
        key=key
    ).order_by('-valid_from')

    # Group by entity/global
    global_parameters = parameters.filter(is_global=True)
    entity_parameters = parameters.filter(is_global=False)

    # Group entity parameters by entity type
    entity_groups = {}
    for param in entity_parameters:
        entity_key = (param.content_type_id, param.object_id)
        if entity_key not in entity_groups:
            # Try to get entity name
            try:
                model_class = param.content_type.model_class()
                entity = model_class.objects.get(pk=param.object_id)
                entity_name = str(entity)
            except:
                entity_name = f"{param.content_type.model} #{param.object_id} (not found)"

            entity_groups[entity_key] = {
                'type': param.content_type.model,
                'id': param.object_id,
                'name': entity_name,
                'parameters': []
            }

        entity_groups[entity_key]['parameters'].append(param)

    context = {
        'parameter_key': key,
        'global_parameters': global_parameters,
        'entity_groups': entity_groups.values(),
    }

    return render(request, 'components/finance/parameter_history.html', context)


@login_required
@user_passes_test(lambda u: is_finance(u) or is_management(u))
def approve_parameter(request, pk):
    """Approve a parameter (separate view for approval-only)"""
    parameter = get_object_or_404(FinancialParameter, pk=pk)

    # Check if user can approve - either finance or management
    if not (is_finance(request.user) or is_management(request.user)):
        messages.error(request, 'You do not have permission to approve parameters.')
        return redirect('aps_finance:financial_parameter_detail', pk=parameter.pk)

    # Check if already approved
    if parameter.is_approved:
        messages.warning(request, 'This parameter is already approved.')
        return redirect('aps_finance:financial_parameter_detail', pk=parameter.pk)

    if request.method == 'POST':
        parameter.approve(request.user)
        messages.success(request, 'Parameter approved successfully.')
        return redirect('aps_finance:financial_parameter_detail', pk=parameter.pk)

    return render(request, 'components/finance/parameter_approve.html', {
        'parameter': parameter,
    })

@login_required
@user_passes_test(is_finance)
def expense_entry(request):
    """Handle daily expense entry and listing with advanced filtering and calculations"""
    if request.method == 'POST':
        try:
            # Create expense
            expense = DailyExpense.objects.create(
                expense_id=f"EXP-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                department_id=request.POST.get('department'),
                date=request.POST.get('date'),
                category=request.POST.get('category'),
                description=request.POST.get('description'),
                amount=request.POST.get('amount'),
                paid_by=request.user,
                status='draft',
                attachments=request.FILES.get('attachments')
            )

            messages.success(request, 'Expense recorded successfully')
            return JsonResponse({'status': 'success', 'expense_id': expense.expense_id})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    # Advanced filtering and aggregation
    filters = {}
    if request.GET.get('start_date'):
        filters['date__gte'] = request.GET.get('start_date')
    if request.GET.get('end_date'):
        filters['date__lte'] = request.GET.get('end_date')
    if request.GET.get('category'):
        filters['category'] = request.GET.get('category')
    if request.GET.get('status'):
        filters['status'] = request.GET.get('status')

    expenses = DailyExpense.objects.filter(**filters).order_by('-date')

    # Calculate statistics
    stats = expenses.aggregate(
        total_amount=Coalesce(Sum('amount'), Value(0, output_field=DecimalField(max_digits=15, decimal_places=2))),
        avg_amount=Coalesce(
            Sum('amount') / Cast(Count('id'), DecimalField(max_digits=15, decimal_places=2)),
            Value(0),
            output_field=DecimalField(max_digits=15, decimal_places=2)
        ),
        count=Count('id')
    )

    # Monthly trend analysis
    monthly_trend = expenses.annotate(
        month=ExtractMonth('date'),
        year=ExtractYear('date')
    ).values('month', 'year').annotate(
        total=Sum('amount')
    ).order_by('year', 'month')

    context = {
        'expenses': expenses,
        'expense_categories': dict(DailyExpense.EXPENSE_CATEGORIES),
        'expense_statuses': dict(DailyExpense.EXPENSE_STATUS),
        'stats': stats,
        'monthly_trend': list(monthly_trend),
        'filters': request.GET
    }
    return render(request, 'components/finance/expense_entry.html', context)

@login_required
@user_passes_test(is_finance)
def voucher_entry(request):
    """Handle voucher creation and management with validation and auto-calculations"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            # Validate debit/credit equality
            total_debit = sum(Decimal(entry['debit']) for entry in data['entries'])
            total_credit = sum(Decimal(entry['credit']) for entry in data['entries'])

            if total_debit != total_credit:
                raise ValueError("Total debit must equal total credit")

            # Create voucher with transaction
            voucher = Voucher.objects.create(
                voucher_number=f"V-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                type=data['type'],
                date=data['date'],
                party_name=data['party_name'],
                purpose=data['purpose'],
                amount=total_debit,
                status='draft'
            )

            # Create voucher details with validation
            for entry in data['entries']:
                if Decimal(entry['debit']) > 0 and Decimal(entry['credit']) > 0:
                    raise ValueError("An entry cannot have both debit and credit")

                VoucherDetail.objects.create(
                    voucher=voucher,
                    account_id=entry['account'],
                    debit_amount=Decimal(entry['debit']),
                    credit_amount=Decimal(entry['credit'])
                )

            return JsonResponse({'status': 'success', 'voucher_id': voucher.id})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    # Get vouchers with filters
    filters = {}
    if request.GET.get('type'):
        filters['type'] = request.GET.get('type')
    if request.GET.get('status'):
        filters['status'] = request.GET.get('status')
    if request.GET.get('date_from'):
        filters['date__gte'] = request.GET.get('date_from')
    if request.GET.get('date_to'):
        filters['date__lte'] = request.GET.get('date_to')

    vouchers = Voucher.objects.filter(**filters).order_by('-date')
    accounts = ChartOfAccount.objects.filter(is_active=True)

    # Calculate voucher statistics
    stats = vouchers.aggregate(
        total_amount=Sum('amount'),
        count=Count('id'),
        pending_approval=Count('id', filter=Q(status='pending_approval'))
    )

    context = {
        'vouchers': vouchers,
        'accounts': accounts,
        'voucher_types': dict(Voucher.VOUCHER_TYPES),
        'voucher_statuses': dict(Voucher.VOUCHER_STATUS),
        'stats': stats,
        'filters': request.GET
    }
    return render(request, 'components/finance/voucher_entry.html', context)



@login_required
@user_passes_test(is_finance)
def invoice_generation(request):
    """Handle client invoice generation with calculations"""
    if request.method == 'POST':
        try:
            # Calculate invoice amounts
            rate = Decimal(request.POST.get('rate'))

            # Get the appropriate count based on billing model
            billing_model = request.POST.get('billing_model')
            if billing_model == 'per_order':
                count = Decimal(request.POST.get('order_count') or 0)
                fte_count = None
                order_count = count
            else:  # per_fte
                count = Decimal(request.POST.get('fte_count') or 0)
                fte_count = count
                order_count = None

            subtotal = rate * count

            tax_rate = Decimal(request.POST.get('tax_rate', '0'))
            tax_amount = (subtotal * tax_rate) / 100

            discount = Decimal(request.POST.get('discount', '0'))
            total_amount = subtotal + tax_amount - discount

            invoice = ClientInvoice.objects.create(
                invoice_number=f"INV-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                client_id=request.POST.get('client'),
                billing_model=billing_model,
                billing_cycle_start=request.POST.get('cycle_start'),
                billing_cycle_end=request.POST.get('cycle_end'),
                order_count=order_count,
                fte_count=fte_count,
                rate=rate,
                subtotal=subtotal,
                tax_amount=tax_amount,
                discount=discount,
                total_amount=total_amount,
                due_date=request.POST.get('due_date'),
                status='draft'
            )

            messages.success(request, 'Invoice generated successfully')
            return redirect('aps_finance:invoice_detail', invoice_id=invoice.id)
        except Exception as e:
            messages.error(request, f'Error generating invoice: {str(e)}')

    # Get invoices with filters
    filters = {}
    if request.GET.get('status'):
        filters['status'] = request.GET.get('status')
    if request.GET.get('client'):
        filters['client'] = request.GET.get('client')
    if request.GET.get('date_from'):
        filters['billing_cycle_start__gte'] = request.GET.get('date_from')

    invoices = ClientInvoice.objects.filter(**filters).order_by('-created_at')

    # Calculate invoice statistics
    stats = {
        'total_amount': invoices.aggregate(total=Sum('total_amount'))['total'] or 0,
        'total_pending': invoices.filter(status='pending_approval').aggregate(total=Sum('total_amount'))['total'] or 0,
        'count': invoices.count()
    }

    # Get all clients from Client group
    clients = User.objects.filter(groups__name='Client')

    context = {
        'invoices': invoices,
        'clients': clients,
        'billing_models': dict(ClientInvoice.BILLING_MODELS),
        'invoice_statuses': dict(ClientInvoice.INVOICE_STATUS),
        'stats': stats,
        'filters': request.GET
    }
    return render(request, 'components/finance/invoice_generation.html', context)


@login_required
@user_passes_test(is_finance)
def invoice_detail(request, invoice_id):
    """Show detailed view of an invoice"""
    invoice = get_object_or_404(ClientInvoice, id=invoice_id)

    context = {
        'invoice': invoice,
        'billing_models': dict(ClientInvoice.BILLING_MODELS),
        'invoice_statuses': dict(ClientInvoice.INVOICE_STATUS),
        'user_is_finance': is_finance(request.user)
    }
    return render(request, 'components/finance/invoice_detail.html', context)


@login_required
@user_passes_test(is_finance)
def invoice_print(request, invoice_id):
    """Generate printable invoice view"""
    invoice = get_object_or_404(ClientInvoice, id=invoice_id)

    context = {
        'invoice': invoice,
        'billing_models': dict(ClientInvoice.BILLING_MODELS),
        'invoice_statuses': dict(ClientInvoice.INVOICE_STATUS),
    }
    return render(request, 'components/finance/invoice_print.html', context)


@login_required
@user_passes_test(is_finance)
def invoice_update_status(request, invoice_id):
    """Update invoice status"""
    if request.method == 'POST':
        invoice = get_object_or_404(ClientInvoice, id=invoice_id)
        new_status = request.POST.get('status')

        if new_status in dict(ClientInvoice.INVOICE_STATUS).keys():
            invoice.status = new_status
            invoice.save()
            messages.success(request, f'Invoice status updated to {dict(ClientInvoice.INVOICE_STATUS)[new_status]}')
        else:
            messages.error(request, 'Invalid status value')

    return redirect('aps_finance:invoice_detail', invoice_id=invoice_id)


@login_required
@user_passes_test(is_finance)
def invoice_edit(request, invoice_id):
    """Edit an existing invoice"""
    invoice = get_object_or_404(ClientInvoice, id=invoice_id)

    # Only draft invoices can be edited
    if invoice.status != 'draft':
        messages.error(request, 'Only draft invoices can be edited')
        return redirect('aps_finance:invoice_detail', invoice_id=invoice.id)

    if request.method == 'POST':
        try:
            # Update invoice data
            rate = Decimal(request.POST.get('rate'))

            # Get the appropriate count based on billing model
            billing_model = request.POST.get('billing_model')
            if billing_model == 'per_order':
                count = Decimal(request.POST.get('order_count') or 0)
                invoice.fte_count = None
                invoice.order_count = count
            else:  # per_fte
                count = Decimal(request.POST.get('fte_count') or 0)
                invoice.fte_count = count
                invoice.order_count = None

            subtotal = rate * count

            tax_rate = Decimal(request.POST.get('tax_rate', '0'))
            tax_amount = (subtotal * tax_rate) / 100

            discount = Decimal(request.POST.get('discount', '0'))
            total_amount = subtotal + tax_amount - discount

            # Update invoice fields
            invoice.client_id = request.POST.get('client')
            invoice.billing_model = billing_model
            invoice.billing_cycle_start = request.POST.get('cycle_start')
            invoice.billing_cycle_end = request.POST.get('cycle_end')
            invoice.rate = rate
            invoice.subtotal = subtotal
            invoice.tax_amount = tax_amount
            invoice.discount = discount
            invoice.total_amount = total_amount
            invoice.due_date = request.POST.get('due_date')
            invoice.save()

            messages.success(request, 'Invoice updated successfully')
            return redirect('aps_finance:invoice_detail', invoice_id=invoice.id)
        except Exception as e:
            messages.error(request, f'Error updating invoice: {str(e)}')

    # Get all clients from Client group
    clients = User.objects.filter(groups__name='Client')

    context = {
        'invoice': invoice,
        'clients': clients,
        'billing_models': dict(ClientInvoice.BILLING_MODELS),
    }
    return render(request, 'components/finance/invoice_edit.html', context)


@login_required
@user_passes_test(is_finance)
def finance_dashboard(request):
    """Finance module dashboard with key metrics and reports"""
    today = timezone.now().date()
    start_date = today - timedelta(days=30)

    # Expense metrics
    expense_metrics = DailyExpense.objects.filter(
        date__gte=start_date
    ).aggregate(
        total=Sum('amount'),
        count=Count('id'),
        pending=Count('id', filter=Q(status='pending_approval'))
    )

    # Invoice metrics
    invoice_metrics = ClientInvoice.objects.filter(
        created_at__date__gte=start_date
    ).aggregate(
        total=Sum('total_amount'),
        pending=Sum('total_amount', filter=Q(status='pending_approval')),
        overdue=Count('id', filter=Q(status='overdue'))
    )

    # Monthly trends
    monthly_expenses = DailyExpense.objects.filter(
        date__gte=start_date
    ).annotate(
        month=ExtractMonth('date')
    ).values('month').annotate(
        total=Sum('amount')
    ).order_by('month')

    monthly_invoices = ClientInvoice.objects.filter(
        created_at__date__gte=start_date
    ).annotate(
        month=ExtractMonth('created_at')
    ).values('month').annotate(
        total=Sum('total_amount')
    ).order_by('month')

    context = {
        'expense_metrics': expense_metrics,
        'invoice_metrics': invoice_metrics,
        'monthly_expenses': list(monthly_expenses),
        'monthly_invoices': list(monthly_invoices),
        'date_range': {
            'start': start_date,
            'end': today
        }
    }
    return render(request, 'components/finance/dashboard.html', context)



# List and Create Bank Payments
@login_required
@user_passes_test(is_finance)
def bank_payment_list(request):
    """View for listing bank payments with filters"""
    # Get payments with filters
    filters = {}
    if request.GET.get('status'):
        filters['status'] = request.GET.get('status')
    if request.GET.get('bank_account'):
        filters['bank_account'] = request.GET.get('bank_account')
    if request.GET.get('date_from'):
        filters['payment_date__gte'] = request.GET.get('date_from')
    if request.GET.get('date_to'):
        filters['payment_date__lte'] = request.GET.get('date_to')
    if request.GET.get('payee_name'):
        filters['party_name__icontains'] = request.GET.get('payee_name')

    payments = BankPayment.objects.filter(**filters).order_by('-payment_date')
    bank_accounts = BankAccount.objects.filter(is_active=True)

    # Calculate payment statistics
    stats = payments.aggregate(
        total_amount=Sum('amount'),
        pending_amount=Sum('amount', filter=Q(status='pending')),
        approved_amount=Sum('amount', filter=Q(status='approved')),
        executed_amount=Sum('amount', filter=Q(status='executed')),
        count=Count('id')
    )

    context = {
        'payments': payments,
        'bank_accounts': bank_accounts,
        'payment_statuses': dict(BankPayment.PAYMENT_STATUS),
        'stats': stats,
        'filters': request.GET
    }
    return render(request, 'components/finance/bank_payment_list.html', context)

@login_required
@user_passes_test(is_finance)
@require_http_methods(["GET", "POST"])
def bank_payment_create(request):
    """Handle bank payment creation"""
    if request.method == 'GET':
        bank_accounts = BankAccount.objects.filter(is_active=True)
        context = {
            'bank_accounts': bank_accounts
        }
        return render(request, 'components/finance/bank_payment_form.html', context)

    try:
        # Create bank payment record
        payment = BankPayment.objects.create(
            payment_id=f"BP-{timezone.now().strftime('%Y%m%d%H%M%S')}",
            bank_account_id=request.POST.get('bank_account'),
            party_name=request.POST.get('payee_name'),
            payment_reason=request.POST.get('purpose'),
            amount=request.POST.get('amount'),
            reference_number=request.POST.get('reference_number'),
            payment_date=request.POST.get('payment_date'),
            status='pending',
            created_by=request.user,
        )

        # Handle file upload
        if 'attachments' in request.FILES:
            payment.attachments = request.FILES.get('attachments')
            payment.save()

        messages.success(request, 'Bank payment created successfully')

        # Return JSON for AJAX requests, redirect for form submissions
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'payment_id': payment.payment_id})
        return redirect('aps_finance:bank_payment_detail', payment_id=payment.payment_id)

    except Exception as e:
        messages.error(request, f'Error creating payment: {str(e)}')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': str(e)})
        return render(request, 'components/finance/bank_payment_form.html', {'error': str(e)})

@login_required
@user_passes_test(is_finance)
def bank_payment_detail(request, payment_id):
    """Display details of a specific bank payment"""
    payment = get_object_or_404(BankPayment, payment_id=payment_id)
    bank_accounts = BankAccount.objects.filter(is_active=True)

    context = {
        'payment': payment,
        'bank_accounts': bank_accounts,
        'payment_statuses': dict(BankPayment.PAYMENT_STATUS)
    }
    return render(request, 'components/finance/bank_payment_detail.html', context)


@login_required
@user_passes_test(is_finance)
@require_http_methods(["POST"])
def bank_payment_update(request, payment_id):
    """Handle bank payment update"""
    try:
        payment = get_object_or_404(BankPayment, payment_id=payment_id)

        # Check if payment can be updated (only pending payments)
        if payment.status not in ['pending']:
            messages.error(request, 'Only pending payments can be updated')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': 'Only pending payments can be updated'})
            return render(request, 'components/finance/bank_payment_detail.html', {'payment': payment})

        # Update payment details
        payment.bank_account_id = request.POST.get('bank_account', payment.bank_account_id)
        payment.party_name = request.POST.get('payee_name', payment.party_name)
        payment.payment_reason = request.POST.get('purpose', payment.payment_reason)
        payment.amount = request.POST.get('amount', payment.amount)
        payment.reference_number = request.POST.get('reference_number', payment.reference_number)
        payment.payment_date = request.POST.get('payment_date', payment.payment_date)

        if 'attachments' in request.FILES:
            payment.attachments = request.FILES.get('attachments')

        payment.save()

        messages.success(request, 'Bank payment updated successfully')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success'})
        return render(request, 'components/finance/bank_payment_detail.html', {'payment': payment})

    except BankPayment.DoesNotExist:
        messages.error(request, 'Payment not found')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': 'Payment not found'})
        return render(request, 'components/finance/bank_payment_list.html')

    except Exception as e:
        messages.error(request, f'Error updating payment: {str(e)}')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': str(e)})
        return render(request, 'components/finance/bank_payment_detail.html', {'payment': payment})

@login_required
@user_passes_test(is_finance)
@require_http_methods(["POST"])
def bank_payment_delete(request, payment_id):
    """Handle bank payment deletion"""
    try:
        payment = get_object_or_404(BankPayment, payment_id=payment_id)

        # Check if payment can be deleted (only pending payments)
        if payment.status not in ['pending']:
            messages.error(request, 'Only pending payments can be deleted')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': 'Only pending payments can be deleted'})
            return redirect('aps_finance:bank_payment_detail', payment_id=payment_id)

        payment.delete()

        messages.success(request, 'Bank payment deleted successfully')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success'})
        return redirect('aps_finance:bank_payment_list')

    except BankPayment.DoesNotExist:
        messages.error(request, 'Payment not found')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': 'Payment not found'})
        return redirect('aps_finance:bank_payment_list')

    except Exception as e:
        messages.error(request, f'Error deleting payment: {str(e)}')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': str(e)})
        return redirect('aps_finance:bank_payment_detail', payment_id=payment_id)

# Payment Workflow Functions
@login_required
@user_passes_test(is_finance)
@require_http_methods(["POST"])
def bank_payment_verify(request, payment_id):
    """Mark a payment as verified"""
    try:
        payment = get_object_or_404(BankPayment, payment_id=payment_id)

        # Check if payment can be verified
        if payment.status != 'pending':
            messages.error(request, 'Only pending payments can be verified')
            return JsonResponse({'status': 'error', 'message': 'Only pending payments can be verified'}) \
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest' else redirect('bank_payment_detail', payment_id=payment_id)

        payment.status = 'verified'
        payment.verified_by = request.user
        payment.save()

        messages.success(request, 'Payment verified successfully')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success'})
        return redirect('aps_finance:bank_payment_detail', payment_id=payment_id)

    except Exception as e:
        messages.error(request, f'Error verifying payment: {str(e)}')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': str(e)})
        return redirect('aps_finance:bank_payment_detail', payment_id=payment_id)

@login_required
@user_passes_test(is_finance)
@require_http_methods(["POST"])
def bank_payment_approve(request, payment_id):
    """Mark a payment as approved"""
    try:
        payment = get_object_or_404(BankPayment, payment_id=payment_id)

        # Check if payment can be approved
        if payment.status != 'verified':
            messages.error(request, 'Only verified payments can be approved')
            return JsonResponse({'status': 'error', 'message': 'Only verified payments can be approved'}) \
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest' else redirect('bank_payment_detail', payment_id=payment_id)

        payment.status = 'approved'
        payment.approved_by = request.user
        payment.save()

        messages.success(request, 'Payment approved successfully')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success'})
        return redirect('aps_finance:bank_payment_detail', payment_id=payment_id)

    except Exception as e:
        messages.error(request, f'Error approving payment: {str(e)}')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': str(e)})
        return redirect('aps_finance:bank_payment_detail', payment_id=payment_id)

@login_required
@user_passes_test(is_finance)
@require_http_methods(["POST"])
def bank_payment_execute(request, payment_id):
    """Mark a payment as executed and update bank balance"""
    try:
        payment = get_object_or_404(BankPayment, payment_id=payment_id)

        # Check if payment can be executed
        if payment.status != 'approved':
            messages.error(request, 'Only approved payments can be executed')
            return JsonResponse({'status': 'error', 'message': 'Only approved payments can be executed'}) \
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest' else redirect('bank_payment_detail', payment_id=payment_id)

        # Update bank account balance
        bank_account = payment.bank_account
        if bank_account.current_balance < payment.amount:
            messages.error(request, 'Insufficient balance in bank account')
            return JsonResponse({'status': 'error', 'message': 'Insufficient balance in bank account'}) \
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest' else redirect('bank_payment_detail', payment_id=payment_id)

        bank_account.current_balance -= payment.amount
        bank_account.save()

        payment.status = 'executed'
        payment.save()

        messages.success(request, f'Payment executed successfully. Bank balance updated: {bank_account.current_balance}')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'new_balance': bank_account.current_balance})
        return redirect('aps_finance:bank_payment_detail', payment_id=payment_id)

    except Exception as e:
        messages.error(request, f'Error executing payment: {str(e)}')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': str(e)})
        return redirect('aps_finance:bank_payment_detail', payment_id=payment_id)

@login_required
@user_passes_test(is_finance)
@require_http_methods(["POST"])
def bank_payment_mark_failed(request, payment_id):
    """Mark a payment as failed"""
    try:
        payment = get_object_or_404(BankPayment, payment_id=payment_id)

        # Only approved payments can be marked as failed
        if payment.status not in ['approved', 'pending', 'verified']:
            messages.error(request, 'Only pending, verified or approved payments can be marked as failed')
            return JsonResponse({'status': 'error', 'message': 'Invalid status for marking as failed'}) \
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest' else redirect('bank_payment_detail', payment_id=payment_id)

        payment.status = 'failed'
        payment.save()

        messages.success(request, 'Payment marked as failed')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success'})
        return redirect('aps_finance:bank_payment_detail', payment_id=payment_id)

    except Exception as e:
        messages.error(request, f'Error marking payment as failed: {str(e)}')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'message': str(e)})
        return redirect('aps_finance:bank_payment_detail', payment_id=payment_id)

# Bank Account Views
@login_required
@user_passes_test(is_finance)
def bank_account_list(request):
    """List all bank accounts"""
    bank_accounts = BankAccount.objects.all().order_by('-is_active', 'bank_name')

    # Calculate account statistics
    stats = bank_accounts.aggregate(
        total_balance=Sum('current_balance', filter=Q(is_active=True)),
        active_accounts=Count('id', filter=Q(is_active=True)),
        total_accounts=Count('id')
    )

    context = {
        'bank_accounts': bank_accounts,
        'stats': stats
    }
    return render(request, 'components/finance/bank_account_list.html', context)

@login_required
@user_passes_test(is_finance)
@require_http_methods(["GET", "POST"])
def bank_account_create(request):
    """Create a new bank account"""
    if request.method == 'POST':
        try:
            bank_account = BankAccount.objects.create(
                name=request.POST.get('name'),
                account_number=request.POST.get('account_number'),
                bank_name=request.POST.get('bank_name'),
                branch=request.POST.get('branch'),
                ifsc_code=request.POST.get('ifsc_code'),
                current_balance=request.POST.get('current_balance', 0),
                is_active=bool(request.POST.get('is_active', True))
            )

            messages.success(request, 'Bank account created successfully')

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success', 'account_id': bank_account.id})
            return redirect('aps_finance:bank_account_list')

        except Exception as e:
            messages.error(request, f'Error creating bank account: {str(e)}')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': str(e)})
            return redirect('aps_finance:bank_account_create')

    return render(request, 'components/finance/bank_account_form.html')

@login_required
@user_passes_test(is_finance)
@require_http_methods(["GET", "POST"])
def bank_account_update(request, account_id):
    """Update a bank account"""
    bank_account = get_object_or_404(BankAccount, id=account_id)

    if request.method == 'POST':
        try:
            bank_account.name = request.POST.get('name', bank_account.name)
            bank_account.bank_name = request.POST.get('bank_name', bank_account.bank_name)
            bank_account.branch = request.POST.get('branch', bank_account.branch)
            bank_account.ifsc_code = request.POST.get('ifsc_code', bank_account.ifsc_code)
            bank_account.current_balance = request.POST.get('current_balance', bank_account.current_balance)
            bank_account.is_active = bool(request.POST.get('is_active', bank_account.is_active))
            bank_account.save()

            messages.success(request, 'Bank account updated successfully')

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success'})
            return redirect('aps_finance:bank_account_list')

        except Exception as e:
            messages.error(request, f'Error updating bank account: {str(e)}')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': str(e)})
            return redirect('aps_finance:bank_account_update', account_id=account_id)

    context = {
        'bank_account': bank_account
    }
    return render(request, 'components/finance/bank_account_form.html', context)

# Dashboard and Reports
@login_required
@user_passes_test(is_finance)
def bank_payment_dashboard(request):
    """Display dashboard with bank payment statistics"""
    # Get payment statistics
    payments = BankPayment.objects.all()

    # Overall stats
    overall_stats = payments.aggregate(
        total_amount=Sum('amount'),
        total_count=Count('id'),
        pending_amount=Sum('amount', filter=Q(status='pending')),
        verified_amount=Sum('amount', filter=Q(status='verified')),
        approved_amount=Sum('amount', filter=Q(status='approved')),
        executed_amount=Sum('amount', filter=Q(status='executed')),
        failed_amount=Sum('amount', filter=Q(status='failed'))
    )

    # Current month stats
    current_month = timezone.now().replace(day=1)
    monthly_stats = payments.filter(payment_date__gte=current_month).aggregate(
        total_amount=Sum('amount'),
        total_count=Count('id'),
        executed_amount=Sum('amount', filter=Q(status='executed'))
    )

    # Bank account stats
    bank_accounts = BankAccount.objects.filter(is_active=True)
    bank_stats = []

    for account in bank_accounts:
        account_payments = payments.filter(bank_account=account)
        bank_stats.append({
            'account': account,
            'total_payments': account_payments.count(),
            'total_amount': account_payments.aggregate(Sum('amount'))['amount__sum'] or 0,
            'executed_amount': account_payments.filter(status='executed').aggregate(Sum('amount'))['amount__sum'] or 0
        })

    context = {
        'overall_stats': overall_stats,
        'monthly_stats': monthly_stats,
        'bank_stats': bank_stats,
        'recent_payments': payments.order_by('-created_at')[:10]
    }
    return render(request, 'components/finance/bank_payment_dashboard.html', context)

@login_required
@user_passes_test(is_finance)
def bank_payment_report(request):
    """Generate bank payment reports"""
    report_type = request.GET.get('report_type', 'monthly')

    # Date filters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    filters = {}
    if date_from:
        filters['payment_date__gte'] = date_from
    if date_to:
        filters['payment_date__lte'] = date_to

    if request.GET.get('bank_account'):
        filters['bank_account'] = request.GET.get('bank_account')
    if request.GET.get('status'):
        filters['status'] = request.GET.get('status')

    payments = BankPayment.objects.filter(**filters)

    # Generate report data based on report_type
    report_data = None

    if report_type == 'monthly':
        # Group payments by month
        report_data = {}
        for payment in payments:
            month_key = payment.payment_date.strftime('%Y-%m')
            if month_key not in report_data:
                report_data[month_key] = {
                    'month': payment.payment_date.strftime('%B %Y'),
                    'total_amount': 0,
                    'executed_amount': 0,
                    'pending_amount': 0,
                    'count': 0
                }

            report_data[month_key]['total_amount'] += payment.amount
            report_data[month_key]['count'] += 1

            if payment.status == 'executed':
                report_data[month_key]['executed_amount'] += payment.amount
            elif payment.status in ['pending', 'verified', 'approved']:
                report_data[month_key]['pending_amount'] += payment.amount

        # Convert to list and sort
        report_data = sorted(report_data.values(), key=lambda x: x['month'], reverse=True)

    elif report_type == 'bank_account':
        # Group payments by bank account
        report_data = {}
        for payment in payments:
            account_key = str(payment.bank_account.id)
            if account_key not in report_data:
                report_data[account_key] = {
                    'account_name': str(payment.bank_account),
                    'total_amount': 0,
                    'executed_amount': 0,
                    'pending_amount': 0,
                    'count': 0
                }

            report_data[account_key]['total_amount'] += payment.amount
            report_data[account_key]['count'] += 1

            if payment.status == 'executed':
                report_data[account_key]['executed_amount'] += payment.amount
            elif payment.status in ['pending', 'verified', 'approved']:
                report_data[account_key]['pending_amount'] += payment.amount

        # Convert to list
        report_data = list(report_data.values())

    context = {
        'report_type': report_type,
        'report_data': report_data,
        'payments': payments,
        'filters': request.GET,
        'bank_accounts': BankAccount.objects.filter(is_active=True),
        'payment_statuses': dict(BankPayment.PAYMENT_STATUS)
    }
    return render(request, 'components/finance/bank_payment_report.html', context)



@login_required
@user_passes_test(is_finance)
def subscription_payment_entry(request):
    """Handle recurring subscription payments"""
    if request.method == 'POST':
        try:
            # Create subscription record
            subscription = Subscription.objects.create(
                subscription_id=f"SUB-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                vendor_name=request.POST.get('vendor_name'),
                subscription_type=request.POST.get('subscription_type'),
                amount=request.POST.get('amount'),
                frequency=request.POST.get('frequency'),
                due_date=request.POST.get('due_date'),
                auto_renew=request.POST.get('auto_renew') == 'on',
                bank_account_id=request.POST.get('bank_account'),
                status='active',
                created_by=request.user
            )

            messages.success(request, 'Subscription created successfully')
            return JsonResponse({'status': 'success', 'subscription_id': subscription.subscription_id})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    # Get subscriptions with filters
    filters = {}
    if request.GET.get('status'):
        filters['status'] = request.GET.get('status')
    if request.GET.get('subscription_type'):
        filters['subscription_type'] = request.GET.get('subscription_type')

    subscriptions = Subscription.objects.filter(**filters).order_by('due_date')
    bank_accounts = BankAccount.objects.filter(is_active=True)

    # Get upcoming payments
    today = timezone.now().date()
    upcoming_payments = subscriptions.filter(
        due_date__gte=today,
        due_date__lte=today + timedelta(days=5),
        status='active'
    )

    # Calculate subscription statistics
    stats = {
        'total_monthly': subscriptions.filter(frequency='monthly').aggregate(Sum('amount'))['amount__sum'] or 0,
        'total_yearly': subscriptions.filter(frequency='yearly').aggregate(Sum('amount'))['amount__sum'] or 0,
        'active_count': subscriptions.filter(status='active').count()
    }

    context = {
        'subscriptions': subscriptions,
        'bank_accounts': bank_accounts,
        'subscription_types': dict(Subscription.SUBSCRIPTION_TYPES),
        'frequencies': dict(Subscription.FREQUENCIES),
        'statuses': dict(Subscription.STATUS_CHOICES),
        'upcoming_payments': upcoming_payments,
        'stats': stats,
        'filters': request.GET
    }
    return render(request, 'components/finance/subscription_payment.html', context)


'''----------------------------------- Entertainment AREA -----------------------------------'''

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponseForbidden, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.db.models import Q, F, ExpressionWrapper, FloatField, Case, When, Value
from .models import TicTacToeGame, PlayerStats, GameIcon, GameSpectator, Notification
from django.views import View
from django.utils.decorators import method_decorator
# Custom decorator for checking user groups
def user_passes_test_groups(test_func):
    def decorator(view_func):
        @login_required
        def wrapped_view(request, *args, **kwargs):
            if test_func(request.user):
                return view_func(request, *args, **kwargs)
            return HttpResponseForbidden("You don't have permission to access this page.")
        return wrapped_view
    return decorator

def is_manager_or_hr_or_employee(user):
    """Check if user belongs to Manager, HR or Employee group"""
    return user.groups.filter(name__in=["Manager", "HR", "Employee"]).exists()

def is_hr(user):
    """Check if user belongs to HR group"""
    return user.groups.filter(name="HR").exists()

def is_admin(user):
    """Check if user belongs to Admin group"""
    return user.groups.filter(name="Admin").exists()

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import GameIcon, TicTacToeGame, PlayerStats


@login_required
@user_passes_test_groups(is_admin)
def entertainment_control(request):
    """Admin control panel for entertainment features"""

    # Get statistics
    stats = {
        'total_games': TicTacToeGame.objects.count(),
        'active_games': TicTacToeGame.objects.filter(status='active').count(),
        'total_players': PlayerStats.objects.count(),
        'total_icons': GameIcon.objects.count()
    }

    # Get recent games
    recent_games = TicTacToeGame.objects.select_related('creator', 'opponent')\
                                      .order_by('-created_at')[:10]

    # Get game icons
    icons = GameIcon.objects.select_related('created_by')\
                          .order_by('-created_at')

    # Get top players
    top_players = PlayerStats.objects.select_related('player')\
                                   .order_by('-games_won')[:10]

    context = {
        'stats': stats,
        'recent_games': recent_games,
        'icons': icons,
        'top_players': top_players,
        'title': 'Entertainment Control Panel'
    }

    return render(request, 'components/entertainment/control_panel.html', context)

@login_required
@user_passes_test_groups(is_manager_or_hr_or_employee)
def entertainment_dashboard(request):
    """
    Main entertainment dashboard view that shows available games
    and entertainment options.

    This function calls the games() function to get games content.
    """
    # Get games dashboard content by calling games()
    games_response = games(request)

    # Additional entertainment options


    context = {
        'title': 'Entertainment Dashboard',
        'games_dashboard': games_response,

    }

    return render(request, 'components/entertainment/dashboard.html', context)

@login_required
@user_passes_test(is_manager_or_hr_or_employee)
def games(request):
    """
    View that displays all available games with descriptions
    and links to play them.
    """
    # Get active games for this user
    active_games = TicTacToeGame.objects.filter(
        Q(creator=request.user) | Q(opponent=request.user),
        status__in=['active', 'pending']
    ).order_by('-updated_at')

    # Get user stats if available
    try:
        user_stats = PlayerStats.objects.get(user=request.user)
    except PlayerStats.DoesNotExist:
        user_stats = None

    # List of available games with descriptions
    available_games = [
        {
            'name': 'Tic-Tac-Toe',
            'description': 'Classic game of X and O. Challenge colleagues to a match!',
            'url': reverse('aps_entertainment:game_list'),
            'icon': 'fa-gamepad',
            'active_count': active_games.count()
        }
        # Add more games here as they become available
    ]

    context = {
        'title': 'Games Center',
        'available_games': available_games,
        'active_games': active_games[:5],  # Show 5 most recent active games
        'user_stats': user_stats,
    }

    return render(request, 'components/entertainment/games/games.html', context)

# Views for Tic-Tac-Toe Game
class TicTacToeGameView(View):
    """
    Class-based view to handle all game-related actions:
    - Display the game board
    - Process player moves
    - Handle game invitations
    - Handle spectator functionality
    """

    @method_decorator(login_required)
    @method_decorator(user_passes_test(is_manager_or_hr_or_employee))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, request, game_id=None):
        """Handle GET requests for either game listing or game detail"""
        if game_id:
            return self._game_detail(request, game_id)
        else:
            return self._game_list(request)

    def post(self, request, game_id=None):
        """Handle POST requests for game actions"""
        try:
            action = request.POST.get('action', '')

            # Map actions to their handler methods
            action_handlers = {
                'move': self._handle_move,
                'forfeit': self._handle_forfeit,
                'accept': self._handle_accept_invitation,
                'decline': self._handle_decline_invitation,
                'create': self._handle_create_game,
                'add_spectator': self._handle_add_spectator,
                'remove_spectator': self._handle_remove_spectator,
                'send_message': self._handle_game_message,
            }

            # If no action is specified but position is, default to move action
            if not action and request.POST.get('position') and game_id:
                return self._handle_move(request, game_id)

            # Execute the appropriate handler or return error
            if action in action_handlers:
                if game_id and action not in ['create']:
                    return action_handlers[action](request, game_id)
                elif action == 'create':
                    return action_handlers[action](request)
                else:
                    messages.error(request, "Invalid request")
                    return redirect('aps_entertainment:game_list')
            else:
                messages.error(request, "Invalid action specified")
                return redirect('aps_entertainment:game_list')

        except Exception as e:
            # Global exception handler
            import logging
            logging.error(f"Unexpected error in game view: {str(e)}")
            messages.error(request, "An unexpected error occurred. Please try again later.")
            return redirect('aps_entertainment:game_list')

    def _game_list(self, request):
        """Display list of active and pending games"""
        active_games = TicTacToeGame.objects.filter(
            Q(creator=request.user) | Q(opponent=request.user),
            status__in=['active', 'pending']
        ).order_by('-updated_at')

        completed_games = TicTacToeGame.objects.filter(
            Q(creator=request.user) | Q(opponent=request.user),
            status='completed'
        ).order_by('-updated_at')[:10]

        # Check for timed out games and update their status
        self._check_timeouts(request, active_games)

        # Get available game icons for creating new games
        game_icons = GameIcon.objects.filter(is_active=True)

        # Get spectating games
        spectating_games = TicTacToeGame.objects.filter(
            spectators__user=request.user,
            status='active',
            allow_spectators=True
        ).distinct()

        # ✅ Step 1: Get IDs of users with 'active' employment status
        active_user_ids = UserDetails.objects.filter(
            employment_status='active'
        ).values_list('user_id', flat=True)

        # ✅ Step 2: Get all users except current user and only those active
        user_list = User.objects.exclude(id=request.user.id).filter(
            id__in=active_user_ids
        )

        context = {
            'active_games': active_games,
            'completed_games': completed_games,
            'game_icons': game_icons,
            'user_list': user_list,
            'spectating_games': spectating_games,
            'unread_notifications': Notification.objects.filter(
                recipient=request.user,
                is_read=False
            ).count()
        }

        return render(request, 'components/entertainment/games/ttt/game_list.html', context)

    def _game_detail(self, request, game_id):
        """Display a specific game and handle game actions"""
        try:
            game = get_object_or_404(TicTacToeGame, id=game_id)

            # Check if user is authorized to view this game
            is_player = request.user == game.creator or request.user == game.opponent
            is_spectator = game.allow_spectators and game.spectators.filter(user=request.user).exists()

            if not (is_player or is_spectator):
                # If game allows spectators, add user as spectator
                if game.allow_spectators and game.status == 'active':
                    GameSpectator.objects.get_or_create(game=game, user=request.user)
                    is_spectator = True
                else:
                    return HttpResponseForbidden("You are not authorized to view this game.")

            # Mark notifications as read
            if is_player:
                Notification.objects.filter(
                    recipient=request.user,
                    game=game,
                    is_read=False
                ).update(is_read=True)

            # Check for timeout
            if game.status == 'active' and game.is_timeout():
                self._handle_timeout(game)

            # Prepare game data for template
            board_display = [game.board[i:i+3] for i in range(0, 9, 3)]

            context = {
                'game': game,
                'board_display': board_display,
                'is_player': is_player,
                'is_spectator': is_spectator,
                'is_creator': request.user == game.creator,
                'is_opponent': request.user == game.opponent,
                'is_your_turn': game.current_turn == request.user,
                'spectator_count': game.spectators.count(),
                'can_move': game.status == 'active' and game.current_turn == request.user,
            }

            return render(request, 'components/entertainment/games/ttt/game_detail.html', context)

        except Exception as e:
            messages.error(request, f"Error loading game: {str(e)}")
            return redirect('aps_entertainment:game_list')

    def _handle_move(self, request, game_id):
        """Handle a player making a move"""
        game = get_object_or_404(TicTacToeGame, id=game_id)

        # Ensure the user is a player in this game
        if request.user != game.creator and request.user != game.opponent:
            messages.error(request, "You are not a player in this game")
            return redirect('aps_entertainment:game_list')

        try:
            position = int(request.POST.get('position'))
            success, message = game.make_move(request.user, position)

            if success:
                # Send websocket notification to update other clients
                self._notify_game_update(game)

                # Update stats if game is completed
                if game.status == 'completed':
                    PlayerStats.update_stats(game)

                    # Create notification for the opponent
                    other_player = game.opponent if request.user == game.creator else game.creator
                    Notification.objects.create(
                        recipient=other_player,
                        message=f"Game finished! {game.winner.username} won." if game.winner else "Game ended in a draw.",
                        notification_type='game_ended',
                        game=game
                    )
            else:
                messages.error(request, message)

            return redirect('aps_entertainment:game_detail', game_id=game_id)

        except (ValueError, TypeError):
            messages.error(request, "Invalid position value")
            return redirect('aps_entertainment:game_detail', game_id=game_id)

    def _handle_forfeit(self, request, game_id):
        """Handle a player forfeiting the game"""
        game = get_object_or_404(TicTacToeGame, id=game_id)

        # Ensure user is a player in this game
        if request.user != game.creator and request.user != game.opponent:
            messages.error(request, "You are not a player in this game")
            return redirect('aps_entertainment:game_list')

        success, message = game.forfeit_game(request.user)

        if success:
            # Update stats
            PlayerStats.update_stats(game)

            # Create notification for the opponent
            other_player = game.opponent if request.user == game.creator else game.creator
            Notification.objects.create(
                recipient=other_player,
                message=f"{request.user.username} has forfeited the game. You win!",
                notification_type='game_forfeit',
                game=game
            )

            # Send websocket notification to update other clients
            self._notify_game_update(game)

            messages.success(request, "You have forfeited the game.")
        else:
            messages.error(request, message)

        return redirect('aps_entertainment:game_detail', game_id=game_id)

    def _handle_accept_invitation(self, request, game_id):
        """Handle accepting a game invitation"""
        game = get_object_or_404(TicTacToeGame, id=game_id, opponent=request.user, status='pending')

        success, message = game.accept_game()

        if success:
            # Create notification for the creator
            Notification.objects.create(
                recipient=game.creator,
                message=f"{request.user.username} has accepted your game invitation!",
                notification_type='game_accepted',
                game=game
            )

            # Send websocket notification to update other clients
            self._notify_game_update(game)

            messages.success(request, "Game accepted! It's your opponent's turn.")
        else:
            messages.error(request, message)

        return redirect('aps_entertainment:game_detail', game_id=game_id)

    def _handle_decline_invitation(self, request, game_id):
        """Handle declining a game invitation"""
        game = get_object_or_404(TicTacToeGame, id=game_id, opponent=request.user, status='pending')

        success, message = game.decline_game()

        if success:
            # Create notification for the creator
            Notification.objects.create(
                recipient=game.creator,
                message=f"{request.user.username} has declined your game invitation.",
                notification_type='game_declined',
                game=game
            )

            messages.success(request, "Game invitation declined.")
        else:
            messages.error(request, message)

        return redirect('aps_entertainment:game_list')

    def _handle_create_game(self, request):
        """Handle creating a new game"""
        opponent_id = request.POST.get('opponent_id')
        creator_icon_id = request.POST.get('creator_icon_id')
        opponent_icon_id = request.POST.get('opponent_icon_id')
        allow_spectators = request.POST.get('allow_spectators') == 'on'

        try:
            # Validate inputs
            from django.contrib.auth import get_user_model
            User = get_user_model()

            opponent = User.objects.get(id=opponent_id)
            creator_icon = GameIcon.objects.get(id=creator_icon_id, is_active=True)
            opponent_icon = GameIcon.objects.get(id=opponent_icon_id, is_active=True)

            # Make sure user is not playing against themselves
            if opponent == request.user:
                messages.error(request, "You cannot play against yourself!")
                return redirect('aps_entertainment:game_list')

            # Create the game
            game = TicTacToeGame.objects.create(
                creator=request.user,
                opponent=opponent,
                creator_icon=creator_icon,
                opponent_icon=opponent_icon,
                allow_spectators=allow_spectators
            )

            # Create notification for the opponent
            Notification.objects.create(
                recipient=opponent,
                message=f"{request.user.username} has invited you to play Tic-Tac-Toe!",
                notification_type='game_invite',
                game=game
            )

            messages.success(request, f"Game invitation sent to {opponent.username}!")
            return redirect('aps_entertainment:game_detail', game_id=game.id)

        except (User.DoesNotExist, GameIcon.DoesNotExist) as e:
            messages.error(request, f"Error creating game: {str(e)}")
            return redirect('aps_entertainment:game_list')

    def _handle_add_spectator(self, request, game_id):
        """Handle adding a spectator to the game"""
        game = get_object_or_404(TicTacToeGame, id=game_id, allow_spectators=True, status='active')

        # Get the user to add as spectator
        from django.contrib.auth import get_user_model
        User = get_user_model()

        try:
            spectator_username = request.POST.get('spectator_username')
            spectator = User.objects.get(username=spectator_username)

            # Don't add players as spectators
            if spectator == game.creator or spectator == game.opponent:
                messages.error(request, "Players cannot be added as spectators")
                return redirect('aps_entertainment:game_detail', game_id=game_id)

            # Create the spectator entry
            GameSpectator.objects.get_or_create(game=game, user=spectator)

            # Create notification for the spectator
            Notification.objects.create(
                recipient=spectator,
                message=f"{request.user.username} has invited you to watch a Tic-Tac-Toe game!",
                notification_type='spectator_invite',
                game=game
            )

            messages.success(request, f"{spectator.username} has been invited to watch the game")
            return redirect('aps_entertainment:game_detail', game_id=game_id)

        except User.DoesNotExist:
            messages.error(request, "User not found")
            return redirect('aps_entertainment:game_detail', game_id=game_id)

    def _handle_remove_spectator(self, request, game_id):
        """Handle removing a spectator from the game"""
        game = get_object_or_404(TicTacToeGame, id=game_id)

        # Only creator or opponent can remove spectators
        if request.user != game.creator and request.user != game.opponent:
            messages.error(request, "Only players can remove spectators")
            return redirect('aps_entertainment:game_detail', game_id=game_id)

        try:
            spectator_id = request.POST.get('spectator_id')
            spectator = GameSpectator.objects.get(id=spectator_id, game=game)
            spectator.delete()

            messages.success(request, "Spectator removed")
            return redirect('aps_entertainment:game_detail', game_id=game_id)

        except GameSpectator.DoesNotExist:
            messages.error(request, "Spectator not found")
            return redirect('aps_entertainment:game_detail', game_id=game_id)

    def _handle_game_message(self, request, game_id):
        """Handle sending a message in the game chat"""
        game = get_object_or_404(TicTacToeGame, id=game_id)

        # Check if user is authorized to send messages
        is_player = request.user == game.creator or request.user == game.opponent
        is_spectator = game.allow_spectators and game.spectators.filter(user=request.user).exists()

        if not (is_player or is_spectator):
            messages.error(request, "You are not authorized to send messages in this game")
            return redirect('aps_entertainment:game_list')

        message = request.POST.get('message', '').strip()
        if message:
            # Send message via websocket
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f'game_{game_id}',
                {
                    'type': 'game_message',
                    'message': message,
                    'username': request.user.username,
                    'timestamp': datetime.now().isoformat(),
                }
            )

        return redirect('aps_entertainment:game_detail', game_id=game_id)

    def _handle_timeout(self, game):
        """Handle game timeout"""
        game.status = 'timeout'
        # Determine the winner (the player who didn't time out)
        if game.current_turn:
            game.winner = game.opponent if game.current_turn == game.creator else game.creator
            game.save()
            PlayerStats.update_stats(game)

            # Create notification for the player who timed out
            Notification.objects.create(
                recipient=game.current_turn,
                message=f"Your game has timed out. {game.winner.username} wins.",
                notification_type='game_timeout',
                game=game
            )

            # Send websocket notification to update other clients
            self._notify_game_update(game)

    def _check_timeouts(self, request, active_games):
        """Check for timed out games and update their status"""
        for game in active_games:
            if game.status == 'active' and game.is_timeout():
                self._handle_timeout(game)

    def _notify_game_update(self, game):
        """Send websocket notification for game updates"""
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f'game_{game.id}',
            {
                'type': 'game_update',
                'game_id': game.id,
                'status': game.status,
                'board': game.board,
                'current_turn': game.current_turn.username if game.current_turn else None,
                'winner': game.winner.username if game.winner else None,
                'updated_at': game.updated_at.isoformat(),
                'spectator_count': game.spectators.count()
            }
        )


class NotificationView(View):
    """Class-based view to handle notifications"""

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, request):
        """Display user notifications"""
        notifications = Notification.objects.filter(recipient=request.user).order_by('-created_at')

        context = {
            'notifications': notifications,
            'unread_count': notifications.filter(is_read=False).count()
        }

        return render(request, 'components/entertainment/games//notification_list.html', context)

    def post(self, request):
        """Mark notifications as read"""
        action = request.POST.get('action')

        if action == 'mark_read':
            notification_ids = request.POST.getlist('notification_ids')
            if notification_ids:
                Notification.objects.filter(id__in=notification_ids, recipient=request.user).update(is_read=True)
                messages.success(request, "Notifications marked as read.")

        elif action == 'mark_all_read':
            Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
            messages.success(request, "All notifications marked as read.")

        return redirect('aps_entertainment:notifications')


class LeaderboardView(View):
    """Class-based view to display the game leaderboard"""

    @method_decorator(login_required)
    @method_decorator(require_GET)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, request):
        """Display the game leaderboard"""
        # Get top players by win percentage (minimum 5 games played)
        top_players = PlayerStats.objects.filter(games_played__gte=5).order_by('-win_percentage')[:20]

        # Get most active players
        most_active = PlayerStats.objects.order_by('-games_played')[:20]

        # Get user's rank if they have stats
        user_rank = None
        try:
            user_stats = PlayerStats.objects.get(user=request.user)
            if user_stats.games_played >= 5:
                # Count how many players have better win percentage
                user_rank = PlayerStats.objects.filter(
                    games_played__gte=5,
                    win_percentage__gt=user_stats.win_percentage
                ).count() + 1
        except PlayerStats.DoesNotExist:
            user_stats = None

        context = {
            'top_players': top_players,
            'most_active': most_active,
            'user_stats': user_stats,
            'user_rank': user_rank
        }

        return render(request, 'components/entertainment/games/tttleaderboard.html', context)


class GameIconView(View):
    """Class-based view to handle game icons"""

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, request):
        """Display game icons"""
        # Only show user's created icons and any active icons
        user_icons = GameIcon.objects.filter(Q(created_by=request.user) | Q(is_active=True)).distinct()

        context = {
            'icons': user_icons
        }

        return render(request, 'components/entertainment/games/game_icons.html', context)

    def post(self, request):
        """Create a new game icon"""
        name = request.POST.get('name')
        symbol = request.POST.get('symbol')

        if not name or not symbol:
            messages.error(request, "Both name and symbol are required.")
            return redirect('aps_entertainment:game_icons')

        # Validate symbol length
        if len(symbol) > 10:
            messages.error(request, "Symbol must be at most 10 characters.")
            return redirect('aps_entertainment:game_icons')

        # Create the icon
        GameIcon.objects.create(
            name=name,
            symbol=symbol,
            created_by=request.user,
            is_active=False  # Default to inactive until approved by admin
        )

        messages.success(request, "Icon created successfully! It will be available after review.")
        return redirect('aps_entertainment:game_icons')


class GameHistoryView(View):
    """Class-based view to display game history"""

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, request, user_id=None):
        """Display game history for a user"""
        from django.contrib.auth import get_user_model
        User = get_user_model()

        if user_id:
            user = get_object_or_404(User, id=user_id)
        else:
            user = request.user

        # Get all completed games for the user
        games = TicTacToeGame.objects.filter(
            Q(creator=user) | Q(opponent=user),
            status__in=['completed', 'timeout']
        ).order_by('-updated_at')

        # Get user stats
        try:
            stats = PlayerStats.objects.get(user=user)
        except PlayerStats.DoesNotExist:
            stats = None

        context = {
            'profile_user': user,
            'games': games,
            'stats': stats,
            'is_own_profile': user == request.user
        }

        return render(request, 'components/entertainment/games/tttgame_history.html', context)

# def handle_move(request):
#     """Handle a player making a move on the board"""
#     try:
#         game_id = request.POST.get('game_id')
#         position = request.POST.get('position')

#         if not game_id or not position:
#             return JsonResponse({
#                 'success': False,
#                 'message': "Missing required parameters"
#             }, status=400)

#         try:
#             position = int(position)
#         except ValueError:
#             return JsonResponse({
#                 'success': False,
#                 'message': "Position must be a number"
#             }, status=400)

#         with transaction.atomic():
#             game = get_object_or_404(TicTacToeGame, id=game_id)

#             # Check if game is active
#             if game.status != 'active':
#                 return JsonResponse({
#                     'success': False,
#                     'message': f"Game is not active. Current status: {game.status}"
#                 }, status=400)

#             # Check if user is a participant and it's their turn
#             if request.user != game.current_turn:
#                 return JsonResponse({
#                     'success': False,
#                     'message': "Not your turn or you're not a participant"
#                 }, status=403)

#             # Handle timeout check
#             if game.is_timeout():
#                 game.handle_timeout()
#                 return JsonResponse({
#                     'success': False,
#                     'message': "Game timed out due to inactivity",
#                     'status': game.status,
#                     'winner': game.winner.username if game.winner else None
#                 }, status=400)

#             # Try to make the move
#             success, message = game.make_move(request.user, position)

#             if success:
#                 response_data = {
#                     'success': True,
#                     'board': game.board,
#                     'status': game.status,
#                     'current_turn': game.current_turn.username if game.current_turn else None,
#                     'last_move_time': game.last_move_time.isoformat() if game.last_move_time else None,
#                     'move_history': game.move_history,
#                 }

#                 # If game is now complete, update player stats
#                 if game.status == 'completed':
#                     PlayerStats.update_stats(game)
#                     response_data['winner'] = game.winner.username if game.winner else None

#                     # Create notification for the opponent
#                     if game.winner:
#                         Notification.objects.create(
#                             recipient=game.opponent if game.winner == game.creator else game.creator,
#                             message=f"Game over! {game.winner.username} has won the game.",
#                             notification_type='game_completed',
#                             game=game
#                         )
#                     else:
#                         # It's a draw
#                         for player in [game.creator, game.opponent]:
#                             Notification.objects.create(
#                                 recipient=player,
#                                 message="Game over! The game ended in a draw.",
#                                 notification_type='game_completed',
#                                 game=game
#                             )

#                 # Notify spectators of the move
#                 for spectator in game.spectators.all():
#                     Notification.objects.create(
#                         recipient=spectator.user,
#                         message=f"{request.user.username} made a move in a game you're watching",
#                         notification_type='game_update',
#                         game=game
#                     )

#                 return JsonResponse(response_data)
#             else:
#                 return JsonResponse({'success': False, 'message': message}, status=400)

#     except TicTacToeGame.DoesNotExist:
#         return JsonResponse({'success': False, 'message': "Game not found"}, status=404)
#     except Exception as e:
#         return JsonResponse({'success': False, 'message': str(e)}, status=500)


# def handle_forfeit(request):
#     """Handle a player forfeiting the game"""
#     try:
#         game_id = request.POST.get('game_id')
#         if not game_id:
#             return JsonResponse({'success': False, 'message': "Game ID is required"}, status=400)

#         with transaction.atomic():
#             game = get_object_or_404(TicTacToeGame, id=game_id)

#             # Check if user is a participant
#             if request.user != game.creator and request.user != game.opponent:
#                 return JsonResponse({
#                     'success': False,
#                     'message': "You're not a participant in this game"
#                 }, status=403)

#             # Check if game can be forfeited
#             if game.status != 'active' and game.status != 'pending':
#                 return JsonResponse({
#                     'success': False,
#                     'message': f"Cannot forfeit a game with status: {game.status}"
#                 }, status=400)

#             success, message = game.forfeit_game(request.user)

#             if success:
#                 # Update player stats
#                 PlayerStats.update_stats(game)

#                 # Create notification for the winner
#                 Notification.objects.create(
#                     recipient=game.winner,
#                     message=f"{request.user.username} has forfeited the game. You win!",
#                     notification_type='game_forfeit',
#                     game=game
#                 )

#                 # Notify spectators
#                 for spectator in game.spectators.all():
#                     Notification.objects.create(
#                         recipient=spectator.user,
#                         message=f"{request.user.username} has forfeited the game you're watching",
#                         notification_type='game_update',
#                         game=game
#                     )

#                 return JsonResponse({
#                     'success': True,
#                     'status': game.status,
#                     'winner': game.winner.username,
#                     'message': message
#                 })
#             else:
#                 return JsonResponse({'success': False, 'message': message}, status=400)

#     except TicTacToeGame.DoesNotExist:
#         return JsonResponse({'success': False, 'message': "Game not found"}, status=404)
#     except Exception as e:
#         return JsonResponse({'success': False, 'message': str(e)}, status=500)


# def handle_accept_invitation(request):
#     """Handle accepting a game invitation"""
#     try:
#         game_id = request.POST.get('game_id')
#         if not game_id:
#             return JsonResponse({'success': False, 'message': "Game ID is required"}, status=400)

#         with transaction.atomic():
#             game = get_object_or_404(TicTacToeGame, id=game_id, opponent=request.user, status='pending')
#             success, message = game.accept_game()

#             if success:
#                 # Create notification for the creator
#                 Notification.objects.create(
#                     recipient=game.creator,
#                     message=f"{request.user.username} has accepted your game invitation",
#                     notification_type='game_invitation_accepted',
#                     game=game
#                 )

#                 return JsonResponse({
#                     'success': True,
#                     'message': "Game invitation accepted!",
#                     'game_id': str(game.id),
#                     'status': game.status
#                 })
#             else:
#                 return JsonResponse({'success': False, 'message': message}, status=400)

#     except TicTacToeGame.DoesNotExist:
#         return JsonResponse({
#             'success': False,
#             'message': "Game invitation not found or already processed"
#         }, status=404)
#     except Exception as e:
#         return JsonResponse({'success': False, 'message': str(e)}, status=500)


# def handle_decline_invitation(request):
#     """Handle declining a game invitation"""
#     try:
#         game_id = request.POST.get('game_id')
#         if not game_id:
#             return JsonResponse({'success': False, 'message': "Game ID is required"}, status=400)

#         with transaction.atomic():
#             game = get_object_or_404(TicTacToeGame, id=game_id, opponent=request.user, status='pending')
#             success, message = game.decline_game()

#             if success:
#                 # Create notification for the creator
#                 Notification.objects.create(
#                     recipient=game.creator,
#                     message=f"{request.user.username} has declined your game invitation",
#                     notification_type='game_invitation_declined',
#                     game=game
#                 )

#                 return JsonResponse({'success': True, 'message': "Game invitation declined."})
#             else:
#                 return JsonResponse({'success': False, 'message': message}, status=400)

#     except TicTacToeGame.DoesNotExist:
#         return JsonResponse({
#             'success': False,
#             'message': "Game invitation not found or already processed"
#         }, status=404)
#     except Exception as e:
#         return JsonResponse({'success': False, 'message': str(e)}, status=500)


# def handle_create_game(request):
#     """Handle creating a new game"""
#     try:
#         opponent_id = request.POST.get('opponent_id')
#         creator_icon_id = request.POST.get('creator_icon')
#         opponent_icon_id = request.POST.get('opponent_icon')
#         allow_spectators = request.POST.get('allow_spectators', 'off') == 'on'
#         time_limit = request.POST.get('time_limit', 86400)  # Default 24 hours in seconds

#         if not opponent_id:
#             return JsonResponse({'success': False, 'message': "Opponent is required"}, status=400)

#         # Try to parse time limit
#         try:
#             time_limit = int(time_limit)
#             if time_limit < 60 or time_limit > 604800:  # Between 1 minute and 7 days
#                 return JsonResponse({
#                     'success': False,
#                     'message': "Time limit must be between 1 minute and 7 days"
#                 }, status=400)
#         except ValueError:
#             time_limit = 86400  # Default to 24 hours if invalid

#         # Validate opponent exists
#         try:
#             with transaction.atomic():
#                 opponent = User.objects.get(id=opponent_id)

#                 # Don't allow inviting yourself
#                 if opponent == request.user:
#                     return JsonResponse({
#                         'success': False,
#                         'message': "You cannot play against yourself."
#                     }, status=400)

#                 # Check if there's already an active or pending game with this opponent
#                 existing_game = TicTacToeGame.objects.filter(
#                     (Q(creator=request.user) & Q(opponent=opponent)) |
#                     (Q(creator=opponent) & Q(opponent=request.user)),
#                     status__in=['active', 'pending']
#                 ).first()

#                 if existing_game:
#                     return JsonResponse({
#                         'success': False,
#                         'message': f"You already have an {existing_game.status} game with this opponent"
#                     }, status=400)

#                 # Use default icons if not specified
#                 try:
#                     default_icon = GameIcon.objects.filter(is_active=True).first()
#                     creator_icon = GameIcon.objects.get(id=creator_icon_id) if creator_icon_id else default_icon
#                     opponent_icon = GameIcon.objects.get(id=opponent_icon_id) if opponent_icon_id else default_icon

#                     # Make sure icons are different
#                     if creator_icon == opponent_icon:
#                         # Try to find a different icon
#                         alternate_icon = GameIcon.objects.filter(
#                             is_active=True
#                         ).exclude(id=creator_icon.id).first()

#                         if alternate_icon:
#                             opponent_icon = alternate_icon

#                 except (GameIcon.DoesNotExist, ValueError):
#                     # Use defaults if specified icons don't exist
#                     default_icons = GameIcon.objects.filter(is_active=True)[:2]
#                     if len(default_icons) >= 2:
#                         creator_icon = default_icons[0]
#                         opponent_icon = default_icons[1]
#                     else:
#                         # Create basic icons if none exist
#                         creator_icon = GameIcon.objects.create(
#                             name="X", symbol="X",
#                             css_class="text-blue-500",
#                             is_active=True
#                         )
#                         opponent_icon = GameIcon.objects.create(
#                             name="O", symbol="O",
#                             css_class="text-red-500",
#                             is_active=True
#                         )

#                 # Create the game
#                 game = TicTacToeGame.objects.create(
#                     creator=request.user,
#                     opponent=opponent,
#                     creator_icon=creator_icon,
#                     opponent_icon=opponent_icon,
#                     allow_spectators=allow_spectators,
#                     time_limit=time_limit
#                 )

#                 # Create notification for the opponent
#                 Notification.objects.create(
#                     recipient=opponent,
#                     message=f"{request.user.username} has invited you to play Tic-Tac-Toe",
#                     notification_type='game_invite',
#                     game=game
#                 )

#                 return JsonResponse({
#                     'success': True,
#                     'message': f"Game invitation sent to {opponent.username}",
#                     'game_id': str(game.id)
#                 })

#         except User.DoesNotExist:
#             return JsonResponse({'success': False, 'message': "Invalid opponent selected."}, status=400)
#         except Exception as e:
#             return JsonResponse({'success': False, 'message': str(e)}, status=500)

#     except Exception as e:
#         return JsonResponse({'success': False, 'message': str(e)}, status=500)


# def handle_replay_game(request):
#     """Handle creating a rematch from an existing game"""
#     try:
#         original_game_id = request.POST.get('original_game_id')
#         if not original_game_id:
#             return JsonResponse({'success': False, 'message': "Original game ID is required"}, status=400)

#         with transaction.atomic():
#             original_game = get_object_or_404(TicTacToeGame, id=original_game_id)

#             # Check if user is a participant
#             if request.user != original_game.creator and request.user != original_game.opponent:
#                 return JsonResponse({
#                     'success': False,
#                     'message': "You're not a participant in this game"
#                 }, status=403)

#             # Check if game is completed
#             if original_game.status != 'completed':
#                 return JsonResponse({
#                     'success': False,
#                     'message': "Can only create a rematch for completed games"
#                 }, status=400)

#             # Determine opponent (the other player)
#             opponent = original_game.opponent if request.user == original_game.creator else original_game.creator

#             # Create a new game but swap creator and opponent if the rematch requester was the opponent
#             if request.user == original_game.opponent:
#                 new_creator = original_game.opponent
#                 new_opponent = original_game.creator
#                 new_creator_icon = original_game.opponent_icon
#                 new_opponent_icon = original_game.creator_icon
#             else:
#                 new_creator = original_game.creator
#                 new_opponent = original_game.opponent
#                 new_creator_icon = original_game.creator_icon
#                 new_opponent_icon = original_game.opponent_icon

#             # Create the rematch game
#             new_game = TicTacToeGame.objects.create(
#                 creator=new_creator,
#                 opponent=new_opponent,
#                 creator_icon=new_creator_icon,
#                 opponent_icon=new_opponent_icon,
#                 allow_spectators=original_game.allow_spectators,
#                 time_limit=original_game.time_limit,
#                 is_rematch=True,
#                 original_game=original_game
#             )

#             # Create notification for the opponent
#             Notification.objects.create(
#                 recipient=new_opponent,
#                 message=f"{new_creator.username} has invited you to a rematch",
#                 notification_type='game_invite',
#                 game=new_game
#             )

#             return JsonResponse({
#                 'success': True,
#                 'message': f"Rematch invitation sent to {new_opponent.username}",
#                 'game_id': str(new_game.id)
#             })

#     except TicTacToeGame.DoesNotExist:
#         return JsonResponse({'success': False, 'message': "Original game not found"}, status=404)
#     except Exception as e:
#         return JsonResponse({'success': False, 'message': str(e)}, status=500)


# def handle_add_spectator(request):
#     """Handle adding a spectator to a game"""
#     try:
#         game_id = request.POST.get('game_id')
#         if not game_id:
#             return JsonResponse({'success': False, 'message': "Game ID is required"}, status=400)

#         with transaction.atomic():
#             game = get_object_or_404(TicTacToeGame, id=game_id)

#             # Check if spectating is allowed
#             if not game.allow_spectators:
#                 return JsonResponse({
#                     'success': False,
#                     'message': "This game does not allow spectators"
#                 }, status=403)

#             # Check if game is active
#             if game.status != 'active':
#                 return JsonResponse({
#                     'success': False,
#                     'message': f"Cannot spectate a game with status: {game.status}"
#                 }, status=400)

#             # Check if user is already a participant
#             if request.user == game.creator or request.user == game.opponent:
#                 return JsonResponse({
#                     'success': False,
#                     'message': "You are already a participant in this game"
#                 }, status=400)

#             # Add user as spectator
#             spectator, created = GameSpectator.objects.get_or_create(
#                 game=game,
#                 user=request.user
#             )

#             # Notify game participants of new spectator
#             if created:
#                 for player in [game.creator, game.opponent]:
#                     Notification.objects.create(
#                         recipient=player,
#                         message=f"{request.user.username} is now spectating your game",
#                         notification_type='game_spectator_added',
#                         game=game
#                     )

#             return JsonResponse({
#                 'success': True,
#                 'message': "You are now spectating this game"
#             })

#     except TicTacToeGame.DoesNotExist:
#         return JsonResponse({'success': False, 'message': "Game not found"}, status=404)
#     except Exception as e:
#         return JsonResponse({'success': False, 'message': str(e)}, status=500)


# def handle_remove_spectator(request):
#     """Handle removing a spectator from a game"""
#     try:
#         game_id = request.POST.get('game_id')
#         if not game_id:
#             return JsonResponse({'success': False, 'message': "Game ID is required"}, status=400)

#         with transaction.atomic():
#             game = get_object_or_404(TicTacToeGame, id=game_id)

#             # Remove spectator
#             deleted, _ = GameSpectator.objects.filter(
#                 game=game,
#                 user=request.user
#             ).delete()

#             if deleted:
#                 return JsonResponse({
#                     'success': True,
#                     'message': "You are no longer spectating this game"
#                 })
#             else:
#                 return JsonResponse({
#                     'success': False,
#                     'message': "You are not spectating this game"
#                 }, status=400)

#     except TicTacToeGame.DoesNotExist:
#         return JsonResponse({'success': False, 'message': "Game not found"}, status=404)
#     except Exception as e:
#         return JsonResponse({'success': False, 'message': str(e)}, status=500)


# def handle_game_message(request):
#     """Handle sending in-game messages"""
#     try:
#         game_id = request.POST.get('game_id')
#         message = request.POST.get('message')

#         if not game_id or not message:
#             return JsonResponse({
#                 'success': False,
#                 'message': "Game ID and message are required"
#             }, status=400)

#         with transaction.atomic():
#             game = get_object_or_404(TicTacToeGame, id=game_id)

#             # Check if user is a participant or spectator
#             is_participant = request.user == game.creator or request.user == game.opponent
#             is_spectator = game.spectators.filter(user=request.user).exists()

#             if not (is_participant or is_spectator):
#                 return JsonResponse({
#                     'success': False,
#                     'message': "You must be a participant or spectator to send messages"
#                 }, status=403)

#             # Add message to game chat
#             from .models import GameMessage
#             game_message = GameMessage.objects.create(
#                 game=game,
#                 sender=request.user,
#                 message=message
#             )

#             # Notify other participants
#             recipients = []
#             if is_participant:
#                 # Notify the other player
#                 other_player = game.opponent if request.user == game.creator else game.creator
#                 recipients.append(other_player)

#                 # And all spectators
#                 for spectator in game.spectators.all():
#                     recipients.append(spectator.user)
#             else:
#                 # Notify both players
#                 recipients.extend([game.creator, game.opponent])

#                 # And other spectators
#                 for spectator in game.spectators.exclude(user=request.user):
#                     recipients.append(spectator.user)

#             # Create notifications
#             for recipient in recipients:
#                 Notification.objects.create(
#                     recipient=recipient,
#                     message=f"New message from {request.user.username} in your game",
#                     notification_type='game_message',
#                     game=game
#                 )

#             return JsonResponse({
#                 'success': True,
#                 'message': "Message sent",
#                 'chat_message': {
#                     'id': str(game_message.id),
#                     'sender': game_message.sender.username,
#                     'message': game_message.message,
#                     'timestamp': game_message.created_at.isoformat()
#                 }
#             })

#     except TicTacToeGame.DoesNotExist:
#         return JsonResponse({'success': False, 'message': "Game not found"}, status=404)
#     except Exception as e:
#         return JsonResponse({'success': False, 'message': str(e)}, status=500)


# def handle_get_request(request):
#     """Handle GET requests to display game data"""
#     try:
#         context = {}

#         # Get active games where the user is a participant
#         context['active_games'] = TicTacToeGame.objects.filter(
#             (Q(creator=request.user) | Q(opponent=request.user)),
#             status__in=['active', 'pending']
#         ).order_by('-updated_at')

#         # Get completed games where the user is a participant
#         context['completed_games'] = TicTacToeGame.objects.filter(
#             (Q(creator=request.user) | Q(opponent=request.user)),
#             status='completed'
#         ).order_by('-updated_at')[:10]  # Limit to recent 10 games

#         # Games user can spectate (active games where user is not a participant)
#         context['spectatable_games'] = TicTacToeGame.objects.filter(
#             status='active',
#             allow_spectators=True
#         ).exclude(
#             Q(creator=request.user) | Q(opponent=request.user)
#         ).order_by('-updated_at')

#         # Get pending invitations
#         context['invitations'] = TicTacToeGame.objects.filter(
#             opponent=request.user,
#             status='pending'
#         ).order_by('-created_at')

#         # Get unread notifications
#         context['notifications'] = Notification.objects.filter(
#             recipient=request.user,
#             is_read=False
#         ).order_by('-created_at')[:5]

#         # Get player stats
#         user_stats, created = PlayerStats.objects.get_or_create(
#             user=request.user,
#             defaults={
#                 'games_played': 0,
#                 'games_won': 0,
#                 'games_lost': 0,
#                 'games_tied': 0
#             }
#         )
#         context['user_stats'] = user_stats

#         # Get top players for leaderboard
#         top_players = PlayerStats.objects.annotate(
#             win_rate=ExpressionWrapper(
#                 Case(
#                     When(games_played__gt=0,
#                         then=F('games_won') * 100.0 / F('games_played')),
#                     default=Value(0.0)
#                 ),
#                 output_field=FloatField()
#             )
#         ).order_by('-games_won', '-win_rate')[:10]

#         context['top_players'] = top_players

#         # Get available icons for the game
#         context['icons'] = GameIcon.objects.filter(is_active=True)

#         # Get list of users who can be invited (in Manager, HR or Employee groups)
#         allowed_groups = Group.objects.filter(name__in=["Manager", "HR", "Employee"])
#         context['potential_opponents'] = User.objects.filter(
#             groups__in=allowed_groups
#         ).exclude(id=request.user.id)

#         # Get available time limits for games
#         context['time_limits'] = [
#             {'seconds': 3600, 'display': '1 hour'},
#             {'seconds': 7200, 'display': '2 hours'},
#             {'seconds': 14400, 'display': '4 hours'},
#             {'seconds': 28800, 'display': '8 hours'},
#             {'seconds': 86400, 'display': '24 hours'},
#             {'seconds': 172800, 'display': '2 days'},
#             {'seconds': 604800, 'display': '7 days'},
#         ]

#         # Check for specific game_id parameter
#         game_id = request.GET.get('game_id')
#         if game_id:
#             return handle_game_detail(request, game_id, context)

#         # Render game list view by default
#         return render(request, 'components/entertainment/games/ttt/game_list.html', context)

#     except Exception as e:
#         import logging
#         logging.error(f"Error in handle_get_request: {str(e)}")
#         context = {'error_message': "An error occurred while loading game data. Please try again."}
#         return render(request, 'components/entertainment/games/ttt/error.html', context)


# def handle_game_detail(request, game_id, context):
#     """Handle displaying detailed game view"""
#     try:
#         game = get_object_or_404(TicTacToeGame, id=game_id)

#         # Determine if user is allowed to view this game
#         is_participant = request.user == game.creator or request.user == game.opponent
#         is_spectator = game.spectators.filter(user=request.user).exists()

#         if not (is_participant or (is_spectator and game.allow_spectators)):
#             # If not a participant or spectator, check if can become spectator
#             if game.status == 'active' and game.allow_spectators:
#                 # Add as spectator if allowed
#                 GameSpectator.objects.get_or_create(game=game, user=request.user)
#                 is_spectator = True
#             else:
#                 return HttpResponseForbidden("You don't have permission to view this game.")

#         # Check for timeout
#         if game.status == 'active' and game.is_timeout():
#             game.handle_timeout()

#         # Mark notifications related to this game as read
#         if is_participant or is_spectator:
#             Notification.objects.filter(
#                 recipient=request.user,
#                 game=game,
#                 is_read=False
#             ).update(is_read=True)

#         # Get spectators list
#         spectators = game.spectators.all()

#         # Get game messages
#         from .models import GameMessage
#         messages = GameMessage.objects.filter(game=game).order_by('created_at')

#         # Get move history
#         move_history = game.get_move_history()

#         # Get rematch history
#         if game.is_rematch:
#             rematch_chain = [game]
#             current_game = game
#             while current_game.original_game:
#                 rematch_chain.append(current_game.original_game)
#                 current_game = current_game.original_game
#             rematch_chain.reverse()  # Show oldest game first
#         else:
#             rematch_chain = []

#         # Check if there are rematches of this game
#         rematches = TicTacToeGame.objects.filter(original_game=game)

#         # Update context with game details
#         context.update({
#             'game': game,
#             'is_participant': is_participant,
#             'is_spectator': is_spectator,
#             'spectators': spectators,
#             'messages': messages,
#             'move_history': move_history,
#             'rematch_chain': rematch_chain,
#             'rematches': rematches,
#             'can_play': game.status == 'active' and request.user == game.current_turn,
#             'can_forfeit': game.status == 'active' and is_participant,
#             'can_request_rematch': game.status == 'completed' and is_participant,
#             'can_add_spectator': is_participant and game.allow_spectators,
#             'current_time': timezone.now(),
#             'time_remaining': game.get_time_remaining() if game.time_limit else None,
#         })

#         return render(request, 'components/entertainment/games/ttt/game_detail.html', context)

#     except Exception as e:
#         import logging
#         logging.error(f"Error in handle_game_detail: {str(e)}")
#         context['error_message'] = "An error occurred while loading the game. Please try again."
#         return render(request, 'components/entertainment/games/ttt/error.html', context)




# @login_required
# @user_passes_test_groups(is_manager_or_hr_or_employee)
# def tictactoe(request):
#     """
#     View for the Tic-Tac-Toe game.
#     - GET request: Returns the game board data
#     - POST request: Processes player moves and AI responses
#     This is called by the games() function to include the game data.
#     """
#     # Handle POST requests (making a move)
#     if request.method == 'POST':
#         game_id = request.POST.get('game_id')
#         position = request.POST.get('position')

#         if game_id and position:
#             game = get_object_or_404(TicTacToeGame, id=game_id)

#             # Check if user is a participant and it's their turn
#             if request.user != game.current_turn:
#                 return JsonResponse({'success': False, 'message': "Not your turn or you're not a participant"})

#             try:
#                 position = int(position)
#                 success, message = game.make_move(request.user, position)

#                 if success:
#                     # If game is now complete, update player stats
#                     if game.status == 'completed':
#                         PlayerStats.update_stats(game)

#                     return JsonResponse({
#                         'success': True,
#                         'board': game.board,
#                         'status': game.status,
#                         'winner': game.winner.username if game.winner else None,
#                         'current_turn': game.current_turn.username if game.current_turn else None
#                     })
#                 else:
#                     return JsonResponse({'success': False, 'message': message})
#             except ValueError:
#                 return JsonResponse({'success': False, 'message': "Invalid position"})

#         # Handle forfeit action
#         if request.POST.get('action') == 'forfeit':
#             game_id = request.POST.get('game_id')
#             game = get_object_or_404(TicTacToeGame, id=game_id)

#             # Check if user is a participant
#             if request.user != game.creator and request.user != game.opponent:
#                 return JsonResponse({'success': False, 'message': "You're not a participant in this game"})

#             success, message = game.forfeit_game(request.user)

#             if success:
#                 # Update player stats
#                 PlayerStats.update_stats(game)

#                 return JsonResponse({
#                     'success': True,
#                     'status': game.status,
#                     'winner': game.winner.username
#                 })
#             else:
#                 return JsonResponse({'success': False, 'message': message})

#         # Handle accept/decline invitation
#         if request.POST.get('action') == 'accept':
#             game_id = request.POST.get('game_id')
#             game = get_object_or_404(TicTacToeGame, id=game_id, opponent=request.user, status='pending')
#             success, message = game.accept_game()

#             if success:
#                 return JsonResponse({'success': True, 'message': "Game invitation accepted!"})
#             else:
#                 return JsonResponse({'success': False, 'message': message})

#         if request.POST.get('action') == 'decline':
#             game_id = request.POST.get('game_id')
#             game = get_object_or_404(TicTacToeGame, id=game_id, opponent=request.user, status='pending')
#             success, message = game.decline_game()

#             if success:
#                 return JsonResponse({'success': True, 'message': "Game invitation declined."})
#             else:
#                 return JsonResponse({'success': False, 'message': message})

#         # Handle create game
#         if request.POST.get('action') == 'create':
#             opponent_id = request.POST.get('opponent_id')
#             creator_icon_id = request.POST.get('creator_icon')
#             opponent_icon_id = request.POST.get('opponent_icon')
#             allow_spectators = request.POST.get('allow_spectators', False) == 'on'

#             # Validate opponent exists
#             try:
#                 opponent = User.objects.get(id=opponent_id)

#                 # Don't allow inviting yourself
#                 if opponent == request.user:
#                     return JsonResponse({'success': False, 'message': "You cannot play against yourself."})

#                 # Use default icons if not specified
#                 default_icon = GameIcon.objects.filter(is_active=True).first()
#                 creator_icon = GameIcon.objects.get(id=creator_icon_id) if creator_icon_id else default_icon
#                 opponent_icon = GameIcon.objects.get(id=opponent_icon_id) if opponent_icon_id else default_icon

#                 # Create the game
#                 game = TicTacToeGame.objects.create(
#                     creator=request.user,
#                     opponent=opponent,
#                     creator_icon=creator_icon,
#                     opponent_icon=opponent_icon,
#                     allow_spectators=allow_spectators
#                 )

#                 # Create notification for the opponent
#                 Notification.objects.create(
#                     recipient=opponent,
#                     message=f"{request.user.username} has invited you to play Tic-Tac-Toe",
#                     notification_type='game_invite',
#                     game=game
#                 )

#                 return JsonResponse({
#                     'success': True,
#                     'message': f"Game invitation sent to {opponent.username}",
#                     'game_id': str(game.id)
#                 })

#             except User.DoesNotExist:
#                 return JsonResponse({'success': False, 'message': "Invalid opponent selected."})

#     # Handle GET requests (display game data)
#     else:
#         context = {}

#         # Get active games where the user is a participant
#         context['active_games'] = TicTacToeGame.objects.filter(
#             (Q(creator=request.user) | Q(opponent=request.user)),
#             status__in=['active', 'pending']
#         ).order_by('-updated_at')

#         # Get completed games where the user is a participant
#         context['completed_games'] = TicTacToeGame.objects.filter(
#             (Q(creator=request.user) | Q(opponent=request.user)),
#             status='completed'
#         ).order_by('-updated_at')[:10]  # Limit to recent 10 games

#         # Games user can spectate (active games where user is not a participant)
#         context['spectatable_games'] = TicTacToeGame.objects.filter(
#             status='active',
#             allow_spectators=True
#         ).exclude(
#             Q(creator=request.user) | Q(opponent=request.user)
#         ).order_by('-updated_at')

#         # Get pending invitations
#         context['invitations'] = TicTacToeGame.objects.filter(
#             opponent=request.user,
#             status='pending'
#         ).order_by('-created_at')

#         # Get unread notifications
#         context['notifications'] = Notification.objects.filter(
#             recipient=request.user,
#             is_read=False
#         ).order_by('-created_at')[:5]

#         # Get top players for leaderboard - WITH FIX FOR win_rate

#         top_players = PlayerStats.objects.annotate(
#             win_rate=ExpressionWrapper(
#                 Case(
#                     When(games_played__gt=0,
#                         then=F('games_won') * 100.0 / F('games_played')),
#                     default=Value(0.0)
#                 ),
#                 output_field=FloatField()
#             )
#         ).order_by('-games_won', '-win_rate')[:10]

#         context['top_players'] = top_players

#         # Get available icons for the game
#         context['icons'] = GameIcon.objects.filter(is_active=True)

#         # Get list of users who can be invited (in Manager, HR or Employee groups)
#         allowed_groups = Group.objects.filter(name__in=["Manager", "HR", "Employee"])
#         context['potential_opponents'] = User.objects.filter(groups__in=allowed_groups).exclude(id=request.user.id)

#         # Check for specific game_id parameter
#         game_id = request.GET.get('game_id')
#         if game_id:
#             game = get_object_or_404(TicTacToeGame, id=game_id)

#             # Determine if user is allowed to view this game
#             is_participant = request.user == game.creator or request.user == game.opponent
#             is_spectator = game.spectators.filter(user=request.user).exists()

#             if not (is_participant or (is_spectator and game.allow_spectators)):
#                 # If not a participant or spectator, check if can become spectator
#                 if game.status == 'active' and game.allow_spectators:
#                     # Add as spectator if allowed
#                     GameSpectator.objects.get_or_create(game=game, user=request.user)
#                     is_spectator = True
#                 else:
#                     return HttpResponseForbidden("You don't have permission to view this game.")

#             # Check for timeout
#             if game.status == 'active' and game.is_timeout():
#                 game.status = 'timeout'
#                 # The player whose turn it is loses due to timeout
#                 game.winner = game.opponent if game.current_turn == game.creator else game.creator
#                 game.save()

#                 # Create timeout notification
#                 Notification.objects.create(
#                     recipient=game.current_turn,
#                     message=f"Your game has timed out due to inactivity. {game.winner.username} wins.",
#                     notification_type='game_timeout',
#                     game=game
#                 )

#                 # Update player stats
#                 PlayerStats.update_stats(game)

#             # Mark notifications related to this game as read
#             if is_participant:
#                 Notification.objects.filter(
#                     recipient=request.user,
#                     game=game,
#                     is_read=False
#                 ).update(is_read=True)

#             # Get spectators list
#             spectators = game.spectators.all()

#             # Add game specific context
#             context['game'] = game
#             context['is_participant'] = is_participant
#             context['is_spectator'] = is_spectator
#             context['is_creator'] = request.user == game.creator
#             context['is_opponent'] = request.user == game.opponent
#             context['is_my_turn'] = game.current_turn == request.user if game.current_turn else False
#             context['spectators'] = spectators
#             context['creator_symbol'] = game.creator_icon.symbol if game.creator_icon else 'X'
#             context['opponent_symbol'] = game.opponent_icon.symbol if game.opponent_icon else 'O'

#             # Render detailed game view
#             return render(request, 'components/entertainment/games/ttt/game_detail.html', context)

#         # Render game list view by default
#         return render(request, 'components/entertainment/games/ttt/game_list.html', context)

# @login_required
# @user_passes_test_groups(is_manager_or_hr_or_employee)
# def game_list(request):
#     """Display list of active and past games for the user"""
#     # Get active games where the user is a participant
#     active_games = TicTacToeGame.objects.filter(
#         (Q(creator=request.user) | Q(opponent=request.user)),
#         status__in=['active', 'pending']
#     ).order_by('-updated_at')

#     # Get completed games where the user is a participant
#     completed_games = TicTacToeGame.objects.filter(
#         (Q(creator=request.user) | Q(opponent=request.user)),
#         status='completed'
#     ).order_by('-updated_at')[:10]  # Limit to recent 10 games

#     # Games user can spectate (active games where user is not a participant)
#     spectatable_games = TicTacToeGame.objects.filter(
#         status='active',
#         allow_spectators=True
#     ).exclude(
#         Q(creator=request.user) | Q(opponent=request.user)
#     ).order_by('-updated_at')

#     # Get pending invitations
#     invitations = TicTacToeGame.objects.filter(
#         opponent=request.user,
#         status='pending'
#     ).order_by('-created_at')

#     # Get unread notifications
#     notifications = Notification.objects.filter(
#         recipient=request.user,
#         is_read=False
#     ).order_by('-created_at')[:5]

#     # Get top players for leaderboard
#     top_players = PlayerStats.objects.all().order_by('-games_won', '-win_percentage')[:10]

#     # Get available icons for the game
#     icons = GameIcon.objects.filter(is_active=True)

#     return render(request, 'games/game_list.html', {
#         'active_games': active_games,
#         'completed_games': completed_games,
#         'spectatable_games': spectatable_games,
#         'invitations': invitations,
#         'notifications': notifications,
#         'top_players': top_players,
#         'icons': icons
#     })

# @login_required
# @user_passes_test_groups(is_manager_or_hr_or_employee)
# def create_game(request):
#     """Create a new game invitation"""
#     if request.method == 'POST':
#         opponent_id = request.POST.get('opponent_id')
#         creator_icon_id = request.POST.get('creator_icon', None)
#         opponent_icon_id = request.POST.get('opponent_icon', None)
#         allow_spectators = request.POST.get('allow_spectators', False) == 'on'

#         # Validate opponent exists
#         try:
#             from django.contrib.auth.models import User
#             opponent = User.objects.get(id=opponent_id)

#             # Don't allow inviting yourself
#             if opponent == request.user:
#                 messages.error(request, "You cannot play against yourself.")
#                 return redirect('game_list')

#             # Use default icons if not specified
#             default_icon = GameIcon.objects.filter(is_active=True).first()
#             creator_icon = GameIcon.objects.get(id=creator_icon_id) if creator_icon_id else default_icon
#             opponent_icon = GameIcon.objects.get(id=opponent_icon_id) if opponent_icon_id else default_icon

#             # Create the game
#             game = TicTacToeGame.objects.create(
#                 creator=request.user,
#                 opponent=opponent,
#                 creator_icon=creator_icon,
#                 opponent_icon=opponent_icon,
#                 allow_spectators=allow_spectators
#             )

#             # Create notification for the opponent
#             Notification.objects.create(
#                 recipient=opponent,
#                 message=f"{request.user.username} has invited you to play Tic-Tac-Toe",
#                 notification_type='game_invite',
#                 game=game
#             )

#             messages.success(request, f"Game invitation sent to {opponent.username}")
#             return redirect('game_detail', game_id=game.id)

#         except User.DoesNotExist:
#             messages.error(request, "Invalid opponent selected.")
#             return redirect('game_list')

#     # GET request - show form to create game
#     # Get list of users who can be invited (in Manager, HR or Employee groups)
#     from django.contrib.auth.models import Group
#     allowed_groups = Group.objects.filter(name__in=["Manager", "HR", "Employee"])
#     potential_opponents = User.objects.filter(groups__in=allowed_groups).exclude(id=request.user.id)

#     # Get available icons
#     icons = GameIcon.objects.filter(is_active=True)

#     return render(request, 'games/create_game.html', {
#         'potential_opponents': potential_opponents,
#         'icons': icons
#     })



# @login_required
# @user_passes_test_groups(is_manager_or_hr_or_employee)
# def game_detail(request, game_id):
#     """Display game details and board for playing"""
#     game = get_object_or_404(TicTacToeGame, id=game_id)

#     # Determine if user is allowed to view this game
#     is_participant = request.user == game.creator or request.user == game.opponent
#     is_spectator = game.spectators.filter(user=request.user).exists()

#     if not (is_participant or (is_spectator and game.allow_spectators)):
#         # If not a participant or spectator, check if can become spectator
#         if game.status == 'active' and game.allow_spectators:
#             # Add as spectator if allowed
#             GameSpectator.objects.get_or_create(game=game, user=request.user)
#             is_spectator = True
#         else:
#             return HttpResponseForbidden("You don't have permission to view this game.")

#     # Check for timeout
#     if game.status == 'active' and game.is_timeout():
#         game.status = 'timeout'
#         # The player whose turn it is loses due to timeout
#         game.winner = game.opponent if game.current_turn == game.creator else game.creator
#         game.save()

#         # Create timeout notification
#         Notification.objects.create(
#             recipient=game.current_turn,
#             message=f"Your game has timed out due to inactivity. {game.winner.username} wins.",
#             notification_type='game_timeout',
#             game=game
#         )

#         # Update player stats
#         PlayerStats.update_stats(game)

#     # Mark notifications related to this game as read
#     if is_participant:
#         Notification.objects.filter(
#             recipient=request.user,
#             game=game,
#             is_read=False
#         ).update(is_read=True)

#     # Get spectators list
#     spectators = game.spectators.all()

#     return render(request, 'games/game_detail.html', {
#         'game': game,
#         'is_participant': is_participant,
#         'is_spectator': is_spectator,
#         'is_creator': request.user == game.creator,
#         'is_opponent': request.user == game.opponent,
#         'is_my_turn': game.current_turn == request.user if game.current_turn else False,
#         'spectators': spectators,
#         'creator_symbol': game.creator_icon.symbol if game.creator_icon else 'X',
#         'opponent_symbol': game.opponent_icon.symbol if game.opponent_icon else 'O',
#     })

# @login_required
# @require_POST
# def accept_game(request, game_id):
#     """Accept a game invitation"""
#     game = get_object_or_404(TicTacToeGame, id=game_id, opponent=request.user, status='pending')
#     success, message = game.accept_game()

#     if success:
#         messages.success(request, "Game invitation accepted!")
#         return redirect('game_detail', game_id=game.id)
#     else:
#         messages.error(request, message)
#         return redirect('game_list')

# @login_required
# @require_POST
# def decline_game(request, game_id):
#     """Decline a game invitation"""
#     game = get_object_or_404(TicTacToeGame, id=game_id, opponent=request.user, status='pending')
#     success, message = game.decline_game()

#     if success:
#         messages.success(request, "Game invitation declined.")
#         return redirect('game_list')
#     else:
#         messages.error(request, message)
#         return redirect('game_list')

# @login_required
# @require_POST
# def make_move(request, game_id):
#     """Make a move in the game"""
#     game = get_object_or_404(TicTacToeGame, id=game_id)

#     # Check if user is a participant and it's their turn
#     if request.user != game.current_turn:
#         return JsonResponse({'success': False, 'message': "Not your turn or you're not a participant"})

#     try:
#         position = int(request.POST.get('position'))
#         success, message = game.make_move(request.user, position)

#         if success:
#             # If game is now complete, update player stats
#             if game.status == 'completed':
#                 PlayerStats.update_stats(game)

#             return JsonResponse({
#                 'success': True,
#                 'board': game.board,
#                 'status': game.status,
#                 'winner': game.winner.username if game.winner else None,
#                 'current_turn': game.current_turn.username if game.current_turn else None
#             })
#         else:
#             return JsonResponse({'success': False, 'message': message})
#     except ValueError:
#         return JsonResponse({'success': False, 'message': "Invalid position"})

# @login_required
# @require_POST
# def forfeit_game(request, game_id):
#     """Forfeit an active game"""
#     game = get_object_or_404(TicTacToeGame, id=game_id)

#     # Check if user is a participant
#     if request.user != game.creator and request.user != game.opponent:
#         return JsonResponse({'success': False, 'message': "You're not a participant in this game"})

#     success, message = game.forfeit_game(request.user)

#     if success:
#         # Update player stats
#         PlayerStats.update_stats(game)

#         return JsonResponse({
#             'success': True,
#             'status': game.status,
#             'winner': game.winner.username
#         })
#     else:
#         return JsonResponse({'success': False, 'message': message})

# @login_required
# @require_GET
# def check_game_status(request, game_id):
#     """Check for updates in game status (for AJAX polling)"""
#     game = get_object_or_404(TicTacToeGame, id=game_id)

#     # Check if game timed out
#     if game.status == 'active' and game.is_timeout():
#         game.status = 'timeout'
#         game.winner = game.opponent if game.current_turn == game.creator else game.creator
#         game.save()

#         # Create timeout notification
#         Notification.objects.create(
#             recipient=game.current_turn,
#             message=f"Your game has timed out due to inactivity. {game.winner.username} wins.",
#             notification_type='game_timeout',
#             game=game
#         )

#         # Update player stats
#         PlayerStats.update_stats(game)

#     return JsonResponse({
#         'id': str(game.id),
#         'board': game.board,
#         'status': game.status,
#         'current_turn': game.current_turn.username if game.current_turn else None,
#         'winner': game.winner.username if game.winner else None,
#         'last_updated': game.updated_at.isoformat()
#     })

# @login_required
# @require_GET
# def get_notifications(request):
#     """Get unread notifications for the current user"""
#     notifications = Notification.objects.filter(
#         recipient=request.user,
#         is_read=False
#     ).order_by('-created_at')[:5]

#     return JsonResponse({
#         'notifications': [
#             {
#                 'id': notification.id,
#                 'message': notification.message,
#                 'type': notification.notification_type,
#                 'created_at': notification.created_at.isoformat(),
#                 'game_id': str(notification.game.id) if notification.game else None
#             }
#             for notification in notifications
#         ]
#     })

# @login_required
# @require_POST
# def mark_notification_read(request, notification_id):
#     """Mark a notification as read"""
#     notification = get_object_or_404(Notification, id=notification_id, recipient=request.user)
#     notification.is_read = True
#     notification.save()

#     return JsonResponse({'success': True})

# @login_required
# @user_passes_test_groups(is_hr)
# def manage_icons(request):
#     """View for HR to manage game icons"""
#     if request.method == 'POST':
#         action = request.POST.get('action')

#         if action == 'create':
#             name = request.POST.get('name')
#             symbol = request.POST.get('symbol')

#             if name and symbol:
#                 GameIcon.objects.create(
#                     name=name,
#                     symbol=symbol,
#                     created_by=request.user,
#                     is_active=True
#                 )
#                 messages.success(request, f"Created new game icon: {name} ({symbol})")
#             else:
#                 messages.error(request, "Name and symbol are required")

#         elif action == 'toggle':
#             icon_id = request.POST.get('icon_id')
#             icon = get_object_or_404(GameIcon, id=icon_id)
#             icon.is_active = not icon.is_active
#             icon.save()
#             status = "activated" if icon.is_active else "deactivated"
#             messages.success(request, f"Icon {icon.name} has been {status}")

#         elif action == 'delete':
#             icon_id = request.POST.get('icon_id')
#             icon = get_object_or_404(GameIcon, id=icon_id)
#             name = icon.name
#             icon.delete()
#             messages.success(request, f"Icon {name} has been deleted")

#     # Get all icons
#     icons = GameIcon.objects.all().order_by('-created_at')

#     return render(request, 'games/manage_icons.html', {
#         'icons': icons
#     })

# @login_required
# @user_passes_test_groups(is_manager_or_hr_or_employee)
# def leaderboard(request):
#     """View the leaderboard of top players"""
#     top_players = PlayerStats.objects.all().order_by(
#         '-games_won', '-win_percentage', '-games_played'
#     )[:20]

#     return render(request, 'games/leaderboard.html', {
#         'top_players': top_players
#     })
